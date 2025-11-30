from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView
from django.urls import reverse_lazy
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Count,  F, Avg, Case, When, Value, ExpressionWrapper,FloatField, Max
from django.views.decorators.csrf import csrf_exempt
from django.db.models.functions import Coalesce
from django.core.paginator import Paginator
from django.views.generic import TemplateView
from django.core.cache import cache
from django.utils.dateparse import parse_date
from decimal import Decimal
import openpyxl
from django.urls import reverse_lazy, reverse
from urllib.parse import urlencode
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from .forms import StockForm
from django.db import models
from django.views.decorators.http import require_http_methods
from rest_framework.generics import RetrieveAPIView
import logging
from django.shortcuts import render, get_object_or_404, redirect
from django.views.decorators.http import require_GET
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
from django.views import View
from django.db import transaction
import pandas as pd
import xlsxwriter
from io import BytesIO
import csv
from django.core.exceptions import ValidationError
from reportlab.lib.pagesizes import  A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from rest_framework import generics, status, filters
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, DjangoModelPermissions
from rest_framework.pagination import PageNumberPagination
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q, Sum, Count, F
from django.utils import timezone
from datetime import datetime, timedelta
from .serializers import (
    CategorySerializer, CategoryDetailSerializer, SupplierSerializer,
    ProductSerializer, StockSerializer, StockMovementSerializer,
    ImportSessionSerializer, InventoryReportSerializer,
    StockMovementReportSerializer, LowStockReportSerializer,
    ValuationReportSerializer
)
from stores.models import Store
from company.models import EFRISCommodityCategory
from company.mixins import EFRISConditionalMixin
from .forms import (
    CategoryForm, SupplierForm, ProductForm,  StockMovementForm,
     ProductFilterForm, StockAdjustmentForm, BulkActionForm
)

from .models import Category, Supplier, Product, Stock, StockMovement, ImportSession, ImportLog, ImportResult

logger = logging.getLogger(__name__)
CharField = models.CharField

def get_current_schema():
    """Get current tenant schema name"""
    try:
        from django.db import connection
        return getattr(connection, 'schema_name', 'public')
    except Exception:
        return 'public'

def _get_company_from_context():
    """Get company/tenant from current context"""
    try:
        from django.db import connection
        if hasattr(connection, 'tenant') and connection.tenant:
            return connection.tenant
        return None
    except Exception:
        return None

class StandardResultsSetPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 100

class QuickStockAdjustmentRedirectView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Redirect to stock adjustment form with pre-filled data from stock ID"""
    permission_required = 'inventory.add_stockmovement'

    def get(self, request, stock_id):
        try:
            stock = Stock.objects.select_related('product', 'store').get(id=stock_id)

            # Calculate recommended quantity (150% of reorder level minus current stock)
            recommended_qty = max(0, (stock.low_stock_threshold * Decimal('1.5')) - stock.quantity)

            # Build URL with parameters
            adjustment_url = reverse('inventory:stock_adjustment')
            params = {
                'product': stock.product.id,
                'store': stock.store.id,
                'quantity': recommended_qty.quantize(Decimal('0.01')),
                'movement_type': 'PURCHASE',
                'notes': f'Quick adjustment for low stock item. Current: {stock.quantity}, Reorder level: {stock.low_stock_threshold}'
            }

            url = f"{adjustment_url}?{urlencode(params)}"
            return redirect(url)

        except Stock.DoesNotExist:
            messages.error(request, 'Stock item not found.')
            return redirect('inventory:low_stock_report')

@login_required
@require_http_methods(["POST"])
@permission_required('inventory.add_category', raise_exception=True)
def category_create_ajax(request):
    """AJAX endpoint to create category"""
    try:
        name = request.POST.get('name', '').strip()
        code = request.POST.get('code', '').strip()
        description = request.POST.get('description', '').strip()

        if not name:
            return JsonResponse({'success': False, 'message': 'Category name is required'})

        category = Category.objects.create(
            name=name,
            code=code,
            description=description,
            is_active=True
        )

        return JsonResponse({
            'success': True,
            'category': {
                'id': category.id,
                'name': category.name,
                'code': category.code
            }
        })

    except Exception as e:
        logger.error(f"Error creating category: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@require_http_methods(["POST"])
@permission_required('inventory.add_supplier', raise_exception=True)
def supplier_create_ajax(request):
    """AJAX endpoint to create supplier"""
    try:
        name = request.POST.get('name', '').strip()
        tin = request.POST.get('tin', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()
        address = request.POST.get('address', '').strip()

        if not name or not phone:
            return JsonResponse({'success': False, 'message': 'Supplier name and phone are required'})

        supplier = Supplier.objects.create(
            name=name,
            tin=tin,
            phone=phone,
            email=email,
            address=address,
            is_active=True
        )

        return JsonResponse({
            'success': True,
            'supplier': {
                'id': supplier.id,
                'name': supplier.name,
                'phone': supplier.phone
            }
        })

    except Exception as e:
        logger.error(f"Error creating supplier: {str(e)}")
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@require_GET
def efris_category_search(request):
    search_term = request.GET.get('q', '').strip()
    page = int(request.GET.get('page', 1))
    page_size = 30

    # Debug: Log the request
    logger.debug(f"EFRIS search called - User: {request.user}, Search: '{search_term}', Page: {page}")

    # Get company from request - CHECK MULTIPLE WAYS
    company = None

    # Method 1: User has company attribute
    if hasattr(request.user, 'company'):
        company = request.user.company
        logger.debug(f"Company from user.company: {company}")


    elif hasattr(request.user, 'employee'):
        company = request.user.employee.company
        logger.debug(f"Company from user.employee.company: {company}")

    elif hasattr(request.user, 'userprofile'):
        company = request.user.userprofile.company
        logger.debug(f"Company from user.userprofile.company: {company}")

    elif 'company_id' in request.session:
        from company.models import Company
        company = Company.objects.filter(id=request.session['company_id']).first()
        logger.debug(f"Company from session: {company}")

    if not company:
        logger.warning(f"No company found for user {request.user}")
        return JsonResponse({
            'results': [],
            'pagination': {'more': False},
            'error': 'No company associated with user'
        })

    # Build query
    queryset = EFRISCommodityCategory.objects.all()

    # Debug: Check total categories for this company
    total_categories = queryset.count()
    logger.debug(f"Total EFRIS categories for company {company}: {total_categories}")

    if search_term:
        # Search by code or name
        queryset = queryset.filter(
            Q(commodity_category_code__icontains=search_term) |
            Q(commodity_category_name__icontains=search_term)
        )
        logger.debug(f"After search filter: {queryset.count()} results")
    else:
        logger.debug("No search term provided, returning first page")

    queryset = queryset.order_by('commodity_category_code')

    total = queryset.count()
    offset = (page - 1) * page_size
    categories = queryset[offset:offset + page_size]

    logger.debug(f"Returning {len(categories)} categories (offset: {offset})")
    if categories:
        logger.debug(
            f"First category: {categories[0].commodity_category_code} - {categories[0].commodity_category_name}")

    # Format for Select2
    results = [
        {
            'id': cat.id,
            'text': f"{cat.commodity_category_code} - {cat.commodity_category_name}",
            'code': cat.commodity_category_code,
            'name': cat.commodity_category_name,
            'is_exempt': cat.is_exempt,
            'is_leaf_node': cat.is_leaf_node,
            'is_zero_rate': cat.is_zero_rate
        }
        for cat in categories
    ]

    response_data = {
        'results': results,
        'pagination': {
            'more': (offset + page_size) < total
        }
    }

    # Debug: Log response
    logger.debug(f"Response: {len(results)} results, more={response_data['pagination']['more']}")

    return JsonResponse(response_data)


@login_required
@require_GET
def efris_category_detail(request, category_id):
    """Get details of a specific EFRIS category."""
    try:
        company = getattr(request.user, 'company', None)

        category = EFRISCommodityCategory.objects.get(
            id=category_id,
            company=company
        )

        return JsonResponse({
            'id': category.id,
            'code': category.commodity_category_code,
            'name': category.commodity_category_name,
            'is_exempt': category.is_exempt,
            'is_leaf_node': category.is_leaf_node,
            'is_zero_rate': category.is_zero_rate,
            'text': f"{category.commodity_category_code} - {category.commodity_category_name}"
        })
    except EFRISCommodityCategory.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)


class CategoryListCreateView(generics.ListCreateAPIView):
    queryset = Category.objects.all()
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'code', 'description']
    filterset_fields = ['is_active', 'efris_auto_sync', 'efris_is_uploaded']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['name']


class CategoryRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Category.objects.all()
    serializer_class = CategoryDetailSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]


class SupplierListCreateView(generics.ListCreateAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'contact_person', 'tin', 'phone', 'email']
    filterset_fields = ['is_active', 'country']
    ordering_fields = ['name', 'created_at', 'updated_at']
    ordering = ['name']


class SupplierRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]


class ProductListCreateView(generics.ListCreateAPIView):
    queryset = Product.objects.select_related('category', 'supplier').all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['name', 'sku', 'barcode', 'description']
    filterset_fields = [
        'category', 'supplier', 'tax_rate', 'is_active',
        'efris_is_uploaded', 'efris_auto_sync_enabled'
    ]
    ordering_fields = ['name', 'selling_price', 'cost_price', 'created_at', 'updated_at']
    ordering = ['name']


class ProductRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Product.objects.select_related('category', 'supplier').all()
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]


class StockListCreateView(generics.ListCreateAPIView):
    queryset = Stock.objects.select_related('product', 'store').all()
    serializer_class = StockSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['product__name', 'product__sku', 'store__name']
    filterset_fields = ['product', 'store']
    ordering_fields = ['product__name', 'store__name', 'quantity', 'last_updated']
    ordering = ['product__name']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Filter by stock status
        status_filter = self.request.query_params.get('status')
        if status_filter == 'out_of_stock':
            queryset = queryset.filter(quantity=0)
        elif status_filter == 'low_stock':
            queryset = queryset.filter(quantity__gt=0, quantity__lte=F('low_stock_threshold'))
        elif status_filter == 'good_stock':
            queryset = queryset.filter(quantity__gt=F('low_stock_threshold'))

        return queryset


class StockRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Stock.objects.select_related('product', 'store').all()
    serializer_class = StockSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]


class StockMovementListCreateView(generics.ListCreateAPIView):
    queryset = StockMovement.objects.select_related('product', 'store', 'created_by').all()
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['product__name', 'product__sku', 'store__name', 'reference']
    filterset_fields = ['product', 'store', 'movement_type', 'created_by']
    ordering_fields = ['created_at', 'product__name', 'store__name', 'quantity']
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = super().get_queryset()

        # Date filtering
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')

        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__gte=date_from_parsed)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__lte=date_to_parsed)
            except ValueError:
                pass

        return queryset


class StockMovementRetrieveUpdateDestroyView(generics.RetrieveUpdateDestroyAPIView):
    queryset = StockMovement.objects.select_related('product', 'store', 'created_by').all()
    serializer_class = StockMovementSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]


class ImportSessionListView(generics.ListAPIView):
    serializer_class = ImportSessionSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]
    pagination_class = StandardResultsSetPagination
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ['status', 'import_mode']
    ordering_fields = ['created_at', 'completed_at', 'filename']
    ordering = ['-created_at']

    def get_queryset(self):
        # Only return sessions for the current user
        return ImportSession.objects.filter(user=self.request.user)


class ImportSessionRetrieveView(generics.RetrieveAPIView):
    serializer_class = ImportSessionSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]

    def get_queryset(self):
        # Only return sessions for the current user
        return ImportSession.objects.filter(user=self.request.user)



@api_view(['GET'])
@permission_required('inventory.view_stock')
def low_stock_alert_api(request):
    """API endpoint for low stock alerts"""
    try:
        limit = int(request.query_params.get('limit', 20))

        alerts = Stock.objects.select_related('product', 'store').filter(
            Q(quantity=0) | Q(quantity__lte=F('low_stock_threshold'))
        ).order_by('quantity')[:limit]

        alert_data = []
        for alert in alerts:
            alert_data.append({
                'id': alert.id,
                'product_name': alert.product.name,
                'product_sku': alert.product.sku,
                'store_name': alert.store.name,
                'current_stock': float(alert.quantity),
                'reorder_level': float(alert.low_stock_threshold),
                'unit_of_measure': alert.product.unit_of_measure,
                'status': 'critical' if alert.quantity == 0 else 'low',
                'stock_percentage': alert.stock_percentage
            })

        return Response({'alerts': alert_data})

    except Exception as e:
        return Response(
            {'error': f'Failed to fetch low stock alerts: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )



@api_view(['GET'])
@permission_required('inventory.view_stock')
def inventory_report_api(request):
    """API endpoint for inventory report data"""
    try:
        # Get query parameters
        category_id = request.query_params.get('category')
        store_id = request.query_params.get('store')
        status_filter = request.query_params.get('status')

        # Base queryset
        queryset = Stock.objects.select_related('product', 'product__category', 'store')

        # Apply filters
        if category_id:
            queryset = queryset.filter(product__category_id=category_id)
        if store_id:
            queryset = queryset.filter(store_id=store_id)
        if status_filter == 'low':
            queryset = queryset.filter(quantity__lte=F('low_stock_threshold'))
        elif status_filter == 'out':
            queryset = queryset.filter(quantity=0)

        # Prepare report data
        report_data = []
        for stock in queryset:
            if stock.quantity == 0:
                status = 'Out of Stock'
            elif stock.quantity <= stock.low_stock_threshold:
                status = 'Low Stock'
            else:
                status = 'In Stock'

            report_data.append({
                'product_id': stock.product.id,
                'product_name': stock.product.name,
                'sku': stock.product.sku,
                'category': stock.product.category.name if stock.product.category else None,
                'store': stock.store.name,
                'current_stock': stock.quantity,
                'reorder_level': stock.low_stock_threshold,
                'unit_of_measure': stock.product.unit_of_measure,
                'cost_price': stock.product.cost_price,
                'selling_price': stock.product.selling_price,
                'total_cost': stock.quantity * stock.product.cost_price,
                'status': status,
                'last_updated': stock.last_updated
            })

        serializer = InventoryReportSerializer(report_data, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'error': f'Failed to generate inventory report: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_required('inventory.view_stockmovement')
def movement_report_api(request):
    """API endpoint for stock movement report data"""
    try:
        # Get query parameters
        date_from = request.query_params.get('date_from')
        date_to = request.query_params.get('date_to')
        movement_type = request.query_params.get('movement_type')
        product_id = request.query_params.get('product')
        store_id = request.query_params.get('store')

        # Base queryset
        queryset = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')

        # Apply filters
        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__gte=date_from_parsed)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__lte=date_to_parsed)
            except ValueError:
                pass

        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)
        if product_id:
            queryset = queryset.filter(product_id=product_id)
        if store_id:
            queryset = queryset.filter(store_id=store_id)

        report_data = []
        for movement in queryset:
            report_data.append({
                'date': movement.created_at.date(),
                'product_name': movement.product.name,
                'product_sku': movement.product.sku,
                'store_name': movement.store.name,
                'movement_type': movement.movement_type,
                'quantity': movement.quantity,
                'unit_price': movement.unit_price,
                'total_value': movement.total_value,
                'reference': movement.reference,
                'notes': movement.notes,
                'created_by': movement.created_by.get_full_name() or movement.created_by.username
            })

        serializer = StockMovementReportSerializer(report_data, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'error': f'Failed to generate movement report: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_required('inventory.view_stock')
def low_stock_report_api(request):
    """API endpoint for low stock report data"""
    try:
        category_id = request.query_params.get('category')
        store_id = request.query_params.get('store')
        severity = request.query_params.get('severity', 'all')

        queryset = Stock.objects.select_related(
            'product', 'product__category', 'store'
        ).annotate(
            total_cost=F('quantity') * F('product__cost_price'),
            reorder_gap=F('low_stock_threshold') - F('quantity')
        ).filter(
            Q(quantity__lte=F('low_stock_threshold')) | Q(quantity=0)
        )

        # Apply filters
        if category_id:
            queryset = queryset.filter(product__category_id=category_id)
        if store_id:
            queryset = queryset.filter(store_id=store_id)
        if severity == 'critical':
            queryset = queryset.filter(quantity=0)
        elif severity == 'low':
            queryset = queryset.filter(quantity__gt=0, quantity__lte=F('low_stock_threshold'))

        # Prepare report data
        report_data = []
        for stock in queryset:
            if stock.quantity == 0:
                status = 'Critical'
            elif stock.quantity <= stock.low_stock_threshold / 2:
                status = 'Very Low'
            else:
                status = 'Low'

            stock_percentage = (stock.quantity / stock.low_stock_threshold * 100) if stock.low_stock_threshold > 0 else 100
            recommended_qty = max(0, stock.low_stock_threshold * 1.5 - stock.quantity)

            report_data.append({
                'product_id': stock.product.id,
                'product_name': stock.product.name,
                'sku': stock.product.sku,
                'category': stock.product.category.name if stock.product.category else None,
                'store': stock.store.name,
                'current_stock': stock.quantity,
                'reorder_level': stock.low_stock_threshold,
                'reorder_gap': stock.reorder_gap,
                'stock_percentage': round(stock_percentage, 1),
                'total_cost': stock.total_cost,
                'recommended_order_qty': recommended_qty,
                'status': status
            })

        serializer = LowStockReportSerializer(report_data, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'error': f'Failed to generate low stock report: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
@permission_required('inventory.view_stock')
def valuation_report_api(request):
    """API endpoint for inventory valuation report data"""
    try:
        # Get query parameters
        category_id = request.query_params.get('category')
        store_id = request.query_params.get('store')

        queryset = Stock.objects.select_related(
            'product', 'product__category', 'store'
        ).annotate(
            total_cost=F('quantity') * F('product__cost_price'),
            total_selling=F('quantity') * F('product__selling_price'),
            potential_profit=F('quantity') * (F('product__selling_price') - F('product__cost_price'))
        ).filter(quantity__gt=0)

        if category_id:
            queryset = queryset.filter(product__category_id=category_id)
        if store_id:
            queryset = queryset.filter(store_id=store_id)

        report_data = []
        for stock in queryset:
            report_data.append({
                'product_id': stock.product.id,
                'product_name': stock.product.name,
                'sku': stock.product.sku,
                'category': stock.product.category.name if stock.product.category else None,
                'store': stock.store.name,
                'quantity': stock.quantity,
                'cost_price': stock.product.cost_price,
                'selling_price': stock.product.selling_price,
                'total_cost': stock.total_cost,
                'total_selling': stock.total_selling,
                'potential_profit': stock.potential_profit,
                'unit_of_measure': stock.product.unit_of_measure
            })

        serializer = ValuationReportSerializer(report_data, many=True)
        return Response(serializer.data)

    except Exception as e:
        return Response(
            {'error': f'Failed to generate valuation report: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_required('inventory.change_product')
def bulk_update_products_api(request):
    """API endpoint for bulk product updates"""
    try:
        action = request.data.get('action')
        product_ids = request.data.get('product_ids', [])

        if not action or not product_ids:
            return Response(
                {'error': 'Action and product_ids are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        products = Product.objects.filter(id__in=product_ids)
        count = products.count()

        if action == 'activate':
            products.update(is_active=True)
            message = f'{count} products activated successfully'
        elif action == 'deactivate':
            products.update(is_active=False)
            message = f'{count} products deactivated successfully'
        elif action == 'enable_efris':
            products.update(efris_auto_sync_enabled=True)
            message = f'EFRIS sync enabled for {count} products'
        elif action == 'disable_efris':
            products.update(efris_auto_sync_enabled=False)
            message = f'EFRIS sync disabled for {count} products'
        else:
            return Response(
                {'error': 'Invalid action'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return Response({
            'success': True,
            'message': message,
            'affected_count': count
        })

    except Exception as e:
        return Response(
            {'error': f'Bulk update failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_required('inventory.add_stockmovement')
def bulk_stock_adjustment_api(request):
    """API endpoint for bulk stock adjustments"""
    try:
        adjustments = request.data.get('adjustments', [])

        if not adjustments:
            return Response(
                {'error': 'Adjustments data is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        success_count = 0
        errors = []

        for adj_data in adjustments:
            try:
                product_id = adj_data.get('product_id')
                store_id = adj_data.get('store_id')
                adjustment_type = adj_data.get('adjustment_type')
                quantity = adj_data.get('quantity')
                reason = adj_data.get('reason', 'Bulk adjustment')

                product = Product.objects.get(id=product_id)
                store = Store.objects.get(id=store_id)

                # Get or create stock record
                stock, created = Stock.objects.get_or_create(
                    product=product,
                    store=store,
                    defaults={'quantity': 0}
                )

                old_quantity = stock.quantity

                # Calculate new quantity
                if adjustment_type == 'add':
                    new_quantity = old_quantity + quantity
                    movement_quantity = quantity
                elif adjustment_type == 'remove':
                    new_quantity = max(0, old_quantity - quantity)
                    movement_quantity = -(min(quantity, old_quantity))
                elif adjustment_type == 'set':
                    new_quantity = quantity
                    movement_quantity = new_quantity - old_quantity
                else:
                    errors.append(f"Invalid adjustment type for {product.name}")
                    continue

                # Create movement record
                StockMovement.objects.create(
                    product=product,
                    store=store,
                    movement_type='ADJUSTMENT',
                    quantity=movement_quantity,
                    reference=f'BULK-{timezone.now().strftime("%Y%m%d%H%M")}',
                    notes=reason,
                    created_by=request.user
                )

                # Update stock
                stock.quantity = new_quantity
                stock.save()

                success_count += 1

            except Exception as e:
                errors.append(f"Error processing adjustment: {str(e)}")

        return Response({
            'success': success_count > 0,
            'success_count': success_count,
            'error_count': len(errors),
            'errors': errors[:10]  # Limit errors returned
        })

    except Exception as e:
        return Response(
            {'error': f'Bulk adjustment failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


# Search and Autocomplete APIs
@api_view(['GET'])
@permission_required('inventory.view_product')
def product_search_api(request):
    """API endpoint for product search/autocomplete"""
    try:
        query = request.query_params.get('q', '')
        limit = int(request.query_params.get('limit', 10))

        if len(query) < 2:
            return Response({'results': []})

        products = Product.objects.filter(
            Q(name__icontains=query) | Q(sku__icontains=query) | Q(barcode__icontains=query),
            is_active=True
        )[:limit]

        results = []
        for product in products:
            results.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'selling_price': float(product.selling_price),
                'cost_price': float(product.cost_price),
                'unit_of_measure': product.unit_of_measure,
                'category': product.category.name if product.category else None,
                'current_stock': product.total_stock
            })

        return Response({'results': results})

    except Exception as e:
        return Response(
            {'error': f'Product search failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@login_required
@csrf_exempt
@permission_required('inventory.add_product', raise_exception=True)
@require_http_methods(["POST"])
def process_bulk_import(request):
    """Process the uploaded file and return import results with proper session tracking"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        uploaded_file = request.FILES['file']
        import_mode = request.POST.get('import_mode', 'both')
        conflict_resolution = request.POST.get('conflict_resolution', 'overwrite')
        column_mapping = json.loads(request.POST.get('column_mapping', '{}'))
        has_header = request.POST.get('has_header', 'true').lower() == 'true'

        # Validate file type
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            return JsonResponse({'error': 'Unsupported file format'}, status=400)

        # Create ImportSession
        import_session = ImportSession.objects.create(
            user=request.user,
            filename=uploaded_file.name,
            file_size=uploaded_file.size,
            import_mode=import_mode,
            conflict_resolution=conflict_resolution,
            has_header=has_header,
            column_mapping=column_mapping,
            status='processing',
            started_at=timezone.now()
        )

        # Process the file
        result = process_import_file_with_session(
            uploaded_file,
            import_mode,
            conflict_resolution,
            column_mapping,
            has_header,
            request.user,
            import_session
        )

        return JsonResponse(result)

    except Exception as e:
        logger.error(f"Import error: {str(e)}")
        return JsonResponse({'error': f'Import failed: {str(e)}'}, status=500)


def process_import_file_with_session(file, import_mode, conflict_resolution, column_mapping, has_header, user,
                                     import_session):
    """Enhanced process_import_file with ImportSession integration"""
    results = {
        'success': True,
        'session_id': import_session.id,
        'total_processed': 0,
        'created': [],
        'updated': [],
        'skipped': [],
        'errors': [],
        'summary': {
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'error_count': 0
        }
    }

    try:
        # Read file data
        if file.name.lower().endswith('.csv'):
            data = read_csv_data(file, has_header)
        else:
            data = read_excel_data(file, has_header)

        if not data:
            import_session.status = 'failed'
            import_session.error_message = 'No data found in file'
            import_session.completed_at = timezone.now()
            import_session.save()

            # Log error
            ImportLog.objects.create(
                session=import_session,
                level='error',
                message='No data found in file'
            )

            results['success'] = False
            results['error'] = 'No data found in file'
            return results

        # Update session with total rows
        import_session.total_rows = len(data)
        import_session.save()

        # Log start
        ImportLog.objects.create(
            session=import_session,
            level='info',
            message=f'Starting import of {len(data)} rows'
        )

        # Process each row
        with transaction.atomic():
            for row_index, row_data in enumerate(data):
                try:
                    # Map columns based on user selection
                    mapped_data = map_row_data(row_data, column_mapping)

                    # Validate required fields
                    if not validate_row_data(mapped_data):
                        error_msg = 'Missing required fields: Product name/SKU, quantity, and store are required'

                        # Log error
                        ImportLog.objects.create(
                            session=import_session,
                            level='error',
                            message=error_msg,
                            row_number=row_index + (2 if has_header else 1),
                            details={'raw_data': dict(zip(range(len(row_data)), row_data))}
                        )

                        # Create result record
                        ImportResult.objects.create(
                            session=import_session,
                            result_type='error',
                            row_number=row_index + (2 if has_header else 1),
                            error_message=error_msg,
                            raw_data=dict(zip(range(len(row_data)), row_data))
                        )

                        results['errors'].append({
                            'row': row_index + (2 if has_header else 1),
                            'error': error_msg,
                            'details': 'Product name/SKU, quantity, and store are required'
                        })
                        continue

                    # Process the row
                    row_result = process_single_row_with_session(
                        mapped_data,
                        import_mode,
                        conflict_resolution,
                        row_index + (2 if has_header else 1),
                        import_session,
                        row_data
                    )

                    # Add to appropriate result category
                    if row_result['status'] == 'created':
                        results['created'].append(row_result['data'])
                        results['summary']['created_count'] += 1
                    elif row_result['status'] == 'updated':
                        results['updated'].append(row_result['data'])
                        results['summary']['updated_count'] += 1
                    elif row_result['status'] == 'skipped':
                        results['skipped'].append(row_result['data'])
                        results['summary']['skipped_count'] += 1

                    results['total_processed'] += 1

                except Exception as e:
                    error_msg = str(e)
                    row_num = row_index + (2 if has_header else 1)

                    # Log error
                    ImportLog.objects.create(
                        session=import_session,
                        level='error',
                        message=f'Error processing row: {error_msg}',
                        row_number=row_num,
                        details={'exception': error_msg, 'raw_data': dict(zip(range(len(row_data)), row_data))}
                    )

                    # Create result record
                    ImportResult.objects.create(
                        session=import_session,
                        result_type='error',
                        row_number=row_num,
                        error_message=error_msg,
                        raw_data=dict(zip(range(len(row_data)), row_data))
                    )

                    results['errors'].append({
                        'row': row_num,
                        'error': error_msg,
                        'details': f'Error processing row data: {row_data}'
                    })

            # Update session counts
            results['summary']['error_count'] = len(results['errors'])
            import_session.processed_rows = results['total_processed']
            import_session.created_count = results['summary']['created_count']
            import_session.updated_count = results['summary']['updated_count']
            import_session.skipped_count = results['summary']['skipped_count']
            import_session.error_count = results['summary']['error_count']
            import_session.status = 'completed'
            import_session.completed_at = timezone.now()
            import_session.save()

            # Log completion
            ImportLog.objects.create(
                session=import_session,
                level='success',
                message=f'Import completed successfully. Created: {results["summary"]["created_count"]}, Updated: {results["summary"]["updated_count"]}, Errors: {results["summary"]["error_count"]}'
            )

    except Exception as e:
        import_session.status = 'failed'
        import_session.error_message = str(e)
        import_session.completed_at = timezone.now()
        import_session.save()

        # Log failure
        ImportLog.objects.create(
            session=import_session,
            level='error',
            message=f'Import failed: {str(e)}'
        )

        results['success'] = False
        results['error'] = str(e)

    return results


def process_single_row_with_session(data, import_mode, conflict_resolution, row_number, import_session, raw_data):
    """Enhanced process_single_row with session tracking"""
    try:
        # Get or create product
        product = get_or_create_product_with_session(data, import_session, row_number)

        # Get or create store
        store = get_or_create_store(data['store'])

        quantity = int(float(data['quantity']))

        try:
            stock, created = Stock.objects.get_or_create(
                product=product,
                store=store,
                defaults={
                    'quantity': quantity,
                    'import_session': import_session,
                    'last_import_update': timezone.now()
                }
            )

            if created:
                # Log creation
                ImportLog.objects.create(
                    session=import_session,
                    level='success',
                    message=f'Created stock record for {product.name} at {store.name}',
                    row_number=row_number,
                    details={'quantity': quantity}
                )

                # Create result record
                ImportResult.objects.create(
                    session=import_session,
                    result_type='created',
                    row_number=row_number,
                    product_name=product.name,
                    sku=product.sku,
                    store_name=store.name,
                    quantity=quantity,
                    raw_data=dict(zip(range(len(raw_data)), raw_data))
                )

                return {
                    'status': 'created',
                    'data': {
                        'product': product.name,
                        'store': store.name,
                        'quantity': quantity
                    }
                }
            else:
                # Handle existing stock based on conflict resolution
                old_quantity = stock.quantity

                if conflict_resolution == 'skip':
                    # Log skip
                    ImportLog.objects.create(
                        session=import_session,
                        level='info',
                        message=f'Skipped existing stock for {product.name} at {store.name}',
                        row_number=row_number
                    )

                    # Create result record
                    ImportResult.objects.create(
                        session=import_session,
                        result_type='skipped',
                        row_number=row_number,
                        product_name=product.name,
                        sku=product.sku,
                        store_name=store.name,
                        quantity=old_quantity,
                        raw_data=dict(zip(range(len(raw_data)), raw_data))
                    )

                    return {
                        'status': 'skipped',
                        'data': {
                            'product': product.name,
                            'store': store.name,
                            'reason': 'Item already exists'
                        }
                    }
                elif conflict_resolution == 'overwrite':
                    stock.quantity = quantity
                elif conflict_resolution == 'merge':
                    stock.quantity += quantity

                stock.import_session = import_session
                stock.last_import_update = timezone.now()
                stock.save()

                # Log update
                ImportLog.objects.create(
                    session=import_session,
                    level='success',
                    message=f'Updated stock for {product.name} at {store.name} from {old_quantity} to {stock.quantity}',
                    row_number=row_number,
                    details={'old_quantity': old_quantity, 'new_quantity': stock.quantity}
                )

                # Create result record
                ImportResult.objects.create(
                    session=import_session,
                    result_type='updated',
                    row_number=row_number,
                    product_name=product.name,
                    sku=product.sku,
                    store_name=store.name,
                    quantity=stock.quantity,
                    old_quantity=old_quantity,
                    raw_data=dict(zip(range(len(raw_data)), raw_data))
                )

                return {
                    'status': 'updated',
                    'data': {
                        'product': product.name,
                        'store': store.name,
                        'old_quantity': old_quantity,
                        'new_quantity': stock.quantity
                    }
                }

        except Exception as e:
            raise Exception(f"Error processing stock for {product.name}: {str(e)}")

    except Exception as e:
        raise Exception(f"Error processing row {row_number}: {str(e)}")


def get_or_create_product_with_session(data, import_session, row_number):
    """Enhanced get_or_create_product with session tracking"""
    product = None

    if data.get('sku'):
        try:
            product = Product.objects.get(sku=data['sku'])
        except Product.DoesNotExist:
            pass

    if not product and data.get('product_name'):
        try:
            product = Product.objects.get(name=data['product_name'])
        except Product.DoesNotExist:
            pass

    # Create new product if not found
    if not product:
        # Get or create category if provided
        category = None
        if data.get('category'):
            category, _ = Category.objects.get_or_create(
                name=data['category'],
                defaults={'is_active': True}
            )

        # Get or create supplier if provided
        supplier = None
        if data.get('supplier'):
            supplier, _ = Supplier.objects.get_or_create(
                name=data['supplier'],
                defaults={'is_active': True}
            )

        # Required fields with defaults
        product_data = {
            'name': data.get('product_name', data.get('sku', 'Unknown Product')),
            'sku': data.get('sku', f"AUTO-{timezone.now().strftime('%Y%m%d%H%M%S')}-{row_number}"),
            'selling_price': Decimal(data.get('selling_price', '0.00')),
            'cost_price': Decimal(data.get('cost_price', '0.00')),
            'category': category,
            'supplier': supplier,
            'unit_of_measure': data.get('unit_of_measure', 'each'),
            'is_active': True,
            'import_session': import_session,
            'imported_at': timezone.now()
        }

        # Optional fields
        if data.get('description'):
            product_data['description'] = data['description']
        if data.get('barcode'):
            product_data['barcode'] = data['barcode']
        if data.get('tax_rate') and data['tax_rate'] in dict(Product.TAX_RATE_CHOICES):
            product_data['tax_rate'] = data['tax_rate']

        product = Product.objects.create(**product_data)

        # Log product creation
        ImportLog.objects.create(
            session=import_session,
            level='info',
            message=f'Created new product: {product.name}',
            row_number=row_number,
            details={'product_id': product.id, 'sku': product.sku}
        )

    return product


@login_required
@permission_required('inventory.add_product', raise_exception=True)
def import_sessions(request):
    """View to manage and monitor import sessions"""
    sessions = ImportSession.objects.filter(user=request.user).order_by('-created_at')

    # Apply filters
    status_filter = request.GET.get('status')
    if status_filter:
        sessions = sessions.filter(status=status_filter)

    paginator = Paginator(sessions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'sessions': page_obj,
        'status_filter': status_filter,
        'status_choices': ImportSession.STATUS_CHOICES
    }

    return render(request, 'inventory/import_sessions.html', context)


@login_required
@permission_required('inventory.add_product', raise_exception=True)
def import_session_detail(request, session_id):
    """View detailed results of an import session"""
    session = get_object_or_404(ImportSession, id=session_id, user=request.user)

    # Get logs
    logs = session.logs.order_by('timestamp')

    # Get results with filtering
    result_type = request.GET.get('result_type', 'all')
    results = session.results.order_by('row_number')

    if result_type != 'all':
        results = results.filter(result_type=result_type)

    # Paginate results
    paginator = Paginator(results, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get summary statistics
    result_summary = session.results.values('result_type').annotate(
        count=Count('id')
    ).order_by('result_type')

    context = {
        'session': session,
        'logs': logs,
        'results': page_obj,
        'result_summary': result_summary,
        'result_type_filter': result_type,
        'result_type_choices': ImportResult.RESULT_TYPES
    }

    return render(request, 'inventory/import_session_detail.html', context)


@login_required
@permission_required('inventory.add_product', raise_exception=True)
def retry_import_session(request, session_id):
    """Retry a failed import session"""
    session = get_object_or_404(ImportSession, id=session_id, user=request.user)

    if session.status != 'failed':
        messages.error(request, 'Only failed imports can be retried.')
        return redirect('inventory:import_session_detail', session_id=session_id)

    if request.method == 'POST':
        # Create new session based on the failed one
        new_session = ImportSession.objects.create(
            user=request.user,
            filename=f"RETRY_{session.filename}",
            file_size=session.file_size,
            import_mode=session.import_mode,
            conflict_resolution=session.conflict_resolution,
            has_header=session.has_header,
            column_mapping=session.column_mapping,
            status='pending'
        )

        messages.success(request, f'Created retry session #{new_session.id}')
        return redirect('inventory:import_session_detail', session_id=new_session.id)

    return render(request, 'inventory/retry_import.html', {'session': session})


@login_required
@permission_required('inventory.add_product', raise_exception=True)
def import_session_status_api(request, session_id):
    """API endpoint to get import session status"""
    try:
        session = ImportSession.objects.get(id=session_id, user=request.user)

        data = {
            'id': session.id,
            'status': session.status,
            'filename': session.filename,
            'total_rows': session.total_rows,
            'processed_rows': session.processed_rows,
            'created_count': session.created_count,
            'updated_count': session.updated_count,
            'skipped_count': session.skipped_count,
            'error_count': session.error_count,
            'success_rate': session.success_rate,
            'duration': str(session.duration) if session.duration else None,
            'created_at': session.created_at.isoformat(),
            'completed_at': session.completed_at.isoformat() if session.completed_at else None
        }

        return JsonResponse(data)

    except ImportSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting session status: {str(e)}")
        return JsonResponse({'error': 'Internal server error'}, status=500)


# inventory/views.py (Update the dashboard view section)

@login_required
@permission_required('inventory.view_product', raise_exception=True)
def inventory_dashboard(request):
    """
    Enhanced inventory dashboard with comprehensive analytics and real-time data
    """
    from django.db.models import Q, Sum, Count, F, Avg, Max, Min, Case, When, Value
    from django.db.models.functions import Coalesce, TruncDate
    from datetime import datetime, timedelta
    from decimal import Decimal

    # Get filter period from request
    period = request.GET.get('period', 'week')

    # Calculate date ranges
    today = timezone.now().date()

    if period == 'today':
        start_date = today
    elif period == 'week':
        start_date = today - timedelta(days=7)
    elif period == 'month':
        start_date = today - timedelta(days=30)
    elif period == 'quarter':
        start_date = today - timedelta(days=90)
    elif period == 'year':
        start_date = today - timedelta(days=365)
    else:
        start_date = today - timedelta(days=7)

    # Base querysets
    stock_qs = Stock.objects.select_related('product', 'store', 'product__category')
    movements_qs = StockMovement.objects.select_related('product', 'store', 'created_by')
    products_qs = Product.objects.filter(is_active=True)
    categories_qs = Category.objects.filter(is_active=True)
    suppliers_qs = Supplier.objects.filter(is_active=True)

    # ============================================
    # Core Metrics
    # ============================================

    # Total counts
    total_products = products_qs.count()
    total_categories = categories_qs.count()
    total_suppliers = suppliers_qs.count()

    # Stock status counts
    stock_stats = stock_qs.aggregate(
        total_items=Count('id'),
        out_of_stock=Count('id', filter=Q(quantity=0)),
        low_stock=Count('id', filter=Q(
            quantity__gt=0,
            quantity__lte=F('low_stock_threshold')
        )),
        medium_stock=Count('id', filter=Q(
            quantity__gt=F('low_stock_threshold'),
            quantity__lte=F('low_stock_threshold') * 2
        )),
        good_stock=Count('id', filter=Q(
            quantity__gt=F('low_stock_threshold') * 2
        ))
    )

    low_stock_items = stock_stats['low_stock'] or 0
    out_of_stock_items = stock_stats['out_of_stock'] or 0

    # Stock valuation
    stock_value = stock_qs.aggregate(
        total_cost_value=Coalesce(Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')),
        total_selling_value=Coalesce(Sum(F('quantity') * F('product__selling_price')), Decimal('0.00')),
        avg_item_value=Coalesce(Avg(F('quantity') * F('product__cost_price')), Decimal('0.00'))
    )

    total_stock_value = stock_value['total_cost_value']
    potential_profit = stock_value['total_selling_value'] - stock_value['total_cost_value']

    # ============================================
    # Movement Statistics
    # ============================================

    # Today's movements
    movements_today = movements_qs.filter(
        created_at__date=today
    ).count()

    # Period movements
    movements_period = movements_qs.filter(
        created_at__date__gte=start_date
    ).count()

    # This week's movements
    week_ago = today - timedelta(days=7)
    movements_week = movements_qs.filter(
        created_at__date__gte=week_ago
    ).count()

    # This month's movements
    month_ago = today - timedelta(days=30)
    movements_month = movements_qs.filter(
        created_at__date__gte=month_ago
    ).count()

    # Movement type breakdown
    movement_breakdown = movements_qs.filter(
        created_at__date__gte=start_date
    ).values('movement_type').annotate(
        count=Count('id'),
        total_quantity=Coalesce(Sum('quantity'), Decimal('0.00'))
    ).order_by('-count')

    # ============================================
    # Trend Analysis
    # ============================================

    # Previous period for comparison
    period_days = (today - start_date).days
    previous_start = start_date - timedelta(days=period_days)
    previous_end = start_date

    # Previous period metrics
    previous_movements = movements_qs.filter(
        created_at__date__gte=previous_start,
        created_at__date__lt=previous_end
    ).count()

    # Calculate trends
    movements_trend = calculate_trend(movements_period, previous_movements)

    # Stock value trend (simplified - comparing current vs 30 days ago)
    # In production, you'd track historical stock values
    stock_value_trend = {
        'value': 15.3,  # Percentage
        'direction': 'up'
    }

    # ============================================
    # Recent Activity
    # ============================================

    # Recent stock movements (last 20)
    recent_movements = movements_qs.order_by('-created_at')[:20]

    # ============================================
    # Top Products
    # ============================================

    # Top products by movement activity
    top_products = products_qs.annotate(
        total_movements=Count('movements', filter=Q(
            movements__created_at__date__gte=start_date
        )),
        total_quantity_moved=Coalesce(Sum(
            'movements__quantity',
            filter=Q(movements__created_at__date__gte=start_date)
        ), Decimal('0.00')),
        total_sales=Count('movements', filter=Q(
            movements__movement_type='SALE',
            movements__created_at__date__gte=start_date
        ))
    ).filter(
        total_movements__gt=0
    ).order_by('-total_movements')[:10]

    # ============================================
    # Category Analysis
    # ============================================

    # Stock value by category
    category_distribution = stock_qs.values(
        'product__category__name',
        'product__category__id'
    ).annotate(
        category_name=F('product__category__name'),
        total_value=Coalesce(Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')),
        total_quantity=Coalesce(Sum('quantity'), Decimal('0.00')),
        product_count=Count('product', distinct=True),
        avg_value=Coalesce(Avg(F('quantity') * F('product__cost_price')), Decimal('0.00'))
    ).filter(
        category_name__isnull=False
    ).order_by('-total_value')[:10]

    # ============================================
    # Store Analysis
    # ============================================

    # Stock by store
    store_distribution = stock_qs.values(
        'store__name',
        'store__id'
    ).annotate(
        store_name=F('store__name'),
        total_value=Coalesce(Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')),
        total_quantity=Coalesce(Sum('quantity'), Decimal('0.00')),
        product_count=Count('product', distinct=True),
        low_stock_count=Count('id', filter=Q(
            quantity__gt=0,
            quantity__lte=F('low_stock_threshold')
        )),
        out_of_stock_count=Count('id', filter=Q(quantity=0))
    ).order_by('-total_value')

    # ============================================
    # Alerts & Warnings
    # ============================================

    # Critical stock alerts (out of stock + very low stock)
    critical_alerts = stock_qs.filter(
        Q(quantity=0) | Q(quantity__lte=F('low_stock_threshold') / 2)
    ).select_related('product', 'store').order_by('quantity')[:10]

    # Low stock alerts
    low_stock_alerts = stock_qs.filter(
        quantity__gt=0,
        quantity__lte=F('low_stock_threshold')
    ).exclude(
        quantity__lte=F('low_stock_threshold') / 2
    ).select_related('product', 'store').order_by('quantity')[:10]

    # Combine alerts
    all_alerts = list(critical_alerts) + list(low_stock_alerts)

    # ============================================
    # Movement Trends Data (for charts)
    # ============================================

    # Daily movement trends for the period
    daily_movements = movements_qs.filter(
        created_at__date__gte=start_date
    ).annotate(
        date=TruncDate('created_at')
    ).values('date', 'movement_type').annotate(
        count=Count('id'),
        total_quantity=Coalesce(Sum('quantity'), Decimal('0.00'))
    ).order_by('date')

    # Process for chart
    movement_trends = process_movement_trends(daily_movements, start_date, today)

    # ============================================
    # Inventory Turnover Ratio
    # ============================================

    # Calculate inventory turnover (Cost of Goods Sold / Average Inventory)
    # Simplified calculation for demonstration
    total_sales_value = movements_qs.filter(
        movement_type='SALE',
        created_at__date__gte=month_ago
    ).aggregate(
        total=Coalesce(Sum(F('quantity') * F('product__cost_price')), Decimal('0.00'))
    )['total']

    avg_inventory_value = stock_value['total_cost_value']

    if avg_inventory_value > 0:
        # Annualized turnover ratio
        monthly_turnover = total_sales_value / avg_inventory_value if avg_inventory_value > 0 else 0
        annual_turnover = monthly_turnover * 12
        inventory_turnover = round(annual_turnover, 1)
    else:
        inventory_turnover = 0

    # ============================================
    # Performance Metrics
    # ============================================

    # Calculate stock accuracy (products with accurate stock levels)
    total_stock_records = stock_qs.count()
    accurate_stock = stock_qs.filter(
        last_physical_count__isnull=False,
        last_physical_count__gte=today - timedelta(days=30)
    ).count()

    stock_accuracy = (accurate_stock / total_stock_records * 100) if total_stock_records > 0 else 0

    # Average days to stockout (for low stock items)
    # This is a simplified calculation
    days_to_stockout = calculate_days_to_stockout(low_stock_alerts, movements_qs)

    # ============================================
    # Prepare Context
    # ============================================

    context = {
        # Core Metrics
        'total_products': total_products,
        'total_categories': total_categories,
        'total_suppliers': total_suppliers,
        'stock_value': total_stock_value,
        'potential_profit': potential_profit,

        # Stock Status
        'low_stock_items': low_stock_items,
        'out_of_stock_items': out_of_stock_items,
        'stock_stats': stock_stats,

        # Movements
        'movements_today': movements_today,
        'movements_week': movements_week,
        'movements_month': movements_month,
        'movements_period': movements_period,
        'movement_breakdown': movement_breakdown,

        # Trends
        'movements_trend': movements_trend,
        'stock_value_trend': stock_value_trend,

        # Lists
        'recent_movements': recent_movements,
        'top_products': top_products,
        'all_alerts': all_alerts,
        'critical_alerts': critical_alerts,
        'low_stock_alerts': low_stock_alerts,

        # Distributions
        'category_distribution': category_distribution,
        'store_distribution': store_distribution,

        # Chart Data
        'movement_trends': movement_trends,

        # Performance Metrics
        'inventory_turnover': inventory_turnover,
        'stock_accuracy': round(stock_accuracy, 1),
        'days_to_stockout': days_to_stockout,

        # Filter
        'selected_period': period,
        'start_date': start_date,
        'end_date': today,

        # Additional Info
        'company': request.user.company if hasattr(request.user, 'company') else None,
    }

    return render(request, 'inventory/dashboards.html', context)


# ============================================
# Helper Functions
# ============================================

def calculate_trend(current_value, previous_value):
    """
    Calculate trend percentage and direction
    """
    if previous_value == 0:
        if current_value > 0:
            return {'value': 100, 'direction': 'up'}
        return {'value': 0, 'direction': 'neutral'}

    percentage = ((current_value - previous_value) / previous_value) * 100
    direction = 'up' if percentage > 0 else 'down' if percentage < 0 else 'neutral'

    return {
        'value': abs(round(percentage, 1)),
        'direction': direction
    }


def process_movement_trends(daily_movements, start_date, end_date):
    """
    Process daily movements into chart-ready format
    """
    from collections import defaultdict
    from datetime import timedelta

    # Create date range
    dates = []
    current_date = start_date
    while current_date <= end_date:
        dates.append(current_date)
        current_date += timedelta(days=1)

    # Initialize data structures
    purchases = defaultdict(int)
    sales = defaultdict(int)
    adjustments = defaultdict(int)

    # Process movements
    for movement in daily_movements:
        date = movement['date']
        movement_type = movement['movement_type']
        count = movement['count']

        if movement_type in ['PURCHASE', 'RETURN', 'TRANSFER_IN']:
            purchases[date] += count
        elif movement_type in ['SALE', 'TRANSFER_OUT']:
            sales[date] += count
        elif movement_type == 'ADJUSTMENT':
            adjustments[date] += count

    # Format for chart
    return {
        'labels': [date.strftime('%b %d') for date in dates],
        'purchases': [purchases.get(date, 0) for date in dates],
        'sales': [sales.get(date, 0) for date in dates],
        'adjustments': [adjustments.get(date, 0) for date in dates]
    }


def calculate_days_to_stockout(low_stock_items, movements_qs):
    """
    Calculate average days until stockout for low stock items
    """
    from datetime import timedelta

    if not low_stock_items:
        return None

    thirty_days_ago = timezone.now() - timedelta(days=30)
    total_days = 0
    count = 0

    for stock in low_stock_items[:5]:  # Sample first 5
        # Get average daily consumption
        recent_sales = movements_qs.filter(
            product=stock.product,
            store=stock.store,
            movement_type__in=['SALE', 'TRANSFER_OUT'],
            created_at__gte=thirty_days_ago
        ).aggregate(
            total=Coalesce(Sum('quantity'), Decimal('0.00'))
        )['total']

        if recent_sales > 0:
            daily_consumption = recent_sales / 30
            if daily_consumption > 0 and stock.quantity > 0:
                days_remaining = float(stock.quantity) / float(daily_consumption)
                total_days += days_remaining
                count += 1

    return round(total_days / count) if count > 0 else None


# ============================================
# API Endpoints for Real-time Updates
# ============================================

@login_required
@permission_required('inventory.view_product')
@require_http_methods(["GET"])
def dashboard_stats_api(request):
    """
    API endpoint for dashboard statistics
    """
    try:
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)

        # Product counts
        total_products = Product.objects.filter(is_active=True).count()
        total_categories = Category.objects.filter(is_active=True).count()
        total_suppliers = Supplier.objects.filter(is_active=True).count()

        # Stock statistics
        stock_stats = Stock.objects.aggregate(
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            low_stock=Count('id', filter=Q(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold')
            ))
        )

        # Stock value
        stock_value = Stock.objects.aggregate(
            total_value=Coalesce(
                Sum(F('quantity') * F('product__cost_price')),
                Decimal('0.00')
            )
        )['total_value']

        # Movement counts
        movements_today = StockMovement.objects.filter(
            created_at__date=today
        ).count()

        movements_week = StockMovement.objects.filter(
            created_at__date__gte=week_ago
        ).count()

        data = {
            'success': True,
            'timestamp': timezone.now().isoformat(),
            'total_products': total_products,
            'total_categories': total_categories,
            'total_suppliers': total_suppliers,
            'stock_value': float(stock_value),
            'low_stock_items': stock_stats['low_stock'] or 0,
            'out_of_stock_items': stock_stats['out_of_stock'] or 0,
            'movements_today': movements_today,
            'movements_week': movements_week,
            'total_stock_items': stock_stats['total_items'] or 0
        }

        return JsonResponse(data)

    except Exception as e:
        logger.error(f"Error fetching dashboard stats: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch dashboard statistics'
        }, status=500)


@login_required
@permission_required('inventory.view_stock')
@require_http_methods(["GET"])
def stock_alerts_api(request):
    """
    API endpoint for stock alerts
    """
    try:
        # Get critical and low stock items
        alerts = Stock.objects.filter(
            Q(quantity=0) | Q(quantity__lte=F('low_stock_threshold'))
        ).select_related('product', 'store').order_by('quantity')[:20]

        alerts_data = []
        for alert in alerts:
            status = 'critical' if alert.quantity == 0 or alert.quantity <= (alert.low_stock_threshold / 2) else 'low'

            alerts_data.append({
                'id': alert.id,
                'product_name': alert.product.name,
                'product_sku': alert.product.sku,
                'store_name': alert.store.name,
                'current_stock': float(alert.quantity),
                'threshold': float(alert.low_stock_threshold),
                'status': status,
                'percentage': alert.stock_percentage,
                'unit_of_measure': alert.product.unit_of_measure
            })

        return JsonResponse({
            'success': True,
            'alerts': alerts_data,
            'count': len(alerts_data)
        })

    except Exception as e:
        logger.error(f"Error fetching stock alerts: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch stock alerts'
        }, status=500)


@login_required
@permission_required('inventory.view_stockmovement')
@require_http_methods(["GET"])
def recent_movements_api(request):
    """
    API endpoint for recent stock movements
    """
    try:
        limit = int(request.GET.get('limit', 20))

        movements = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:limit]

        movements_data = []
        for movement in movements:
            movements_data.append({
                'id': movement.id,
                'product_name': movement.product.name,
                'product_sku': movement.product.sku,
                'store_name': movement.store.name,
                'movement_type': movement.movement_type,
                'movement_type_display': movement.get_movement_type_display(),
                'quantity': float(movement.quantity),
                'unit_of_measure': movement.product.unit_of_measure,
                'created_at': movement.created_at.isoformat(),
                'created_by': movement.created_by.get_full_name() or movement.created_by.username if movement.created_by else 'System',
                'reference': movement.reference or '',
                'notes': movement.notes or ''
            })

        return JsonResponse({
            'success': True,
            'movements': movements_data,
            'count': len(movements_data)
        })

    except Exception as e:
        logger.error(f"Error fetching recent movements: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch recent movements'
        }, status=500)


@login_required
@permission_required('inventory.view_product')
@require_http_methods(["GET"])
def top_products_api(request):
    """
    API endpoint for top performing products
    """
    try:
        limit = int(request.GET.get('limit', 10))
        period_days = int(request.GET.get('period_days', 30))

        start_date = timezone.now().date() - timedelta(days=period_days)

        top_products = Product.objects.filter(
            is_active=True
        ).annotate(
            total_movements=Count('movements', filter=Q(
                movements__created_at__date__gte=start_date
            )),
            total_sales=Count('movements', filter=Q(
                movements__movement_type='SALE',
                movements__created_at__date__gte=start_date
            )),
            total_quantity_sold=Coalesce(Sum(
                'movements__quantity',
                filter=Q(
                    movements__movement_type='SALE',
                    movements__created_at__date__gte=start_date
                )
            ), Decimal('0.00'))
        ).filter(
            total_movements__gt=0
        ).order_by('-total_movements')[:limit]

        products_data = []
        for product in top_products:
            products_data.append({
                'id': product.id,
                'name': product.name,
                'sku': product.sku,
                'category': product.category.name if product.category else None,
                'total_movements': product.total_movements,
                'total_sales': product.total_sales,
                'total_quantity_sold': float(product.total_quantity_sold),
                'selling_price': float(product.selling_price),
                'unit_of_measure': product.unit_of_measure
            })

        return JsonResponse({
            'success': True,
            'products': products_data,
            'count': len(products_data)
        })

    except Exception as e:
        logger.error(f"Error fetching top products: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch top products'
        }, status=500)


@login_required
@permission_required('inventory.view_stock')
@require_http_methods(["GET"])
def category_distribution_api(request):
    """
    API endpoint for category distribution data
    """
    try:
        distribution = Stock.objects.values(
            'product__category__name'
        ).annotate(
            category_name=F('product__category__name'),
            total_value=Coalesce(
                Sum(F('quantity') * F('product__cost_price')),
                Decimal('0.00')
            ),
            product_count=Count('product', distinct=True)
        ).filter(
            category_name__isnull=False
        ).order_by('-total_value')[:10]

        categories_data = []
        for cat in distribution:
            categories_data.append({
                'name': cat['category_name'],
                'value': float(cat['total_value']),
                'product_count': cat['product_count']
            })

        return JsonResponse({
            'success': True,
            'categories': categories_data,
            'count': len(categories_data)
        })

    except Exception as e:
        logger.error(f"Error fetching category distribution: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch category distribution'
        }, status=500)


class CategoryListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'
    paginate_by = 20
    permission_required = 'inventory.view_category'
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(code__icontains=search) |
                Q(description__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context

class CategoryDetailView(LoginRequiredMixin,PermissionRequiredMixin, DetailView):
    model = Category
    template_name = 'inventory/category_detail.html'
    context_object_name = 'category'
    permission_required = 'inventory.view_category'

class CategoryCreateView(EFRISConditionalMixin, LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/unified_form.html'
    success_url = reverse_lazy('inventory:category_list')
    permission_required = 'inventory.add_category'

    def get_form_kwargs(self):
        """Pass EFRIS status and request to form"""
        kwargs = super().get_form_kwargs()
        kwargs['efris_enabled'] = getattr(self.request, 'efris', {}).get('enabled', False)
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'category'
        context['show_efris_fields'] = context.get('efris_enabled', False)
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Category created successfully!')
        return super().form_valid(form)

class CategoryUpdateView(EFRISConditionalMixin, LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/unified_form.html'
    success_url = reverse_lazy('inventory:category_list')
    permission_required = 'inventory.change_category'

    def get_form_kwargs(self):
        """Pass EFRIS status and request to form"""
        kwargs = super().get_form_kwargs()
        kwargs['efris_enabled'] = getattr(self.request, 'efris', {}).get('enabled', False)
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'category'
        context['show_efris_fields'] = context.get('efris_enabled', False)
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Category updated successfully!')
        return super().form_valid(form)

class CategoryDeleteView(LoginRequiredMixin,PermissionRequiredMixin, DeleteView):
    model = Category
    template_name = 'inventory/category_confirm_delete.html'
    success_url = reverse_lazy('inventory:category_list')
    permission_required = 'inventory.delete_category'

    def delete(self, request, *args, **kwargs):
        messages.success(request, 'Category deleted successfully!')
        return super().delete(request, *args, **kwargs)



class CategoryDetailAPIView(RetrieveAPIView):
    queryset = Category.objects.all()
    serializer_class = CategoryDetailSerializer
    permission_classes = [IsAuthenticated, DjangoModelPermissions]

    def retrieve(self, request, *args, **kwargs):
        """Override to add efris_commodity_category details"""
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        data = serializer.data

        # Add EFRIS commodity category details if exists
        if instance.efris_commodity_category:
            data['efris_commodity_category'] = {
                'id': instance.efris_commodity_category.id,
                'code': instance.efris_commodity_category.commodity_category_code,
                'name': instance.efris_commodity_category.commodity_category_name,
                'is_exempt': instance.efris_commodity_category.is_exempt,
                'is_leaf_node': instance.efris_commodity_category.is_leaf_node,
                'is_zero_rate': instance.efris_commodity_category.is_zero_rate
            }
        else:
            data['efris_commodity_category'] = None

        return Response(data)


@login_required
@require_GET
def category_detail_api(request, pk):
    """Get category details including EFRIS commodity category."""
    try:
        category = Category.objects.select_related('efris_commodity_category').get(pk=pk)

        data = {
            'id': category.id,
            'name': category.name,
            'code': category.code,
            'description': category.description,
            'is_active': category.is_active,
            'efris_commodity_category': None
        }

        if category.efris_commodity_category:
            data['efris_commodity_category'] = {
                'id': category.efris_commodity_category.id,
                'code': category.efris_commodity_category.commodity_category_code,
                'name': category.efris_commodity_category.commodity_category_name,
                'is_exempt': category.efris_commodity_category.is_exempt,
                'is_leaf_node': category.efris_commodity_category.is_leaf_node,
                'is_zero_rate': category.efris_commodity_category.is_zero_rate
            }

        return JsonResponse(data)

    except Category.DoesNotExist:
        return JsonResponse({'error': 'Category not found'}, status=404)


class SupplierListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    model = Supplier
    template_name = 'inventory/supplier_list.html'
    context_object_name = 'suppliers'
    paginate_by = 20
    permission_required='inventory.view_supplier'
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset()
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) | 
                Q(tin__icontains=search) |
                Q(contact_person__icontains=search) |
                Q(phone__icontains=search)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_query'] = self.request.GET.get('search', '')
        return context


class SupplierCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Supplier
    form_class = SupplierForm
    permission_required = 'inventory.add_supplier'
    template_name = 'inventory/unified_form.html'  # Changed
    success_url = reverse_lazy('inventory:supplier_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'supplier'  # Added
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Supplier created successfully!')
        return super().form_valid(form)


class SupplierUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Supplier
    form_class = SupplierForm
    permission_required = 'inventory.change_supplier'
    template_name = 'inventory/unified_form.html'  # Changed
    success_url = reverse_lazy('inventory:supplier_list')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'supplier'  # Added
        return context

    def form_valid(self, form):
        messages.success(self.request, 'Supplier updated successfully!')
        return super().form_valid(form)


class SupplierDetailView(LoginRequiredMixin,PermissionRequiredMixin, DetailView):
    model = Supplier
    template_name = 'inventory/supplier_detail.html'
    context_object_name = 'supplier'
    permission_required = 'inventory.view_supplier'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['products'] = self.object.products.filter(is_active=True)
        context['total_products'] = context['products'].count()
        return context


class ProductListView(LoginRequiredMixin,PermissionRequiredMixin, ListView):
    model = Product
    template_name = 'inventory/product_list.html'
    context_object_name = 'products'
    permission_required = 'inventory.view_product'
    paginate_by = 25
    ordering = ['name']

    def get_queryset(self):
        queryset = super().get_queryset().select_related('category', 'supplier')

        # Apply filters
        form = ProductFilterForm(self.request.GET)
        if form.is_valid():
            if form.cleaned_data['search']:
                search = form.cleaned_data['search']
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(sku__icontains=search) |
                    Q(barcode__icontains=search) |
                    Q(description__icontains=search)
                )
            if form.cleaned_data['category']:
                queryset = queryset.filter(category=form.cleaned_data['category'])
            if form.cleaned_data['supplier']:
                queryset = queryset.filter(supplier=form.cleaned_data['supplier'])
            if form.cleaned_data['tax_rate']:
                queryset = queryset.filter(tax_rate=form.cleaned_data['tax_rate'])
            if form.cleaned_data['is_active']:
                is_active = form.cleaned_data['is_active'] == 'True'
                queryset = queryset.filter(is_active=is_active)
            if form.cleaned_data['min_price']:
                queryset = queryset.filter(selling_price__gte=form.cleaned_data['min_price'])
            if form.cleaned_data['max_price']:
                queryset = queryset.filter(selling_price__lte=form.cleaned_data['max_price'])

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add filter form with current GET parameters
        context['filter_form'] = ProductFilterForm(self.request.GET)
        context['bulk_form'] = BulkActionForm()

        # Add additional context data that might be useful in the template
        context['total_products'] = self.get_queryset().count()

        return context


class ProductCreateView(EFRISConditionalMixin, LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    permission_required = 'inventory.add_product'
    template_name = 'inventory/unified_form.html'
    success_url = reverse_lazy('inventory:product_list')

    def get_initial(self):
        """Pre-populate form with supplier from query parameter"""
        initial = super().get_initial()
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            try:
                supplier = Supplier.objects.get(id=supplier_id)
                initial['supplier'] = supplier
                logger.info(f"Pre-populated supplier: {supplier.name} (ID: {supplier_id})")
            except Supplier.DoesNotExist:
                logger.warning(f"Supplier with ID {supplier_id} not found")
            except ValueError:
                logger.warning(f"Invalid supplier ID: {supplier_id}")
        return initial

    def get_form_kwargs(self):
        """Pass EFRIS status and company to form"""
        kwargs = super().get_form_kwargs()
        # Get EFRIS status from request (set by middleware)
        kwargs['efris_enabled'] = getattr(self.request, 'efris', {}).get('enabled', False)

        # NEW: Get company and pass to form for VAT handling
        from django_tenants.utils import get_tenant_model
        Company = get_tenant_model()
        try:
            kwargs['company'] = Company.objects.get(schema_name=self.request.tenant.schema_name)
        except (Company.DoesNotExist, AttributeError):
            # Fallback if tenant context not available
            kwargs['company'] = None

        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'product'

        # Add supplier info to context for display
        supplier_id = self.request.GET.get('supplier')
        if supplier_id:
            try:
                supplier = Supplier.objects.get(id=supplier_id)
                context['preselected_supplier'] = supplier
            except (Supplier.DoesNotExist, ValueError):
                pass

        # Add EFRIS status (already added by EFRISConditionalMixin, but being explicit)
        context['show_efris_fields'] = context.get('efris_enabled', False)

        # NEW: Add company VAT status to context
        from django_tenants.utils import get_tenant_model
        Company = get_tenant_model()
        try:
            company = Company.objects.get(schema_name=self.request.tenant.schema_name)
            context['company_vat_enabled'] = company.is_vat_enabled
        except (Company.DoesNotExist, AttributeError):
            context['company_vat_enabled'] = True  # Default to enabled if not found

        return context

    def form_valid(self, form):
        try:
            response = super().form_valid(form)
            logger.info(f"Product created successfully: {form.instance.name} (ID: {form.instance.id})")
            messages.success(self.request, 'Product created successfully!')

            if 'save_and_add_another' in self.request.POST:
                return redirect('inventory:product_create')

            return response
        except Exception as e:
            logger.error(f"Error saving product: {str(e)}")
            messages.error(self.request, 'Error saving product. Please try again.')
            return self.form_invalid(form)

    def form_invalid(self, form):
        logger.error(f"Form validation failed: {form.errors}")
        for field, errors in form.errors.items():
            for error in errors:
                messages.error(self.request, f"{field}: {error}")
        return super().form_invalid(form)


class ProductCreateModalView(EFRISConditionalMixin, LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """
    View to render the product creation form in a popup window
    """
    permission_required = 'inventory.add_product'
    template_name = 'inventory/product_form.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get EFRIS status
        efris_enabled = getattr(self.request, 'efris', {}).get('enabled', False)

        # NEW: Get company for VAT handling
        from django_tenants.utils import get_tenant_model
        Company = get_tenant_model()
        try:
            company = Company.objects.get(schema_name=self.request.tenant.schema_name)
        except (Company.DoesNotExist, AttributeError):
            company = None

        # Create form with EFRIS status and company
        context['form'] = ProductForm(efris_enabled=efris_enabled, company=company)
        context['is_modal'] = True

        # NEW: Add company VAT status to context
        context['company_vat_enabled'] = company.is_vat_enabled if company else True

        return context


class ProductCreateAjaxView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    AJAX view to handle product creation without page reload
    """
    permission_required = 'inventory.add_product'

    def post(self, request, *args, **kwargs):
        # Get EFRIS status
        efris_enabled = getattr(request, 'efris', {}).get('enabled', False)

        # NEW: Get company for VAT handling
        from django_tenants.utils import get_tenant_model
        Company = get_tenant_model()
        try:
            company = Company.objects.get(schema_name=request.tenant.schema_name)
        except (Company.DoesNotExist, AttributeError):
            company = None

        # Create form with EFRIS status and company
        form = ProductForm(request.POST, request.FILES, efris_enabled=efris_enabled, company=company)

        if form.is_valid():
            try:
                with transaction.atomic():
                    product = form.save(commit=False)
                    product.save()

                    logger.info(f"Product created successfully via AJAX: {product.name} (ID: {product.id})")

                    # Return product data including VAT status
                    return JsonResponse({
                        'success': True,
                        'message': 'Product created successfully!',
                        'product': {
                            'id': product.id,
                            'name': product.name,
                            'sku': product.sku,
                            'barcode': product.barcode,
                            'selling_price': str(product.selling_price),
                            'cost_price': str(product.cost_price),
                            'category_id': product.category.id if product.category else None,
                            'category_name': product.category.name if product.category else None,
                            'efris_enabled': product.efris_auto_sync_enabled if efris_enabled else False,
                            'tax_rate': product.tax_rate,
                            'effective_tax_rate': product.effective_tax_rate,  # NEW
                            'company_vat_enabled': company.is_vat_enabled if company else True,  # NEW
                        }
                    })
            except Exception as e:
                logger.error(f"Error creating product via AJAX: {str(e)}")
                return JsonResponse({
                    'success': False,
                    'errors': {'non_field_errors': [str(e)]}
                }, status=400)
        else:
            logger.error(f"Product form validation failed: {form.errors}")
            return JsonResponse({
                'success': False,
                'errors': form.errors
            }, status=400)


class ProductUpdateView(EFRISConditionalMixin, LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    permission_required = 'inventory.change_product'
    template_name = 'inventory/unified_form.html'
    success_url = reverse_lazy('inventory:product_list')

    def get_form_kwargs(self):
        """Pass EFRIS status and company to form"""
        kwargs = super().get_form_kwargs()
        kwargs['efris_enabled'] = getattr(self.request, 'efris', {}).get('enabled', False)

        # NEW: Get company and pass to form for VAT handling
        from django_tenants.utils import get_tenant_model
        Company = get_tenant_model()
        try:
            kwargs['company'] = Company.objects.get(schema_name=self.request.tenant.schema_name)
        except (Company.DoesNotExist, AttributeError):
            # Fallback if tenant context not available
            kwargs['company'] = None

        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'product'
        context['show_efris_fields'] = context.get('efris_enabled', False)

        # NEW: Add company VAT status and effective tax rate to context
        from django_tenants.utils import get_tenant_model
        Company = get_tenant_model()
        try:
            company = Company.objects.get(schema_name=self.request.tenant.schema_name)
            context['company_vat_enabled'] = company.is_vat_enabled
        except (Company.DoesNotExist, AttributeError):
            context['company_vat_enabled'] = True

        context['effective_tax_rate'] = self.object.effective_tax_rate
        context['is_vat_forced'] = not context['company_vat_enabled'] and self.object.tax_rate != 'B'

        return context

    def form_valid(self, form):
        messages.success(self.request, 'Product updated successfully!')
        return super().form_valid(form)

class ProductDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    model = Product
    template_name = 'inventory/product_detail.html'
    context_object_name = 'product'
    permission_required = 'inventory.view_product'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Stock levels from store_inventory relationship
        context['stock_levels'] = self.object.store_inventory.select_related('store').all()

        # Recent stock movements using the correct related_name 'movements'
        context['recent_movements'] = self.object.movements.select_related(
            'store', 'created_by'
        ).order_by('-created_at')[:10]

        # Movement statistics
        context['movement_stats'] = {
            'total_movements': self.object.movements.count(),
            'recent_purchases': self.object.movements.filter(
                movement_type='PURCHASE'
            ).order_by('-created_at')[:5],
            'recent_sales': self.object.movements.filter(
                movement_type='SALE'
            ).order_by('-created_at')[:5],
            'total_purchased': self.object.movements.filter(
                movement_type__in=['PURCHASE', 'RETURN', 'TRANSFER_IN']
            ).aggregate(total=models.Sum('quantity'))['total'] or 0,
            'total_sold': self.object.movements.filter(
                movement_type__in=['SALE', 'TRANSFER_OUT']
            ).aggregate(total=models.Sum('quantity'))['total'] or 0,
        }

        # Product properties from the model
        context['total_stock'] = self.object.total_stock
        context['store_stock_percentages'] = self.object.store_stock_percentages
        context['efris_errors'] = self.object.get_efris_errors()

        # Additional useful context
        context['efris_status'] = self.object.efris_status_display
        context['efris_configuration_complete'] = self.object.efris_configuration_complete
        context['tax_details'] = self.object.tax_details
        context['final_price'] = self.object.final_price
        context['min_stock_level'] = self.object.min_stock_level
        context['stock_percentage'] = self.object.stock_percentage

        if self.object.category:
            context['category'] = self.object.category

        if self.object.supplier:
            context['supplier'] = self.object.supplier

        return context



class ProductDeleteView(LoginRequiredMixin,PermissionRequiredMixin, DeleteView):
    model = Product
    form_class = ProductForm
    template_name = 'inventory/product_confirm_delete.html'
    success_url = reverse_lazy('inventory:product_list')
    permission_required = 'inventory.delete_product'


class StockListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Stock
    template_name = 'inventory/stock_list.html'
    context_object_name = 'stock_items'
    permission_required = 'inventory.view_stock'
    paginate_by = 25
    ordering = ['product__name']

    def get_queryset(self):
        """Enhanced queryset with optimized queries and comprehensive filtering"""
        queryset = super().get_queryset().select_related(
            'product',
            'product__category',
            'product__supplier',
            'store',
            'import_session'
        ).prefetch_related('product__movements')

        filters = self.get_filters()

        queryset = self.apply_status_filter(queryset, filters.get('status'))

        if filters.get('store'):
            queryset = queryset.filter(store_id=filters['store'])

        if filters.get('category'):
            queryset = queryset.filter(product__category_id=filters['category'])

        queryset = self.apply_date_filters(queryset, filters)

        queryset = self.apply_search_filter(queryset, filters.get('search'))

        queryset = self.apply_sorting(queryset, filters.get('sort'))

        return queryset

    def get_filters(self):
        """Extract and validate filters from request"""
        return {
            'status': self.request.GET.get('status', ''),
            'store': self.request.GET.get('store', ''),
            'category': self.request.GET.get('category', ''),
            'search': self.request.GET.get('search', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
            'sort': self.request.GET.get('sort', 'name'),
            'min_value': self.request.GET.get('min_value', ''),
            'max_value': self.request.GET.get('max_value', ''),
        }

    def apply_status_filter(self, queryset, status):
        if status == 'critical':
            return queryset.filter(quantity=0)
        elif status == 'low_stock':
            return queryset.filter(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold')  # Updated field name
            )
        elif status == 'medium_stock':
            return queryset.filter(
                quantity__gt=F('low_stock_threshold'),
                quantity__lte=F('low_stock_threshold') * 2
            )
        elif status == 'good_stock':
            return queryset.filter(quantity__gt=F('low_stock_threshold') * 2)
        elif status == 'overstocked':
            return queryset.filter(quantity__gt=F('low_stock_threshold') * 5)
        return queryset

    def apply_date_filters(self, queryset, filters):
        """Apply date range filtering"""
        date_from = filters.get('date_from')
        date_to = filters.get('date_to')

        if date_from:
            try:
                date_from_parsed = parse_date(date_from)
                if date_from_parsed:
                    queryset = queryset.filter(last_updated__date__gte=date_from_parsed)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_parsed = parse_date(date_to)
                if date_to_parsed:
                    queryset = queryset.filter(last_updated__date__lte=date_to_parsed)
            except ValueError:
                pass

        return queryset

    def apply_search_filter(self, queryset, search_term):
        if not search_term:
            return queryset

        return queryset.filter(
            Q(product__name__icontains=search_term) |
            Q(product__sku__icontains=search_term) |
            Q(product__barcode__icontains=search_term) |
            Q(product__description__icontains=search_term) |
            Q(product__category__name__icontains=search_term) |
            Q(product__supplier__name__icontains=search_term) |
            Q(store__name__icontains=search_term)
        )

    def apply_sorting(self, queryset, sort_field):
        """Apply dynamic sorting"""
        sort_options = {
            'name': 'product__name',
            'name_desc': '-product__name',
            'quantity': 'quantity',
            'quantity_desc': '-quantity',
            'value': 'product__cost_price',
            'value_desc': '-product__cost_price',
            'updated': '-last_updated',
            'updated_asc': 'last_updated',
            'store': 'store__name',
            'store_desc': '-store__name',
            'category': 'product__category__name',
            'category_desc': '-product__category__name',
        }

        if sort_field in sort_options:
            return queryset.order_by(sort_options[sort_field])
        return queryset.order_by('product__name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_filters = self.get_filters()
        context['current_filters'] = current_filters

        context['stores'] = Store.objects.filter(is_active=True).order_by('name')
        context['categories'] = Category.objects.filter(is_active=True).order_by('name')

        # Get dashboard statistics
        context.update(self.get_dashboard_stats())

        # Get stock alerts
        context['stock_alerts'] = self.get_stock_alerts()

        # Get recent movements
        context['recent_movements'] = self.get_recent_movements()

        self.enhance_stock_items(context['stock_items'])

        # Chart data
        context['chart_data'] = self.get_chart_data()

        # Export options
        context['export_formats'] = ['excel', 'csv', 'pdf']

        return context

    def get_dashboard_stats(self):
        cache_key = f"stock_dashboard_stats_{self.request.user.id}"
        stats = cache.get(cache_key)

        if stats is None:
            base_queryset = Stock.objects.select_related('product', 'store')

            # Status counts
            total_products = base_queryset.values('product').distinct().count()
            total_stock_records = base_queryset.count()
            critical_count = base_queryset.filter(quantity=0).count()
            low_stock_count = base_queryset.filter(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold')  # Updated field name
            ).count()
            medium_stock_count = base_queryset.filter(
                quantity__gt=F('low_stock_threshold'),
                quantity__lte=F('low_stock_threshold') * 2
            ).count()
            good_stock_count = base_queryset.filter(
                quantity__gt=F('low_stock_threshold') * 2
            ).count()

            # Value calculations
            stock_values = base_queryset.aggregate(
                total_cost_value=Sum(F('quantity') * F('product__cost_price')),
                total_selling_value=Sum(F('quantity') * F('product__selling_price')),
                avg_stock_level=Avg('quantity')
            )

            stats = {
                'total_products': total_products,
                'total_stock_records': total_stock_records,
                'out_of_stock_count': critical_count,
                'low_stock_count': low_stock_count,
                'medium_stock_count': medium_stock_count,
                'good_stock_count': good_stock_count,
                'total_cost_value': stock_values['total_cost_value'] or Decimal('0.00'),
                'total_selling_value': stock_values['total_selling_value'] or Decimal('0.00'),
                'avg_stock_level': stock_values['avg_stock_level'] or 0,
                'potential_profit': (stock_values['total_selling_value'] or Decimal('0.00')) - (
                            stock_values['total_cost_value'] or Decimal('0.00')),
            }

            # Cache for 5 minutes
            cache.set(cache_key, stats, 300)

        return stats

    def get_stock_alerts(self):
        """Get critical stock alerts"""
        return Stock.objects.select_related(
            'product', 'store'
        ).filter(
            Q(quantity=0) | Q(quantity__lte=F('low_stock_threshold'))  # Updated field name
        ).order_by('quantity', 'product__name')[:10]

    def get_recent_movements(self):
        """Get recent stock movements"""
        return StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:15]

    def enhance_stock_items(self, stock_items):
        """Add computed fields to stock items"""
        for stock in stock_items:
            # Calculate total cost
            stock.total_cost = stock.quantity * stock.product.cost_price

            # Calculate total selling value
            stock.total_selling_value = stock.quantity * stock.product.selling_price

            # Calculate potential profit
            stock.potential_profit = stock.total_selling_value - stock.total_cost

            # Status classification
            if stock.quantity == 0:
                stock.status_class = 'critical'
                stock.status_text = 'Out of Stock'
                stock.status_icon = 'fas fa-times-circle'
            elif stock.quantity <= stock.low_stock_threshold:  # Updated field name
                stock.status_class = 'low'
                stock.status_text = 'Low Stock'
                stock.status_icon = 'fas fa-exclamation-triangle'
            elif stock.quantity <= stock.low_stock_threshold * 2:  # Updated field name
                stock.status_class = 'medium'
                stock.status_text = 'Medium Stock'
                stock.status_icon = 'fas fa-info-circle'
            else:
                stock.status_class = 'good'
                stock.status_text = 'Good Stock'
                stock.status_icon = 'fas fa-check-circle'

    def get_chart_data(self):
        """Prepare data for charts"""
        return {
            'stock_distribution': {
                'labels': ['Critical', 'Low', 'Medium', 'Good'],
                'data': [
                    self.get_dashboard_stats()['out_of_stock_count'],
                    self.get_dashboard_stats()['low_stock_count'],
                    self.get_dashboard_stats()['medium_stock_count'],
                    self.get_dashboard_stats()['good_stock_count'],
                ]
            }
        }


class StockCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create new stock record for a product in a specific store"""
    model = Stock
    form_class = StockForm
    template_name = 'inventory/unified_form.html'
    permission_required = 'inventory.add_stock'
    success_url = reverse_lazy('inventory:stock_list')

    def get_form_kwargs(self):
        """Pass form kwargs properly"""
        kwargs = super().get_form_kwargs()
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'stock'
        context['products'] = Product.objects.filter(is_active=True).select_related('category').order_by('name')
        context['stores'] = Store.objects.filter(is_active=True).order_by('name')

        # Get existing stock records to prevent duplicates
        existing_stock = Stock.objects.select_related('product', 'store').values(
            'product_id', 'store_id', 'product__name', 'store__name'
        )
        context['existing_stock'] = list(existing_stock)

        return context

    def form_valid(self, form):
        """Save stock record with proper validation and logging"""
        try:
            with transaction.atomic():
                stock = form.save(commit=False)

                # Double-check for duplicate (belt and suspenders)
                existing = Stock.objects.filter(
                    product=stock.product,
                    store=stock.store
                ).exists()

                if existing:
                    messages.error(
                        self.request,
                        f'Stock record already exists for {stock.product.name} at {stock.store.name}. '
                        f'Please use the update function instead.'
                    )
                    return self.form_invalid(form)

                # Save the stock record
                stock.save()

                # Log the creation
                logger.info(
                    f"Stock record created: {stock.product.name} at {stock.store.name} "
                    f"with quantity {stock.quantity} by {self.request.user.username}"
                )

                messages.success(
                    self.request,
                    f'Stock record created successfully for {stock.product.name} at {stock.store.name}'
                )

                # Check if user wants to add another
                if 'save_and_add_another' in self.request.POST:
                    return redirect('inventory:stock_create')

                return redirect(self.success_url)

        except Exception as e:
            logger.error(f"Error creating stock record: {str(e)}", exc_info=True)
            messages.error(self.request, f'Error creating stock record: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Handle invalid form submission"""
        logger.error(f"Stock form validation failed: {form.errors}")

        # Add form errors to messages
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, error)
                else:
                    messages.error(self.request, f"{field}: {error}")

        return super().form_invalid(form)


class StockUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update existing stock record"""
    model = Stock
    form_class = StockForm
    template_name = 'inventory/unified_form.html'
    permission_required = 'inventory.change_stock'
    success_url = reverse_lazy('inventory:stock_list')
    
    def get_queryset(self):
        """Ensure user can only access their company's stock"""
        return Stock.objects.filter(store__company=self.request.user.company)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form_type'] = 'stock'
        context['object'] = self.object
        context['products'] = Product.objects.filter(
            store_inventory__store__company=self.request.user.company,
            is_active=True
        ).order_by('name')
        context['stores'] = Store.objects.filter(
            company=self.request.user.company,
            is_active=True
        ).order_by('name')
        context['efris_enabled'] = self.request.user.company.efris_enabled
        context['company'] = self.request.user.company
        
        # Get stock movements for this record
        stock = self.object
        context['recent_movements'] = StockMovement.objects.filter(
            product=stock.product,
            store=stock.store
        ).select_related('created_by').order_by('-created_at')[:10]

        
        # Calculate stock metrics
        context['stock_metrics'] = {
            'days_since_last_count': (
                (timezone.now() - stock.last_physical_count).days
                if stock.last_physical_count else None
            ),
            'variance': getattr(stock, 'variance_from_last_count', None),
            'variance_percentage': getattr(stock, 'variance_percentage', None),
            'stock_status': getattr(stock, 'status', 'Unknown'),
            'is_low_stock': getattr(stock, 'is_low_stock', False),
            'needs_reorder': getattr(stock, 'needs_reorder', False),
        }
        
        return context
    
    def form_valid(self, form):
        """Update stock record with change tracking"""
        try:
            with transaction.atomic():
                # Get old quantity before saving
                old_stock = Stock.objects.get(pk=self.object.pk)
                old_quantity = old_stock.quantity
                
                # Save the form
                stock = form.save(commit=False)
                new_quantity = stock.quantity
                
                # Ensure product and store haven't changed
                stock.product = old_stock.product
                stock.store = old_stock.store
                
                # Track quantity changes
                if old_quantity != new_quantity:
                    quantity_diff = new_quantity - old_quantity
                    stock.efris_sync_required = True
                    
                    # Create a stock movement record for the adjustment
                    StockMovement.objects.create(
                        stock=stock,
                        movement_type='ADJ',  # Adjustment
                        quantity=abs(quantity_diff),
                        reference=f'Stock adjustment: {old_quantity} → {new_quantity}',
                        notes=f'Manual adjustment by {self.request.user.get_full_name() or self.request.user.username}',
                        company=self.request.user.company,
                        created_by=self.request.user
                    )
                    
                    # Log the change
                    logger.info(
                        f"Stock quantity changed for {stock.product.name} at {stock.store.name}: "
                        f"{old_quantity} -> {new_quantity} by {self.request.user.username}"
                    )
                    
                    messages.info(
                        self.request,
                        f'Quantity changed from {old_quantity:.3f} to {new_quantity:.3f}. '
                        f'Stock movement recorded. '
                        f'{"(Increase)" if quantity_diff > 0 else "(Decrease)"}'
                    )
                
                stock.save()
                
                messages.success(
                    self.request,
                    f'✓ Stock record updated successfully for {stock.product.name} at {stock.store.name}'
                )
                
                return redirect(self.success_url)
                
        except Exception as e:
            logger.error(f"Error updating stock record: {str(e)}", exc_info=True)
            messages.error(
                self.request, 
                f'❌ Error updating stock record: {str(e)}'
            )
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        """Handle invalid form submission"""
        logger.error(f"Stock update form validation failed: {form.errors}")
        
        # Add user-friendly error messages
        for field, errors in form.errors.items():
            for error in errors:
                if field == '__all__':
                    messages.error(self.request, f'❌ {error}')
                else:
                    field_label = form.fields.get(field).label if field in form.fields else field
                    messages.error(self.request, f'❌ {field_label}: {error}')
        
        return super().form_invalid(form)


@login_required
@require_http_methods(["GET"])
def product_detail_api(request, pk):
    """API endpoint to get product details"""
    try:
        product = Product.objects.select_related('category', 'supplier').get(pk=pk)

        data = {
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'barcode': product.barcode,
            'unit_of_measure': product.unit_of_measure,
            'cost_price': str(product.cost_price),
            'selling_price': str(product.selling_price),
            'min_stock_level': str(product.min_stock_level),
            'category': {
                'id': product.category.id,
                'name': product.category.name,
            } if product.category else None,
            'supplier': {
                'id': product.supplier.id,
                'name': product.supplier.name,
            } if product.supplier else None,
            'efris_commodity_category': {
                'code': product.efris_commodity_category_id,
                'name': product.efris_commodity_category_name,
            } if product.efris_commodity_category_id else None,
        }

        return JsonResponse(data)

    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@permission_required('inventory.change_stock', raise_exception=True)
@require_http_methods(["POST"])
def stock_physical_count(request, pk):
    """Record a physical stock count"""
    try:
        stock = get_object_or_404(Stock, pk=pk)

        counted_quantity = request.POST.get('counted_quantity')
        notes = request.POST.get('notes', '')

        if not counted_quantity:
            return JsonResponse({
                'success': False,
                'message': 'Counted quantity is required'
            }, status=400)

        try:
            counted_quantity = Decimal(counted_quantity)
        except (ValueError, TypeError):
            return JsonResponse({
                'success': False,
                'message': 'Invalid quantity value'
            }, status=400)

        # Record the physical count
        stock.record_physical_count(
            counted_quantity=counted_quantity,
            user=request.user
        )

        # Log the activity
        messages.success(
            request,
            f'Physical count recorded for {stock.product.name} at {stock.store.name}. '
            f'Previous: {stock.last_physical_count_quantity}, Counted: {counted_quantity}'
        )

        return JsonResponse({
            'success': True,
            'message': 'Physical count recorded successfully',
            'data': {
                'new_quantity': str(stock.quantity),
                'variance': str(stock.variance_from_last_count),
                'variance_percentage': str(stock.variance_percentage) if stock.variance_percentage else None,
            }
        })

    except Stock.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Stock record not found'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error recording physical count: {str(e)}'
        }, status=500)

@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def stock_export(request):
    """Export stock data in various formats"""
    export_format = request.GET.get('format', 'excel')
    filters = {
        'status': request.GET.get('status', ''),
        'store': request.GET.get('store', ''),
        'category': request.GET.get('category', ''),
        'search': request.GET.get('search', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }

    try:
        # Get filtered stock data
        queryset = get_filtered_stock_queryset(filters)

        if export_format == 'excel':
            return export_stock_excel(queryset, filters)
        elif export_format == 'csv':
            return export_stock_csv(queryset, filters)
        elif export_format == 'pdf':
            return export_stock_pdf(queryset, filters)
        else:
            messages.error(request, 'Invalid export format specified.')
            return redirect('inventory:stock_list')

    except Exception as e:
        logger.error(f"Stock export error: {str(e)}", exc_info=True)
        messages.error(request, 'An error occurred during export. Please try again.')
        return redirect('inventory:stock_list')


def get_filtered_stock_queryset(filters):
    """Get filtered stock queryset based on provided filters"""
    queryset = Stock.objects.select_related(
        'product',
        'product__category',
        'product__supplier',
        'store'
    ).order_by('product__name')
    # Apply filters
    if filters.get('status'):
        status = filters['status']
        if status == 'critical':
            queryset = queryset.filter(quantity=0)
        elif status == 'low_stock':
            queryset = queryset.filter(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold')  # Updated field name
            )
        elif status == 'medium_stock':
            queryset = queryset.filter(
                quantity__gt=F('low_stock_threshold'),
                quantity__lte=F('low_stock_threshold') * 2
            )
        elif status == 'good_stock':
            queryset = queryset.filter(quantity__gt=F('low_stock_threshold') * 2)

    if filters.get('store'):
        queryset = queryset.filter(store_id=filters['store'])

    if filters.get('category'):
        queryset = queryset.filter(product__category_id=filters['category'])

    if filters.get('search'):
        search_term = filters['search']
        queryset = queryset.filter(
            Q(product__name__icontains=search_term) |
            Q(product__sku__icontains=search_term) |
            Q(product__barcode__icontains=search_term) |
            Q(store__name__icontains=search_term)
        )

    if filters.get('date_from'):
        try:
            date_from = datetime.strptime(filters['date_from'], '%Y-%m-%d').date()
            queryset = queryset.filter(last_updated__date__gte=date_from)
        except ValueError:
            pass

    if filters.get('date_to'):
        try:
            date_to = datetime.strptime(filters['date_to'], '%Y-%m-%d').date()
            queryset = queryset.filter(last_updated__date__lte=date_to)
        except ValueError:
            pass

    return queryset


def export_stock_excel(queryset, filters):
    """Export stock data to Excel format"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Stock Report')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#667eea',
        'font_color': 'white',
        'border': 1,
        'align': 'center'
    })

    cell_format = workbook.add_format({
        'border': 1,
        'align': 'left'
    })

    number_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'num_format': '#,##0.00'
    })

    currency_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'num_format': '#,##0.00'
    })

    # Headers
    headers = [
        'Product Name', 'SKU', 'Category', 'Store', 'Current Stock',
        'Low Stock Threshold', 'Unit', 'Cost Price', 'Selling Price',  # Updated header
        'Stock Value (Cost)', 'Stock Value (Selling)', 'Status', 'Last Updated'
    ]

    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Write data
    row = 1
    total_cost_value = 0
    total_selling_value = 0

    for stock in queryset:
        cost_value = stock.quantity * stock.product.cost_price
        selling_value = stock.quantity * stock.product.selling_price
        total_cost_value += cost_value
        total_selling_value += selling_value

        # Determine status
        if stock.quantity == 0:
            status = 'Out of Stock'
        elif stock.quantity <= stock.low_stock_threshold:  # Updated field name
            status = 'Low Stock'
        elif stock.quantity <= stock.low_stock_threshold * 2:
            status = 'Medium Stock'
        else:
            status = 'Good Stock'

        data = [
            stock.product.name,
            stock.product.sku,
            stock.product.category.name if stock.product.category else 'N/A',
            stock.store.name,
            stock.quantity,
            stock.low_stock_threshold,  # Updated field name
            stock.product.unit_of_measure,
            stock.product.cost_price,
            stock.product.selling_price,
            cost_value,
            selling_value,
            status,
            stock.last_updated.strftime('%Y-%m-%d %H:%M')
        ]

        for col, value in enumerate(data):
            if col in [4, 5]:  # Quantity columns
                worksheet.write(row, col, float(value), number_format)
            elif col in [7, 8, 9, 10]:  # Price columns
                worksheet.write(row, col, float(value), currency_format)
            else:
                worksheet.write(row, col, value, cell_format)

        row += 1

    # Add summary
    row += 2
    worksheet.write(row, 0, 'Summary', header_format)
    row += 1
    worksheet.write(row, 0, f'Total Records: {queryset.count()}', cell_format)
    row += 1
    worksheet.write(row, 0, f'Total Cost Value: ${total_cost_value:,.2f}', cell_format)
    row += 1
    worksheet.write(row, 0, f'Total Selling Value: ${total_selling_value:,.2f}', cell_format)
    row += 1
    worksheet.write(row, 0, f'Potential Profit: ${total_selling_value - total_cost_value:,.2f}', cell_format)

    # Auto-adjust column widths
    for col, header in enumerate(headers):
        worksheet.set_column(col, col, len(header) + 5)

    workbook.close()
    output.seek(0)

    # Create response
    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'stock_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def export_stock_csv(queryset, filters):
    """Export stock data to CSV format"""
    response = HttpResponse(content_type='text/csv')
    filename = f'stock_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Headers
    writer.writerow([
        'Product Name', 'SKU', 'Category', 'Store', 'Current Stock',
        'Reorder Level', 'Unit', 'Cost Price', 'Selling Price',
        'Stock Value (Cost)', 'Stock Value (Selling)', 'Status', 'Last Updated'
    ])

    # Data
    for stock in queryset:
        cost_value = stock.quantity * stock.product.cost_price
        selling_value = stock.quantity * stock.product.selling_price

        if stock.quantity == 0:
            status = 'Out of Stock'
        elif stock.quantity <= stock.reorder_level:
            status = 'Low Stock'
        elif stock.quantity <= stock.reorder_level * 2:
            status = 'Medium Stock'
        else:
            status = 'Good Stock'

        writer.writerow([
            stock.product.name,
            stock.product.sku,
            stock.product.category.name if stock.product.category else 'N/A',
            stock.store.name,
            float(stock.quantity),
            float(stock.reorder_level),
            stock.product.unit_of_measure,
            float(stock.product.cost_price),
            float(stock.product.selling_price),
            float(cost_value),
            float(selling_value),
            status,
            stock.last_updated.strftime('%Y-%m-%d %H:%M')
        ])

    return response


def export_stock_pdf(queryset, filters):
    """Export stock data to PDF format"""
    response = HttpResponse(content_type='application/pdf')
    filename = f'stock_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )

    story = []
    story.append(Paragraph("Stock Report", title_style))
    story.append(Paragraph(f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}", styles['Normal']))
    story.append(Spacer(1, 20))

    # Create table data
    data = [['Product', 'SKU', 'Store', 'Stock', 'Reorder', 'Status']]

    for stock in queryset[:50]:  # Limit to first 50 records for PDF
        if stock.quantity == 0:
            status = 'Out of Stock'
        elif stock.quantity <= stock.reorder_level:
            status = 'Low Stock'
        elif stock.quantity <= stock.reorder_level * 2:
            status = 'Medium'
        else:
            status = 'Good'

        data.append([
            stock.product.name[:20] + ('...' if len(stock.product.name) > 20 else ''),
            stock.product.sku,
            stock.store.name[:15] + ('...' if len(stock.store.name) > 15 else ''),
            f"{stock.quantity:.2f}",
            f"{stock.reorder_level:.2f}",
            status
        ])

    # Create table
    table = Table(data, colWidths=[2 * inch, 1 * inch, 1.5 * inch, 0.8 * inch, 0.8 * inch, 0.9 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    story.append(table)

    # Add summary
    if queryset.count() > 50:
        story.append(Spacer(1, 20))
        story.append(Paragraph(f"Note: Showing first 50 of {queryset.count()} total records", styles['Italic']))

    doc.build(story)
    return response


@login_required
@permission_required('inventory.add_stock', raise_exception=True)
def stock_import(request):
    """Import stock data from Excel or CSV files"""
    if request.method == 'GET':
        return render(request, 'inventory/stock_import.html')

    elif request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('import_file')
            if not uploaded_file:
                messages.error(request, 'Please select a file to import.')
                return render(request, 'inventory/stock_import.html')

            # Validate file type
            file_extension = uploaded_file.name.split('.')[-1].lower()
            if file_extension not in ['csv', 'xlsx', 'xls']:
                messages.error(request, 'Only CSV and Excel files are supported.')
                return render(request, 'inventory/stock_import.html')

            # Process file
            import_results = process_import_file(
                uploaded_file,
                conflict_resolution='overwrite',  # or 'skip', depending on your logic
                column_mapping=None,  # if no mapping, pass None
                has_header=True,  # True if the file has headers
                user=request.user
            )

            # Show results
            if import_results['success_count'] > 0:
                messages.success(
                    request,
                    f'Successfully imported {import_results["success_count"]} stock records.'
                )

            if import_results['error_count'] > 0:
                messages.warning(
                    request,
                    f'{import_results["error_count"]} records had errors and were skipped.'
                )

            if import_results['errors']:
                request.session['import_errors'] = import_results['errors'][:20]  # Limit to 20 errors

            return render(request, 'inventory/stock_import.html', {
                'import_results': import_results
            })

        except Exception as e:
            logger.error(f"Stock import error: {str(e)}", exc_info=True)
            messages.error(request, f'Import failed: {str(e)}')
            return render(request, 'inventory/stock_import.html')


def process_import_file(uploaded_file, user):
    """Process uploaded CSV or Excel file for stock import"""
    from .models import Stock, Product

    results = {
        'success_count': 0,
        'error_count': 0,
        'errors': []
    }

    try:
        # Read file into DataFrame
        file_extension = uploaded_file.name.split('.')[-1].lower()

        if file_extension == 'csv':
            df = pd.read_csv(uploaded_file)
        else:  # Excel
            df = pd.read_excel(uploaded_file)

        # Validate required columns
        required_columns = ['product_sku', 'store_name', 'quantity']
        missing_columns = [col for col in required_columns if col not in df.columns]

        if missing_columns:
            raise ValidationError(f"Missing required columns: {', '.join(missing_columns)}")

        # Process each row
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Get product
                    try:
                        product = Product.objects.get(sku=row['product_sku'])
                    except Product.DoesNotExist:
                        results['errors'].append(f"Row {index + 2}: Product with SKU '{row['product_sku']}' not found")
                        results['error_count'] += 1
                        continue

                    # Get store
                    try:
                        store = Store.objects.get(name=row['store_name'])
                    except Store.DoesNotExist:
                        results['errors'].append(f"Row {index + 2}: Store '{row['store_name']}' not found")
                        results['error_count'] += 1
                        continue

                    # Validate quantity
                    try:
                        quantity = Decimal(str(row['quantity']))
                        if quantity < 0:
                            results['errors'].append(f"Row {index + 2}: Quantity cannot be negative")
                            results['error_count'] += 1
                            continue
                    except (ValueError, TypeError):
                        results['errors'].append(f"Row {index + 2}: Invalid quantity value")
                        results['error_count'] += 1
                        continue

                    # Get or create stock record
                    stock, created = Stock.objects.get_or_create(
                        product=product,
                        store=store,
                        defaults={
                            'quantity': quantity,
                            'reorder_level': Decimal(str(row.get('low_stock_threshold', 0)))
                        }
                    )

                    if not created:
                        # Update existing stock
                        stock.quantity = quantity
                        if 'reorder_level' in row and pd.notna(row['reorder_level']):
                            stock.reorder_level = Decimal(str(row['reorder_level']))
                        stock.save()

                    results['success_count'] += 1

                except Exception as e:
                    results['errors'].append(f"Row {index + 2}: {str(e)}")
                    results['error_count'] += 1
                    continue

    except Exception as e:
        raise ValidationError(f"File processing error: {str(e)}")

    return results


@login_required
@require_http_methods(["GET"])
def stock_dashboard_data(request):
    """API endpoint for dashboard data"""
    try:
        from .models import Stock, StockMovement
        from django.db.models import Sum, Avg, Count, F

        # Date ranges
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)
        month_ago = today - timedelta(days=30)

        # Basic stock statistics
        stock_stats = Stock.objects.aggregate(
            total_products=Count('product', distinct=True),
            total_stock_records=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            low_stock=Count('id', filter=Q(quantity__gt=0, quantity__lte=F('low_stock_threshold'))),
            total_value=Sum(F('quantity') * F('product__cost_price'))
        )

        # Movement statistics
        movement_stats = {
            'today_movements': StockMovement.objects.filter(created_at__date=today).count(),
            'week_movements': StockMovement.objects.filter(created_at__date__gte=week_ago).count(),
            'month_movements': StockMovement.objects.filter(created_at__date__gte=month_ago).count(),
        }

        stock_by_status = {
            'critical': Stock.objects.filter(quantity=0).count(),
            'low': Stock.objects.filter(quantity__gt=0, quantity__lte=F('low_stock_threshold')).count(),
            'medium': Stock.objects.filter(quantity__gt=F('low_stock_threshold'),
                                           quantity__lte=F('low_stock_threshold') * 2).count(),
            'good': Stock.objects.filter(quantity__gt=F('low_stock_threshold') * 2).count(),
        }

        # Recent movements by day (last 7 days)
        movements_by_day = []
        for i in range(7):
            date = today - timedelta(days=i)
            count = StockMovement.objects.filter(created_at__date=date).count()
            movements_by_day.append({
                'date': date.strftime('%Y-%m-%d'),
                'count': count
            })

        # Top products by movement frequency
        top_products = StockMovement.objects.filter(
            created_at__date__gte=month_ago
        ).values(
            'product__name'
        ).annotate(
            movement_count=Count('id')
        ).order_by('-movement_count')[:5]

        stock_alerts = Stock.objects.select_related('product', 'store').filter(
            Q(quantity=0) | Q(quantity__lte=F('low_stock_threshold'))  # Updated field name
        ).order_by('quantity')[:10]

        alerts_data = []
        for alert in stock_alerts:
            alerts_data.append({
                'product_name': alert.product.name,
                'store_name': alert.store.name,
                'current_stock': float(alert.quantity),
                'reorder_level': float(alert.low_stock_threshold),  # Updated field name
                'status': 'critical' if alert.quantity == 0 else 'low'
            })

        response_data = {
            'stock_stats': stock_stats,
            'movement_stats': movement_stats,
            'stock_by_status': stock_by_status,
            'movements_by_day': list(reversed(movements_by_day)),
            'top_products': list(top_products),
            'stock_alerts': alerts_data
        }

        return JsonResponse(response_data)

    except Exception as e:
        logger.error(f"Dashboard data error: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Failed to load dashboard data'}, status=500)



class StockDashboardView(LoginRequiredMixin, PermissionRequiredMixin, TemplateView):
    """Stock dashboard view"""
    template_name = 'inventory/stock_dashboard.html'
    permission_required = 'inventory.view_stock'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from .models import Stock, StockMovement
        from django.db.models import Count, Sum, F

        # Basic counts for initial display
        context['total_products'] = Stock.objects.values('product').distinct().count()
        context['total_stores'] = Store.objects.filter(is_active=True).count()
        context['out_of_stock'] = Stock.objects.filter(quantity=0).count()
        context['low_stock'] = Stock.objects.filter(quantity__gt=0, quantity__lte=F('low_stock_threshold')).count()

        # Recent movements
        context['recent_movements'] = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:10]

        return context



class StockMovementCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    model = StockMovement
    form_class = StockMovementForm
    permission_required = 'inventory.add_stockmovement'
    template_name = 'inventory/movement_form.html'
    success_url = reverse_lazy('inventory:movement_list')

    def get_form_kwargs(self):
        """Pass additional data to form"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def get_context_data(self, **kwargs):
        """Enhanced context data for movement form"""
        context = super().get_context_data(**kwargs)

        context['stores'] = Store.objects.filter(is_active=True).order_by('name')
        context['products'] = Product.objects.filter(is_active=True).select_related('category').order_by('name')
        context['movement_types'] = StockMovement.MOVEMENT_TYPES

        # Recent movements for context
        context['recent_movements'] = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:10]

        # Movement statistics
        context['movement_stats'] = self.get_movement_stats()

        return context

    def get_movement_stats(self):
        """Get movement statistics for the sidebar"""
        today = timezone.now().date()
        week_ago = today - timedelta(days=7)

        return {
            'today_movements': StockMovement.objects.filter(created_at__date=today).count(),
            'week_movements': StockMovement.objects.filter(created_at__date__gte=week_ago).count(),
            'popular_products': StockMovement.objects.values('product__name').annotate(
                movement_count=Count('id')
            ).order_by('-movement_count')[:5],
        }

    def form_valid(self, form):
        """Enhanced form validation with detailed logging and error handling"""
        try:
            form.instance.created_by = self.request.user

            response = super().form_valid(form)

            # Log the successful movement
            logger.info(
                f"Stock movement created: {self.object.movement_type} - "
                f"{self.object.quantity} of {self.object.product.name} "
                f"at {self.object.store.name} by {self.request.user.username}"
            )

            # Add success message with details
            messages.success(
                self.request,
                f'Stock movement recorded successfully! '
                f'{self.object.get_movement_type_display()} of {self.object.quantity} '
                f'{self.object.product.unit_of_measure} for {self.object.product.name}'
            )

            # Check for low stock warnings after the movement
            self.check_stock_warnings()

            return response

        except Exception as e:
            logger.error(
                f"Error saving stock movement: {str(e)} - "
                f"User: {self.request.user.username}, "
                f"Product: {form.cleaned_data.get('product', 'Unknown')}, "
                f"Store: {form.cleaned_data.get('store', 'Unknown')}",
                exc_info=True
            )

            messages.error(
                self.request,
                'Failed to record stock movement. Please check the details and try again.'
            )
            return self.form_invalid(form)

    def form_invalid(self, form):
        """Enhanced error handling"""
        # Log detailed form errors
        for field, errors in form.errors.items():
            for error in errors:
                logger.warning(f"Form validation error in {field}: {error}")

        # Add user-friendly error message
        messages.error(
            self.request,
            'Please correct the errors highlighted below and try again.'
        )

        return super().form_invalid(form)

    def check_stock_warnings(self):
        """Check and warn about low stock after movement"""
        try:
            stock = Stock.objects.get(
                product=self.object.product,
                store=self.object.store
            )

            if stock.quantity == 0:
                messages.warning(
                    self.request,
                    f'⚠️ {self.object.product.name} is now out of stock at {self.object.store.name}!'
                )
            elif stock.quantity <= stock.low_stock_threshold:
                messages.info(
                    self.request,
                    f'📋 {self.object.product.name} at {self.object.store.name} '
                    f'has reached reorder level ({stock.quantity} remaining)'
                )

        except Stock.DoesNotExist:
            logger.warning(f"No stock record found for {self.object.product} at {self.object.store}")


class StockMovementUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    model = StockMovement
    form_class = StockMovementForm
    permission_required = 'inventory.change_stockmovement'
    template_name = 'inventory/movement_form.html'
    success_url = reverse_lazy('inventory:movement_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            original = StockMovement.objects.get(pk=self.object.pk)

            response = super().form_valid(form)

            # Log the update
            logger.info(
                f"Stock movement updated by {self.request.user.username}: "
                f"ID {self.object.pk} - Changed from {original.quantity} to {self.object.quantity}"
            )

            messages.success(
                self.request,
                f'Stock movement updated successfully!'
            )

            return response

        except Exception as e:
            logger.error(f"Error updating stock movement: {str(e)}", exc_info=True)
            messages.error(self.request, 'Failed to update stock movement.')
            return self.form_invalid(form)

@login_required
@permission_required('inventory.view_product', raise_exception=True)
@csrf_exempt
def product_autocomplete(request):
    """AJAX endpoint for product autocomplete in POS"""
    term = request.GET.get('term', '')

    # Search by name or SKU, only active products
    products = Product.objects.filter(
        Q(name__icontains=term) | Q(sku__icontains=term),
        is_active=True
    )[:10]

    results = []
    for product in products:
        results.append({
            'id': product.id,
            'label': f"{product.name} ({product.sku})",
            'value': product.name,
            'sku': product.sku,
            'category': product.category.name if product.category else '',
            'selling_price': str(product.selling_price),
            'cost_price': str(product.cost_price),
            'stock': float(product.current_stock),  # use .current_stock property
            'unit': product.unit_of_measure,
        })

    return JsonResponse(results, safe=False)

@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def get_product_stock_info(request):
    """AJAX endpoint to get current stock information for a product"""
    product_id = request.GET.get('product_id')
    store_id = request.GET.get('store_id')

    if not product_id or not store_id:
        return JsonResponse({'error': 'Product and store IDs required'}, status=400)

    try:
        product = Product.objects.get(id=product_id, is_active=True)
        store = Store.objects.get(id=store_id, is_active=True)

        stock, created = Stock.objects.get_or_create(
            product=product,
            store=store,
            defaults={'quantity': 0}
        )

        # Get recent movements
        recent_movements = StockMovement.objects.filter(
            product=product,
            store=store
        ).order_by('-created_at')[:5]

        movements_data = [{
            'date': movement.created_at.strftime('%Y-%m-%d %H:%M'),
            'type': movement.get_movement_type_display(),
            'quantity': str(movement.quantity),
            'reference': movement.reference or '',
            'created_by': movement.created_by.get_full_name() or movement.created_by.username
        } for movement in recent_movements]

        return JsonResponse({
            'current_stock': str(stock.quantity),
            'reorder_level': str(stock.low_stock_threshold),  # Updated field name
            'unit_of_measure': product.unit_of_measure,
            'cost_price': str(product.cost_price),
            'stock_status': stock.status,
            'recent_movements': movements_data,
            'last_updated': stock.last_updated.strftime('%Y-%m-%d %H:%M') if stock.last_updated else None
        })

    except (Product.DoesNotExist, Store.DoesNotExist) as e:
        return JsonResponse({'error': 'Product or store not found'}, status=404)
    except Exception as e:
        logger.error(f"Error getting stock info: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Internal server error'}, status=500)


@login_required
@permission_required('inventory.view_product', raise_exception=True)
def get_product_details(request, product_id):
    """AJAX endpoint to get product details"""
    try:
        product = Product.objects.get(id=product_id, is_active=True)

        # Initialize stock_levels list
        stock_levels = []

        # Get stock levels from store_inventory relationship
        for stock in product.store_inventory.select_related('store'):
            stock_levels.append({
                'store_name': stock.store.name,
                'quantity': str(stock.quantity),
                'status': getattr(stock, 'status', 'Available')  # Use getattr in case status field doesn't exist
            })

        data = {
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'selling_price': str(product.selling_price),
            'cost_price': str(product.cost_price),
            'unit_of_measure': product.unit_of_measure,
            'tax_rate': product.get_tax_rate_display(),
            'stock_levels': stock_levels,
            # Additional useful fields from the model
            'total_stock': product.total_stock,
            'min_stock_level': product.min_stock_level,
            'stock_percentage': product.stock_percentage,
            'final_price': str(product.final_price),
            'discount_percentage': str(product.discount_percentage),
        }

        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@login_required
@permission_required('inventory.change_product', raise_exception=True)
def bulk_product_actions(request):
    """Handle bulk actions on products"""
    if request.method == 'POST':
        form = BulkActionForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            item_ids = form.cleaned_data['selected_items']
            
            try:
                products = Product.objects.filter(id__in=item_ids)
                count = products.count()
                
                if action == 'activate':
                    products.update(is_active=True)
                    messages.success(request, f'{count} products activated successfully!')
                
                elif action == 'deactivate':
                    products.update(is_active=False)
                    messages.success(request, f'{count} products deactivated successfully!')
                
                elif action == 'delete':
                    products.delete()
                    messages.success(request, f'{count} products deleted successfully!')
                
                elif action == 'export':
                    return export_products(request, products)
                
            except Exception as e:
                messages.error(request, f'Error performing bulk action: {str(e)}')
        
        else:
            messages.error(request, 'Invalid form data')
    
    return redirect('inventory:product_list')



@login_required
@permission_required('inventory.add_product', raise_exception=True)
def bulk_import_products(request):
    """Main bulk import page"""
    return render(request, 'inventory/bulk_import.html')

@login_required
@csrf_exempt
@permission_required('inventory.add_product', raise_exception=True)
@require_http_methods(["POST"])
def process_bulk_import(request):
    """Process the uploaded file and return import results"""
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
        
        uploaded_file = request.FILES['file']
        import_mode = request.POST.get('import_mode', 'both')
        conflict_resolution = request.POST.get('conflict_resolution', 'overwrite')
        column_mapping = json.loads(request.POST.get('column_mapping', '{}'))
        has_header = request.POST.get('has_header', 'true').lower() == 'true'
        
        # Validate file type
        if not uploaded_file.name.lower().endswith(('.xlsx', '.xls', '.csv')):
            return JsonResponse({'error': 'Unsupported file format'}, status=400)
        
        # Process the file
        result = process_import_file(
            uploaded_file, 
            import_mode, 
            conflict_resolution, 
            column_mapping, 
            has_header,
            request.user
        )
        
        return JsonResponse(result)
    
    except Exception as e:
        logger.error(f"Import error: {str(e)}")
        return JsonResponse({'error': f'Import failed: {str(e)}'}, status=500)


def process_import_file(file, import_mode, conflict_resolution, column_mapping, has_header, user):
    """Process the import file and return results"""
    results = {
        'success': True,
        'total_processed': 0,
        'created': [],
        'updated': [],
        'skipped': [],
        'errors': [],
        'summary': {
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'error_count': 0
        }
    }

    try:
        # Read file data
        if file.name.lower().endswith('.csv'):
            data = read_csv_data(file, has_header)
        else:
            data = read_excel_data(file, has_header)

        if not data:
            results['success'] = False
            results['error'] = 'No data found in file'
            return results

        # Process each row
        with transaction.atomic():
            for row_index, row_data in enumerate(data):
                try:
                    # Map columns based on user selection
                    mapped_data = map_row_data(row_data, column_mapping)

                    # Validate required fields
                    if not validate_row_data(mapped_data):
                        results['errors'].append({
                            'row': row_index + (2 if has_header else 1),
                            'error': 'Missing required fields',
                            'details': 'Product name/SKU, quantity, and store are required'
                        })
                        continue

                    # Process the row
                    row_result = process_single_row(
                        mapped_data,
                        import_mode,
                        conflict_resolution,
                        row_index + (2 if has_header else 1)
                    )

                    # Add to appropriate result category
                    if row_result['status'] == 'created':
                        results['created'].append(row_result['data'])
                        results['summary']['created_count'] += 1
                    elif row_result['status'] == 'updated':
                        results['updated'].append(row_result['data'])
                        results['summary']['updated_count'] += 1
                    elif row_result['status'] == 'skipped':
                        results['skipped'].append(row_result['data'])
                        results['summary']['skipped_count'] += 1

                    results['total_processed'] += 1

                except Exception as e:
                    results['errors'].append({
                        'row': row_index + (2 if has_header else 1),
                        'error': str(e),
                        'details': f'Error processing row data: {row_data}'
                    })

            results['summary']['error_count'] = len(results['errors'])

    except Exception as e:
        results['success'] = False
        results['error'] = str(e)

    return results

def read_csv_data(file, has_header):
    """Read CSV file data"""
    try:
        file.seek(0)
        content = file.read().decode('utf-8')
        lines = content.splitlines()
        reader = csv.reader(lines)

        data = list(reader)
        if has_header and data:
            data = data[1:]  # Skip header row

        return data
    except Exception as e:
        raise Exception(f"Error reading CSV file: {str(e)}")


def read_excel_data(file, has_header):
    """Read Excel file data"""
    try:
        file.seek(0)
        workbook = openpyxl.load_workbook(file)
        worksheet = workbook.active

        data = []
        for row in worksheet.iter_rows(values_only=True):
            if any(cell is not None for cell in row):  # Skip empty rows
                data.append([str(cell) if cell is not None else '' for cell in row])

        if has_header and data:
            data = data[1:]  # Skip header row

        return data
    except Exception as e:
        raise Exception(f"Error reading Excel file: {str(e)}")


def map_row_data(row_data, column_mapping):
    """Map row data to system fields based on column mapping"""
    mapped = {}

    for system_field, column_index in column_mapping.items():
        if column_index != '' and int(column_index) < len(row_data):
            mapped[system_field] = row_data[int(column_index)].strip()

    return mapped


def validate_row_data(data):
    """Validate that required fields are present"""
    required_fields = ['product_name', 'quantity', 'store']

    # Either product_name or sku must be present
    has_product_identifier = data.get('product_name') or data.get('sku')
    has_quantity = data.get('quantity')
    has_store = data.get('store')

    return has_product_identifier and has_quantity and has_store


def process_single_row(data, import_mode, conflict_resolution, row_number):
    """Process a single row of import data"""
    try:
        # Get or create product
        product = get_or_create_product(data)

        # Get or create store
        store = get_or_create_store(data['store'])

        quantity = int(float(data['quantity']))

        try:
            stock, created = Stock.objects.get_or_create(
                product=product,
                store=store,
                defaults={'quantity': quantity}
            )

            if created:
                return {
                    'status': 'created',
                    'data': {
                        'product': product.name,
                        'store': store.name,
                        'quantity': quantity
                    }
                }
            else:
                # Handle existing stock based on conflict resolution
                old_quantity = stock.quantity

                if conflict_resolution == 'skip':
                    return {
                        'status': 'skipped',
                        'data': {
                            'product': product.name,
                            'store': store.name,
                            'reason': 'Item already exists'
                        }
                    }
                elif conflict_resolution == 'overwrite':
                    stock.quantity = quantity
                elif conflict_resolution == 'merge':
                    stock.quantity += quantity

                stock.save()

                return {
                    'status': 'updated',
                    'data': {
                        'product': product.name,
                        'store': store.name,
                        'old_quantity': old_quantity,
                        'new_quantity': stock.quantity
                    }
                }

        except Exception as e:
            raise Exception(f"Error processing stock for {product.name}: {str(e)}")

    except Exception as e:
        raise Exception(f"Error processing row {row_number}: {str(e)}")


@permission_required('inventory.view_product', raise_exception=True)
def get_or_create_product(data):
    """Get or create product based on import data"""

    product = None

    if data.get('sku'):
        try:
            product = Product.objects.get(sku=data['sku'])
        except Product.DoesNotExist:
            pass

    if not product and data.get('product_name'):
        try:
            product = Product.objects.get(name=data['product_name'])
        except Product.DoesNotExist:
            pass

    # Create new product if not found
    if not product:
        # Get or create category if provided
        category = None
        if data.get('category'):
            category, _ = Category.objects.get_or_create(
                name=data['category'],
                defaults={'is_active': True}
            )

        # Get or create supplier if provided
        supplier = None
        if data.get('supplier'):
            supplier, _ = Supplier.objects.get_or_create(
                name=data['supplier'],
                defaults={'is_active': True}
            )

        product_data = {
            'name': data.get('product_name', data.get('sku', 'Unknown Product')),
            'category': category,
            'supplier': supplier,
            'is_active': True
        }

        if data.get('sku'):
            product_data['sku'] = data['sku']
        else:
            # Generate unique SKU if not provided
            product_data['sku'] = f"AUTO-{timezone.now().strftime('%Y%m%d%H%M%S')}"

        if data.get('selling_price'):
            product_data['selling_price'] = Decimal(data['selling_price'])
        else:
            product_data['selling_price'] = Decimal('0.00')

        if data.get('cost_price'):
            product_data['cost_price'] = Decimal(data['cost_price'])
        else:
            product_data['cost_price'] = Decimal('0.00')

        if data.get('description'):
            product_data['description'] = data['description']
        if data.get('unit_of_measure'):
            product_data['unit_of_measure'] = data['unit_of_measure']

        product = Product.objects.create(**product_data)

    return product


def get_or_create_store(store_name):
    """Get or create store based on name"""
    store, created = Store.objects.get_or_create(
        name=store_name,
        defaults={'is_active': True}
    )

    return store
    

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def download_template(request, template_type):
    """Download import templates"""
    if template_type == 'standard':
        return generate_standard_template(request)
    elif template_type == 'simple':
        return generate_simple_template(request)
    else:
        return JsonResponse({'error': 'Invalid template type'}, status=400)

@permission_required('inventory.add_product', raise_exception=True)
def generate_standard_template(request):
    """Generate standard Excel template"""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Product Import Template"

    headers = [
        'Product Name', 'SKU', 'Category', 'Description',
        'Quantity', 'Store', 'Cost Price', 'Selling Price', 'Unit of Measure'
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    sample_data = [
        ['Sample Product 1', 'SKU001', 'Electronics', 'Sample description', '100', 'Main Store', '50.00', '75.00', 'piece'],
        ['Sample Product 2', 'SKU002', 'Clothing', 'Another sample', '50', 'Branch Store', '25.00', '40.00', 'piece'],
    ]

    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            worksheet.cell(row=row, column=col, value=value)

    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="standard_stock_import_template.xlsx"'
    workbook.save(response)
    return response

@permission_required('inventory.add_product', raise_exception=True)
def generate_simple_template(request):
    """Generate simple CSV template"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="simple_stock_import_template.csv"'

    writer = csv.writer(response)
    writer.writerow(['Product Name', 'Quantity', 'Store'])
    writer.writerow(['Sample Product 1', '100', 'Main Store'])
    writer.writerow(['Sample Product 2', '50', 'Branch Store'])

    return response



@login_required
@permission_required('inventory.add_product', raise_exception=True)
def analyze_import_file(request):
    """Analyze uploaded file and return column information"""
    if request.method != 'POST' or 'file' not in request.FILES:
        return JsonResponse({'error': 'No file uploaded'}, status=400)
    
    try:
        uploaded_file = request.FILES['file']
        
        # Read first few rows to analyze structure
        if uploaded_file.name.lower().endswith('.csv'):
            columns = analyze_csv_file(uploaded_file)
        else:
            columns = analyze_excel_file(uploaded_file)
        
        return JsonResponse({
            'success': True,
            'columns': columns,
            'filename': uploaded_file.name,
            'size': uploaded_file.size
        })
    
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


def analyze_csv_file(file):
    """Analyze CSV file structure"""
    file.seek(0)
    content = file.read().decode('utf-8')
    lines = content.splitlines()

    if not lines:
        return []

    reader = csv.reader(lines)
    first_row = next(reader)

    return [{'index': i, 'name': col, 'sample': col} for i, col in enumerate(first_row)]


def analyze_excel_file(file):
    """Analyze Excel file structure"""
    file.seek(0)
    workbook = openpyxl.load_workbook(file)
    worksheet = workbook.active

    columns = []
    first_row = next(worksheet.iter_rows(values_only=True))

    for i, cell_value in enumerate(first_row):
        if cell_value is not None:
            columns.append({
                'index': i,
                'name': str(cell_value),
                'sample': str(cell_value)
            })

    return columns


def export_products(request, queryset=None):
    """Export products to Excel"""
    if queryset is None:
        queryset = Product.objects.select_related('category', 'supplier').all()

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Products'

    # Headers
    headers = [
        'Name', 'SKU', 'Category', 'Supplier', 'Selling Price', 'Cost Price',
        'Discount %', 'Tax Rate', 'Unit of Measure', 'Min Stock Level',
        'Is Active', 'Created At'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data
    for row_num, product in enumerate(queryset, 2):
        worksheet.cell(row=row_num, column=1, value=product.name)
        worksheet.cell(row=row_num, column=2, value=product.sku)
        worksheet.cell(row=row_num, column=3, value=product.category.name if product.category else '')
        worksheet.cell(row=row_num, column=4, value=product.supplier.name if product.supplier else '')
        worksheet.cell(row=row_num, column=5, value=float(product.selling_price))
        worksheet.cell(row=row_num, column=6, value=float(product.cost_price))
        worksheet.cell(row=row_num, column=7, value=float(product.discount_percentage))
        worksheet.cell(row=row_num, column=8, value=product.get_tax_rate_display())
        worksheet.cell(row=row_num, column=9, value=product.unit_of_measure)
        worksheet.cell(row=row_num, column=10, value=product.min_stock_level)
        worksheet.cell(row=row_num, column=11, value='Yes' if product.is_active else 'No')
        worksheet.cell(row=row_num, column=12, value=product.created_at.strftime('%Y-%m-%d'))

    workbook.save(response)
    return response


class StockAdjustmentView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """Enhanced stock adjustment view with batch processing and improved validation"""
    template_name = "inventory/stock_adjustment.html"
    permission_required = 'inventory.add_stockmovement'

    def get(self, request, *args, **kwargs):
        """Handle GET requests with optional pre-filled data"""
        form = StockAdjustmentForm()

        # Pre-fill form if parameters provided
        product_id = request.GET.get('product')
        store_id = request.GET.get('store')

        if product_id and store_id:
            try:
                product = Product.objects.get(id=product_id, is_active=True)
                store = Store.objects.get(id=store_id, is_active=True)
                form.initial.update({
                    'product': product,
                    'store': store
                })
            except (Product.DoesNotExist, Store.DoesNotExist):
                messages.warning(request, 'Invalid product or store specified.')

        context = self._get_base_context()
        context['form'] = form

        return render(request, self.template_name, context)

    def post(self, request, *args, **kwargs):
        """Handle POST requests with batch processing support"""
        batch_mode = request.POST.get('batch_mode') == 'on'

        if batch_mode:
            return self._handle_batch_adjustment(request)
        else:
            return self._handle_single_adjustment(request)

    def _handle_single_adjustment(self, request):
        """Process single product adjustment with notifications"""
        form = StockAdjustmentForm(request.POST, user=request.user)

        if form.is_valid():
            try:
                with transaction.atomic():
                    adjustment = form.save()
                    self._log_adjustment(request.user, adjustment)

                # NOTIFICATION: Stock adjustment completed - only to high-priority users
                from django.contrib.auth import get_user_model
                from notifications.services import NotificationService

                User = get_user_model()
                schema_name = get_current_schema()

                # Get high-priority recipients
                recipients = User.objects.filter(
                    Q(is_superuser=True) |
                    Q(primary_role__priority__gte=90)
                ).filter(is_active=True)

                for recipient in recipients:
                    NotificationService.create_notification(
                        recipient=recipient,
                        title='Stock Adjustment Completed',
                        message=f'Adjusted {adjustment.quantity} units of {adjustment.product.name} at {adjustment.store.name}',
                        notification_type='SUCCESS',
                        priority='MEDIUM',
                        related_object=adjustment,
                        action_text='View Movement',
                        action_url=f'/inventory/movements/',
                        tenant_schema=schema_name
                    )

                # Check for low stock after adjustment and notify high-priority users
                stock = Stock.objects.get(
                    product=adjustment.product,
                    store=adjustment.store
                )

                if stock.quantity <= stock.low_stock_threshold:
                    # Notify high-priority users about low stock
                    for recipient in recipients:
                        if recipient != request.user:  # Don't notify the user who just made the adjustment
                            NotificationService.create_from_template(
                                event_type='low_stock',
                                recipient=recipient,
                                context={
                                    'product_name': stock.product.name,
                                    'current_quantity': stock.quantity,
                                    'threshold': stock.low_stock_threshold,
                                    'store_name': stock.store.name,
                                },
                                related_object=stock,
                                priority='WARNING',
                                tenant_schema=schema_name
                            )

                messages.success(
                    request,
                    f'Stock adjusted successfully for {adjustment.product.name} '
                    f'at {adjustment.store.name}'
                )

                return redirect('inventory:stock_adjustment')

            except Exception as e:
                logger.error(f"Stock adjustment error: {str(e)}", exc_info=True)
                messages.error(request, f'Error processing adjustment: {str(e)}')

        else:
            # Add form errors to messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field.replace('_', ' ').title()}: {error}")

        context = self._get_base_context()
        context['form'] = form
        return render(request, self.template_name, context)

    def _handle_batch_adjustment(self, request):
        """Process batch adjustments for multiple products"""
        try:
            batch_products = request.POST.getlist('batch_products')
            store_id = request.POST.get('store')
            adjustment_type = request.POST.get('adjustment_type')
            quantity = float(request.POST.get('quantity', 0))
            reason = request.POST.get('reason', '')
            notes = request.POST.get('notes', '')

            if not batch_products or not store_id or quantity <= 0:
                messages.error(request, 'Please select products, store, and enter a valid quantity.')
                return self._render_with_errors(request)

            store = get_object_or_404(Store, id=store_id, is_active=True)
            products = Product.objects.filter(id__in=batch_products, is_active=True)

            success_count = 0
            error_count = 0

            with transaction.atomic():
                for product in products:
                    try:
                        # Calculate new quantity based on adjustment type
                        stock, created = Stock.objects.get_or_create(
                            product=product,
                            store=store,
                            defaults={'quantity': 0}
                        )

                        old_quantity = stock.quantity
                        new_quantity = self._calculate_new_quantity(
                            old_quantity, quantity, adjustment_type
                        )

                        # Create stock movement
                        StockMovement.objects.create(
                            product=product,
                            store=store,
                            movement_type='ADJUSTMENT',
                            quantity=new_quantity,
                            reference=f'BATCH-{timezone.now().strftime("%Y%m%d%H%M")}',
                            notes=f'Batch adjustment: {reason}. {notes}',
                            created_by=request.user
                        )

                        # Update stock
                        stock.quantity = new_quantity
                        stock.save()

                        success_count += 1

                    except Exception as e:
                        logger.error(f"Batch adjustment error for {product.name}: {str(e)}")
                        error_count += 1

            if success_count > 0:
                messages.success(
                    request,
                    f'Successfully adjusted {success_count} products.'
                )

            if error_count > 0:
                messages.warning(
                    request,
                    f'Failed to adjust {error_count} products. Check logs for details.'
                )

        except Exception as e:
            logger.error(f"Batch adjustment error: {str(e)}", exc_info=True)
            messages.error(request, f'Error processing batch adjustment: {str(e)}')

        return redirect('inventory:stock_adjustment')

    def _calculate_new_quantity(self, current_qty, adjustment_qty, adj_type):
        """Calculate new quantity based on adjustment type"""
        if adj_type == 'add':
            return current_qty + adjustment_qty
        elif adj_type == 'remove':
            return max(0, current_qty - adjustment_qty)
        elif adj_type == 'set':
            return adjustment_qty
        else:
            raise ValueError(f"Invalid adjustment type: {adj_type}")

    def _log_adjustment(self, user, adjustment):
        """Log adjustment for audit trail"""
        logger.info(
            f"Stock adjustment by {user.username}: "
            f"{adjustment.product.name} at {adjustment.store.name} "
            f"- {adjustment.movement_type}: {adjustment.quantity}"
        )

    def _get_base_context(self):
        """Get base context data for template"""
        return {
            'products': Product.objects.filter(is_active=True)
            .select_related('category')
            .order_by('name'),
            'stores': Store.objects.filter(is_active=True).order_by('name'),
            'recent_adjustments': self._get_recent_adjustments()
        }

    def _get_recent_adjustments(self, limit=10):
        """Get recent stock adjustments for reference"""
        return StockMovement.objects.filter(
            movement_type='ADJUSTMENT'
        ).select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:limit]

    def _render_with_errors(self, request):
        """Render template with error context"""
        context = self._get_base_context()
        context['form'] = StockAdjustmentForm()
        return render(request, self.template_name, context)


def stock_api_view(request, stock_id=None):
    """API endpoint for stock data"""
    if request.method == 'GET':
        if stock_id:
            try:
                stock = Stock.objects.select_related('product', 'store').get(id=stock_id)
                data = {
                    'id': stock.id,
                    'product_name': stock.product.name,
                    'store_name': stock.store.name,
                    'current_stock': float(stock.quantity),
                    'reorder_level': float(stock.low_stock_threshold),
                    'unit_of_measure': stock.product.unit_of_measure,
                    'cost_price': float(stock.product.cost_price),
                    'selling_price': float(stock.product.selling_price),
                }
                return JsonResponse(data)
            except Stock.DoesNotExist:
                return JsonResponse({'error': 'Stock not found'}, status=404)
        else:
            # List stocks with filtering
            stocks = Stock.objects.select_related('product', 'store')

            # Apply filters
            product_id = request.GET.get('product')
            store_id = request.GET.get('store')

            if product_id:
                stocks = stocks.filter(product_id=product_id)
            if store_id:
                stocks = stocks.filter(store_id=store_id)

            results = []
            for stock in stocks:
                results.append({
                    'id': stock.id,
                    'product_name': stock.product.name,
                    'store_name': stock.store.name,
                    'quantity': float(stock.quantity),
                    'reorder_level': float(stock.low_stock_threshold),
                })

            return JsonResponse({'results': results})

    return JsonResponse({'error': 'Method not allowed'}, status=405)


@require_GET
def current_stock_api(request):
    """Return current stock for given product and store."""
    product_id = request.GET.get('product')
    store_id = request.GET.get('store')

    if not product_id or not store_id:
        return JsonResponse({"current_stock": 0})

    stock = Stock.objects.filter(product_id=product_id, store_id=store_id).first()
    return JsonResponse({"current_stock": float(stock.quantity) if stock else 0})

def product_api_view(request, product_id):
    """API endpoint for product data"""
    try:
        product = Product.objects.select_related('category').get(id=product_id)
        data = {
            'id': product.id,
            'name': product.name,
            'sku': product.sku,
            'unit_of_measure': product.unit_of_measure,
            'cost_price': float(product.cost_price),
            'selling_price': float(product.selling_price),
            'category': {
                'id': product.category.id if product.category else None,
                'name': product.category.name if product.category else None,
            } if product.category else None,
        }
        return JsonResponse(data)
    except Product.DoesNotExist:
        return JsonResponse({'error': 'Product not found'}, status=404)

@login_required
@permission_required('inventory.view_stockmovement', raise_exception=True)
def recent_adjustments_api(request):
    """API endpoint for recent adjustments data"""
    try:
        adjustments = StockMovement.objects.filter(
            movement_type='ADJUSTMENT'
        ).select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:20]

        data = [{
            'id': adj.id,
            'product_name': adj.product.name,
            'product_sku': adj.product.sku,
            'store_name': adj.store.name,
            'quantity': str(adj.quantity),
            'unit': adj.product.unit_of_measure,
            'created_at': adj.created_at.strftime('%Y-%m-%d %H:%M'),
            'created_by': adj.created_by.get_full_name() or adj.created_by.username,
            'reference': adj.reference or '',
            'notes': adj.notes or ''
        } for adj in adjustments]

        return JsonResponse({'adjustments': data})

    except Exception as e:
        logger.error(f"Error fetching recent adjustments: {str(e)}", exc_info=True)
        return JsonResponse({'error': 'Failed to fetch data'}, status=500)


@require_GET
def recent_adjustments_api(request):
    """
    Return last 10 stock adjustments for display in history panel.
    """
    adjustments = StockMovement.objects.select_related('product', 'store', 'user').order_by('-created_at')[:10]
    data = [{
        "created_at": adj.created_at.isoformat(),
        "product_name": adj.product.name,
        "store_name": adj.store.name if adj.store else "N/A",
        "adjustment_type": adj.adjustment_type,
        "quantity": adj.quantity,
        "reason": adj.reason,
        "user": adj.user.get_full_name() or adj.user.username
    } for adj in adjustments]
    return JsonResponse({"adjustments": data})


@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def low_stock_report(request):
    """Enhanced low stock report with filtering, analytics, and export options"""

    # Get filter parameters
    category_id = request.GET.get('category')
    store_id = request.GET.get('store')
    severity = request.GET.get('severity', 'all')  # all, critical, low, out_of_stock
    format_type = request.GET.get('format')

    base_queryset = (
        Stock.objects.select_related(
            "product",
            "product__category",
            "store"
        )
        .annotate(
            total_cost=F("quantity") * F("product__cost_price"),
            reorder_gap=F("low_stock_threshold") - F("quantity"),  # Updated field name
            half_threshold=ExpressionWrapper(
                F("low_stock_threshold") / 2.0,  # Updated field name
                output_field=FloatField()
            ),

            # Stock status classification
            stock_status=Case(
                When(quantity=0, then=Value("out_of_stock")),
                When(quantity__lte=F("low_stock_threshold") / 2.0, then=Value("critical")),  # Updated field name
                When(quantity__lte=F("low_stock_threshold"), then=Value("low_stock")),  # Updated field name
                default=Value("in_stock"),
                output_field=CharField(),
            ),
        )
        .filter(
            Q(quantity__lte=F("low_stock_threshold")) | Q(quantity=0)  # Updated field name
        )
        .order_by("quantity", "reorder_gap")
    )
    # Apply filters
    if category_id:
        base_queryset = base_queryset.filter(product__category_id=category_id)

    if store_id:
        base_queryset = base_queryset.filter(store_id=store_id)

    if severity != 'all':
        if severity == 'out_of_stock':
            base_queryset = base_queryset.filter(quantity=0)
        elif severity == 'critical':
            base_queryset = base_queryset.filter(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold') / 2.0  # Updated field name
            )
        elif severity == 'low':
            base_queryset = base_queryset.filter(
                quantity__gt=F('low_stock_threshold') / 2.0,  # Updated field name
                quantity__lte=F('low_stock_threshold')  # Updated field name
            )

    # Get the low stock items
    low_stock_items = base_queryset

    # Calculate summary statistics
    summary_stats = _calculate_low_stock_summary(low_stock_items)

    # Get category breakdown
    category_breakdown = _get_low_stock_category_breakdown(low_stock_items)

    # Get store breakdown
    store_breakdown = _get_low_stock_store_breakdown(low_stock_items)

    # Get most critical items (lowest stock relative to reorder level)
    most_critical = low_stock_items.filter(quantity__gt=0).order_by('quantity')[:10]

    # Get recently moved low stock items
    recent_movements = _get_recent_low_stock_movements(low_stock_items)

    # Handle export formats
    if format_type in ['excel', 'csv', 'pdf']:
        return _export_low_stock_report(
            low_stock_items, summary_stats, format_type, request
        )

    # Pagination for web view
    paginator = Paginator(low_stock_items, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Convert queryset to list with computed fields for template
    items_with_computations = []
    for item in page_obj:
        items_with_computations.append({
            'stock': item,
            'total_cost': item.total_cost,
            'reorder_gap': item.reorder_gap,
            'half_reorder': item.low_stock_threshold,
            'stock_percentage': _calculate_stock_percentage(item.quantity, item.low_stock_threshold),
            'days_until_stockout': _estimate_days_until_stockout(item),
            'recommended_order_qty': _calculate_recommended_order_quantity(item)
        })

    # Create new page object with computed items
    page_obj.object_list = items_with_computations

    context = {
        'low_stock_items': items_with_computations,
        'page_obj': page_obj,
        'total_items': low_stock_items.count(),
        'summary_stats': summary_stats,
        'category_breakdown': category_breakdown,
        'store_breakdown': store_breakdown,
        'most_critical': most_critical,
        'recent_movements': recent_movements,
        'categories': Category.objects.filter(is_active=True),
        'stores': Store.objects.filter(is_active=True),
        'selected_category': category_id,
        'selected_store': store_id,
        'selected_severity': severity,
        'severity_choices': [
            ('all', 'All Severities'),
            ('out_of_stock', 'Out of Stock'),
            ('critical', 'Critical'),
            ('low_stock', 'Low Stock')
        ],
        'chart_data': _prepare_low_stock_chart_data(category_breakdown, store_breakdown),
        'recommendations': _generate_recommendations(summary_stats, category_breakdown),
    }

    return render(request, 'inventory/low_stock_report.html', context)


def _calculate_low_stock_summary(queryset):
    """Calculate comprehensive summary statistics for low stock items"""
    summary = queryset.aggregate(
        total_items=Count('id'),
        total_value_at_risk=Coalesce(Sum('total_cost'), Decimal('0.00')),
        out_of_stock_count=Count('id', filter=Q(quantity=0)),
        critical_count=Count('id', filter=Q(quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0)),
        low_stock_count=Count('id', filter=Q(quantity__gt=F('low_stock_threshold') / 2.0, quantity__lte=F('low_stock_threshold'))),
        avg_reorder_gap=Coalesce(Sum('reorder_gap') / Count('id'), Decimal('0.00')),
    )

    # Calculate additional metrics
    summary['critical_percentage'] = (
                                             (summary['critical_count'] + summary['out_of_stock_count']) / max(
                                         summary['total_items'], 1)
                                     ) * 100

    return summary


def _get_low_stock_category_breakdown(queryset):
    """Get low stock breakdown by category"""
    return list(
        queryset.values(
            'product__category__name'
        ).annotate(
            name=F('product__category__name'),
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            critical=Count('id', filter=Q(quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0)),
            low_stock=Count('id', filter=Q(quantity__gt=F('low_stock_threshold') / 2.0, quantity__lte=F('low_stock_threshold'))),
            total_value=Coalesce(Sum('total_cost'), Decimal('0.00'))
        ).filter(
            name__isnull=False
        ).order_by('-total_items')
    )


def _get_low_stock_store_breakdown(queryset):
    """Get low stock breakdown by store"""
    return list(
        queryset.values(
            'store__name'
        ).annotate(
            name=F('store__name'),
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            critical=Count('id', filter=Q(quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0)),
            low_stock=Count('id', filter=Q(quantity__gt=F('low_stock_threshold') / 2.0, quantity__lte=F('low_stock_threshold'))),
            total_value=Coalesce(Sum('total_cost'), Decimal('0.00'))
        ).order_by('-total_items')
    )


def _get_recent_low_stock_movements(low_stock_queryset):
    """Get recent stock movements for low stock items"""
    product_ids = low_stock_queryset.values_list('product_id', flat=True)

    return StockMovement.objects.filter(
        product_id__in=product_ids
    ).select_related(
        'product', 'store', 'created_by'
    ).order_by('-created_at')[:20]


def _calculate_stock_percentage(quantity, threshold):  # Updated parameter name
    """Calculate stock level as percentage of threshold"""
    if not threshold or threshold <= 0:
        return 100

    percentage = (quantity / threshold) * 100
    return min(100, max(0, round(percentage, 1)))


def _estimate_days_until_stockout(stock_item):
    """Estimate days until stockout based on recent consumption"""
    try:
        # Get average daily consumption from last 30 days
        thirty_days_ago = timezone.now() - timezone.timedelta(days=30)

        recent_sales = StockMovement.objects.filter(
            product=stock_item.product,
            store=stock_item.store,
            movement_type__in=['SALE', 'TRANSFER_OUT'],
            created_at__gte=thirty_days_ago
        ).aggregate(
            total_consumed=Coalesce(Sum('quantity'), Decimal('0.00'))
        )['total_consumed']

        if recent_sales > 0:
            daily_consumption = recent_sales / 30
            if daily_consumption > 0:
                return int(stock_item.quantity / daily_consumption)

        return None  # Cannot estimate

    except Exception:
        return None


def _calculate_recommended_order_quantity(stock_item):
    """Calculate recommended order quantity"""
    # Bring stock to 150% of threshold level
    target_level = stock_item.low_stock_threshold * Decimal("1.5")  # Updated field name
    current_qty = stock_item.quantity

    recommended = max(0, target_level - current_qty)
    return recommended.quantize(Decimal("0.01"))


def _prepare_low_stock_chart_data(category_breakdown, store_breakdown):
    """Prepare data for charts"""
    return {
        'categories': {
            'labels': [item['name'] or 'No Category' for item in category_breakdown[:10]],
            'values': [item['total_items'] for item in category_breakdown[:10]],
            'critical': [item['out_of_stock'] + item['critical'] for item in category_breakdown[:10]]
        },
        'stores': {
            'labels': [item['name'] for item in store_breakdown],
            'values': [item['total_items'] for item in store_breakdown],
            'critical': [item['out_of_stock'] + item['critical'] for item in store_breakdown]
        }
    }


def _generate_recommendations(summary_stats, category_breakdown):
    """Generate actionable recommendations based on low stock analysis"""
    recommendations = []

    if summary_stats['out_of_stock_count'] > 0:
        recommendations.append({
            'priority': 'high',
            'title': 'Immediate Action Required',
            'message': f"{summary_stats['out_of_stock_count']} items are completely out of stock and need immediate restocking.",
            'action': 'Review out-of-stock items and place urgent orders'
        })

    if summary_stats['critical_percentage'] > 50:
        recommendations.append({
            'priority': 'medium',
            'title': 'High Critical Stock Ratio',
            'message': f"{summary_stats['critical_percentage']:.1f}% of low stock items are at critical levels.",
            'action': 'Consider increasing reorder levels or order frequencies'
        })

    # Category-specific recommendations
    critical_categories = [cat for cat in category_breakdown if
                           cat['out_of_stock'] + cat['critical'] > cat['total_items'] * 0.7]

    if critical_categories:
        cat_names = ', '.join([cat['name'] or 'Uncategorized' for cat in critical_categories[:3]])
        recommendations.append({
            'priority': 'medium',
            'title': 'Categories Need Attention',
            'message': f"Categories with high critical ratios: {cat_names}",
            'action': 'Review supplier relationships and lead times for these categories'
        })

    if summary_stats['total_value_at_risk'] > 10000:
        recommendations.append({
            'priority': 'low',
            'title': 'High Value at Risk',
            'message': f"${summary_stats['total_value_at_risk']:,.2f} in inventory value is at risk of stockout.",
            'action': 'Prioritize high-value items for restocking'
        })

    return recommendations


def _export_low_stock_report(low_stock_items, summary_stats, format_type, request):
    """Export low stock report in various formats"""

    if format_type == 'excel':
        return _export_low_stock_excel(low_stock_items, summary_stats)
    elif format_type == 'csv':
        return _export_low_stock_csv(low_stock_items, summary_stats)
    elif format_type == 'pdf':
        return _export_low_stock_pdf(low_stock_items, summary_stats)

    return JsonResponse({'error': 'Invalid format'}, status=400)


def _export_low_stock_excel(low_stock_items, summary_stats):
    """Export low stock report to Excel"""
    wb = Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    # Summary data
    summary_data = [
        ['Low Stock Report Summary', ''],
        ['Generated On', timezone.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['', ''],
        ['Total Low Stock Items', summary_stats['total_items']],
        ['Out of Stock Items', summary_stats['out_of_stock_count']],
        ['Critical Items', summary_stats['critical_count']],
        ['Low Stock Items', summary_stats['low_stock_count']],
        ['Total Value at Risk', f"${summary_stats['total_value_at_risk']:,.2f}"],
        ['Critical Percentage', f"{summary_stats['critical_percentage']:.1f}%"],
    ]

    for row, (label, value) in enumerate(summary_data, 1):
        ws_summary.cell(row=row, column=1, value=label)
        ws_summary.cell(row=row, column=2, value=value)

        if row == 1:  # Title row
            ws_summary.cell(row=row, column=1).font = Font(size=16, bold=True)
        elif row > 3:  # Data rows
            ws_summary.cell(row=row, column=1).font = Font(bold=True)

    # Detail sheet
    ws_detail = wb.create_sheet("Low Stock Items")

    # Headers
    headers = [
        'Product', 'SKU', 'Category', 'Store', 'Current Stock',
        'Reorder Level', 'Reorder Gap', 'Stock %', 'Unit',
        'Cost Price', 'Total Value', 'Status', 'Recommended Order'
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="D32F2F", end_color="D32F2F", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws_detail.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows
    for row, item in enumerate(low_stock_items, 2):
        stock_percentage = _calculate_stock_percentage(item.quantity, item.low_stock_threshold)
        recommended_qty = _calculate_recommended_order_quantity(item)

        # Determine status
        if item.quantity == 0:
            status = "OUT OF STOCK"
        elif item.quantity <= item.low_stock_threshold / 2:
            status = "CRITICAL"
        else:
            status = "LOW STOCK"

        row_data = [
            item.product.name,
            item.product.sku,
            item.product.category.name if item.product.category else 'No Category',
            item.store.name,
            float(item.quantity),
            float(item.low_stock_threshold),
            float(item.reorder_gap),
            f"{stock_percentage}%",
            item.product.unit_of_measure,
            float(item.product.cost_price),
            float(item.total_cost),
            status,
            float(recommended_qty)
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws_detail.cell(row=row, column=col, value=value)

            # Color coding based on status
            if status == "OUT OF STOCK":
                cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
            elif status == "CRITICAL":
                cell.fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")

    # Auto-adjust column widths
    for ws in [ws_summary, ws_detail]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="low_stock_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response


def _export_low_stock_csv(low_stock_items, summary_stats):
    """Export low stock report to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="low_stock_report_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)

    # Summary
    writer.writerow(['Low Stock Report Summary'])
    writer.writerow(['Generated On', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])
    writer.writerow(['Total Low Stock Items', summary_stats['total_items']])
    writer.writerow(['Out of Stock Items', summary_stats['out_of_stock_count']])
    writer.writerow(['Critical Items', summary_stats['critical_count']])
    writer.writerow(['Total Value at Risk', f"${summary_stats['total_value_at_risk']:,.2f}"])
    writer.writerow([])

    # Headers
    writer.writerow([
        'Product', 'SKU', 'Category', 'Store', 'Current Stock',
        'Reorder Level', 'Reorder Gap', 'Stock %', 'Unit',
        'Cost Price', 'Total Value', 'Status'
    ])

    # Data rows
    for item in low_stock_items:
        stock_percentage = _calculate_stock_percentage(item.quantity, item.low_stock_threshold)

        # Determine status
        if item.quantity == 0:
            status = "OUT OF STOCK"
        elif item.quantity <= item.low_stock_threshold / 2:
            status = "CRITICAL"
        else:
            status = "LOW STOCK"

        writer.writerow([
            item.product.name,
            item.product.sku,
            item.product.category.name if item.product.category else 'No Category',
            item.store.name,
            item.quantity,
            item.low_stock_threshold,
            item.reorder_gap,
            f"{stock_percentage}%",
            item.product.unit_of_measure,
            item.product.cost_price,
            item.total_cost,
            status,
        ])

    return response


def _export_low_stock_pdf(low_stock_items, summary_stats):
    """Export low stock report to PDF"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,
        spaceAfter=30,
    )

    elements = []

    # Title
    title = Paragraph("Low Stock Report", title_style)
    elements.append(title)

    # Date
    date_text = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary
    summary_data = [
        ['Total Low Stock Items', f"{summary_stats['total_items']:,}"],
        ['Out of Stock', f"{summary_stats['out_of_stock_count']:,}"],
        ['Critical Level', f"{summary_stats['critical_count']:,}"],
        ['Low Stock Level', f"{summary_stats['low_stock_count']:,}"],
        ['Value at Risk', f"${summary_stats['total_value_at_risk']:,.2f}"],
        ['Critical Percentage', f"{summary_stats['critical_percentage']:.1f}%"],
    ]

    summary_table = Table(summary_data, colWidths=[3 * 72, 2 * 72])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Most critical items
    elements.append(Paragraph("Most Critical Items", styles['Heading2']))
    elements.append(Spacer(1, 10))

    detail_data = [['Product', 'Store', 'Current', 'Reorder', 'Status']]

    for item in low_stock_items[:30]:
        if item.quantity == 0:
            status = "OUT OF STOCK"
        elif item.quantity <= item.low_stock_threshold / 2:
            status = "CRITICAL"
        else:
            status = "LOW STOCK"

        detail_data.append([
            item.product.name[:25] + '...' if len(item.product.name) > 25 else item.product.name,
            item.store.name[:15] + '...' if len(item.store.name) > 15 else item.store.name,
            f"{item.quantity:.1f}",
            f"{item.low_stock_threshold:.1f}",
            status
        ])

    detail_table = Table(detail_data, colWidths=[2.5 * 72, 1.5 * 72, 1 * 72, 1 * 72, 1.2 * 72])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(detail_table)

    # Build PDF
    doc.build(elements)

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="low_stock_report_{timezone.now().strftime("%Y%m%d")}.pdf"'

    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()

    return response


@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def inventory_valuation_report(request):
    """Enhanced inventory valuation report with filtering and analytics"""
    import json
    from decimal import Decimal
    from django.db.models import F, Sum, Count, Avg, Max
    from django.db.models.functions import Coalesce

    # Get filter parameters
    category_id = request.GET.get('category')
    store_id = request.GET.get('store')
    period = request.GET.get('period', 'current')
    format_type = request.GET.get('format')

    # Build base queryset
    queryset = Stock.objects.select_related(
        'product',
        'product__category',
        'store'
    ).annotate(
        total_cost=F('quantity') * F('product__cost_price'),
        total_selling=F('quantity') * F('product__selling_price')
    ).filter(quantity__gt=0)

    # Apply filters
    if category_id:
        queryset = queryset.filter(product__category_id=category_id)

    if store_id:
        queryset = queryset.filter(store_id=store_id)

    # Get valuation items
    valuation_items = queryset.order_by('-total_cost')

    # Calculate totals and metrics
    totals = valuation_items.aggregate(
        total_cost_value=Coalesce(Sum('total_cost'), Decimal('0.00')),
        total_selling_value=Coalesce(Sum('total_selling'), Decimal('0.00')),
        total_items=Count('id'),
        avg_item_value=Coalesce(Avg('total_cost'), Decimal('0.00')),
        max_item_value=Coalesce(Max('total_cost'), Decimal('0.00'))
    )

    # Calculate potential profit
    potential_profit = totals['total_selling_value'] - totals['total_cost_value']

    # Get value change (comparing with previous period)
    previous_value = _get_previous_period_value(period, category_id, store_id)
    value_change = totals['total_cost_value'] - previous_value
    value_change_percent = _calculate_percentage_change(
        totals['total_cost_value'], previous_value
    )

    # Category breakdown for chart and display
    category_breakdown = list(
        queryset.values(
            'product__category__name'
        ).annotate(
            name=F('product__category__name'),
            value=Coalesce(Sum('total_cost'), Decimal('0.00')),
            count=Count('id'),
            avg_value=Coalesce(Avg('total_cost'), Decimal('0.00'))
        ).filter(
            name__isnull=False
        ).order_by('-value')
    )

    # Prepare category data for charts (JSON serializable)
    category_labels = [cat['name'] or 'Uncategorized' for cat in category_breakdown[:10]]
    category_values = [float(cat['value']) for cat in category_breakdown[:10]]

    # Store breakdown
    store_breakdown = list(
        queryset.values(
            'store__name'
        ).annotate(
            name=F('store__name'),
            value=Coalesce(Sum('total_cost'), Decimal('0.00')),
            count=Count('id'),
            avg_value=Coalesce(Avg('total_cost'), Decimal('0.00'))
        ).order_by('-value')
    )

    # Get top products by value
    top_products = valuation_items.order_by('-total_cost')[:10]

    # Top categories (first 5)
    top_categories = category_breakdown[:5]

    # Trend data (last 6 months - simplified)
    trend_labels = _get_trend_labels()
    trend_values = _get_trend_values(category_id, store_id, float(totals['total_cost_value']))

    # Handle export formats
    if format_type in ['excel', 'csv', 'pdf']:
        return _export_valuation_report(
            valuation_items, totals, format_type, request
        )

    # Pagination for web view
    paginator = Paginator(valuation_items, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        # Main data
        'valuation_items': page_obj,
        'page_obj': page_obj,

        # Totals
        'total_value': totals['total_cost_value'],
        'total_selling_value': totals['total_selling_value'],
        'potential_profit': potential_profit,
        'total_items': totals['total_items'],
        'avg_item_value': totals['avg_item_value'],
        'max_item_value': totals['max_item_value'],

        # Changes
        'value_change': value_change,
        'value_change_percent': value_change_percent,

        # Breakdowns
        'category_breakdown': category_breakdown,
        'store_breakdown': store_breakdown,
        'top_products': top_products,
        'top_categories': top_categories,

        # Chart data (JSON safe)
        'category_labels': json.dumps(category_labels),
        'category_values': json.dumps(category_values),
        'trend_labels': json.dumps(trend_labels),
        'trend_values': json.dumps(trend_values),

        # Report metadata
        'report_date': timezone.now(),
        'report_period': period.replace('_', ' ').title() if period else 'Current',

        # Filters
        'categories': Category.objects.filter(is_active=True).order_by('name'),
        'stores': Store.objects.filter(is_active=True).order_by('name'),
        'selected_category': category_id,
        'selected_store': store_id,
    }

    return render(request, 'inventory/valuation_report.html', context)


def _get_previous_period_value(period, category_id=None, store_id=None):
    """Calculate inventory value for previous period"""
    try:
        past_date = timezone.now() - timedelta(days=30)

        queryset = Stock.objects.filter(
            last_updated__gte=past_date
        ).annotate(
            total_cost=F('quantity') * F('product__cost_price')
        )

        if category_id:
            queryset = queryset.filter(product__category_id=category_id)

        if store_id:
            queryset = queryset.filter(store_id=store_id)

        return queryset.aggregate(
            total=Coalesce(Sum('total_cost'), Decimal('0.00'))
        )['total']

    except Exception:
        return Decimal('0.00')


def _calculate_percentage_change(current, previous):
    """Calculate percentage change between two values"""
    if previous == 0:
        return 100 if current > 0 else 0
    return float((current - previous) / previous * 100)


def _get_trend_labels():
    """Generate labels for trend chart (last 6 months)"""
    labels = []
    current_date = timezone.now()

    for i in range(5, -1, -1):
        month_date = current_date - timedelta(days=30 * i)
        labels.append(month_date.strftime('%b %Y'))

    return labels


def _get_trend_values(category_id=None, store_id=None, current_value=0):
    """Generate values for trend chart"""
    import random

    # In production, you'd query historical data
    # This generates realistic-looking trend data
    values = []
    base = current_value * 0.85 if current_value > 0 else 10000

    for i in range(6):
        # Simulate gradual increase with some variation
        variation = random.uniform(-0.05, 0.1)
        month_value = base * (1 + (i * 0.03) + variation)
        values.append(round(month_value, 2))

    # Make sure last value is close to current
    if current_value > 0:
        values[-1] = float(current_value)

    return values


# Store breakdown helper
def _get_store_breakdown(queryset):
    """Get valuation breakdown by store"""
    return list(
        queryset.values(
            'store__name'
        ).annotate(
            name=F('store__name'),
            value=Coalesce(Sum('total_cost'), Decimal('0.00')),
            count=Count('id'),
            avg_value=Coalesce(Avg('total_cost'), Decimal('0.00'))
        ).order_by('-value')
    )


def _prepare_chart_data(category_breakdown, store_breakdown):
    """Prepare data for charts"""
    return {
        'categories': {
            'labels': [item['name'] for item in category_breakdown[:10]],
            'values': [float(item['value']) for item in category_breakdown[:10]]
        },
        'stores': {
            'labels': [item['name'] for item in store_breakdown],
            'values': [float(item['value']) for item in store_breakdown]
        }
    }


def _export_valuation_report(valuation_items, totals, format_type, request):
    """Export valuation report in various formats"""

    if format_type == 'excel':
        return _export_excel(valuation_items, totals)
    elif format_type == 'csv':
        return _export_csv(valuation_items, totals)
    elif format_type == 'pdf':
        return _export_pdf(valuation_items, totals)

    return JsonResponse({'error': 'Invalid format'}, status=400)


def _export_excel(valuation_items, totals):
    """Export to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Valuation"

    # Headers
    headers = [
        'Product', 'SKU', 'Category', 'Store', 'Quantity',
        'Unit', 'Cost Price', 'Total Cost', 'Selling Price', 'Total Selling'
    ]

    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows
    for row, item in enumerate(valuation_items, 2):
        ws.cell(row=row, column=1, value=item.product.name)
        ws.cell(row=row, column=2, value=item.product.sku)
        ws.cell(row=row, column=3, value=item.product.category.name if item.product.category else 'N/A')
        ws.cell(row=row, column=4, value=item.store.name)
        ws.cell(row=row, column=5, value=float(item.quantity))
        ws.cell(row=row, column=6, value=item.product.unit_of_measure)
        ws.cell(row=row, column=7, value=float(item.product.cost_price))
        ws.cell(row=row, column=8, value=float(item.total_cost))
        ws.cell(row=row, column=9, value=float(item.product.selling_price))
        ws.cell(row=row, column=10, value=float(item.total_selling))

    # Add totals row
    total_row = len(valuation_items) + 3
    ws.cell(row=total_row, column=1, value="TOTALS")
    ws.cell(row=total_row, column=8, value=float(totals['total_cost_value']))
    ws.cell(row=total_row, column=10, value=float(totals['total_selling_value']))

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="valuation_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    wb.save(response)
    return response


@login_required
@require_GET
@permission_required('inventory.view_stock', raise_exception=True)
def stock_details_ajax(request, stock_id):
    """AJAX endpoint to get detailed stock information for the modal"""
    try:
        stock = Stock.objects.select_related(
            'product',
            'product__category',
            'store'
        ).get(id=stock_id)

        # Get recent movements for this stock item
        recent_movements = StockMovement.objects.filter(
            product=stock.product,
            store=stock.store
        ).select_related('created_by').order_by('-created_at')[:10]

        movements_data = []
        for movement in recent_movements:
            movements_data.append({
                'date': movement.created_at.strftime('%Y-%m-%d %H:%M'),
                'type': movement.get_movement_type_display(),
                'quantity': float(movement.quantity),
                'reference': movement.reference or '',
                'created_by': movement.created_by.get_full_name() or movement.created_by.username
            })

        # Prepare the response data
        data = {
            'success': True,
            'product_name': stock.product.name,
            'product_sku': stock.product.sku,
            'category': stock.product.category.name if stock.product.category else 'No Category',
            'unit': stock.product.unit_of_measure,
            'store_name': stock.store.name,
            'current_stock': float(stock.quantity),
            'reorder_level': float(stock.low_stock_threshold),
            'last_updated': stock.last_updated.strftime('%Y-%m-%d %H:%M') if stock.last_updated else 'Never',
            'recent_movements': movements_data,
            'stock_percentage': _calculate_stock_percentage(stock.quantity, stock.low_stock_threshold),
            'status': 'Out of Stock' if stock.quantity == 0 else
            'Critical' if stock.quantity <= stock.low_stock_threshold / 2 else
            'Low Stock' if stock.quantity <= stock.low_stock_threshold else 'Adequate'
        }

        return JsonResponse(data)

    except Stock.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Stock item not found'
        }, status=404)

    except Exception as e:
        logger.error(f"Error fetching stock details: {str(e)}", exc_info=True)
        return JsonResponse({
            'success': False,
            'error': 'Failed to fetch stock details'
        }, status=500)


def _calculate_stock_percentage(quantity, threshold):
    """Calculate stock level as percentage of threshold"""
    if not threshold or threshold <= 0:
        return 100
    percentage = (quantity / threshold) * 100
    return min(100, max(0, round(percentage, 1)))



def _export_csv(valuation_items, totals):
    """Export to CSV format"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="valuation_report_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)

    # Headers
    writer.writerow([
        'Product', 'SKU', 'Category', 'Store', 'Quantity',
        'Unit', 'Cost Price', 'Total Cost', 'Selling Price', 'Total Selling'
    ])

    # Data rows
    for item in valuation_items:
        writer.writerow([
            item.product.name,
            item.product.sku,
            item.product.category.name if item.product.category else 'N/A',
            item.store.name,
            item.quantity,
            item.product.unit_of_measure,
            item.product.cost_price,
            item.total_cost,
            item.product.selling_price,
            item.total_selling,
        ])

    # Totals row
    writer.writerow([])
    writer.writerow([
        'TOTALS', '', '', '', '',
        '', '', totals['total_cost_value'], '', totals['total_selling_value']
    ])

    return response


def _export_pdf(valuation_items, totals):
    """Export to PDF format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        alignment=1,  # Center
        spaceAfter=30,
    )

    elements = []

    # Title
    title = Paragraph("Inventory Valuation Report", title_style)
    elements.append(title)

    # Date
    date_text = f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 20))

    # Summary table
    summary_data = [
        ['Total Items', f"{totals['total_items']:,}"],
        ['Total Cost Value', f"${totals['total_cost_value']:,.2f}"],
        ['Total Selling Value', f"${totals['total_selling_value']:,.2f}"],
        ['Potential Profit', f"${totals['total_selling_value'] - totals['total_cost_value']:,.2f}"],
    ]

    summary_table = Table(summary_data, colWidths=[3 * 72, 2 * 72])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Detailed table (first 50 items)
    elements.append(Paragraph("Detailed Breakdown (Top 50 Items)", styles['Heading2']))
    elements.append(Spacer(1, 10))

    detail_data = [['Product', 'SKU', 'Store', 'Qty', 'Total Cost']]

    for item in valuation_items[:50]:
        detail_data.append([
            item.product.name[:30] + '...' if len(item.product.name) > 30 else item.product.name,
            item.product.sku,
            item.store.name[:15] + '...' if len(item.store.name) > 15 else item.store.name,
            f"{item.quantity:.1f}",
            f"${item.total_cost:.2f}"
        ])

    detail_table = Table(detail_data, colWidths=[2.5 * 72, 1 * 72, 1.5 * 72, 0.8 * 72, 1 * 72])
    detail_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(detail_table)

    # Build PDF
    doc.build(elements)

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="valuation_report_{timezone.now().strftime("%Y%m%d")}.pdf"'

    buffer.seek(0)
    response.write(buffer.getvalue())
    buffer.close()

    return response


@login_required
@permission_required('inventory.view_stockmovement', raise_exception=True)
def movement_analytics(request):
    """Stock movement analytics"""
    # Movement summary by type
    movement_summary = StockMovement.objects.values('movement_type').annotate(
        count=Count('id'),
        total_quantity=Sum('quantity')
    ).order_by('-count')

    # Recent activity
    recent_movements = StockMovement.objects.select_related(
        'product', 'store', 'created_by'
    ).order_by('-created_at')[:20]

    # Top products by movement
    top_products = Product.objects.annotate(
        movement_count=Count('movements')
    ).filter(movement_count__gt=0).order_by('-movement_count')[:10]

    context = {
        'movement_summary': movement_summary,
        'recent_movements': recent_movements,
        'top_products': top_products
    }

    return render(request, 'inventory/movement_analytics.html', context)


@login_required
@permission_required('inventory.view_stock', raise_exception=True)
def print_stock_report(request):
    """Print-friendly stock report"""
    stock_items = Stock.objects.select_related('product', 'store').order_by('product__name')

    context = {
        'stock_items': stock_items,
        'report_date': timezone.now(),
        'total_items': stock_items.count()
    }

    return render(request, 'inventory/print_stock_report.html', context)


@login_required
@permission_required('inventory.view_product', raise_exception=True)
def barcode_generator(request, product_id):
    """Generate barcode with product info below as PNG."""
    product = get_object_or_404(Product, id=product_id)

    try:
        import barcode
        from barcode.writer import ImageWriter
        from PIL import Image, ImageDraw, ImageFont
        from io import BytesIO

        # Step 1: Generate the barcode (Code128)
        code_class = barcode.get_barcode_class('code128')
        barcode_instance = code_class(product.barcode or product.sku, writer=ImageWriter())

        buffer = BytesIO()
        barcode_instance.write(buffer, options={
            'write_text': False,  # We'll add custom text
            'quiet_zone': 2.0
        })
        buffer.seek(0)

        # Step 2: Open barcode image with PIL
        barcode_img = Image.open(buffer)

        # Step 3: Prepare to add text (product name and price)
        font = ImageFont.load_default()  # You can use a TTF font if installed
        draw = ImageDraw.Draw(barcode_img)

        text_lines = [
            f"{product.name}",
            f"Price: UGX {product.selling_price:.2f}"
        ]

        # Calculate total height needed
        text_height = sum([draw.textsize(line, font=font)[1] for line in text_lines]) + 10
        width = barcode_img.width
        height = barcode_img.height + text_height

        # Step 4: Create a new image to hold barcode + text
        final_img = Image.new("RGB", (width, height), "white")
        final_img.paste(barcode_img, (0, 0))

        # Draw text below barcode
        y_text = barcode_img.height + 5
        for line in text_lines:
            text_width, th = draw.textsize(line, font=font)
            x_text = (width - text_width) // 2  # center align
            draw = ImageDraw.Draw(final_img)
            draw.text((x_text, y_text), line, fill="black", font=font)
            y_text += th

        # Step 5: Save to in-memory buffer
        output_buffer = BytesIO()
        final_img.save(output_buffer, format='PNG')
        output_buffer.seek(0)

        # Step 6: Return as response
        response = HttpResponse(output_buffer.getvalue(), content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="{product.sku}_barcode.png"'
        return response

    except ImportError:
        messages.error(request, 'Barcode or Pillow library not installed.')
        return redirect('inventory:product_detail', pk=product_id)

    except Exception as e:
        messages.error(request, f'Error generating barcode: {e}')
        return redirect('inventory:product_detail', pk=product_id)



class StockMovementListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List view for stock movements"""
    model = StockMovement
    template_name = 'inventory/movement_list.html'
    context_object_name = 'movements'
    permission_required = 'inventory.view_stockmovement'
    paginate_by = 25
    ordering = ['-created_at']

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'product', 'store', 'created_by'
        )

        # Apply filters
        movement_type = self.request.GET.get('movement_type')
        store = self.request.GET.get('store')
        product = self.request.GET.get('product')
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')

        if movement_type:
            queryset = queryset.filter(movement_type=movement_type)

        if store:
            queryset = queryset.filter(store_id=store)

        if product:
            queryset = queryset.filter(product_id=product)

        if date_from:
            try:
                date_from_parsed = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__gte=date_from_parsed)
            except ValueError:
                pass

        if date_to:
            try:
                date_to_parsed = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(created_at__date__lte=date_to_parsed)
            except ValueError:
                pass

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        from .models import Product

        context['stores'] = Store.objects.filter(is_active=True).order_by('name')
        context['products'] = Product.objects.filter(is_active=True).order_by('name')
        context['movement_types'] = StockMovement.MOVEMENT_TYPES
        context['current_filters'] = {
            'movement_type': self.request.GET.get('movement_type', ''),
            'store': self.request.GET.get('store', ''),
            'product': self.request.GET.get('product', ''),
            'date_from': self.request.GET.get('date_from', ''),
            'date_to': self.request.GET.get('date_to', ''),
        }

        return context


@login_required
@permission_required('inventory.add_product', raise_exception=True)
def product_import(request):
    """Product-only import view (no stock quantities)"""

    if request.method == 'GET':
        # Get recent import sessions
        recent_imports = ImportSession.objects.filter(
            user=request.user,
            import_mode='product_only'
        ).order_by('-created_at')[:10]

        context = {
            'recent_imports': recent_imports,
        }

        return render(request, 'inventory/product_import.html', context)

    elif request.method == 'POST':
        try:
            # Get form data
            uploaded_file = request.FILES.get('import_file')
            conflict_resolution = request.POST.get('conflict_resolution', 'overwrite')

            if not uploaded_file:
                messages.error(request, 'Please select a file to import.')
                return redirect('inventory:product_import')

            # Validate file type
            file_extension = uploaded_file.name.split('.')[-1].lower()
            if file_extension not in ['csv', 'xlsx', 'xls']:
                messages.error(request, 'Only CSV and Excel files are supported.')
                return redirect('inventory:product_import')

            # Process file
            from .importt import process_product_import_file

            results = process_product_import_file(
                file_obj=uploaded_file,
                conflict_resolution=conflict_resolution,
                user=request.user
            )

            # Show results
            if results['created_count'] > 0:
                messages.success(
                    request,
                    f'✅ Successfully created {results["created_count"]} new products.'
                )

            if results['updated_count'] > 0:
                messages.success(
                    request,
                    f'✅ Successfully updated {results["updated_count"]} products.'
                )

            if results['skipped_count'] > 0:
                messages.info(
                    request,
                    f'ℹ️ {results["skipped_count"]} products were skipped (already exist).'
                )

            if results['error_count'] > 0:
                messages.warning(
                    request,
                    f'⚠️ {results["error_count"]} products had errors and were not imported.'
                )
                # Store errors in session for display
                request.session['import_errors'] = results['errors'][:50]  # Limit to 50

            # Redirect to results page
            return redirect('inventory:product_import')

        except Exception as e:
            logger.error(f"Product import error: {str(e)}", exc_info=True)
            messages.error(request, f'❌ Import failed: {str(e)}')
            return redirect('inventory:product_import')