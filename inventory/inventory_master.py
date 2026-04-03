"""
inventory_master.py
====================
Single unified view file for all inventory pages.
Works with inventory_master.html via the `active_tab` context variable.

URL routing (urls.py) should map to InventoryMasterView like this:

    from .inventory_master import (
        InventoryMasterView,
        dashboard_stats_api,
        stock_alerts_api,
        recent_movements_api,
        top_products_api,
        stock_details_ajax,
    )

    urlpatterns = [
        # ── Master template routes ──────────────────────────────
        path('dashboard/',            InventoryMasterView.as_view(), {'active_tab': 'dashboard'},          name='dashboard'),
        path('dashboard/realtime/',   InventoryMasterView.as_view(), {'active_tab': 'realtime'},           name='realtime_dashboard'),
        path('stock-dashboard/',      InventoryMasterView.as_view(), {'active_tab': 'stock_dashboard'},    name='stock_dashboard'),
        path('reports/inventory/',    InventoryMasterView.as_view(), {'active_tab': 'inventory_report'},   name='inventory_report'),
        path('reports/low-stock/',    InventoryMasterView.as_view(), {'active_tab': 'low_stock'},          name='low_stock_report'),
        path('reports/movements/',    InventoryMasterView.as_view(), {'active_tab': 'movement_report'},    name='stock_movement_report'),
        path('reports/valuation/',    InventoryMasterView.as_view(), {'active_tab': 'valuation'},          name='valuation_report'),
        path('analytics/movements/',  InventoryMasterView.as_view(), {'active_tab': 'movement_analytics'}, name='movement_analytics'),

        # ── AJAX / API endpoints ────────────────────────────────
        path('ajax/dashboard-stats/', dashboard_stats_api,   name='dashboard_stats_api'),
        path('ajax/stock-alerts/',    stock_alerts_api,       name='stock_alerts_api'),
        path('ajax/recent-movements/',recent_movements_api,   name='recent_movements_api'),
        path('ajax/top-products/',    top_products_api,       name='top_products_api'),
        path('ajax/stock-details/<int:stock_id>/', stock_details_ajax, name='stock_details_ajax'),
    ]
"""

import json
import csv
import io
import logging
import random
from datetime import datetime, timedelta
from decimal import Decimal

from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.core.paginator import Paginator
from django.db import models
from django.db.models import (
    Q, F, Sum, Count, Avg, Max, Min,
    Case, When, Value, ExpressionWrapper, FloatField
)
from django.db.models.functions import Coalesce, TruncDate
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.views import View
from django.views.decorators.http import require_http_methods, require_GET

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from stores.models import Store
from .models import Category, Supplier, Product, Stock, StockMovement

logger = logging.getLogger(__name__)


# ══════════════════════════════════════════════════════════════════════════════
# MASTER VIEW
# ══════════════════════════════════════════════════════════════════════════════

class InventoryMasterView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """
    Single class-based view that renders inventory_master.html for every tab.
    The URL conf passes `active_tab` as a keyword argument.
    Each tab's context is built by a dedicated _build_*_context() method.
    """
    template_name = 'inventory/inventory_master.html'
    permission_required = 'inventory.view_stock'

    # Map active_tab → builder method
    TAB_BUILDERS = {
        'dashboard':          '_build_dashboard_context',
        'realtime':           '_build_dashboard_context',
        'stock_dashboard':    '_build_dashboard_context',
        'enhanced_dashboard': '_build_dashboard_context',
        'inventory_report':   '_build_inventory_report_context',
        'low_stock':          '_build_low_stock_context',
        'movement_report':    '_build_movement_report_context',
        'valuation':          '_build_valuation_context',
        'movement_analytics': '_build_movement_analytics_context',
    }

    def get(self, request, *args, **kwargs):
        active_tab = kwargs.get('active_tab', 'dashboard')

        # Base context available on every tab
        context = {
            'active_tab': active_tab,
            'company': getattr(request, 'tenant', None) or (
                request.user.company if hasattr(request.user, 'company') else None
            ),
            # Badge count for nav pill
            'low_stock_count': Stock.objects.filter(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold')
            ).count(),
            'efris_enabled': self._efris_enabled(),
        }

        # Build tab-specific context
        builder_name = self.TAB_BUILDERS.get(active_tab)
        if builder_name:
            builder = getattr(self, builder_name)
            tab_context = builder(request)
            # Handle export responses (CSV/PDF/Excel) returned directly
            if isinstance(tab_context, HttpResponse):
                return tab_context
            context.update(tab_context)

        return render(request, self.template_name, context)

    @staticmethod
    def _efris_enabled():
        try:
            from company.models import Company
            from django.db import connection
            tenant = getattr(connection, 'tenant', None)
            return getattr(tenant, 'efris_enabled', False) if tenant else False
        except Exception:
            return False

    # ──────────────────────────────────────────────────────────────────────────
    # TAB: DASHBOARD  (dashboard / realtime / stock_dashboard / enhanced)
    # ──────────────────────────────────────────────────────────────────────────
    def _build_dashboard_context(self, request):
        period = request.GET.get('period', 'week')
        today = timezone.now().date()

        period_map = {
            'today':   timedelta(days=0),
            'week':    timedelta(days=7),
            'month':   timedelta(days=30),
            'quarter': timedelta(days=90),
            'year':    timedelta(days=365),
        }
        start_date = today - period_map.get(period, timedelta(days=7))

        # ── Querysets ──
        stock_qs      = Stock.objects.select_related('product', 'store', 'product__category')
        movements_qs  = StockMovement.objects.select_related('product', 'store', 'created_by')
        products_qs   = Product.objects.filter(is_active=True)
        categories_qs = Category.objects.filter(is_active=True)
        suppliers_qs  = Supplier.objects.filter(is_active=True)

        # ── Core counts ──
        total_products   = products_qs.count()
        total_categories = categories_qs.count()
        total_suppliers  = suppliers_qs.count()
        total_stores     = Store.objects.filter(is_active=True).count()

        stock_stats = stock_qs.aggregate(
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            low_stock=Count('id', filter=Q(
                quantity__gt=0,
                quantity__lte=F('low_stock_threshold')
            )),
            good_stock=Count('id', filter=Q(
                quantity__gt=F('low_stock_threshold') * 2
            )),
        )

        low_stock    = stock_stats['low_stock'] or 0
        out_of_stock = stock_stats['out_of_stock'] or 0
        in_stock     = max(0, (stock_stats['total_items'] or 0) - low_stock - out_of_stock)

        # ── Valuation ──
        stock_value = stock_qs.aggregate(
            total_cost=Coalesce(
                Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')
            ),
            total_selling=Coalesce(
                Sum(F('quantity') * F('product__selling_price')), Decimal('0.00')
            ),
        )
        total_stock_value = stock_value['total_cost']
        potential_profit  = stock_value['total_selling'] - stock_value['total_cost']

        # ── Movements ──
        movements_today = movements_qs.filter(created_at__date=today).count()
        movements_period = movements_qs.filter(created_at__date__gte=start_date).count()

        period_days      = max((today - start_date).days, 1)
        previous_start   = start_date - timedelta(days=period_days)
        previous_movements = movements_qs.filter(
            created_at__date__gte=previous_start,
            created_at__date__lt=start_date
        ).count()
        movements_trend = _calculate_trend(movements_period, previous_movements)

        # ── Recent movements (for table + activity feed) ──
        recent_movements = movements_qs.order_by('-created_at')[:20]

        # ── Top products ──
        top_products = products_qs.annotate(
            total_movements=Count(
                'movements',
                filter=Q(movements__created_at__date__gte=start_date)
            ),
            movement_count=Count(
                'movements',
                filter=Q(movements__created_at__date__gte=start_date)
            ),
        ).filter(total_movements__gt=0).order_by('-total_movements')[:10]

        # ── Category distribution ──
        category_distribution = stock_qs.values(
            'product__category__name',
            'product__category__id',
        ).annotate(
            name=F('product__category__name'),
            count=Count('product', distinct=True),
            total_value=Coalesce(
                Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')
            ),
        ).filter(name__isnull=False).order_by('-total_value')[:10]

        # ── Store distribution ──
        store_distribution = stock_qs.values('store__name', 'store__id').annotate(
            store_name=F('store__name'),
            total_value=Coalesce(
                Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')
            ),
            total_quantity=Coalesce(Sum('quantity'), Decimal('0.00')),
            product_count=Count('product', distinct=True),
            low_stock_count=Count('id', filter=Q(
                quantity__gt=0, quantity__lte=F('low_stock_threshold')
            )),
            out_of_stock_count=Count('id', filter=Q(quantity=0)),
        ).order_by('-total_value')

        # ── Daily movement trends for chart ──
        daily_movements_qs = movements_qs.filter(
            created_at__date__gte=start_date
        ).annotate(
            date=TruncDate('created_at')
        ).values('date', 'movement_type').annotate(
            count=Count('id'),
            total_quantity=Coalesce(Sum('quantity'), Decimal('0.00')),
        ).order_by('date')
        movement_trends = _process_movement_trends(daily_movements_qs, start_date, today)

        # ── Inventory turnover ──
        month_ago = today - timedelta(days=30)
        total_sales_value = movements_qs.filter(
            movement_type='SALE',
            created_at__date__gte=month_ago
        ).aggregate(
            total=Coalesce(
                Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')
            )
        )['total']
        if total_stock_value > 0:
            inventory_turnover = round(float(total_sales_value / total_stock_value) * 12, 1)
        else:
            inventory_turnover = 0

        # ── EFRIS stats ──
        efris_products     = 0
        stocks_needing_sync = 0
        efris_products_qs  = Product.objects.none()
        if self._efris_enabled():
            efris_products = Product.objects.filter(efris_is_uploaded=True).count()
            stocks_needing_sync = Stock.objects.filter(
                product__efris_is_uploaded=True,
                efris_sync_required=True
            ).count()
            # Products list for EFRIS modals (Increase / Decrease / Query)
            efris_products_qs = Product.objects.filter(
                efris_is_uploaded=True, is_active=True
            ).order_by('name')

        # ── 4-tier stock status (Critical / Low / Medium / Good) ──
        total_stock_items = stock_stats['total_items'] or 0
        medium_stock = Stock.objects.filter(
            quantity__gt=F('low_stock_threshold'),
            quantity__lte=F('low_stock_threshold') * 2,
        ).count()
        good_stock = Stock.objects.filter(
            quantity__gt=F('low_stock_threshold') * 2
        ).count()
        stock_by_status = {
            'critical': out_of_stock,
            'low':      low_stock,
            'medium':   medium_stock,
            'good':     good_stock,
        }

        # ── Daily movement counts for 7-day trend chart (API-compatible) ──
        movements_by_day = []
        for i in range(7):
            day = today - timedelta(days=6 - i)
            count = StockMovement.objects.filter(created_at__date=day).count()
            movements_by_day.append({
                'date':  day.strftime('%Y-%m-%d'),
                'count': count,
            })

        return {
            # counts
            'total_products':    total_products,
            'total_categories':  total_categories,
            'total_suppliers':   total_suppliers,
            'total_stores':      total_stores,
            # stock status (3-tier legacy)
            'low_stock':         low_stock,
            'out_of_stock':      out_of_stock,
            'in_stock':          in_stock,
            'stock_stats':       stock_stats,
            # stock status (4-tier for enhanced chart)
            'stock_by_status':   stock_by_status,
            # valuation
            'total_stock_value': total_stock_value,
            'potential_profit':  potential_profit,
            # movements
            'movements_today':   movements_today,
            'movements_period':  movements_period,
            'movements_trend':   movements_trend,
            'movements_by_day':  movements_by_day,
            # lists
            'recent_movements':  recent_movements,
            'top_products':      top_products,
            # distributions
            'category_distribution': category_distribution,
            'store_distribution':    store_distribution,
            # chart
            'movement_trends':   movement_trends,
            # performance
            'inventory_turnover': inventory_turnover,
            # efris
            'efris_products':      efris_products,
            'stocks_needing_sync': stocks_needing_sync,
            'products':            efris_products_qs,   # for EFRIS modals
            # filter
            'selected_period': period,
            'start_date':      start_date,
            'end_date':        today,
        }

    # ──────────────────────────────────────────────────────────────────────────
    # TAB: INVENTORY REPORT
    # ──────────────────────────────────────────────────────────────────────────
    def _build_inventory_report_context(self, request):
        date_from  = request.GET.get('date_from', '')
        date_to    = request.GET.get('date_to', '')
        search     = request.GET.get('search', '')
        format_type = request.GET.get('format')

        # ── Transfer stats ──
        from .models import StockTransfer  # guard against missing model
        transfer_qs = StockTransfer.objects.all() if hasattr(StockTransfer, 'objects') else None

        if transfer_qs is not None:
            transfer_stats = transfer_qs.aggregate(
                total_transfers=Count('id'),
                completed_transfers=Count('id', filter=Q(status='COMPLETED')),
                pending_transfers=Count('id', filter=Q(status='PENDING')),
                total_items_transferred=Coalesce(Sum('quantity'), Decimal('0')),
            )
        else:
            transfer_stats = {
                'total_transfers': 0, 'completed_transfers': 0,
                'pending_transfers': 0, 'total_items_transferred': 0,
            }

        # ── Warehouse / branch stock ──
        warehouse_stores = Store.objects.filter(is_active=True, store_type='WAREHOUSE') \
            if hasattr(Store, 'store_type') else Store.objects.filter(is_active=True)[:5]
        branch_stores    = Store.objects.filter(is_active=True, store_type='BRANCH') \
            if hasattr(Store, 'store_type') else Store.objects.filter(is_active=True)[5:10]

        product_qs = Product.objects.filter(is_active=True)
        if search:
            product_qs = product_qs.filter(
                Q(name__icontains=search) | Q(sku__icontains=search)
            )

        # Build warehouse inventory rows
        def build_inventory_rows(stores_qs, product_qs):
            rows = []
            stocks = Stock.objects.filter(
                store__in=stores_qs,
            ).select_related('product', 'store')
            if search:
                stocks = stocks.filter(
                    Q(product__name__icontains=search) | Q(product__sku__icontains=search)
                )
            # Group by product
            product_map = {}
            for s in stocks:
                if s.product_id not in product_map:
                    product_map[s.product_id] = {
                        'product': s.product,
                        'stock_by_store': {},
                        'total_stock': Decimal('0'),
                    }
                product_map[s.product_id]['stock_by_store'][s.store_id] = s.quantity
                product_map[s.product_id]['total_stock'] += s.quantity
            return list(product_map.values())

        warehouse_inventory = build_inventory_rows(warehouse_stores, product_qs)
        branch_inventory    = build_inventory_rows(branch_stores, product_qs)

        # ── Movements tab ──
        movements_qs = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')
        if date_from:
            try:
                movements_qs = movements_qs.filter(
                    created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date()
                )
            except ValueError:
                pass
        if date_to:
            try:
                movements_qs = movements_qs.filter(
                    created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date()
                )
            except ValueError:
                pass
        if search:
            movements_qs = movements_qs.filter(product__name__icontains=search)
        stock_movements = movements_qs[:200]

        # ── Top products ──
        top_products = Product.objects.filter(is_active=True).annotate(
            total_movements=Count('movements'),   # used by template
            movement_count=Count('movements'),    # alias for internal use
        ).filter(total_movements__gt=0).order_by('-total_movements')[:20]

        # ── Handle export ──
        if format_type == 'excel':
            return _export_inventory_report_excel(warehouse_inventory, warehouse_stores)
        if format_type == 'csv':
            return _export_inventory_report_csv(warehouse_inventory, warehouse_stores)

        return {
            'transfer_stats':      transfer_stats,
            'warehouse_stores':    warehouse_stores,
            'branch_stores':       branch_stores,
            'warehouse_inventory': warehouse_inventory,
            'branch_inventory':    branch_inventory,
            'stock_movements':     stock_movements,
            'top_products':        top_products,
            'date_from':           date_from,
            'date_to':             date_to,
        }

    # ──────────────────────────────────────────────────────────────────────────
    # TAB: LOW STOCK REPORT
    # ──────────────────────────────────────────────────────────────────────────
    def _build_low_stock_context(self, request):
        category_id  = request.GET.get('category')
        store_id     = request.GET.get('store')
        severity     = request.GET.get('severity', 'all')
        format_type  = request.GET.get('format')

        # Annotated base queryset
        CharField = models.CharField
        base_qs = (
            Stock.objects.select_related(
                'product', 'product__category', 'store'
            ).annotate(
                total_cost=F('quantity') * F('product__cost_price'),
                reorder_gap=F('low_stock_threshold') - F('quantity'),
                half_threshold=ExpressionWrapper(
                    F('low_stock_threshold') / 2.0,
                    output_field=FloatField()
                ),
                stock_status=Case(
                    When(quantity=0,                            then=Value('out_of_stock')),
                    When(quantity__lte=F('low_stock_threshold') / 2.0, then=Value('critical')),
                    When(quantity__lte=F('low_stock_threshold'),       then=Value('low_stock')),
                    default=Value('in_stock'),
                    output_field=CharField(),
                ),
            )
            .filter(Q(quantity__lte=F('low_stock_threshold')) | Q(quantity=0))
            .order_by('quantity', 'reorder_gap')
        )

        if category_id:
            base_qs = base_qs.filter(product__category_id=category_id)
        if store_id:
            base_qs = base_qs.filter(store_id=store_id)
        if severity == 'out_of_stock':
            base_qs = base_qs.filter(quantity=0)
        elif severity == 'critical':
            base_qs = base_qs.filter(quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0)
        elif severity == 'low':
            base_qs = base_qs.filter(
                quantity__gt=F('low_stock_threshold') / 2.0,
                quantity__lte=F('low_stock_threshold')
            )

        summary_stats     = _calculate_low_stock_summary(base_qs)
        category_breakdown = _get_low_stock_category_breakdown(base_qs)
        store_breakdown    = _get_low_stock_store_breakdown(base_qs)
        chart_data         = _prepare_low_stock_chart_data(category_breakdown, store_breakdown)
        recommendations    = _generate_recommendations(summary_stats, category_breakdown)

        # Export
        if format_type in ('excel', 'csv', 'pdf'):
            return _export_low_stock_report(base_qs, summary_stats, format_type, request)

        # Pagination
        paginator  = Paginator(base_qs, 50)
        page_obj   = paginator.get_page(request.GET.get('page'))

        items_with_computations = []
        for item in page_obj:
            items_with_computations.append({
                'stock':               item,
                'product':             item.product,
                'total_cost':          item.total_cost,
                'reorder_gap':         item.reorder_gap,
                'stock_percentage':    _calculate_stock_percentage(item.quantity, item.low_stock_threshold),
                'days_until_stockout': _estimate_days_until_stockout(item),
                'recommended_order_qty': _calculate_recommended_order_quantity(item),
            })
        page_obj.object_list = items_with_computations

        return {
            'low_stock_items':    items_with_computations,
            'page_obj':           page_obj,
            'summary_stats':      summary_stats,
            'out_of_stock_count': summary_stats.get('out_of_stock_count', 0),
            'critical_stock_count': summary_stats.get('critical_count', 0),
            'low_stock_count':    summary_stats.get('low_stock_count', 0),
            'total_low_stock_value': summary_stats.get('total_value_at_risk', Decimal('0')),
            'category_breakdown': category_breakdown,
            'store_breakdown':    store_breakdown,
            'chart_data':         chart_data,
            'recommendations':    recommendations,
            'categories':  Category.objects.filter(is_active=True),
            'stores':      Store.objects.filter(is_active=True),
            'selected_category': category_id,
            'selected_store':    store_id,
            'selected_status':   severity,
        }

    # ──────────────────────────────────────────────────────────────────────────
    # TAB: STOCK MOVEMENT REPORT
    # ──────────────────────────────────────────────────────────────────────────
    def _build_movement_report_context(self, request):
        today    = timezone.now().date()
        date_from = request.GET.get('date_from', str(today - timedelta(days=30)))
        date_to   = request.GET.get('date_to',   str(today))
        store_id  = request.GET.get('stockstore')
        search    = request.GET.get('search', '')
        format_type = request.GET.get('format')

        # Base queryset
        qs = StockMovement.objects.select_related(
            'product', 'product__category', 'store', 'created_by'
        ).order_by('-created_at')

        if date_from:
            try:
                qs = qs.filter(
                    created_at__date__gte=datetime.strptime(date_from, '%Y-%m-%d').date()
                )
            except ValueError:
                pass
        if date_to:
            try:
                qs = qs.filter(
                    created_at__date__lte=datetime.strptime(date_to, '%Y-%m-%d').date()
                )
            except ValueError:
                pass
        if store_id:
            qs = qs.filter(store_id=store_id)
        if search:
            qs = qs.filter(
                Q(product__name__icontains=search) | Q(product__sku__icontains=search)
            )

        # ── Aggregate stats ──
        stats = qs.aggregate(
            total_movements=Count('id'),
            total_inbound=Count('id', filter=Q(
                movement_type__in=['PURCHASE', 'TRANSFER_IN', 'ADJUSTMENT_IN']
            )),
            total_outbound=Count('id', filter=Q(
                movement_type__in=['SALE', 'TRANSFER_OUT', 'ADJUSTMENT_OUT']
            )),
        )
        net_movement = stats['total_inbound'] - stats['total_outbound']

        # ── Movement summary by type ──
        movement_summary = qs.values('movement_type').annotate(
            count=Count('id'),
            total_quantity=Coalesce(Sum('quantity'), Decimal('0.00')),
        ).order_by('-count')

        # ── Daily movements for chart ──
        daily_movements = qs.annotate(
            date=TruncDate('created_at')
        ).values('date').annotate(
            count=Count('id')
        ).order_by('date')

        # ── Top products ──
        top_products = Product.objects.filter(
            movements__in=qs
        ).annotate(
            movement_count=Count('movements', filter=Q(movements__in=qs)),
            total_in=Coalesce(Sum(
                'movements__quantity',
                filter=Q(movements__movement_type__in=['PURCHASE', 'TRANSFER_IN'])
            ), Decimal('0')),
            total_out=Coalesce(Sum(
                'movements__quantity',
                filter=Q(movements__movement_type__in=['SALE', 'TRANSFER_OUT'])
            ), Decimal('0')),
        ).annotate(
            net_change=F('total_in') - F('total_out')
        ).filter(movement_count__gt=0).order_by('-movement_count')[:20]

        # ── Export ──
        if format_type == 'excel':
            return _export_movement_report_excel(qs, stats)
        if format_type == 'csv':
            return _export_movement_report_csv(qs)

        # ── Pagination ──
        paginator = Paginator(qs, 50)
        movements = paginator.get_page(request.GET.get('page'))

        return {
            'movements':         movements,
            'movement_summary':  movement_summary,
            'daily_movements':   daily_movements,
            'top_products':      top_products,
            'total_movements':   stats['total_movements'],
            'total_inbound':     stats['total_inbound'],
            'total_outbound':    stats['total_outbound'],
            'net_movement':      net_movement,
            'stockstores':       Store.objects.filter(is_active=True),
            'selected_stockstore': store_id,
            'date_from':         date_from,
            'date_to':           date_to,
        }

    # ──────────────────────────────────────────────────────────────────────────
    # TAB: VALUATION REPORT
    # ──────────────────────────────────────────────────────────────────────────
    def _build_valuation_context(self, request):
        category_id  = request.GET.get('category')
        store_id     = request.GET.get('store')
        period       = request.GET.get('period', 'current')
        format_type  = request.GET.get('format')

        queryset = Stock.objects.select_related(
            'product', 'product__category', 'store'
        ).annotate(
            total_cost=F('quantity') * F('product__cost_price'),
            total_selling=F('quantity') * F('product__selling_price'),
        ).filter(quantity__gt=0)

        if category_id:
            queryset = queryset.filter(product__category_id=category_id)
        if store_id:
            queryset = queryset.filter(store_id=store_id)

        valuation_items = queryset.order_by('-total_cost')

        totals = valuation_items.aggregate(
            total_cost_value=Coalesce(Sum('total_cost'),    Decimal('0.00')),
            total_selling_value=Coalesce(Sum('total_selling'), Decimal('0.00')),
            total_items=Count('id'),
            avg_item_value=Coalesce(Avg('total_cost'), Decimal('0.00')),
            max_item_value=Coalesce(Max('total_cost'), Decimal('0.00')),
        )

        potential_profit  = totals['total_selling_value'] - totals['total_cost_value']
        previous_value    = _get_previous_period_value(period, category_id, store_id)
        value_change      = totals['total_cost_value'] - previous_value
        value_change_percent = _calculate_percentage_change(totals['total_cost_value'], previous_value)

        # Category breakdown
        category_breakdown = list(
            queryset.values('product__category__name').annotate(
                name=F('product__category__name'),
                value=Coalesce(Sum('total_cost'), Decimal('0.00')),
                count=Count('id'),
                avg_value=Coalesce(Avg('total_cost'), Decimal('0.00')),
            ).filter(name__isnull=False).order_by('-value')
        )
        top_categories = category_breakdown[:5]

        # Store breakdown
        store_breakdown = list(
            queryset.values('store__name').annotate(
                name=F('store__name'),
                value=Coalesce(Sum('total_cost'), Decimal('0.00')),
                count=Count('id'),
            ).order_by('-value')
        )

        top_products    = valuation_items[:10]
        trend_labels    = _get_trend_labels()
        trend_values    = _get_trend_values(category_id, store_id, float(totals['total_cost_value']))
        valuation_trend = [{'date': lbl, 'value': val}
                           for lbl, val in zip(trend_labels, trend_values)]

        # Export
        if format_type in ('excel', 'csv', 'pdf'):
            return _export_valuation_report(valuation_items, totals, format_type, request)

        # Pagination
        paginator = Paginator(valuation_items, 50)
        page_obj  = paginator.get_page(request.GET.get('page'))

        return {
            'valuation_items':     page_obj,
            'page_obj':            page_obj,
            'total_value':         totals['total_cost_value'],
            'total_selling_value': totals['total_selling_value'],
            'potential_profit':    potential_profit,
            'total_items':         totals['total_items'],
            'avg_item_value':      totals['avg_item_value'],
            'max_item_value':      totals['max_item_value'],
            'value_change':        value_change,
            'value_change_percent': value_change_percent,
            'category_breakdown':  category_breakdown,
            'store_breakdown':     store_breakdown,
            'top_products':        top_products,
            'top_categories':      top_categories,
            'valuation_trend':     valuation_trend,
            'category_labels':     json.dumps([c['name'] or 'Uncategorized' for c in category_breakdown[:10]]),
            'category_values':     json.dumps([float(c['value']) for c in category_breakdown[:10]]),
            'trend_labels':        json.dumps(trend_labels),
            'trend_values':        json.dumps(trend_values),
            'report_date':         timezone.now(),
            'categories':  Category.objects.filter(is_active=True).order_by('name'),
            'stores':      Store.objects.filter(is_active=True).order_by('name'),
            'selected_category': category_id,
            'selected_store':    store_id,
        }

    # ──────────────────────────────────────────────────────────────────────────
    # TAB: MOVEMENT ANALYTICS
    # ──────────────────────────────────────────────────────────────────────────
    def _build_movement_analytics_context(self, request):
        movement_summary = StockMovement.objects.values('movement_type').annotate(
            count=Count('id'),
            total_quantity=Coalesce(Sum('quantity'), Decimal('0.00')),
        ).order_by('-count')

        recent_movements = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:20]

        top_products = Product.objects.annotate(
            movement_count=Count('movements')
        ).filter(movement_count__gt=0).order_by('-movement_count')[:10]

        return {
            'movement_summary': movement_summary,
            'recent_movements': recent_movements,
            'top_products':     top_products,
        }


# ══════════════════════════════════════════════════════════════════════════════
# AJAX / API ENDPOINTS
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@permission_required('inventory.view_product')
@require_http_methods(["GET"])
def dashboard_stats_api(request):
    """AJAX: live dashboard KPI numbers for JS auto-refresh."""
    try:
        today    = timezone.now().date()
        week_ago = today - timedelta(days=7)

        stock_stats = Stock.objects.aggregate(
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            low_stock=Count('id', filter=Q(
                quantity__gt=0, quantity__lte=F('low_stock_threshold')
            )),
        )
        stock_value = Stock.objects.aggregate(
            total=Coalesce(
                Sum(F('quantity') * F('product__cost_price')), Decimal('0.00')
            )
        )['total']

        # 4-tier stock status (matches EnhancedStockDashboardView / view.py)
        stock_by_status = {
            'critical': stock_stats['out_of_stock'] or 0,
            'low':      stock_stats['low_stock'] or 0,
            'medium':   Stock.objects.filter(
                quantity__gt=F('low_stock_threshold'),
                quantity__lte=F('low_stock_threshold') * 2,
            ).count(),
            'good':     Stock.objects.filter(
                quantity__gt=F('low_stock_threshold') * 2
            ).count(),
        }

        # 7-day daily movement counts (matches view.py movements_by_day)
        movements_by_day = []
        for i in range(7):
            day = today - timedelta(days=6 - i)
            movements_by_day.append({
                'date':  day.strftime('%Y-%m-%d'),
                'count': StockMovement.objects.filter(created_at__date=day).count(),
            })

        # EFRIS stats
        efris_stats = {
            'products_uploaded':  Product.objects.filter(efris_is_uploaded=True).count(),
            'stocks_needing_sync': Stock.objects.filter(
                product__efris_is_uploaded=True,
                efris_sync_required=True,
            ).count(),
        }

        # Stock alerts (critical + low, matches view.py stock_alerts)
        stock_alerts = []
        for stock in Stock.objects.filter(quantity=0).select_related('product', 'store')[:10]:
            stock_alerts.append({
                'product_name':  stock.product.name,
                'store_name':    stock.store.name,
                'current_stock': float(stock.quantity),
                'reorder_level': float(stock.low_stock_threshold or 0),
                'status':        'critical',
                'is_out':        True,
            })
        for stock in Stock.objects.filter(
            quantity__gt=0, quantity__lte=F('low_stock_threshold')
        ).select_related('product', 'store')[:10]:
            stock_alerts.append({
                'product_name':  stock.product.name,
                'store_name':    stock.store.name,
                'current_stock': float(stock.quantity),
                'reorder_level': float(stock.low_stock_threshold or 0),
                'status':        'low',
                'is_out':        False,
            })

        # Recent movements (matches view.py values() shape)
        recent_movements = list(
            StockMovement.objects.select_related('product', 'store', 'created_by')
            .order_by('-created_at')[:15]
            .values(
                'product__name', 'store__name', 'movement_type',
                'quantity', 'created_at', 'created_by__username',
            )
        )
        # Make datetime JSON-serialisable
        for m in recent_movements:
            if m.get('created_at'):
                m['created_at'] = m['created_at'].isoformat()

        return JsonResponse({
            'success':           True,
            'timestamp':         timezone.now().isoformat(),
            # flat KPIs (legacy keys kept for backward compat)
            'total_products':    Product.objects.filter(is_active=True).count(),
            'total_categories':  Category.objects.filter(is_active=True).count(),
            'total_suppliers':   Supplier.objects.filter(is_active=True).count(),
            'stock_value':       float(stock_value),
            'low_stock_items':   stock_stats['low_stock'] or 0,
            'out_of_stock_items': stock_stats['out_of_stock'] or 0,
            'movements_today':   StockMovement.objects.filter(created_at__date=today).count(),
            'movements_week':    StockMovement.objects.filter(created_at__date__gte=week_ago).count(),
            'total_stock_items': stock_stats['total_items'] or 0,
            # structured (matches EnhancedStockDashboardView response shape)
            'stock_stats':       {
                'total_products':  Product.objects.filter(is_active=True).count(),
                'total_stores':    Store.objects.filter(is_active=True).count(),
                'out_of_stock':    stock_stats['out_of_stock'] or 0,
                'low_stock':       stock_stats['low_stock'] or 0,
                'efris_products':  efris_stats['products_uploaded'],
            },
            'stock_by_status':   stock_by_status,
            'movements_by_day':  movements_by_day,
            'movement_stats':    {'movements_by_day': movements_by_day},
            'stock_alerts':      stock_alerts,
            'recent_movements':  recent_movements,
            'efris_stats':       efris_stats,
        })
    except Exception as e:
        logger.error(f"dashboard_stats_api error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Failed to fetch stats'}, status=500)


@login_required
@permission_required('inventory.view_stock')
@require_http_methods(["GET"])
def stock_alerts_api(request):
    """AJAX: stock alert items for the dashboard alerts panel."""
    try:
        alerts = Stock.objects.filter(
            Q(quantity=0) | Q(quantity__lte=F('low_stock_threshold'))
        ).select_related('product', 'store').order_by('quantity')[:20]

        alerts_data = []
        for a in alerts:
            threshold = a.low_stock_threshold or 0
            is_critical = a.quantity == 0 or (threshold and a.quantity <= threshold / 2)
            alerts_data.append({
                'id':               a.id,
                'product_name':     a.product.name,
                'product_sku':      a.product.sku,
                'store_name':       a.store.name,
                'current_quantity': float(a.quantity),
                'threshold':        float(threshold),
                'status':           'out_of_stock' if a.quantity == 0 else ('critical' if is_critical else 'low'),
                'percentage':       _calculate_stock_percentage(a.quantity, threshold),
                'unit_of_measure':  a.product.unit_of_measure,
            })

        return JsonResponse({'success': True, 'alerts': alerts_data, 'count': len(alerts_data)})
    except Exception as e:
        logger.error(f"stock_alerts_api error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Failed to fetch alerts'}, status=500)


@login_required
@permission_required('inventory.view_stockmovement')
@require_http_methods(["GET"])
def recent_movements_api(request):
    """AJAX: recent movements feed."""
    try:
        limit = min(int(request.GET.get('limit', 20)), 50)
        movements = StockMovement.objects.select_related(
            'product', 'store', 'created_by'
        ).order_by('-created_at')[:limit]

        data = [{
            'id':               m.id,
            'product_name':     m.product.name,
            'product_sku':      m.product.sku,
            'store_name':       m.store.name,
            'movement_type':    m.movement_type,
            'movement_type_display': m.get_movement_type_display(),
            'quantity':         float(m.quantity),
            'unit_of_measure':  m.product.unit_of_measure,
            'created_at':       m.created_at.isoformat(),
            'created_by':       (
                m.created_by.get_full_name() or m.created_by.username
                if m.created_by else 'System'
            ),
            'reference': m.reference or '',
            'notes':     m.notes or '',
        } for m in movements]

        return JsonResponse({'success': True, 'movements': data, 'count': len(data)})
    except Exception as e:
        logger.error(f"recent_movements_api error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Failed to fetch movements'}, status=500)


@login_required
@permission_required('inventory.view_product')
@require_http_methods(["GET"])
def top_products_api(request):
    """AJAX: top products by movement count."""
    try:
        limit       = min(int(request.GET.get('limit', 10)), 50)
        period_days = int(request.GET.get('period_days', 30))
        start_date  = timezone.now().date() - timedelta(days=period_days)

        products = Product.objects.filter(is_active=True).annotate(
            total_movements=Count(
                'movements',
                filter=Q(movements__created_at__date__gte=start_date)
            ),
            total_quantity_moved=Coalesce(Sum(
                'movements__quantity',
                filter=Q(movements__created_at__date__gte=start_date)
            ), Decimal('0.00')),
        ).filter(total_movements__gt=0).order_by('-total_movements')[:limit]

        data = [{
            'id':                  p.id,
            'name':                p.name,
            'sku':                 p.sku,
            'category':            p.category.name if p.category else None,
            'total_movements':     p.total_movements,
            'total_quantity_moved': float(p.total_quantity_moved),
        } for p in products]

        return JsonResponse({'success': True, 'products': data, 'count': len(data)})
    except Exception as e:
        logger.error(f"top_products_api error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Failed to fetch top products'}, status=500)


@login_required
@require_GET
@permission_required('inventory.view_stock', raise_exception=True)
def stock_details_ajax(request, stock_id):
    """AJAX: stock item details for the low-stock modal."""
    try:
        stock = Stock.objects.select_related(
            'product', 'product__category', 'store'
        ).get(id=stock_id)

        recent_movements = stock.get_recent_movements(days=30)[:10] \
            if hasattr(stock, 'get_recent_movements') else \
            StockMovement.objects.filter(
                product=stock.product, store=stock.store
            ).select_related('created_by').order_by('-created_at')[:10]

        movements_data = [{
            'date':       m.created_at.strftime('%Y-%m-%d %H:%M'),
            'type':       m.get_movement_type_display(),
            'quantity':   float(m.quantity),
            'reference':  m.reference or '',
            'created_by': (
                m.created_by.get_full_name() or m.created_by.username
                if m.created_by else 'System'
            ),
        } for m in recent_movements]

        return JsonResponse({
            'success':          True,
            'product_name':     stock.product.name,
            'product_sku':      stock.product.sku,
            'category':         stock.product.category.name if stock.product.category else 'No Category',
            'unit':             getattr(stock.product, 'unit_of_measure', 'N/A'),
            'store_name':       stock.store.name,
            'current_stock':    float(stock.quantity),
            'reorder_level':    float(stock.low_stock_threshold),
            'last_updated':     stock.last_updated.strftime('%Y-%m-%d %H:%M') if stock.last_updated else 'Never',
            'recent_movements': movements_data,
            'stock_percentage': getattr(stock, 'stock_percentage', 0),
            'status':           getattr(stock, 'status', 'unknown'),
        })
    except Stock.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Stock item not found'}, status=404)
    except Exception as e:
        logger.error(f"stock_details_ajax error: {e}", exc_info=True)
        return JsonResponse({'success': False, 'error': 'Failed to fetch stock details'}, status=500)


# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPER FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def _calculate_trend(current_value, previous_value):
    if previous_value == 0:
        return {'value': 100, 'direction': 'up'} if current_value > 0 else {'value': 0, 'direction': 'neutral'}
    pct = ((current_value - previous_value) / previous_value) * 100
    return {
        'value':     round(abs(pct), 1),
        'direction': 'up' if pct > 0 else ('down' if pct < 0 else 'neutral'),
    }


def _process_movement_trends(daily_movements_qs, start_date, end_date):
    """Convert the daily-movements queryset into a chart-friendly dict."""
    date_map = {}
    for row in daily_movements_qs:
        date_str = str(row['date'])
        if date_str not in date_map:
            date_map[date_str] = {'inbound': 0, 'outbound': 0, 'total': 0}
        if row['movement_type'] in ('PURCHASE', 'TRANSFER_IN', 'ADJUSTMENT_IN'):
            date_map[date_str]['inbound'] += row['count']
        elif row['movement_type'] in ('SALE', 'TRANSFER_OUT', 'ADJUSTMENT_OUT'):
            date_map[date_str]['outbound'] += row['count']
        date_map[date_str]['total'] += row['count']

    labels, inbound, outbound, total = [], [], [], []
    current = start_date
    while current <= end_date:
        ds = str(current)
        labels.append(current.strftime('%b %d'))
        day = date_map.get(ds, {})
        inbound.append(day.get('inbound', 0))
        outbound.append(day.get('outbound', 0))
        total.append(day.get('total', 0))
        current += timedelta(days=1)

    return {'labels': labels, 'inbound': inbound, 'outbound': outbound, 'total': total}


# ── Low-stock helpers ──────────────────────────────────────────────────────

def _calculate_low_stock_summary(queryset):
    summary = queryset.aggregate(
        total_items=Count('id'),
        total_value_at_risk=Coalesce(Sum('total_cost'), Decimal('0.00')),
        out_of_stock_count=Count('id', filter=Q(quantity=0)),
        critical_count=Count('id', filter=Q(
            quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0
        )),
        low_stock_count=Count('id', filter=Q(
            quantity__gt=F('low_stock_threshold') / 2.0,
            quantity__lte=F('low_stock_threshold')
        )),
        avg_reorder_gap=Coalesce(
            Sum('reorder_gap') / Count('id'), Decimal('0.00')
        ),
    )
    total = max(summary['total_items'], 1)
    summary['critical_percentage'] = (
        (summary['critical_count'] + summary['out_of_stock_count']) / total * 100
    )
    return summary


def _get_low_stock_category_breakdown(queryset):
    return list(
        queryset.values('product__category__name').annotate(
            name=F('product__category__name'),
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            critical=Count('id', filter=Q(
                quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0
            )),
            low_stock=Count('id', filter=Q(
                quantity__gt=F('low_stock_threshold') / 2.0,
                quantity__lte=F('low_stock_threshold')
            )),
            total_value=Coalesce(Sum('total_cost'), Decimal('0.00')),
        ).filter(name__isnull=False).order_by('-total_items')
    )


def _get_low_stock_store_breakdown(queryset):
    return list(
        queryset.values('store__name').annotate(
            name=F('store__name'),
            total_items=Count('id'),
            out_of_stock=Count('id', filter=Q(quantity=0)),
            critical=Count('id', filter=Q(
                quantity__gt=0, quantity__lte=F('low_stock_threshold') / 2.0
            )),
            low_stock=Count('id', filter=Q(
                quantity__gt=F('low_stock_threshold') / 2.0,
                quantity__lte=F('low_stock_threshold')
            )),
            total_value=Coalesce(Sum('total_cost'), Decimal('0.00')),
        ).order_by('-total_items')
    )


def _prepare_low_stock_chart_data(category_breakdown, store_breakdown):
    return {
        'categories': {
            'labels':   [i['name'] or 'No Category' for i in category_breakdown[:10]],
            'values':   [i['total_items'] for i in category_breakdown[:10]],
            'critical': [i['out_of_stock'] + i['critical'] for i in category_breakdown[:10]],
        },
        'stores': {
            'labels':   [i['name'] for i in store_breakdown],
            'values':   [i['total_items'] for i in store_breakdown],
            'critical': [i['out_of_stock'] + i['critical'] for i in store_breakdown],
        },
    }


def _generate_recommendations(summary_stats, category_breakdown):
    recs = []
    if summary_stats['out_of_stock_count'] > 0:
        recs.append({
            'urgency': 'HIGH',
            'title':   'Immediate Action Required',
            'message': f"{summary_stats['out_of_stock_count']} items are completely out of stock.",
            'action':  'Review out-of-stock items and place urgent orders.',
        })
    if summary_stats.get('critical_percentage', 0) > 50:
        recs.append({
            'urgency': 'MEDIUM',
            'title':   'High Critical Stock Ratio',
            'message': f"{summary_stats['critical_percentage']:.1f}% of low-stock items are at critical levels.",
            'action':  'Increase reorder levels or order frequencies.',
        })
    critical_cats = [
        c for c in category_breakdown
        if (c['out_of_stock'] + c['critical']) > c['total_items'] * 0.7
    ]
    if critical_cats:
        names = ', '.join(c['name'] or 'Uncategorized' for c in critical_cats[:3])
        recs.append({
            'urgency': 'MEDIUM',
            'title':   'Categories Need Attention',
            'message': f"High critical ratios in: {names}.",
            'action':  'Review supplier relationships and lead times.',
        })
    if summary_stats.get('total_value_at_risk', 0) > 10_000:
        recs.append({
            'urgency': 'LOW',
            'title':   'High Value at Risk',
            'message': f"UGX {summary_stats['total_value_at_risk']:,.0f} of inventory value is at risk.",
            'action':  'Prioritise high-value items for restocking.',
        })
    return recs


def _calculate_stock_percentage(quantity, threshold):
    if not threshold or threshold <= 0:
        return 100
    return min(100, max(0, round(float(quantity) / float(threshold) * 100, 1)))


def _estimate_days_until_stockout(stock_item):
    try:
        thirty_days_ago = timezone.now() - timedelta(days=30)
        consumed = StockMovement.objects.filter(
            product=stock_item.product,
            store=stock_item.store,
            movement_type__in=['SALE', 'TRANSFER_OUT'],
            created_at__gte=thirty_days_ago,
        ).aggregate(
            total=Coalesce(Sum('quantity'), Decimal('0.00'))
        )['total']
        if consumed > 0:
            daily = float(consumed) / 30
            return int(float(stock_item.quantity) / daily) if daily > 0 else None
        return None
    except Exception:
        return None


def _calculate_recommended_order_quantity(stock_item):
    target = stock_item.low_stock_threshold * Decimal('1.5')
    return max(Decimal('0'), target - stock_item.quantity).quantize(Decimal('0.01'))


# ── Valuation helpers ──────────────────────────────────────────────────────

def _get_previous_period_value(period, category_id=None, store_id=None):
    try:
        past_date = timezone.now() - timedelta(days=30)
        qs = Stock.objects.filter(last_updated__gte=past_date).annotate(
            total_cost=F('quantity') * F('product__cost_price')
        )
        if category_id:
            qs = qs.filter(product__category_id=category_id)
        if store_id:
            qs = qs.filter(store_id=store_id)
        return qs.aggregate(
            total=Coalesce(Sum('total_cost'), Decimal('0.00'))
        )['total']
    except Exception:
        return Decimal('0.00')


def _calculate_percentage_change(current, previous):
    if previous == 0:
        return 100 if current > 0 else 0
    return float((current - previous) / previous * 100)


def _get_trend_labels():
    now = timezone.now()
    return [
        (now - timedelta(days=30 * i)).strftime('%b %Y')
        for i in range(5, -1, -1)
    ]


def _get_trend_values(category_id=None, store_id=None, current_value=0.0):
    # TODO: Replace this with real historical Stock snapshots or a StockValuationSnapshot
    # model keyed by (date, category_id, store_id). The random jitter below produces
    # plausible-looking but fabricated numbers — do NOT ship this in production for a
    # financial report. Consider a nightly Celery task that records daily valuation totals.
    base = current_value * 0.85 if current_value > 0 else 10_000
    values = [
        round(base * (1 + i * 0.03 + random.uniform(-0.05, 0.1)), 2)
        for i in range(6)
    ]
    if current_value > 0:
        values[-1] = current_value
    return values


# ── Export helpers ─────────────────────────────────────────────────────────

def _export_inventory_report_excel(inventory_rows, stores):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory Report'
    bold_white  = Font(bold=True, color='FFFFFF')
    blue_fill   = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_align = Alignment(horizontal='center')

    headers = ['Product', 'SKU', 'Category'] + [s.name for s in stores] + ['Total', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment = bold_white, blue_fill, center_align

    for row_idx, item in enumerate(inventory_rows, 2):
        p = item['product']
        ws.cell(row=row_idx, column=1, value=p.name)
        ws.cell(row=row_idx, column=2, value=p.sku)
        ws.cell(row=row_idx, column=3, value=p.category.name if p.category else 'N/A')
        for col_idx, store in enumerate(stores, 4):
            ws.cell(row=row_idx, column=col_idx,
                    value=float(item['stock_by_store'].get(store.id, 0)))
        ws.cell(row=row_idx, column=4 + len(list(stores)), value=float(item['total_stock']))

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(resp)
    return resp


def _export_inventory_report_csv(inventory_rows, stores):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = (
        f'attachment; filename="inventory_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    )
    writer = csv.writer(resp)
    writer.writerow(['Product', 'SKU', 'Category'] + [s.name for s in stores] + ['Total'])
    for item in inventory_rows:
        p = item['product']
        row = [p.name, p.sku, p.category.name if p.category else 'N/A']
        row += [float(item['stock_by_store'].get(s.id, 0)) for s in stores]
        row += [float(item['total_stock'])]
        writer.writerow(row)
    return resp


def _export_movement_report_excel(qs, stats):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Stock Movements'
    bold_white  = Font(bold=True, color='FFFFFF')
    blue_fill   = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_align = Alignment(horizontal='center')

    headers = ['Date', 'Product', 'SKU', 'Type', 'Quantity', 'Store', 'Reference', 'By']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment = bold_white, blue_fill, center_align

    for row_idx, m in enumerate(qs, 2):
        ws.cell(row=row_idx, column=1, value=m.created_at.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row_idx, column=2, value=m.product.name)
        ws.cell(row=row_idx, column=3, value=m.product.sku)
        ws.cell(row=row_idx, column=4, value=m.movement_type)
        ws.cell(row=row_idx, column=5, value=float(m.quantity))
        ws.cell(row=row_idx, column=6, value=m.store.name)
        ws.cell(row=row_idx, column=7, value=m.reference or '')
        ws.cell(row=row_idx, column=8,
                value=m.created_by.get_full_name() or m.created_by.username
                      if m.created_by else 'System')

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="movement_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(resp)
    return resp


def _export_movement_report_csv(qs):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = (
        f'attachment; filename="movement_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    )
    writer = csv.writer(resp)
    writer.writerow(['Date', 'Product', 'SKU', 'Type', 'Quantity', 'Store', 'Reference', 'By'])
    for m in qs:
        writer.writerow([
            m.created_at.strftime('%Y-%m-%d %H:%M'),
            m.product.name, m.product.sku, m.movement_type,
            float(m.quantity), m.store.name, m.reference or '',
            m.created_by.get_full_name() or m.created_by.username if m.created_by else 'System',
        ])
    return resp


def _export_low_stock_report(low_stock_items, summary_stats, format_type, request):
    """Delegate to format-specific exporters."""
    if format_type == 'excel':
        return _export_low_stock_excel(low_stock_items, summary_stats)
    if format_type == 'csv':
        return _export_low_stock_csv(low_stock_items, summary_stats)
    if format_type == 'pdf':
        return _export_low_stock_pdf(low_stock_items, summary_stats)
    return JsonResponse({'error': 'Invalid format'}, status=400)


def _export_low_stock_excel(items, summary_stats):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Low Stock Report'
    bold_white   = Font(bold=True, color='FFFFFF')
    red_fill     = PatternFill(start_color='C0392B', end_color='C0392B', fill_type='solid')
    center_align = Alignment(horizontal='center')

    headers = ['Product', 'SKU', 'Category', 'Store', 'Qty', 'Reorder Level', 'Status', 'Cost Value']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment = bold_white, red_fill, center_align

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.product.name)
        ws.cell(row=row_idx, column=2, value=item.product.sku)
        ws.cell(row=row_idx, column=3,
                value=item.product.category.name if item.product.category else 'N/A')
        ws.cell(row=row_idx, column=4, value=item.store.name)
        ws.cell(row=row_idx, column=5, value=float(item.quantity))
        ws.cell(row=row_idx, column=6, value=float(item.low_stock_threshold))
        ws.cell(row=row_idx, column=7,
                value='Out of Stock' if item.quantity == 0 else (
                    'Critical' if item.quantity <= item.low_stock_threshold / 2 else 'Low Stock'
                ))
        ws.cell(row=row_idx, column=8,
                value=float(item.total_cost) if hasattr(item, 'total_cost') else 0)

    # Summary section
    last_row = items.count() + 3
    ws.cell(row=last_row,     column=1, value='SUMMARY')
    ws.cell(row=last_row + 1, column=1, value='Total Items at Risk')
    ws.cell(row=last_row + 1, column=2, value=summary_stats.get('total_items', 0))
    ws.cell(row=last_row + 2, column=1, value='Out of Stock')
    ws.cell(row=last_row + 2, column=2, value=summary_stats.get('out_of_stock_count', 0))
    ws.cell(row=last_row + 3, column=1, value='Total Value at Risk')
    ws.cell(row=last_row + 3, column=2, value=float(summary_stats.get('total_value_at_risk', 0)))

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="low_stock_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(resp)
    return resp


def _export_low_stock_csv(items, summary_stats):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = (
        f'attachment; filename="low_stock_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    )
    writer = csv.writer(resp)
    writer.writerow(['Product', 'SKU', 'Category', 'Store', 'Qty', 'Reorder Level', 'Status'])
    for item in items:
        writer.writerow([
            item.product.name, item.product.sku,
            item.product.category.name if item.product.category else 'N/A',
            item.store.name, float(item.quantity), float(item.low_stock_threshold),
            'Out of Stock' if item.quantity == 0 else 'Low Stock',
        ])
    return resp


def _export_low_stock_pdf(items, summary_stats):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph('Low Stock Report', ParagraphStyle(
            'Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20
        )),
        Paragraph(
            f"Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
            styles['Normal']
        ),
        Spacer(1, 12),
        Paragraph('Summary', styles['Heading2']),
    ]

    summary_data = [
        ['Total Items',          str(summary_stats.get('total_items', 0))],
        ['Out of Stock',         str(summary_stats.get('out_of_stock_count', 0))],
        ['Critical',             str(summary_stats.get('critical_count', 0))],
        ['Value at Risk',        f"UGX {summary_stats.get('total_value_at_risk', 0):,.0f}"],
    ]
    tbl = Table(summary_data, colWidths=[3 * 72, 2 * 72])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID',       (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightyellow),
    ]))
    elements.extend([tbl, Spacer(1, 20)])

    detail_data = [['Product', 'Store', 'Qty', 'Reorder Lvl', 'Status']]
    for item in items[:100]:
        detail_data.append([
            item.product.name[:30],
            item.store.name[:15],
            f"{float(item.quantity):.1f}",
            f"{float(item.low_stock_threshold):.1f}",
            'Out of Stock' if item.quantity == 0 else 'Low Stock',
        ])
    detail_tbl = Table(detail_data, colWidths=[2.2 * 72, 1.5 * 72, 0.7 * 72, 1 * 72, 1 * 72])
    detail_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkred),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightyellow]),
    ]))
    elements.append(detail_tbl)

    doc.build(elements)
    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="low_stock_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    buffer.seek(0)
    resp.write(buffer.getvalue())
    buffer.close()
    return resp


def _export_valuation_report(valuation_items, totals, format_type, request):
    if format_type == 'excel':
        return _export_valuation_excel(valuation_items, totals)
    if format_type == 'csv':
        return _export_valuation_csv(valuation_items, totals)
    if format_type == 'pdf':
        return _export_valuation_pdf(valuation_items, totals)
    return JsonResponse({'error': 'Invalid format'}, status=400)


def _export_valuation_excel(items, totals):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Inventory Valuation'
    bold_white   = Font(bold=True, color='FFFFFF')
    blue_fill    = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    center_align = Alignment(horizontal='center')

    headers = ['Product', 'SKU', 'Category', 'Store', 'Qty', 'Unit',
               'Cost Price', 'Total Cost', 'Sell Price', 'Total Value']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font, cell.fill, cell.alignment = bold_white, blue_fill, center_align

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1,  value=item.product.name)
        ws.cell(row=row_idx, column=2,  value=item.product.sku)
        ws.cell(row=row_idx, column=3,  value=item.product.category.name if item.product.category else 'N/A')
        ws.cell(row=row_idx, column=4,  value=item.store.name)
        ws.cell(row=row_idx, column=5,  value=float(item.quantity))
        ws.cell(row=row_idx, column=6,  value=item.product.unit_of_measure)
        ws.cell(row=row_idx, column=7,  value=float(item.product.cost_price))
        ws.cell(row=row_idx, column=8,  value=float(item.total_cost))
        ws.cell(row=row_idx, column=9,  value=float(item.product.selling_price))
        ws.cell(row=row_idx, column=10, value=float(item.total_selling))

    # Totals
    total_row = items.count() + 3
    ws.cell(row=total_row, column=1, value='TOTALS').font = Font(bold=True)
    ws.cell(row=total_row, column=8,  value=float(totals['total_cost_value']))
    ws.cell(row=total_row, column=10, value=float(totals['total_selling_value']))

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    resp = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    resp['Content-Disposition'] = (
        f'attachment; filename="valuation_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    )
    wb.save(resp)
    return resp


def _export_valuation_csv(items, totals):
    resp = HttpResponse(content_type='text/csv')
    resp['Content-Disposition'] = (
        f'attachment; filename="valuation_report_{timezone.now().strftime("%Y%m%d")}.csv"'
    )
    writer = csv.writer(resp)
    writer.writerow(['Product', 'SKU', 'Category', 'Store', 'Qty', 'Unit',
                     'Cost Price', 'Total Cost', 'Sell Price', 'Total Value'])
    for item in items:
        writer.writerow([
            item.product.name, item.product.sku,
            item.product.category.name if item.product.category else 'N/A',
            item.store.name, float(item.quantity), item.product.unit_of_measure,
            float(item.product.cost_price), float(item.total_cost),
            float(item.product.selling_price), float(item.total_selling),
        ])
    writer.writerow([])
    writer.writerow(['TOTALS', '', '', '', '', '', '',
                     float(totals['total_cost_value']), '', float(totals['total_selling_value'])])
    return resp


def _export_valuation_pdf(items, totals):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph('Inventory Valuation Report', ParagraphStyle(
            'Title', parent=styles['Heading1'], fontSize=18, alignment=1, spaceAfter=20
        )),
        Paragraph(
            f"Generated: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
            styles['Normal']
        ),
        Spacer(1, 20),
    ]

    summary_data = [
        ['Total Items',           f"{totals['total_items']:,}"],
        ['Total Cost Value',      f"UGX {totals['total_cost_value']:,.0f}"],
        ['Total Selling Value',   f"UGX {totals['total_selling_value']:,.0f}"],
        ['Potential Profit',
         f"UGX {totals['total_selling_value'] - totals['total_cost_value']:,.0f}"],
    ]
    tbl = Table(summary_data, colWidths=[3 * 72, 2 * 72])
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID',       (0, 0), (-1, -1), 1, colors.black),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
    ]))
    elements.extend([tbl, Spacer(1, 20)])

    detail_data = [['Product', 'SKU', 'Store', 'Qty', 'Total Cost']]
    for item in items[:50]:
        nm = item.product.name
        detail_data.append([
            nm[:28] + '…' if len(nm) > 30 else nm,
            item.product.sku,
            item.store.name[:14] + '…' if len(item.store.name) > 15 else item.store.name,
            f"{item.quantity:.1f}",
            f"UGX {item.total_cost:,.0f}",
        ])
    detail_tbl = Table(detail_data, colWidths=[2.5 * 72, 1 * 72, 1.5 * 72, 0.8 * 72, 1.2 * 72])
    detail_tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR',  (0, 0), (-1, 0), colors.white),
        ('FONTNAME',   (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 8),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.beige]),
    ]))
    elements.append(detail_tbl)

    doc.build(elements)
    resp = HttpResponse(content_type='application/pdf')
    resp['Content-Disposition'] = (
        f'attachment; filename="valuation_report_{timezone.now().strftime("%Y%m%d")}.pdf"'
    )
    buffer.seek(0)
    resp.write(buffer.getvalue())
    buffer.close()
    return resp