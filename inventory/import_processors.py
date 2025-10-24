import pandas as pd
import xlwings as xw
from openpyxl import load_workbook
import csv
import json
import xml.etree.ElementTree as ET
from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional, Tuple
from decimal import Decimal, InvalidOperation
from django.core.exceptions import ValidationError
from django.utils import timezone
from django.db import transaction
import logging
from io import StringIO, BytesIO

from .models import Product, Category, Supplier, Stock, ImportSession, ImportLog, ImportResult
from stores.models import Store

logger = logging.getLogger(__name__)


class BaseImportProcessor(ABC):
    """Abstract base class for import processors"""

    def __init__(self, import_session: ImportSession, user):
        self.session = import_session
        self.user = user
        self.processed_rows = 0
        self.created_count = 0
        self.updated_count = 0
        self.skipped_count = 0
        self.error_count = 0
        self.errors = []

    @abstractmethod
    def read_file(self, file) -> List[Dict]:
        """Read and parse the file into a list of dictionaries"""
        pass

    @abstractmethod
    def validate_headers(self, headers: List[str]) -> bool:
        """Validate that required headers are present"""
        pass

    def log_message(self, level: str, message: str, row_number: Optional[int] = None, details: Dict = None):
        """Log a message to the import session"""
        ImportLog.objects.create(
            session=self.session,
            level=level,
            message=message,
            row_number=row_number,
            details=details or {}
        )

    def create_result(self, result_type: str, row_number: int, **kwargs):
        """Create an import result record"""
        ImportResult.objects.create(
            session=self.session,
            result_type=result_type,
            row_number=row_number,
            **kwargs
        )

    def update_session_progress(self):
        """Update session with current progress"""
        self.session.processed_rows = self.processed_rows
        self.session.created_count = self.created_count
        self.session.updated_count = self.updated_count
        self.session.skipped_count = self.skipped_count
        self.session.error_count = self.error_count
        self.session.save(update_fields=[
            'processed_rows', 'created_count', 'updated_count',
            'skipped_count', 'error_count'
        ])

    def clean_decimal_value(self, value: Any, default: Decimal = Decimal('0.00')) -> Decimal:
        """Clean and convert value to Decimal"""
        if pd.isna(value) or value == '' or value is None:
            return default

        try:
            # Remove currency symbols and commas
            if isinstance(value, str):
                cleaned = value.replace('$', '').replace(',', '').strip()
                if cleaned == '':
                    return default
                return Decimal(cleaned)
            return Decimal(str(value))
        except (InvalidOperation, ValueError, TypeError):
            return default

    def clean_integer_value(self, value: Any, default: int = 0) -> int:
        """Clean and convert value to integer"""
        if pd.isna(value) or value == '' or value is None:
            return default

        try:
            return int(float(value))
        except (ValueError, TypeError):
            return default

    def clean_string_value(self, value: Any, default: str = '') -> str:
        """Clean and convert value to string"""
        if pd.isna(value) or value is None:
            return default
        return str(value).strip()


class ExcelImportProcessor(BaseImportProcessor):
    """Excel file import processor with advanced features"""

    REQUIRED_HEADERS = ['product_name', 'sku']
    OPTIONAL_HEADERS = [
        'category', 'supplier', 'description', 'selling_price', 'cost_price',
        'quantity', 'store', 'unit_of_measure', 'barcode', 'reorder_level'
    ]

    def read_file(self, file) -> List[Dict]:
        """Read Excel file with multiple sheet support"""
        try:
            # Try to read with openpyxl first (more reliable)
            workbook = load_workbook(file, read_only=True, data_only=True)

            # Check if multiple sheets exist
            if len(workbook.sheetnames) > 1:
                self.log_message('info', f'Found {len(workbook.sheetnames)} sheets: {", ".join(workbook.sheetnames)}')

            # Use first sheet or look for specific sheet names
            sheet_name = self.find_data_sheet(workbook.sheetnames)
            worksheet = workbook[sheet_name]

            # Convert to DataFrame for easier processing
            data = []
            headers = []

            for row_idx, row in enumerate(worksheet.iter_rows(values_only=True)):
                if row_idx == 0:
                    headers = [self.clean_string_value(cell).lower().replace(' ', '_') for cell in row if cell]
                    if not self.validate_headers(headers):
                        raise ValidationError("Required headers not found")
                    continue

                # Skip empty rows
                if not any(cell for cell in row):
                    continue

                row_dict = {}
                for col_idx, cell in enumerate(row):
                    if col_idx < len(headers):
                        row_dict[headers[col_idx]] = cell

                data.append(row_dict)

            self.log_message('info', f'Successfully read {len(data)} rows from Excel file')
            return data

        except Exception as e:
            self.log_message('error', f'Error reading Excel file: {str(e)}')
            raise

    def find_data_sheet(self, sheet_names: List[str]) -> str:
        """Find the most likely data sheet"""
        priority_names = ['products', 'inventory', 'stock', 'data']

        for priority in priority_names:
            for sheet in sheet_names:
                if priority.lower() in sheet.lower():
                    return sheet

        # Return first sheet if no match found
        return sheet_names[0]

    def validate_headers(self, headers: List[str]) -> bool:
        """Validate that required headers are present"""
        missing_headers = []
        for required in self.REQUIRED_HEADERS:
            if required not in headers:
                missing_headers.append(required)

        if missing_headers:
            self.log_message('error', f'Missing required headers: {", ".join(missing_headers)}')
            return False

        return True


class CSVImportProcessor(BaseImportProcessor):
    """CSV file import processor with encoding detection"""

    REQUIRED_HEADERS = ['product_name', 'sku']

    def read_file(self, file) -> List[Dict]:
        """Read CSV file with automatic encoding detection"""
        try:
            # Try to detect encoding
            file.seek(0)
            raw_data = file.read()

            # Try common encodings
            encodings = ['utf-8', 'utf-8-sig', 'iso-8859-1', 'cp1252']
            content = None
            used_encoding = None

            for encoding in encodings:
                try:
                    content = raw_data.decode(encoding)
                    used_encoding = encoding
                    break
                except UnicodeDecodeError:
                    continue

            if content is None:
                raise ValidationError("Could not detect file encoding")

            self.log_message('info', f'Using encoding: {used_encoding}')

            # Parse CSV
            csv_reader = csv.DictReader(StringIO(content))

            # Normalize headers
            fieldnames = [self.clean_string_value(field).lower().replace(' ', '_') for field in csv_reader.fieldnames]

            if not self.validate_headers(fieldnames):
                raise ValidationError("Required headers not found")

            data = []
            for row in csv_reader:
                # Normalize row keys to match normalized fieldnames
                normalized_row = {}
                for old_key, new_key in zip(csv_reader.fieldnames, fieldnames):
                    normalized_row[new_key] = row[old_key]
                data.append(normalized_row)

            self.log_message('info', f'Successfully read {len(data)} rows from CSV file')
            return data

        except Exception as e:
            self.log_message('error', f'Error reading CSV file: {str(e)}')
            raise

    def validate_headers(self, headers: List[str]) -> bool:
        """Validate CSV headers"""
        missing_headers = [h for h in self.REQUIRED_HEADERS if h not in headers]
        if missing_headers:
            self.log_message('error', f'Missing required headers: {", ".join(missing_headers)}')
            return False
        return True


class XMLImportProcessor(BaseImportProcessor):
    """XML file import processor"""

    def read_file(self, file) -> List[Dict]:
        """Read XML file and convert to dictionary format"""
        try:
            file.seek(0)
            content = file.read()

            # Try to decode if bytes
            if isinstance(content, bytes):
                content = content.decode('utf-8')

            root = ET.fromstring(content)
            data = []

            # Look for common XML structures
            items = self.find_items(root)

            for item in items:
                row_dict = self.xml_to_dict(item)
                if row_dict:
                    data.append(row_dict)

            self.log_message('info', f'Successfully read {len(data)} items from XML file')
            return data

        except Exception as e:
            self.log_message('error', f'Error reading XML file: {str(e)}')
            raise

    def find_items(self, root):
        """Find item nodes in XML"""
        possible_names = ['item', 'product', 'record', 'row', 'entry']

        for name in possible_names:
            items = root.findall(f'.//{name}')
            if items:
                return items

        # If no standard names found, use direct children
        return list(root)

    def xml_to_dict(self, element) -> Dict:
        """Convert XML element to dictionary"""
        result = {}

        # Add attributes
        result.update(element.attrib)

        # Add child elements
        for child in element:
            key = child.tag.lower().replace(' ', '_')
            value = child.text or ''
            result[key] = value.strip()

        return result

    def validate_headers(self, headers: List[str]) -> bool:
        """XML doesn't have traditional headers, validate during processing"""
        return True


class JSONImportProcessor(BaseImportProcessor):
    """JSON file import processor"""

    def read_file(self, file) -> List[Dict]:
        """Read JSON file"""
        try:
            file.seek(0)
            data = json.load(file)

            # Handle different JSON structures
            if isinstance(data, list):
                items = data
            elif isinstance(data, dict):
                # Look for common array keys
                array_keys = ['items', 'products', 'data', 'records']
                items = None

                for key in array_keys:
                    if key in data and isinstance(data[key], list):
                        items = data[key]
                        break

                if items is None:
                    # Treat as single item
                    items = [data]
            else:
                raise ValidationError("Invalid JSON structure")

            # Normalize keys
            normalized_items = []
            for item in items:
                if isinstance(item, dict):
                    normalized = {}
                    for key, value in item.items():
                        normalized_key = key.lower().replace(' ', '_')
                        normalized[normalized_key] = value
                    normalized_items.append(normalized)

            self.log_message('info', f'Successfully read {len(normalized_items)} items from JSON file')
            return normalized_items

        except Exception as e:
            self.log_message('error', f'Error reading JSON file: {str(e)}')
            raise

    def validate_headers(self, headers: List[str]) -> bool:
        """JSON validation happens during processing"""
        return True


class AdvancedImportManager:
    """Advanced import manager with multiple format support"""

    PROCESSORS = {
        '.xlsx': ExcelImportProcessor,
        '.xls': ExcelImportProcessor,
        '.csv': CSVImportProcessor,
        '.xml': XMLImportProcessor,
        '.json': JSONImportProcessor,
    }

    def __init__(self, import_session: ImportSession, user):
        self.session = import_session
        self.user = user
        self.processor = None

    def get_processor(self, filename: str) -> BaseImportProcessor:
        """Get appropriate processor for file type"""
        extension = self.get_file_extension(filename)
        processor_class = self.PROCESSORS.get(extension)

        if not processor_class:
            raise ValidationError(f"Unsupported file format: {extension}")

        return processor_class(self.session, self.user)

    def get_file_extension(self, filename: str) -> str:
        """Extract file extension"""
        return filename.lower().split('.')[-1] if '.' in filename else ''

    def process_import(self, file):
        """Process import file with appropriate processor"""
        try:
            # Update session status
            self.session.status = 'processing'
            self.session.started_at = timezone.now()
            self.session.save()

            # Get processor
            processor = self.get_processor(self.session.filename)

            # Read and process file
            data = processor.read_file(file)
            self.session.total_rows = len(data)
            self.session.save()

            # Process each row
            with transaction.atomic():
                self.process_data(processor, data)

            # Update final status
            self.session.status = 'completed' if processor.error_count == 0 else 'failed'
            self.session.completed_at = timezone.now()
            self.session.processed_rows = processor.processed_rows
            self.session.created_count = processor.created_count
            self.session.updated_count = processor.updated_count
            self.session.skipped_count = processor.skipped_count
            self.session.error_count = processor.error_count

            if processor.errors:
                self.session.error_details = processor.errors[:100]  # Store first 100 errors

            self.session.save()

            return {
                'success': True,
                'processed_rows': processor.processed_rows,
                'created_count': processor.created_count,
                'updated_count': processor.updated_count,
                'error_count': processor.error_count
            }

        except Exception as e:
            self.session.status = 'failed'
            self.session.error_message = str(e)
            self.session.completed_at = timezone.now()
            self.session.save()

            logger.error(f"Import processing failed: {str(e)}")
            raise

    def process_data(self, processor: BaseImportProcessor, data: List[Dict]):
        """Process the imported data"""

        for row_idx, row_data in enumerate(data, start=2):  # Start at 2 for header
            try:
                result = self.process_single_row(processor, row_data, row_idx)

                if result['status'] == 'created':
                    processor.created_count += 1
                elif result['status'] == 'updated':
                    processor.updated_count += 1
                elif result['status'] == 'skipped':
                    processor.skipped_count += 1

                processor.processed_rows += 1

                # Update progress every 10 rows
                if processor.processed_rows % 10 == 0:
                    processor.update_session_progress()

            except Exception as e:
                processor.error_count += 1
                processor.errors.append({
                    'row': row_idx,
                    'error': str(e),
                    'data': row_data
                })

                processor.log_message('error', f'Error processing row {row_idx}: {str(e)}', row_idx)
                processor.create_result('error', row_idx, error_message=str(e), raw_data=row_data)

        # Final progress update
        processor.update_session_progress()

    def process_single_row(self, processor: BaseImportProcessor, row_data: Dict, row_number: int) -> Dict:
        """Process a single row of data"""

        # Extract and clean data
        product_name = processor.clean_string_value(row_data.get('product_name'))
        sku = processor.clean_string_value(row_data.get('sku'))

        if not product_name and not sku:
            raise ValidationError("Either product_name or sku is required")

        # Get or create product
        product = self.get_or_create_product(processor, row_data, row_number)

        # Handle stock if quantity and store are provided
        if row_data.get('quantity') is not None and row_data.get('store'):
            self.update_stock(processor, product, row_data, row_number)

        processor.log_message('info', f'Successfully processed product: {product.name}', row_number)

        return {
            'status': 'created',  # or 'updated' based on logic
            'product': product
        }

    def get_or_create_product(self, processor: BaseImportProcessor, row_data: Dict, row_number: int) -> Product:
        """Get or create product from row data"""

        product_name = processor.clean_string_value(row_data.get('product_name'))
        sku = processor.clean_string_value(row_data.get('sku'))

        # Try to find existing product
        product = None
        if sku:
            try:
                product = Product.objects.get(sku=sku)
            except Product.DoesNotExist:
                pass

        if not product and product_name:
            try:
                product = Product.objects.get(name=product_name)
            except Product.DoesNotExist:
                pass

        # Create new product if not found
        if not product:
            product_data = {
                'name': product_name or sku,
                'sku': sku or f"AUTO-{timezone.now().strftime('%Y%m%d%H%M%S')}-{row_number}",
                'selling_price': processor.clean_decimal_value(row_data.get('selling_price')),
                'cost_price': processor.clean_decimal_value(row_data.get('cost_price')),
                'description': processor.clean_string_value(row_data.get('description')),
                'unit_of_measure': processor.clean_string_value(row_data.get('unit_of_measure'), 'each'),
                'min_stock_level': processor.clean_integer_value(row_data.get('min_stock_level'), 5),
                'is_active': True,
                'import_session': self.session,
                'imported_at': timezone.now()
            }

            # Handle barcode
            barcode = processor.clean_string_value(row_data.get('barcode'))
            if barcode:
                product_data['barcode'] = barcode

            # Handle category
            category_name = processor.clean_string_value(row_data.get('category'))
            if category_name:
                category, _ = Category.objects.get_or_create(
                    name=category_name,
                    defaults={'is_active': True}
                )
                product_data['category'] = category

            # Handle supplier
            supplier_name = processor.clean_string_value(row_data.get('supplier'))
            if supplier_name:
                supplier, _ = Supplier.objects.get_or_create(
                    name=supplier_name,
                    defaults={'is_active': True}
                )
                product_data['supplier'] = supplier

            product = Product.objects.create(**product_data)

            processor.create_result(
                'created', row_number,
                product_name=product.name,
                sku=product.sku,
                raw_data=row_data
            )

        return product

    def update_stock(self, processor: BaseImportProcessor, product: Product, row_data: Dict, row_number: int):
        """Update stock levels"""

        store_name = processor.clean_string_value(row_data.get('store'))
        quantity = processor.clean_decimal_value(row_data.get('quantity'))
        reorder_level = processor.clean_decimal_value(row_data.get('low_stock_threshold'))

        if not store_name:
            return

        # Get or create store
        store, _ = Store.objects.get_or_create(
            name=store_name,
            defaults={'is_active': True}
        )

        # Update stock
        stock, created = Stock.objects.get_or_create(
            product=product,
            store=store,
            defaults={
                'quantity': quantity,
                'reorder_level': reorder_level,
                'import_session': self.session,
                'last_import_update': timezone.now()
            }
        )

        if not created:
            # Handle conflict resolution
            if self.session.conflict_resolution == 'overwrite':
                stock.quantity = quantity
                if reorder_level > 0:
                    stock.reorder_level = reorder_level
            elif self.session.conflict_resolution == 'merge':
                stock.quantity += quantity
            # 'skip' means do nothing

            stock.import_session = self.session
            stock.last_import_update = timezone.now()
            stock.save()

        processor.create_result(
            'created' if created else 'updated',
            row_number,
            product_name=product.name,
            sku=product.sku,
            store_name=store.name,
            quantity=float(stock.quantity),
            old_quantity=float(stock.quantity) if not created else None,
            raw_data=row_data
        )