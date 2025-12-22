import csv
import logging
from io import BytesIO, StringIO
from decimal import Decimal, InvalidOperation
from datetime import datetime

import xlsxwriter
import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError
from difflib import get_close_matches

from inventory.models import (
    Product, Stock, Category, Supplier,
    ImportSession, ImportLog, ImportResult, StockMovement
)
from stores.models import Store

logger = logging.getLogger(__name__)


# ============================================================================
# SAMPLE FILE GENERATION VIEWS
# ============================================================================

@login_required
def download_sample_products_csv(request):
    """Generate CSV sample file for products with stock including category fields"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_products_stock.csv"'

    writer = csv.writer(response)

    # Headers with all fields including category-specific
    headers = [
        # Product Basic Info
        'Product Name*', 'SKU*', 'Barcode', 'Description',

        # Category & Supplier
        'Category Name*', 'Category Code', 'Category Description', 'Category Type*',
        'EFRIS Commodity Category Code', 'EFRIS Auto Sync', 'Supplier',

        # Pricing
        'Selling Price*', 'Cost Price*', 'Discount %',

        # Tax Information
        'Tax Rate*', 'Excise Duty Rate',
        'EFRIS Excise Duty Code',

        # Unit and Stock
        'Unit of Measure*', 'Min Stock Level', 'Is Active',

        # Store Stock Information
        'Store Name*', 'Quantity*', 'Low Stock Threshold', 'Reorder Quantity',

        # Additional EFRIS Fields
        'EFRIS Item Code', 'EFRIS Has Piece Unit', 'EFRIS Piece Measure Unit',
        'EFRIS Piece Unit Price', 'EFRIS Goods Code'
    ]
    writer.writerow(headers)

    # Sample data rows
    sample_data = [
        [
            # Product Basic Info
            'Coca Cola 500ml', 'CC-500ML', '5000112345678', 'Refreshing cola drink 500ml bottle',

            # Category & Supplier
            'Beverages', 'BEV-001', 'Soft drinks and beverages', 'product',
            '101113010000000000', 'Yes', 'Century Bottling',

            # Pricing
            '3000', '2000', '0',

            # Tax Information
            'A', '0', '',

            # Unit and Stock
            '102', '20', 'Yes',

            # Store Stock Information
            'Main Store', '100', '20', '50',

            # Additional EFRIS Fields
            '', 'No', '', '', ''
        ],
        [
            # Product Basic Info
            'Samsung Galaxy A54', 'SGH-A54-BLK', '8806094123456', 'Samsung Galaxy A54 128GB Black',

            # Category & Supplier
            'Electronics', 'ELEC-001', 'Mobile phones and accessories', 'product',
            '101113020000000000', 'Yes', 'Samsung Uganda',

            # Pricing
            '1500000', '1200000', '5',

            # Tax Information
            'A', '0', '',

            # Unit and Stock
            '101', '5', 'Yes',

            # Store Stock Information
            'Downtown Branch', '15', '5', '10',

            # Additional EFRIS Fields
            '', 'No', '', '', ''
        ],
        [
            # Product Basic Info
            'Rice 1KG', 'RICE-1KG', '', 'Premium basmati rice 1kg pack',

            # Category & Supplier
            'Food & Groceries', 'FOOD-001', 'Food items and groceries', 'product',
            '101113030000000000', 'Yes', 'Tilda Uganda',

            # Pricing
            '5000', '3500', '0',

            # Tax Information
            'B', '0', '',

            # Unit and Stock
            '103', '50', 'Yes',

            # Store Stock Information
            'Main Store', '200', '50', '100',

            # Additional EFRIS Fields
            '', 'No', '', '', ''
        ],
    ]

    for row in sample_data:
        writer.writerow(row)

    return response


@login_required
def download_sample_products_excel(request):
    """Generate Excel sample file with formatting and instructions including category fields"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)

    # Create worksheets
    data_sheet = workbook.add_worksheet('Products & Stock Data')
    instructions_sheet = workbook.add_worksheet('Instructions')
    reference_sheet = workbook.add_worksheet('Reference Data')
    category_ref_sheet = workbook.add_worksheet('Category Reference')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4F46E5',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    required_format = workbook.add_format({
        'bold': True,
        'bg_color': '#DC2626',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    category_format = workbook.add_format({
        'bold': True,
        'bg_color': '#10B981',  # Green for category fields
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    sample_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    instruction_header = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'font_color': '#1F2937'
    })

    instruction_text = workbook.add_format({
        'text_wrap': True,
        'valign': 'top'
    })

    # ========================================================================
    # DATA SHEET
    # ========================================================================

    headers = [
        # Product Basic Info
        ('Product Name*', True),
        ('SKU*', True),
        ('Barcode', False),
        ('Description', False),

        # Category Fields
        ('Category Name*', True),
        ('Category Code', False),
        ('Category Description', False),
        ('Category Type*', True),
        ('EFRIS Commodity Category Code', False),
        ('EFRIS Auto Sync', False),
        ('Supplier', False),

        # Pricing
        ('Selling Price*', True),
        ('Cost Price*', True),
        ('Discount %', False),

        # Tax Information
        ('Tax Rate*', True),
        ('Excise Duty Rate', False),
        ('EFRIS Excise Duty Code', False),

        # Unit and Stock
        ('Unit of Measure*', True),
        ('Min Stock Level', False),
        ('Is Active', False),

        # Store Stock Information
        ('Store Name*', True),
        ('Quantity*', True),
        ('Low Stock Threshold', False),
        ('Reorder Quantity', False),

        # Additional EFRIS Fields
        ('EFRIS Item Code', False),
        ('EFRIS Has Piece Unit', False),
        ('EFRIS Piece Measure Unit', False),
        ('EFRIS Piece Unit Price', False),
        ('EFRIS Goods Code', False),
    ]

    # Write headers with different colors for categories
    for col, (header, is_required) in enumerate(headers):
        if is_required:
            data_sheet.write(0, col, header, required_format)
        elif header.startswith('Category') or header.startswith('EFRIS'):
            data_sheet.write(0, col, header, category_format)
        else:
            data_sheet.write(0, col, header, header_format)

    # Sample data
    sample_data = [
        [
            # Product Basic Info
            'Coca Cola 500ml', 'CC-500ML', '5000112345678', 'Refreshing cola drink 500ml bottle',

            # Category & Supplier
            'Beverages', 'BEV-001', 'Soft drinks and beverages', 'product',
            '101113010000000000', 'Yes', 'Century Bottling',

            # Pricing
            3000, 2000, 0,

            # Tax Information
            'A', 0, '',

            # Unit and Stock
            '102', 20, 'Yes',

            # Store Stock Information
            'Main Store', 100, 20, 50,

            # Additional EFRIS Fields
            '', 'No', '', '', ''
        ],
        [
            # Product Basic Info
            'Samsung Galaxy A54', 'SGH-A54-BLK', '8806094123456', 'Samsung Galaxy A54 128GB Black',

            # Category & Supplier
            'Electronics', 'ELEC-001', 'Mobile phones and accessories', 'product',
            '101113020000000000', 'Yes', 'Samsung Uganda',

            # Pricing
            1500000, 1200000, 5,

            # Tax Information
            'A', 0, '',

            # Unit and Stock
            '101', 5, 'Yes',

            # Store Stock Information
            'Downtown Branch', 15, 5, 10,

            # Additional EFRIS Fields
            '', 'No', '', '', ''
        ],
        [
            # Product Basic Info
            'Rice 1KG', 'RICE-1KG', '', 'Premium basmati rice 1kg pack',

            # Category & Supplier
            'Food & Groceries', 'FOOD-001', 'Food items and groceries', 'product',
            '101113030000000000', 'Yes', 'Tilda Uganda',

            # Pricing
            5000, 3500, 0,

            # Tax Information
            'B', 0, '',

            # Unit and Stock
            '103', 50, 'Yes',

            # Store Stock Information
            'Main Store', 200, 50, 100,

            # Additional EFRIS Fields
            '', 'No', '', '', ''
        ],
    ]

    for row_idx, row_data in enumerate(sample_data, start=1):
        for col_idx, value in enumerate(row_data):
            data_sheet.write(row_idx, col_idx, value, sample_format)

    # Set column widths
    column_widths = [20, 15, 15, 25, 15, 15, 20, 15, 25, 15, 15, 12, 12, 10, 10, 12, 15, 12, 10, 15, 10, 15, 15, 15, 15,
                     15, 15]
    for col, width in enumerate(column_widths):
        data_sheet.set_column(col, col, width)

    # Freeze first row
    data_sheet.freeze_panes(1, 0)

    # ========================================================================
    # INSTRUCTIONS SHEET
    # ========================================================================

    instructions_sheet.set_column(0, 0, 50)
    instructions_sheet.set_column(1, 1, 60)

    row = 0
    instructions_sheet.write(row, 0, 'STOCK IMPORT INSTRUCTIONS', instruction_header)
    row += 2

    instructions = [
        ('Import Modes:', 'This template supports both:'),
        ('', '1. COMBINED IMPORT: Import new products with stock quantities'),
        ('', '2. STOCK UPDATE: Update quantities for existing products (SKU must exist)'),
        ('', ''),
        ('Required Fields:', 'Fields marked with * (red header) are mandatory'),
        ('', '- Product Name, SKU, Selling Price, Cost Price, Tax Rate, Unit of Measure'),
        ('', '- Category Name, Category Type, Store Name, Quantity'),
        ('', ''),
        ('Category Fields:', 'Green header fields relate to categories:'),
        ('', '- Category Name: Name of existing category or new category to create'),
        ('', '- Category Type: Must be "product" or "service"'),
        ('', '- EFRIS Commodity Category Code: 18-digit URA commodity code'),
        ('', '- EFRIS Auto Sync: Yes/No - Enable automatic EFRIS synchronization'),
        ('', ''),
        ('Category Creation:', 'If category does not exist:'),
        ('', '- System will create new category with provided details'),
        ('', '- EFRIS fields will be set if provided'),
        ('', '- Category will be marked as active'),
        ('', ''),
        ('EFRIS Validation:', 'For EFRIS-enabled categories:'),
        ('', '- EFRIS Commodity Category Code must be valid 18-digit code'),
        ('', '- Category must be a leaf node in EFRIS hierarchy'),
        ('', '- Category type must match EFRIS category type'),
        ('', ''),
        ('Tax Rates:', 'Use these codes:'),
        ('', 'A = Standard rate (18%)'),
        ('', 'B = Zero rate (0%)'),
        ('', 'C = Exempt (Not taxable)'),
        ('', 'D = Deemed rate (18%)'),
        ('', 'E = Excise Duty rate'),
        ('', ''),
        ('Unit of Measure:', 'Common codes:'),
        ('', '101 = Stick/Piece'),
        ('', '102 = Litre'),
        ('', '103 = Kilogram'),
        ('', '(See Reference Data sheet for full list)'),
        ('', ''),
        ('Category Type:', 'Must be one of:'),
        ('', 'product = Product Category'),
        ('', 'service = Service Category'),
        ('', ''),
        ('Suppliers:', 'If supplier doesn\'t exist:'),
        ('', '- System will create new supplier with basic details'),
        ('', '- Phone number will be set to default if not provided'),
        ('', ''),
        ('Store Names:', 'Must match existing store names exactly'),
        ('', '(Case-insensitive, but spelling must match)'),
        ('', ''),
        ('EFRIS Fields:', 'Optional fields for tax compliance:'),
        ('', '- EFRIS Excise Duty Code: For excisable goods'),
        ('', '- EFRIS Item Code: Internal EFRIS item code'),
        ('', '- EFRIS Has Piece Unit: Yes/No - Whether product has piece unit'),
        ('', '- EFRIS Piece Measure Unit: Unit code for piece measurement'),
        ('', '- EFRIS Piece Unit Price: Price per piece unit'),
        ('', '- EFRIS Goods Code: EFRIS-assigned goods code'),
        ('', ''),
        ('Tips:', '- Keep SKU unique for each product'),
        ('', '- Use consistent category names across imports'),
        ('', '- Verify EFRIS commodity codes before importing'),
        ('', '- Test with a few rows first'),
        ('', '- Check validation messages after upload'),
    ]

    for instruction in instructions:
        instructions_sheet.write(row, 0, instruction[0],
                                 instruction_header if instruction[0] and instruction[0].endswith(
                                     ':') else instruction_text)
        instructions_sheet.write(row, 1, instruction[1], instruction_text)
        row += 1

    # ========================================================================
    # REFERENCE DATA SHEET
    # ========================================================================

    reference_sheet.set_column(0, 0, 15)
    reference_sheet.set_column(1, 1, 40)

    row = 0
    reference_sheet.write(row, 0, 'UNIT CODES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Code', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    # Common unit codes
    common_units = [
        ('101', 'Stick/Piece'),
        ('102', 'Litre'),
        ('103', 'Kilogram'),
        ('104', 'User per day of access'),
        ('105', 'Minute'),
        ('106', '1000 sticks'),
        ('107', '50kgs'),
        ('108', '-'),
        ('109', 'Gram'),
        ('110', 'Box'),
        ('111', 'Pair'),
        ('112', 'Yard'),
        ('113', 'Dozen'),
    ]

    for code, desc in common_units:
        reference_sheet.write(row, 0, code, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    # Add tax rates reference
    row += 2
    reference_sheet.write(row, 0, 'TAX RATES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Code', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    tax_rates = [
        ('A', 'Standard rate (18%)'),
        ('B', 'Zero rate (0%)'),
        ('C', 'Exempt (Not taxable)'),
        ('D', 'Deemed rate (18%)'),
        ('E', 'Excise Duty rate'),
    ]

    for code, desc in tax_rates:
        reference_sheet.write(row, 0, code, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    # ========================================================================
    # CATEGORY REFERENCE SHEET
    # ========================================================================

    category_ref_sheet.set_column(0, 0, 25)
    category_ref_sheet.set_column(1, 1, 50)

    row = 0
    category_ref_sheet.write(row, 0, 'CATEGORY FIELD REFERENCE', instruction_header)
    row += 2

    category_fields = [
        ('Category Name*', 'Name of the category. Will be created if it doesn\'t exist.', 'Required'),
        ('Category Code', 'Internal category code (optional).', 'Optional'),
        ('Category Description', 'Description of the category.', 'Optional'),
        ('Category Type*', 'Must be "product" or "service".', 'Required'),
        ('EFRIS Commodity Category Code', '18-digit URA commodity category code.', 'Optional but required for EFRIS'),
        ('EFRIS Auto Sync', 'Yes/No - Enable automatic EFRIS synchronization.', 'Optional'),
    ]

    category_ref_sheet.write(row, 0, 'Field Name', category_format)
    category_ref_sheet.write(row, 1, 'Description', category_format)
    category_ref_sheet.write(row, 2, 'Required', category_format)
    row += 1

    for field_name, description, required in category_fields:
        category_ref_sheet.write(row, 0, field_name, sample_format)
        category_ref_sheet.write(row, 1, description, sample_format)
        category_ref_sheet.write(row, 2, required, sample_format)
        row += 1

    # Add category type reference
    row += 2
    category_ref_sheet.write(row, 0, 'CATEGORY TYPES', instruction_header)
    row += 1
    category_ref_sheet.write(row, 0, 'Type', category_format)
    category_ref_sheet.write(row, 1, 'Description', category_format)
    row += 1

    category_types = [
        ('product', 'Product Category - For physical goods and items'),
        ('service', 'Service Category - For services and intangible offerings'),
    ]

    for type_code, desc in category_types:
        category_ref_sheet.write(row, 0, type_code, sample_format)
        category_ref_sheet.write(row, 1, desc, sample_format)
        row += 1

    # Add EFRIS commodity code examples
    row += 2
    category_ref_sheet.write(row, 0, 'EFRIS COMMODITY CODE EXAMPLES', instruction_header)
    row += 1
    category_ref_sheet.write(row, 0, 'Commodity Code', category_format)
    category_ref_sheet.write(row, 1, 'Description', category_format)
    row += 1

    efris_examples = [
        ('101113010000000000', 'Beverages - Soft Drinks'),
        ('101113020000000000', 'Electronics - Mobile Phones'),
        ('101113030000000000', 'Food & Groceries - Rice & Grains'),
        ('101113040000000000', 'Clothing & Apparel'),
        ('101113050000000000', 'Household Items'),
    ]

    for code, desc in efris_examples:
        category_ref_sheet.write(row, 0, code, sample_format)
        category_ref_sheet.write(row, 1, desc, sample_format)
        row += 1

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_products_stock.xlsx"'

    return response

@login_required
def download_sample_stock_only_csv(request):
    """Generate CSV sample for stock-only updates"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_stock_update.csv"'

    writer = csv.writer(response)

    # Headers for stock-only updates
    headers = ['SKU*', 'Store Name*', 'Quantity*', 'Low Stock Threshold', 'Reorder Quantity']
    writer.writerow(headers)

    # Sample data
    sample_data = [
        ['CC-500ML', 'Main Store', '150', '20', '50'],
        ['SGH-A54-BLK', 'Downtown Branch', '25', '5', '10'],
        ['RICE-1KG', 'Main Store', '300', '50', '100'],
    ]

    for row in sample_data:
        writer.writerow(row)

    return response


@login_required
def download_sample_stock_only_excel(request):
    """Generate Excel sample for stock-only updates"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Stock Updates')

    # Formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4F46E5',
        'font_color': 'white',
        'border': 1,
        'align': 'center'
    })

    required_format = workbook.add_format({
        'bold': True,
        'bg_color': '#DC2626',
        'font_color': 'white',
        'border': 1,
        'align': 'center'
    })

    sample_format = workbook.add_format({'border': 1})

    # Headers
    headers = [
        ('SKU*', True),
        ('Store Name*', True),
        ('Quantity*', True),
        ('Low Stock Threshold', False),
        ('Reorder Quantity', False),
    ]

    for col, (header, is_required) in enumerate(headers):
        worksheet.write(0, col, header, required_format if is_required else header_format)

    # Sample data
    sample_data = [
        ['CC-500ML', 'Main Store', 150, 20, 50],
        ['SGH-A54-BLK', 'Downtown Branch', 25, 5, 10],
        ['RICE-1KG', 'Main Store', 300, 50, 100],
    ]

    for row_idx, row_data in enumerate(sample_data, start=1):
        for col_idx, value in enumerate(row_data):
            worksheet.write(row_idx, col_idx, value, sample_format)

    # Set column widths
    worksheet.set_column(0, 0, 20)
    worksheet.set_column(1, 1, 20)
    worksheet.set_column(2, 2, 15)
    worksheet.set_column(3, 3, 20)
    worksheet.set_column(4, 4, 20)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_stock_update.xlsx"'

    return response


# ============================================================================
# IMPORT PROCESSING LOGIC
# ============================================================================

class ColumnMapper:
    """Smart column mapper with fuzzy matching"""

    # Define standard column mappings
    COLUMN_MAPPINGS = {
        'product_name': ['product name', 'name', 'product', 'item name', 'item', 'goods name'],
        'sku': ['sku', 'sku code', 'product code', 'code', 'item code'],
        'barcode': ['barcode', 'bar code', 'ean', 'upc'],
        'category': ['category', 'cat', 'product category', 'category name'],
        'supplier': ['supplier', 'vendor', 'supplier name'],
        'selling_price': ['selling price', 'price', 'unit price', 'retail price', 'sale price'],
        'cost_price': ['cost price', 'cost', 'purchase price', 'buying price'],
        'discount_percentage': ['discount', 'discount %', 'discount percentage', 'disc %'],
        'tax_rate': ['tax rate', 'tax', 'vat rate', 'tax code'],
        'excise_duty_rate': ['excise duty rate', 'excise rate', 'excise duty', 'excise'],
        'unit_of_measure': ['unit', 'unit of measure', 'uom', 'measurement unit'],
        'min_stock_level': ['min stock', 'minimum stock', 'min stock level', 'minimum'],
        'description': ['description', 'desc', 'notes', 'details'],
        'store_name': ['store', 'store name', 'location', 'branch', 'shop'],
        'quantity': ['quantity', 'qty', 'stock', 'stock quantity', 'amount'],
        'low_stock_threshold': ['low stock threshold', 'reorder level', 'low stock', 'threshold'],
        'reorder_quantity': ['reorder quantity', 'reorder qty', 'reorder', 'order quantity'],
        'efris_commodity_code': ['efris commodity code', 'commodity code', 'efris code'],
        'efris_excise_duty_code': ['efris excise duty code', 'excise duty code', 'efris excise'],
        'efris_auto_sync': ['efris auto sync', 'auto sync', 'efris sync', 'sync'],
    }

    @classmethod
    def map_columns(cls, file_headers):
        """
        Map file headers to standardized column names
        Returns: dict mapping file_header -> standard_name
        """
        mapped = {}
        file_headers_lower = [h.lower().strip() for h in file_headers]

        for standard_name, variations in cls.COLUMN_MAPPINGS.items():
            # Try exact match first
            for variation in variations:
                if variation in file_headers_lower:
                    idx = file_headers_lower.index(variation)
                    mapped[file_headers[idx]] = standard_name
                    break

            # If no exact match, try fuzzy matching
            if standard_name not in mapped.values():
                for header in file_headers:
                    matches = get_close_matches(header.lower(), variations, n=1, cutoff=0.8)
                    if matches:
                        mapped[header] = standard_name
                        break

        return mapped

    @classmethod
    def get_mapping_suggestions(cls, file_headers):
        """Get suggestions for unmapped columns"""
        mapped = cls.map_columns(file_headers)
        suggestions = {}

        for header in file_headers:
            if header not in mapped:
                # Find closest matches
                all_variations = []
                for variations in cls.COLUMN_MAPPINGS.values():
                    all_variations.extend(variations)

                matches = get_close_matches(header.lower(), all_variations, n=3, cutoff=0.6)
                if matches:
                    suggestions[header] = matches

        return suggestions


def parse_uploaded_file(file_obj):
    """Parse CSV or Excel file and return headers + data"""
    file_extension = file_obj.name.split('.')[-1].lower()

    if file_extension == 'csv':
        return parse_csv_file(file_obj)
    elif file_extension in ['xlsx', 'xls']:
        return parse_excel_file(file_obj)
    else:
        raise ValueError(f"Unsupported file type: {file_extension}")


def parse_csv_file(file_obj):
    """Parse CSV file"""
    file_obj.seek(0)
    content = file_obj.read().decode('utf-8-sig')  # Handle BOM
    reader = csv.DictReader(StringIO(content))

    headers = reader.fieldnames
    data = list(reader)

    return headers, data


def parse_excel_file(file_obj):
    """Parse Excel file"""
    file_obj.seek(0)
    workbook = openpyxl.load_workbook(file_obj, read_only=True)
    sheet = workbook.active

    # Get headers from first row
    headers = [cell.value for cell in sheet[1]]

    # Get data rows
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip empty rows
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

    workbook.close()
    return headers, data


def validate_row_data(row_data, mapped_columns, import_mode, row_number):
    """
    Validate a single row of data
    Returns: (is_valid, errors_list, cleaned_data)
    """
    errors = []
    cleaned = {}

    # Required fields based on import mode
    if import_mode == 'combined':
        required_fields = ['product_name', 'sku', 'selling_price', 'cost_price',
                          'tax_rate', 'unit_of_measure', 'store_name', 'quantity']
    else:  # stock_only
        required_fields = ['sku', 'store_name', 'quantity']

    # Check required fields
    for field in required_fields:
        value = None
        for file_header, standard_name in mapped_columns.items():
            if standard_name == field:
                value = row_data.get(file_header)
                break

        if not value or (isinstance(value, str) and not value.strip()):
            field_display = field.replace('_', ' ').title()
            errors.append(f"{field_display} is required")
        else:
            cleaned[field] = value

    # If required fields missing, return early
    if errors:
        return False, errors, cleaned

    # Validate and clean optional fields
    for file_header, standard_name in mapped_columns.items():
        if standard_name in cleaned:
            continue  # Already processed

        value = row_data.get(file_header)
        if value and isinstance(value, str):
            value = value.strip()

        if not value:
            continue

        # Type-specific validation
        try:
            if standard_name in ['selling_price', 'cost_price', 'discount_percentage',
                                'excise_duty_rate', 'quantity', 'min_stock_level',
                                'low_stock_threshold', 'reorder_quantity']:
                cleaned[standard_name] = Decimal(str(value))
                if cleaned[standard_name] < 0:
                    errors.append(f"{standard_name.replace('_', ' ').title()} cannot be negative")

            elif standard_name == 'tax_rate':
                value = str(value).upper().strip()
                if value not in ['A', 'B', 'C', 'D', 'E']:
                    errors.append(f"Invalid tax rate '{value}'. Must be A, B, C, D, or E")
                else:
                    cleaned[standard_name] = value

            elif standard_name == 'efris_auto_sync':
                value_lower = str(value).lower().strip()
                cleaned[standard_name] = value_lower in ['yes', 'true', '1', 'y']

            else:
                cleaned[standard_name] = value

        except (ValueError, InvalidOperation) as e:
            errors.append(f"Invalid {standard_name.replace('_', ' ')}: {value}")

    is_valid = len(errors) == 0
    return is_valid, errors, cleaned


@transaction.atomic
def process_import_file(file_obj, import_mode, conflict_resolution, user, column_mapping=None, has_header=True):
    """
    Main import processing function

    Args:
        file_obj: Uploaded file object
        import_mode: 'combined' or 'stock_only'
        conflict_resolution: 'overwrite' or 'skip'
        user: User performing the import
        column_mapping: Optional manual column mapping dict
        has_header: Whether file has header row

    Returns:
        dict with import results
    """

    # Create import session
    session = ImportSession.objects.create(
        user=user,
        filename=file_obj.name,
        file_size=file_obj.size,
        import_mode=import_mode,
        conflict_resolution=conflict_resolution,
        has_header=has_header,
        status='processing',
        started_at=timezone.now()
    )

    try:
        # Parse file
        headers, data = parse_uploaded_file(file_obj)
        session.total_rows = len(data)
        session.save()

        # Map columns
        if column_mapping:
            mapped_columns = column_mapping
        else:
            mapped_columns = ColumnMapper.map_columns(headers)

        session.column_mapping = mapped_columns
        session.save()

        # Log column mapping
        ImportLog.objects.create(
            session=session,
            level='info',
            message=f'Column mapping completed: {len(mapped_columns)} columns mapped'
        )

        # Process each row
        results = {
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'error_count': 0,
            'errors': [],
            'warnings': []
        }

        for idx, row_data in enumerate(data, start=2):  # Start at 2 for Excel row numbers
            try:
                result = process_single_row(
                    row_data=row_data,
                    mapped_columns=mapped_columns,
                    import_mode=import_mode,
                    conflict_resolution=conflict_resolution,
                    user=user,
                    session=session,
                    row_number=idx
                )

                # Update counts
                if result['status'] == 'created':
                    results['created_count'] += 1
                elif result['status'] == 'updated':
                    results['updated_count'] += 1
                elif result['status'] == 'skipped':
                    results['skipped_count'] += 1
                elif result['status'] == 'error':
                    results['error_count'] += 1
                    results['errors'].append({
                        'row': idx,
                        'errors': result['errors']
                    })

                session.processed_rows += 1

            except Exception as e:
                logger.error(f"Error processing row {idx}: {str(e)}", exc_info=True)
                results['error_count'] += 1
                results['errors'].append({
                    'row': idx,
                    'errors': [f"Unexpected error: {str(e)}"]
                })

                ImportLog.objects.create(
                    session=session,
                    level='error',
                    message=f'Row {idx}: Unexpected error',
                    row_number=idx,
                    details={'error': str(e)}
                )

            # Update session periodically
            if idx % 10 == 0:
                session.created_count = results['created_count']
                session.updated_count = results['updated_count']
                session.skipped_count = results['skipped_count']
                session.error_count = results['error_count']
                session.save()

        # Final session update
        session.created_count = results['created_count']
        session.updated_count = results['updated_count']
        session.skipped_count = results['skipped_count']
        session.error_count = results['error_count']
        session.status = 'completed'
        session.completed_at = timezone.now()
        session.save()

        return results

    except Exception as e:
        logger.error(f"Import failed: {str(e)}", exc_info=True)
        session.status = 'failed'
        session.error_message = str(e)
        session.completed_at = timezone.now()
        session.save()

        raise


def process_single_row(row_data, mapped_columns, import_mode, conflict_resolution, user, session, row_number):
    """Process a single row of import data with proper error handling"""
    from django.db import transaction

    # Validate row
    is_valid, errors, cleaned_data = validate_row_data(row_data, mapped_columns, import_mode, row_number)

    if not is_valid:
        ImportResult.objects.create(
            session=session,
            result_type='error',
            row_number=row_number,
            error_message='; '.join(errors),
            raw_data=row_data
        )
        return {'status': 'error', 'errors': errors}

    # Use savepoint to handle errors gracefully
    try:
        sid = transaction.savepoint()

        if import_mode == 'combined':
            result = process_combined_import(cleaned_data, conflict_resolution, user, session, row_number, row_data)
        else:
            result = process_stock_only_import(cleaned_data, conflict_resolution, user, session, row_number, row_data)

        transaction.savepoint_commit(sid)
        return result

    except ValueError as e:
        # Handle validation errors (like duplicate barcode)
        transaction.savepoint_rollback(sid)
        error_msg = str(e)
        logger.warning(f"Validation error on row {row_number}: {error_msg}")

        ImportResult.objects.create(
            session=session,
            result_type='error',
            row_number=row_number,
            error_message=error_msg,
            raw_data=row_data
        )
        return {'status': 'error', 'errors': [error_msg]}

    except Exception as e:
        # Handle any other errors
        transaction.savepoint_rollback(sid)
        error_msg = f"Unexpected error: {str(e)}"
        logger.error(f"Error processing row {row_number}: {error_msg}", exc_info=True)

        ImportResult.objects.create(
            session=session,
            result_type='error',
            row_number=row_number,
            error_message=error_msg,
            raw_data=row_data
        )
        return {'status': 'error', 'errors': [error_msg]}


def process_combined_import(cleaned_data, conflict_resolution, user, session, row_number, raw_data):
    """Process combined product + stock import"""

    sku = cleaned_data['sku']
    product = None
    created = False

    # Check if product exists
    try:
        product = Product.objects.get(sku=sku)

        if conflict_resolution == 'skip':
            ImportResult.objects.create(
                session=session,
                result_type='skipped',
                row_number=row_number,
                product_name=cleaned_data.get('product_name', ''),
                sku=sku,
                raw_data=raw_data
            )
            return {'status': 'skipped', 'message': 'Product already exists'}

        # Overwrite mode - update product
        product.name = cleaned_data['product_name']
        product.selling_price = cleaned_data['selling_price']
        product.cost_price = cleaned_data['cost_price']
        product.tax_rate = cleaned_data.get('tax_rate', 'A')
        product.unit_of_measure = cleaned_data.get('unit_of_measure', '103')

        # Optional fields
        if 'barcode' in cleaned_data:
            product.barcode = cleaned_data['barcode']
        if 'description' in cleaned_data:
            product.description = cleaned_data['description']
        if 'discount_percentage' in cleaned_data:
            product.discount_percentage = cleaned_data['discount_percentage']
        if 'excise_duty_rate' in cleaned_data:
            product.excise_duty_rate = cleaned_data['excise_duty_rate']
        if 'min_stock_level' in cleaned_data:
            product.min_stock_level = int(cleaned_data['min_stock_level'])

        # EFRIS fields
        if 'efris_commodity_code' in cleaned_data:
            # Try to find category with this commodity code
            try:
                category = Category.objects.get(efris_commodity_category_code=cleaned_data['efris_commodity_code'])
                product.category = category
            except Category.DoesNotExist:
                pass

        if 'efris_excise_duty_code' in cleaned_data:
            product.efris_excise_duty_code = cleaned_data['efris_excise_duty_code']

        if 'efris_auto_sync' in cleaned_data:
            product.efris_auto_sync_enabled = cleaned_data['efris_auto_sync']

        product.import_session = session
        product.save()

    except Product.DoesNotExist:
        # Create new product
        created = True

        product = Product(
            sku=sku,
            name=cleaned_data['product_name'],
            selling_price=cleaned_data['selling_price'],
            cost_price=cleaned_data['cost_price'],
            tax_rate=cleaned_data.get('tax_rate', 'A'),
            unit_of_measure=cleaned_data.get('unit_of_measure', '103'),
            discount_percentage=cleaned_data.get('discount_percentage', 0),
            excise_duty_rate=cleaned_data.get('excise_duty_rate', 0),
            min_stock_level=int(cleaned_data.get('min_stock_level', 5)),
            description=cleaned_data.get('description', ''),
            barcode=cleaned_data.get('barcode', ''),
            import_session=session,
            imported_at=timezone.now()
        )

        # Handle category
        if 'category' in cleaned_data:
            category, _ = Category.objects.get_or_create(
                name=cleaned_data['category'],
                defaults={'is_active': True}
            )
            product.category = category
        elif 'efris_commodity_code' in cleaned_data:
            try:
                category = Category.objects.get(efris_commodity_category_code=cleaned_data['efris_commodity_code'])
                product.category = category
            except Category.DoesNotExist:
                pass

        # Handle supplier
        if 'supplier' in cleaned_data:
            supplier, _ = Supplier.objects.get_or_create(
                name=cleaned_data['supplier'],
                defaults={
                    'phone': '0000000000',
                    'is_active': True
                }
            )
            product.supplier = supplier

        # EFRIS fields
        if 'efris_excise_duty_code' in cleaned_data:
            product.efris_excise_duty_code = cleaned_data['efris_excise_duty_code']

        if 'efris_auto_sync' in cleaned_data:
            product.efris_auto_sync_enabled = cleaned_data['efris_auto_sync']

        product.save()

    # Process stock for the store
    store_name = cleaned_data['store_name']

    try:
        store = Store.objects.get(name__iexact=store_name)
    except Store.DoesNotExist:
        error_msg = f"Store '{store_name}' not found. Please create the store first."
        ImportResult.objects.create(
            session=session,
            result_type='error',
            row_number=row_number,
            product_name=product.name,
            sku=sku,
            store_name=store_name,
            error_message=error_msg,
            raw_data=raw_data
        )
        return {'status': 'error', 'errors': [error_msg]}

    # Get or create stock record with transaction lock
    from django.db import transaction

    with transaction.atomic():
        stock, stock_created = Stock.objects.select_for_update().get_or_create(
            product=product,
            store=store,
            defaults={
                'quantity': Decimal('0'),
                'low_stock_threshold': cleaned_data.get('low_stock_threshold', 5),
                'reorder_quantity': cleaned_data.get('reorder_quantity', 10),
            }
        )

        # Get quantities
        old_quantity = Decimal(str(stock.quantity))
        new_quantity = Decimal(str(cleaned_data['quantity']))

        # Only proceed if quantities are different or new stock record
        if stock_created or old_quantity != new_quantity:
            # Calculate difference
            quantity_diff = new_quantity - old_quantity

            # Create stock movement record for audit trail
            movement = StockMovement(
                product=product,
                store=store,
                movement_type='ADJUSTMENT',
                quantity=quantity_diff,
                reference=f'Import #{session.id}',
                notes=f'Import: Changed quantity from {old_quantity} to {new_quantity} via {session.filename}',
                created_by=user
            )
            # Save movement WITHOUT triggering automatic stock update
            # (we'll update stock directly below)
            super(StockMovement, movement).save()

            # Update stock quantity DIRECTLY
            stock.quantity = new_quantity
            stock.last_import_update = timezone.now()
            stock.import_session = session

            # Update optional stock fields
            if 'low_stock_threshold' in cleaned_data:
                stock.low_stock_threshold = cleaned_data['low_stock_threshold']
            if 'reorder_quantity' in cleaned_data:
                stock.reorder_quantity = cleaned_data['reorder_quantity']

            # Save with specific fields to avoid conflicts
            stock.save(update_fields=['quantity', 'low_stock_threshold', 'reorder_quantity',
                                     'last_import_update', 'import_session'])

            logger.info(f"✅ Stock updated: {product.name} at {store.name} - {old_quantity} → {new_quantity}")

    # Update optional stock fields
    if 'low_stock_threshold' in cleaned_data:
        stock.low_stock_threshold = cleaned_data['low_stock_threshold']
    if 'reorder_quantity' in cleaned_data:
        stock.reorder_quantity = cleaned_data['reorder_quantity']

    stock.last_import_update = timezone.now()
    stock.import_session = session
    stock.save()

    # Create import result
    result_type = 'created' if created else 'updated'
    ImportResult.objects.create(
        session=session,
        result_type=result_type,
        row_number=row_number,
        product_name=product.name,
        sku=sku,
        store_name=store.name,
        quantity=int(new_quantity),
        old_quantity=int(old_quantity) if not created else None,
        raw_data=raw_data
    )

    return {
        'status': result_type,
        'product': product,
        'stock': stock
    }


def process_stock_only_import(cleaned_data, conflict_resolution, user, session, row_number, raw_data):
    """Process stock-only update (no product creation)"""

    sku = cleaned_data['sku']

    # Product must exist
    try:
        product = Product.objects.get(sku=sku)
    except Product.DoesNotExist:
        error_msg = f"Product with SKU '{sku}' not found. Cannot update stock."
        ImportResult.objects.create(
            session=session,
            result_type='error',
            row_number=row_number,
            sku=sku,
            error_message=error_msg,
            raw_data=raw_data
        )
        return {'status': 'error', 'errors': [error_msg]}

    # Store must exist
    store_name = cleaned_data['store_name']
    try:
        store = Store.objects.get(name__iexact=store_name)
    except Store.DoesNotExist:
        error_msg = f"Store '{store_name}' not found."
        ImportResult.objects.create(
            session=session,
            result_type='error',
            row_number=row_number,
            product_name=product.name,
            sku=sku,
            store_name=store_name,
            error_message=error_msg,
            raw_data=raw_data
        )
        return {'status': 'error', 'errors': [error_msg]}

    # Get or create stock record with transaction lock
    from django.db import transaction

    with transaction.atomic():
        stock, created = Stock.objects.select_for_update().get_or_create(
            product=product,
            store=store,
            defaults={
                'quantity': Decimal('0'),
                'low_stock_threshold': cleaned_data.get('low_stock_threshold', 5),
                'reorder_quantity': cleaned_data.get('reorder_quantity', 10),
            }
        )

        # Get quantities
        old_quantity = Decimal(str(stock.quantity))
        new_quantity = Decimal(str(cleaned_data['quantity']))

        # Only proceed if quantities are different or new record
        if created or old_quantity != new_quantity:
            # Calculate difference
            quantity_diff = new_quantity - old_quantity

            # Create stock movement for audit trail
            movement = StockMovement(
                product=product,
                store=store,
                movement_type='ADJUSTMENT',
                quantity=quantity_diff,
                reference=f'Stock Update Import #{session.id}',
                notes=f'Stock update: Changed from {old_quantity} to {new_quantity} via {session.filename}',
                created_by=user
            )
            # Save WITHOUT triggering automatic stock update
            super(StockMovement, movement).save()

            # Update stock quantity DIRECTLY
            stock.quantity = new_quantity
            stock.last_import_update = timezone.now()
            stock.import_session = session

            # Update optional fields
            if 'low_stock_threshold' in cleaned_data:
                stock.low_stock_threshold = cleaned_data['low_stock_threshold']
            if 'reorder_quantity' in cleaned_data:
                stock.reorder_quantity = cleaned_data['reorder_quantity']

            # Save with specific fields
            stock.save(update_fields=['quantity', 'low_stock_threshold', 'reorder_quantity',
                                     'last_import_update', 'import_session'])

            logger.info(f"✅ Stock updated: {product.name} at {store.name} - {old_quantity} → {new_quantity}")

    # Update optional fields
    if 'low_stock_threshold' in cleaned_data:
        stock.low_stock_threshold = cleaned_data['low_stock_threshold']
    if 'reorder_quantity' in cleaned_data:
        stock.reorder_quantity = cleaned_data['reorder_quantity']

    stock.last_import_update = timezone.now()
    stock.import_session = session
    stock.save()

    # Create import result
    ImportResult.objects.create(
        session=session,
        result_type='updated',
        row_number=row_number,
        product_name=product.name,
        sku=sku,
        store_name=store.name,
        quantity=int(new_quantity),
        old_quantity=int(old_quantity),
        raw_data=raw_data
    )

    return {
        'status': 'updated',
        'product': product,
        'stock': stock
    }


# ============================================================================
# MAIN IMPORT VIEWS
# ============================================================================

@login_required
@permission_required('inventory.add_product', raise_exception=True)
def stock_import(request):
    """Main stock import view"""

    if request.method == 'GET':
        # Get recent import sessions
        recent_imports = ImportSession.objects.filter(
            user=request.user
        ).order_by('-created_at')[:10]

        # Get available stores
        stores = Store.objects.filter(is_active=True).order_by('name')

        context = {
            'recent_imports': recent_imports,
            'stores': stores,
        }

        return render(request, 'inventory/stock_import.html', context)

    elif request.method == 'POST':
        try:
            # Get form data
            uploaded_file = request.FILES.get('import_file')
            import_mode = request.POST.get('import_mode', 'combined')
            conflict_resolution = request.POST.get('conflict_resolution', 'overwrite')

            if not uploaded_file:
                messages.error(request, 'Please select a file to import.')
                return redirect('inventory:stock_import')

            # Validate file type
            file_extension = uploaded_file.name.split('.')[-1].lower()
            if file_extension not in ['csv', 'xlsx', 'xls']:
                messages.error(request, 'Only CSV and Excel files are supported.')
                return redirect('inventory:stock_import')

            # Process file
            results = process_import_file(
                file_obj=uploaded_file,
                import_mode=import_mode,
                conflict_resolution=conflict_resolution,
                user=request.user
            )

            # Show results
            if results['created_count'] > 0:
                messages.success(
                    request,
                    f'Successfully created {results["created_count"]} products with stock.'
                )

            if results['updated_count'] > 0:
                messages.success(
                    request,
                    f'Successfully updated {results["updated_count"]} records.'
                )

            if results['skipped_count'] > 0:
                messages.info(
                    request,
                    f'{results["skipped_count"]} records were skipped (already exist).'
                )

            if results['error_count'] > 0:
                messages.warning(
                    request,
                    f'{results["error_count"]} records had errors and were not imported.'
                )
                # Store errors in session for display
                request.session['import_errors'] = results['errors'][:50]  # Limit to 50

            # Redirect to results page
            return redirect('inventory:stock_import')

        except Exception as e:
            logger.error(f"Stock import error: {str(e)}", exc_info=True)
            messages.error(request, f'Import failed: {str(e)}')
            return redirect('inventory:stock_import')


@login_required
def import_session_detail(request, session_id):
    """View details of a specific import session"""

    session = ImportSession.objects.get(id=session_id, user=request.user)

    # Get results
    results = ImportResult.objects.filter(session=session).order_by('row_number')
    logs = ImportLog.objects.filter(session=session).order_by('timestamp')

    # Pagination for results
    from django.core.paginator import Paginator

    paginator = Paginator(results, 50)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'session': session,
        'results': page_obj,
        'logs': logs,
        'error_results': results.filter(result_type='error'),
    }

    return render(request, 'inventory/import_session_detail.html', context)


@login_required
def preview_import(request):
    """Preview import data before processing"""

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        uploaded_file = request.FILES.get('preview_file')
        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        # Parse file
        headers, data = parse_uploaded_file(uploaded_file)

        # Map columns
        mapped_columns = ColumnMapper.map_columns(headers)
        suggestions = ColumnMapper.get_mapping_suggestions(headers)

        # Get preview data (first 10 rows)
        preview_data = data[:10]

        # Validate preview rows
        preview_results = []
        for idx, row in enumerate(preview_data, start=2):
            is_valid, errors, cleaned = validate_row_data(
                row,
                mapped_columns,
                request.POST.get('import_mode', 'combined'),
                idx
            )
            preview_results.append({
                'row_number': idx,
                'data': row,
                'is_valid': is_valid,
                'errors': errors,
                'cleaned': {k: str(v) for k, v in cleaned.items()}
            })

        return JsonResponse({
            'success': True,
            'headers': headers,
            'mapped_columns': mapped_columns,
            'suggestions': suggestions,
            'total_rows': len(data),
            'preview_data': preview_results,
        })

    except Exception as e:
        logger.error(f"Preview error: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def validate_import_data(request):
    """Validate import data without saving"""

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        uploaded_file = request.FILES.get('validate_file')
        import_mode = request.POST.get('import_mode', 'combined')

        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        # Parse file
        headers, data = parse_uploaded_file(uploaded_file)

        # Map columns
        mapped_columns = ColumnMapper.map_columns(headers)

        # Validate all rows
        validation_results = {
            'total_rows': len(data),
            'valid_rows': 0,
            'invalid_rows': 0,
            'errors': []
        }

        for idx, row in enumerate(data, start=2):
            is_valid, errors, cleaned = validate_row_data(
                row,
                mapped_columns,
                import_mode,
                idx
            )

            if is_valid:
                validation_results['valid_rows'] += 1
            else:
                validation_results['invalid_rows'] += 1
                validation_results['errors'].append({
                    'row': idx,
                    'errors': errors,
                    'data': row
                })

        return JsonResponse({
            'success': True,
            'validation': validation_results,
            'mapped_columns': mapped_columns
        })

    except Exception as e:
        logger.error(f"Validation error: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=400)





# ============================================================================
# PRODUCT EXPORT VIEWS
# ============================================================================

@login_required
@permission_required('inventory.view_product', raise_exception=True)
def export_products_selection(request):
    """View to select export options before generating file"""

    if request.method == 'GET':
        # Get filter options
        categories = Category.objects.filter(is_active=True).order_by('name')
        suppliers = Supplier.objects.filter(is_active=True).order_by('name')
        stores = Store.objects.filter(is_active=True).order_by('name')

        # Get product counts
        total_products = Product.objects.filter(is_active=True).count()

        context = {
            'categories': categories,
            'suppliers': suppliers,
            'stores': stores,
            'total_products': total_products,
        }

        return render(request, 'inventory/export_products_selection.html', context)


@login_required
@permission_required('inventory.view_product', raise_exception=True)
def export_products_csv(request):
    """Export products to CSV format compatible with import template"""

    # Get filter parameters
    export_mode = request.GET.get('mode', 'products_only')  # products_only, with_stock
    category_ids = request.GET.getlist('categories')
    supplier_ids = request.GET.getlist('suppliers')
    store_ids = request.GET.getlist('stores')
    include_inactive = request.GET.get('include_inactive', 'false') == 'true'

    # Build queryset
    products = Product.objects.select_related('category', 'supplier')

    if not include_inactive:
        products = products.filter(is_active=True)

    if category_ids:
        products = products.filter(category_id__in=category_ids)

    if supplier_ids:
        products = products.filter(supplier_id__in=supplier_ids)

    products = products.order_by('sku')

    # Create response
    response = HttpResponse(content_type='text/csv')

    if export_mode == 'with_stock':
        response['Content-Disposition'] = 'attachment; filename="products_with_stock_export.csv"'
        return _export_products_with_stock_csv(response, products, store_ids)
    else:
        response['Content-Disposition'] = 'attachment; filename="products_export.csv"'
        return _export_products_only_csv(response, products)


def _export_products_only_csv(response, products):
    """Export products only (no stock) to CSV"""

    writer = csv.writer(response)

    # Headers matching import template
    headers = [
        'Product Name*', 'SKU*', 'Barcode', 'Category', 'Supplier',
        'Selling Price*', 'Cost Price*', 'Discount %',
        'Tax Rate*', 'Excise Duty Rate', 'Unit of Measure*',
        'Min Stock Level', 'Description', 'Is Active',
        'EFRIS Commodity Code', 'EFRIS Excise Duty Code', 'EFRIS Auto Sync',
        'EFRIS Item Code', 'EFRIS Has Piece Unit', 'EFRIS Piece Measure Unit',
        'EFRIS Piece Unit Price', 'EFRIS Goods Code'
    ]
    writer.writerow(headers)

    # Write product data
    for product in products:
        row = [
            product.name,
            product.sku,
            product.barcode or '',
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            str(product.selling_price),
            str(product.cost_price),
            str(product.discount_percentage),
            product.tax_rate,
            str(product.excise_duty_rate),
            product.unit_of_measure,
            str(product.min_stock_level),
            product.description or '',
            'Yes' if product.is_active else 'No',
            # EFRIS fields
            getattr(product.category, 'efris_commodity_category_code', '') if product.category else '',
            product.efris_excise_duty_code or '',
            'Yes' if product.efris_auto_sync_enabled else 'No',
            product.efris_item_code or '',
            'Yes' if product.efris_has_piece_unit else 'No',
            product.efris_piece_measure_unit or '',
            str(product.efris_piece_unit_price) if product.efris_piece_unit_price else '',
            product.efris_goods_code_field or '',
        ]
        writer.writerow(row)

    return response


def _export_products_with_stock_csv(response, products, store_ids):
    """Export products with stock information to CSV"""

    writer = csv.writer(response)

    # Headers matching combined import template
    headers = [
        # Product Basic Info
        'Product Name*', 'SKU*', 'Barcode', 'Description',

        # Category & Supplier
        'Category Name*', 'Category Code', 'Category Description', 'Category Type*',
        'EFRIS Commodity Category Code', 'EFRIS Auto Sync', 'Supplier',

        # Pricing
        'Selling Price*', 'Cost Price*', 'Discount %',

        # Tax Information
        'Tax Rate*', 'Excise Duty Rate', 'EFRIS Excise Duty Code',

        # Unit and Stock
        'Unit of Measure*', 'Min Stock Level', 'Is Active',

        # Store Stock Information
        'Store Name*', 'Quantity*', 'Low Stock Threshold', 'Reorder Quantity',

        # Additional EFRIS Fields
        'EFRIS Item Code', 'EFRIS Has Piece Unit', 'EFRIS Piece Measure Unit',
        'EFRIS Piece Unit Price', 'EFRIS Goods Code'
    ]
    writer.writerow(headers)

    # Get stores to export
    if store_ids:
        stores = Store.objects.filter(id__in=store_ids, is_active=True)
    else:
        stores = Store.objects.filter(is_active=True)

    # Write product data with stock for each store
    for product in products:
        stocks = Stock.objects.filter(product=product, store__in=stores).select_related('store')

        if stocks.exists():
            # Create a row for each store where product has stock
            for stock in stocks:
                row = [
                    # Product Basic Info
                    product.name,
                    product.sku,
                    product.barcode or '',
                    product.description or '',

                    # Category & Supplier
                    product.category.name if product.category else '',
                    getattr(product.category, 'code', '') if product.category else '',
                    getattr(product.category, 'description', '') if product.category else '',
                    getattr(product.category, 'category_type', 'product') if product.category else 'product',
                    getattr(product.category, 'efris_commodity_category_code', '') if product.category else '',
                    'Yes' if getattr(product.category, 'efris_auto_sync_enabled', False) else 'No',
                    product.supplier.name if product.supplier else '',

                    # Pricing
                    str(product.selling_price),
                    str(product.cost_price),
                    str(product.discount_percentage),

                    # Tax Information
                    product.tax_rate,
                    str(product.excise_duty_rate),
                    product.efris_excise_duty_code or '',

                    # Unit and Stock
                    product.unit_of_measure,
                    str(product.min_stock_level),
                    'Yes' if product.is_active else 'No',

                    # Store Stock Information
                    stock.store.name,
                    str(stock.quantity),
                    str(stock.low_stock_threshold),
                    str(stock.reorder_quantity),

                    # Additional EFRIS Fields
                    product.efris_item_code or '',
                    'Yes' if product.efris_has_piece_unit else 'No',
                    product.efris_piece_measure_unit or '',
                    str(product.efris_piece_unit_price) if product.efris_piece_unit_price else '',
                    product.efris_goods_code_field or '',
                ]
                writer.writerow(row)
        else:
            # Product has no stock in any selected store - export with default store and 0 quantity
            default_store = stores.first() if stores.exists() else None
            if default_store:
                row = [
                    # Product Basic Info
                    product.name,
                    product.sku,
                    product.barcode or '',
                    product.description or '',

                    # Category & Supplier
                    product.category.name if product.category else '',
                    getattr(product.category, 'code', '') if product.category else '',
                    getattr(product.category, 'description', '') if product.category else '',
                    getattr(product.category, 'category_type', 'product') if product.category else 'product',
                    getattr(product.category, 'efris_commodity_category_code', '') if product.category else '',
                    'Yes' if getattr(product.category, 'efris_auto_sync_enabled', False) else 'No',
                    product.supplier.name if product.supplier else '',

                    # Pricing
                    str(product.selling_price),
                    str(product.cost_price),
                    str(product.discount_percentage),

                    # Tax Information
                    product.tax_rate,
                    str(product.excise_duty_rate),
                    product.efris_excise_duty_code or '',

                    # Unit and Stock
                    product.unit_of_measure,
                    str(product.min_stock_level),
                    'Yes' if product.is_active else 'No',

                    # Store Stock Information
                    default_store.name,
                    '0',  # Zero quantity
                    '5',  # Default threshold
                    '10',  # Default reorder

                    # Additional EFRIS Fields
                    product.efris_item_code or '',
                    'Yes' if product.efris_has_piece_unit else 'No',
                    product.efris_piece_measure_unit or '',
                    str(product.efris_piece_unit_price) if product.efris_piece_unit_price else '',
                    product.efris_goods_code_field or '',
                ]
                writer.writerow(row)

    return response


@login_required
@permission_required('inventory.view_product', raise_exception=True)
def export_products_excel(request):
    """Export products to Excel format compatible with import template"""

    # Get filter parameters
    export_mode = request.GET.get('mode', 'products_only')
    category_ids = request.GET.getlist('categories')
    supplier_ids = request.GET.getlist('suppliers')
    store_ids = request.GET.getlist('stores')
    include_inactive = request.GET.get('include_inactive', 'false') == 'true'

    # Build queryset
    products = Product.objects.select_related('category', 'supplier')

    if not include_inactive:
        products = products.filter(is_active=True)

    if category_ids:
        products = products.filter(category_id__in=category_ids)

    if supplier_ids:
        products = products.filter(supplier_id__in=supplier_ids)

    products = products.order_by('sku')

    if export_mode == 'with_stock':
        return _export_products_with_stock_excel(products, store_ids)
    else:
        return _export_products_only_excel(products)


def _export_products_only_excel(products):
    """Export products only (no stock) to Excel"""

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Products')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#10B981',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    data_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    # Headers
    headers = [
        'Product Name*', 'SKU*', 'Barcode', 'Category', 'Supplier',
        'Selling Price*', 'Cost Price*', 'Discount %',
        'Tax Rate*', 'Excise Duty Rate', 'Unit of Measure*',
        'Min Stock Level', 'Description', 'Is Active',
        'EFRIS Commodity Code', 'EFRIS Excise Duty Code', 'EFRIS Auto Sync',
        'EFRIS Item Code', 'EFRIS Has Piece Unit', 'EFRIS Piece Measure Unit',
        'EFRIS Piece Unit Price', 'EFRIS Goods Code'
    ]

    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Write product data
    for row_idx, product in enumerate(products, start=1):
        data = [
            product.name,
            product.sku,
            product.barcode or '',
            product.category.name if product.category else '',
            product.supplier.name if product.supplier else '',
            float(product.selling_price),
            float(product.cost_price),
            float(product.discount_percentage),
            product.tax_rate,
            float(product.excise_duty_rate),
            product.unit_of_measure,
            product.min_stock_level,
            product.description or '',
            'Yes' if product.is_active else 'No',
            # EFRIS fields
            getattr(product.category, 'efris_commodity_category_code', '') if product.category else '',
            product.efris_excise_duty_code or '',
            'Yes' if product.efris_auto_sync_enabled else 'No',
            product.efris_item_code or '',
            'Yes' if product.efris_has_piece_unit else 'No',
            product.efris_piece_measure_unit or '',
            float(product.efris_piece_unit_price) if product.efris_piece_unit_price else '',
            product.efris_goods_code_field or '',
        ]

        for col_idx, value in enumerate(data):
            worksheet.write(row_idx, col_idx, value, data_format)

    # Set column widths
    column_widths = [20, 15, 15, 15, 15, 12, 12, 10, 10, 12, 15, 12, 30, 10, 20, 20, 15, 15, 15, 15, 15, 15]
    for col, width in enumerate(column_widths):
        worksheet.set_column(col, col, width)

    # Freeze first row
    worksheet.freeze_panes(1, 0)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_export.xlsx"'

    return response


def _export_products_with_stock_excel(products, store_ids):
    """Export products with stock to Excel"""

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Products & Stock')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4F46E5',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    data_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    # Headers
    headers = [
        'Product Name*', 'SKU*', 'Barcode', 'Description',
        'Category Name*', 'Category Code', 'Category Description', 'Category Type*',
        'EFRIS Commodity Category Code', 'EFRIS Auto Sync', 'Supplier',
        'Selling Price*', 'Cost Price*', 'Discount %',
        'Tax Rate*', 'Excise Duty Rate', 'EFRIS Excise Duty Code',
        'Unit of Measure*', 'Min Stock Level', 'Is Active',
        'Store Name*', 'Quantity*', 'Low Stock Threshold', 'Reorder Quantity',
        'EFRIS Item Code', 'EFRIS Has Piece Unit', 'EFRIS Piece Measure Unit',
        'EFRIS Piece Unit Price', 'EFRIS Goods Code'
    ]

    # Write headers
    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Get stores
    if store_ids:
        stores = Store.objects.filter(id__in=store_ids, is_active=True)
    else:
        stores = Store.objects.filter(is_active=True)

    # Write data
    row_idx = 1
    for product in products:
        stocks = Stock.objects.filter(product=product, store__in=stores).select_related('store')

        if stocks.exists():
            for stock in stocks:
                data = [
                    product.name,
                    product.sku,
                    product.barcode or '',
                    product.description or '',
                    product.category.name if product.category else '',
                    getattr(product.category, 'code', '') if product.category else '',
                    getattr(product.category, 'description', '') if product.category else '',
                    getattr(product.category, 'category_type', 'product') if product.category else 'product',
                    getattr(product.category, 'efris_commodity_category_code', '') if product.category else '',
                    'Yes' if getattr(product.category, 'efris_auto_sync_enabled', False) else 'No',
                    product.supplier.name if product.supplier else '',
                    float(product.selling_price),
                    float(product.cost_price),
                    float(product.discount_percentage),
                    product.tax_rate,
                    float(product.excise_duty_rate),
                    product.efris_excise_duty_code or '',
                    product.unit_of_measure,
                    product.min_stock_level,
                    'Yes' if product.is_active else 'No',
                    stock.store.name,
                    float(stock.quantity),
                    stock.low_stock_threshold,
                    stock.reorder_quantity,
                    product.efris_item_code or '',
                    'Yes' if product.efris_has_piece_unit else 'No',
                    product.efris_piece_measure_unit or '',
                    float(product.efris_piece_unit_price) if product.efris_piece_unit_price else '',
                    product.efris_goods_code_field or '',
                ]

                for col_idx, value in enumerate(data):
                    worksheet.write(row_idx, col_idx, value, data_format)
                row_idx += 1
        else:
            # No stock - export with default store
            default_store = stores.first() if stores.exists() else None
            if default_store:
                data = [
                    product.name,
                    product.sku,
                    product.barcode or '',
                    product.description or '',
                    product.category.name if product.category else '',
                    getattr(product.category, 'code', '') if product.category else '',
                    getattr(product.category, 'description', '') if product.category else '',
                    getattr(product.category, 'category_type', 'product') if product.category else 'product',
                    getattr(product.category, 'efris_commodity_category_code', '') if product.category else '',
                    'Yes' if getattr(product.category, 'efris_auto_sync_enabled', False) else 'No',
                    product.supplier.name if product.supplier else '',
                    float(product.selling_price),
                    float(product.cost_price),
                    float(product.discount_percentage),
                    product.tax_rate,
                    float(product.excise_duty_rate),
                    product.efris_excise_duty_code or '',
                    product.unit_of_measure,
                    product.min_stock_level,
                    'Yes' if product.is_active else 'No',
                    default_store.name,
                    0,
                    5,
                    10,
                    product.efris_item_code or '',
                    'Yes' if product.efris_has_piece_unit else 'No',
                    product.efris_piece_measure_unit or '',
                    float(product.efris_piece_unit_price) if product.efris_piece_unit_price else '',
                    product.efris_goods_code_field or '',
                ]

                for col_idx, value in enumerate(data):
                    worksheet.write(row_idx, col_idx, value, data_format)
                row_idx += 1

    # Set column widths
    column_widths = [20, 15, 15, 25, 15, 15, 20, 15, 25, 15, 15, 12, 12, 10, 10, 12, 15, 12, 10, 15, 15, 10, 15, 15, 15,
                     15, 15, 15, 15]
    for col, width in enumerate(column_widths):
        worksheet.set_column(col, col, width)

    # Freeze first row
    worksheet.freeze_panes(1, 0)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="products_with_stock_export.xlsx"'

    return response

@login_required
def download_sample_products_only_csv(request):
    """Generate CSV sample file for product-only import (no stock)"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_products_only.csv"'

    writer = csv.writer(response)

    # Headers for product-only import with all fields
    headers = [
        'Product Name*', 'SKU*', 'Barcode', 'Category', 'Supplier',
        'Selling Price*', 'Cost Price*', 'Discount %',
        'Tax Rate*', 'Excise Duty Rate', 'Unit of Measure*',
        'Min Stock Level', 'Description', 'Is Active',
        'EFRIS Commodity Code', 'EFRIS Excise Duty Code', 'EFRIS Auto Sync',
        'EFRIS Item Code', 'EFRIS Has Piece Unit', 'EFRIS Piece Measure Unit',
        'EFRIS Piece Unit Price', 'EFRIS Goods Code'
    ]
    writer.writerow(headers)

    # Sample data rows
    sample_data = [
        [
            'Coca Cola 500ml', 'CC-500ML', '5000112345678', 'Beverages', 'Century Bottling',
            '3000', '2000', '0', 'A', '0', '102',
            '20', 'Refreshing cola drink 500ml bottle', 'Yes',
            '101113010000000000', '', 'Yes', '', 'No', '', '', ''
        ],
        [
            'Samsung Galaxy A54', 'SGH-A54-BLK', '8806094123456', 'Electronics', 'Samsung Uganda',
            '1500000', '1200000', '5', 'A', '0', '101',
            '5', 'Samsung Galaxy A54 128GB Black', 'Yes',
            '101113020000000000', '', 'Yes', '', 'No', '', '', ''
        ],
        [
            'Rice 1KG', 'RICE-1KG', '', 'Food & Groceries', 'Tilda Uganda',
            '5000', '3500', '0', 'B', '0', '103',
            '50', 'Premium basmati rice 1kg pack', 'Yes',
            '101113030000000000', '', 'Yes', '', 'No', '', '', ''
        ],
    ]

    for row in sample_data:
        writer.writerow(row)

    return response


@login_required
def download_sample_products_only_excel(request):
    """Generate Excel sample file for product-only import with formatting"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)

    # Create worksheets
    data_sheet = workbook.add_worksheet('Products Data')
    instructions_sheet = workbook.add_worksheet('Instructions')
    reference_sheet = workbook.add_worksheet('Reference Data')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#10B981',  # Green for products
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    required_format = workbook.add_format({
        'bold': True,
        'bg_color': '#DC2626',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    sample_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    instruction_header = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'font_color': '#1F2937'
    })

    instruction_text = workbook.add_format({
        'text_wrap': True,
        'valign': 'top'
    })

    # ========================================================================
    # DATA SHEET
    # ========================================================================

    headers = [
        ('Product Name*', True),
        ('SKU*', True),
        ('Barcode', False),
        ('Category', False),
        ('Supplier', False),
        ('Selling Price*', True),
        ('Cost Price*', True),
        ('Discount %', False),
        ('Tax Rate*', True),
        ('Excise Duty Rate', False),
        ('Unit of Measure*', True),
        ('Min Stock Level', False),
        ('Description', False),
        ('Is Active', False),
        ('EFRIS Commodity Code', False),
        ('EFRIS Excise Duty Code', False),
        ('EFRIS Auto Sync', False),
        ('EFRIS Item Code', False),
        ('EFRIS Has Piece Unit', False),
        ('EFRIS Piece Measure Unit', False),
        ('EFRIS Piece Unit Price', False),
        ('EFRIS Goods Code', False),
    ]

    # Write headers
    for col, (header, is_required) in enumerate(headers):
        if is_required:
            data_sheet.write(0, col, header, required_format)
        else:
            data_sheet.write(0, col, header, header_format)

    # Sample data
    sample_data = [
        [
            'Coca Cola 500ml', 'CC-500ML', '5000112345678', 'Beverages', 'Century Bottling',
            3000, 2000, 0, 'A', 0, '102',
            20, 'Refreshing cola drink 500ml bottle', 'Yes',
            '101113010000000000', '', 'Yes', '', 'No', '', '', ''
        ],
        [
            'Samsung Galaxy A54', 'SGH-A54-BLK', '8806094123456', 'Electronics', 'Samsung Uganda',
            1500000, 1200000, 5, 'A', 0, '101',
            5, 'Samsung Galaxy A54 128GB Black', 'Yes',
            '101113020000000000', '', 'Yes', '', 'No', '', '', ''
        ],
        [
            'Rice 1KG', 'RICE-1KG', '', 'Food & Groceries', 'Tilda Uganda',
            5000, 3500, 0, 'B', 0, '103',
            50, 'Premium basmati rice 1kg pack', 'Yes',
            '101113030000000000', '', 'Yes', '', 'No', '', '', ''
        ],
    ]

    for row_idx, row_data in enumerate(sample_data, start=1):
        for col_idx, value in enumerate(row_data):
            data_sheet.write(row_idx, col_idx, value, sample_format)

    # Set column widths
    column_widths = [20, 15, 15, 15, 15, 12, 12, 10, 10, 12, 15, 12, 30, 10, 20, 20, 15, 15, 15, 15, 15, 15]
    for col, width in enumerate(column_widths):
        data_sheet.set_column(col, col, width)

    # Freeze first row
    data_sheet.freeze_panes(1, 0)

    # ========================================================================
    # INSTRUCTIONS SHEET
    # ========================================================================

    instructions_sheet.set_column(0, 0, 50)
    instructions_sheet.set_column(1, 1, 60)

    row = 0
    instructions_sheet.write(row, 0, 'PRODUCT IMPORT INSTRUCTIONS', instruction_header)
    row += 2

    instructions = [
        ('Purpose:', 'This template is for importing PRODUCTS ONLY (no stock quantities)'),
        ('', 'Use this when you want to:'),
        ('', '- Add new products to your catalog'),
        ('', '- Update existing product information'),
        ('', '- Bulk update product prices or details'),
        ('', ''),
        ('Required Fields:', 'Fields marked with * (red header) are mandatory'),
        ('', '- Product Name: Full name of the product'),
        ('', '- SKU: Unique product code (Stock Keeping Unit)'),
        ('', '- Selling Price: Price you sell to customers'),
        ('', '- Cost Price: Your purchase/cost price'),
        ('', '- Tax Rate: VAT/Tax category (A, B, C, D, or E)'),
        ('', '- Unit of Measure: How product is measured (see reference)'),
        ('', ''),
        ('Optional Fields:', 'All other fields can be left empty'),
        ('', '- Barcode: Unique product barcode (must be unique if provided)'),
        ('', '- Category: Product category (created if doesn\'t exist)'),
        ('', '- Supplier: Product supplier (created if doesn\'t exist)'),
        ('', '- Is Active: Yes/No - whether product is active for sales'),
        ('', ''),
        ('EFRIS Fields:', 'Optional fields for Uganda Revenue Authority compliance:'),
        ('', '- EFRIS Commodity Code: 18-digit URA commodity code'),
        ('', '- EFRIS Excise Duty Code: For excisable goods'),
        ('', '- EFRIS Auto Sync: Yes/No - Auto-sync to EFRIS system'),
        ('', '- EFRIS Item Code: Internal EFRIS item code'),
        ('', '- EFRIS Has Piece Unit: Yes/No - Whether product has piece unit'),
        ('', '- EFRIS Piece Measure Unit: Unit code for piece measurement'),
        ('', '- EFRIS Piece Unit Price: Price per piece unit'),
        ('', '- EFRIS Goods Code: EFRIS-assigned goods code'),
        ('', ''),
        ('SKU Handling:', 'The SKU is the unique identifier for products'),
        ('', '- If SKU exists: Product will be UPDATED with new information'),
        ('', '- If SKU is new: A new product will be CREATED'),
        ('', '- SKU format: Use letters, numbers, hyphens (e.g., PROD-001)'),
        ('', ''),
        ('Tax Rates:', 'Use these codes:'),
        ('', 'A = Standard rate (18%)'),
        ('', 'B = Zero rate (0%)'),
        ('', 'C = Exempt (Not taxable)'),
        ('', 'D = Deemed rate (18%)'),
        ('', 'E = Excise Duty rate'),
        ('', ''),
        ('Unit of Measure:', 'Common codes:'),
        ('', '101 = Stick/Piece'),
        ('', '102 = Litre'),
        ('', '103 = Kilogram'),
        ('', '(See Reference Data sheet for full list)'),
        ('', ''),
        ('Categories & Suppliers:', 'Handling non-existent entries:'),
        ('', '- If category doesn\'t exist, it will be created'),
        ('', '- If supplier doesn\'t exist, it will be created'),
        ('', '- Leave blank to skip (product will have no category/supplier)'),
        ('', ''),
        ('Conflict Resolution:', 'When SKU already exists:'),
        ('', '- OVERWRITE mode: Updates product with new data'),
        ('', '- SKIP mode: Ignores row, keeps existing product'),
        ('', ''),
        ('Best Practices:', '- Keep SKUs consistent and unique'),
        ('', '- Use descriptive product names'),
        ('', '- Verify prices before importing'),
        ('', '- Test with 5-10 products first'),
        ('', '- Review validation messages after upload'),
        ('', '- Back up existing data before large imports'),
    ]

    for instruction in instructions:
        instructions_sheet.write(row, 0, instruction[0],
                               instruction_header if instruction[0] and instruction[0].endswith(':') else instruction_text)
        instructions_sheet.write(row, 1, instruction[1], instruction_text)
        row += 1

    # ========================================================================
    # REFERENCE DATA SHEET
    # ========================================================================

    reference_sheet.set_column(0, 0, 15)
    reference_sheet.set_column(1, 1, 40)

    row = 0
    reference_sheet.write(row, 0, 'UNIT CODES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Code', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    # Common unit codes
    common_units = [
        ('101', 'Stick/Piece'),
        ('102', 'Litre'),
        ('103', 'Kilogram'),
        ('104', 'User per day of access'),
        ('105', 'Minute'),
        ('106', '1000 sticks'),
        ('107', '50kgs'),
        ('108', '-'),
        ('109', 'Gram'),
        ('110', 'Box'),
        ('111', 'Pair'),
        ('112', 'Yard'),
        ('113', 'Dozen'),
        ('200', 'Metre'),
    ]

    for code, desc in common_units:
        reference_sheet.write(row, 0, code, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    # Add tax rates reference
    row += 2
    reference_sheet.write(row, 0, 'TAX RATES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Code', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    tax_rates = [
        ('A', 'Standard rate (18%)'),
        ('B', 'Zero rate (0%)'),
        ('C', 'Exempt (Not taxable)'),
        ('D', 'Deemed rate (18%)'),
        ('E', 'Excise Duty rate'),
    ]

    for code, desc in tax_rates:
        reference_sheet.write(row, 0, code, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    # Add EFRIS reference
    row += 2
    reference_sheet.write(row, 0, 'EFRIS FIELDS', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Field', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    efris_fields = [
        ('EFRIS Auto Sync', 'Yes/No - Enable automatic EFRIS synchronization'),
        ('EFRIS Has Piece Unit', 'Yes/No - Whether product has piece unit pricing'),
        ('EFRIS Piece Measure Unit', 'Unit code for piece measurement (e.g., 101, 102)'),
        ('EFRIS Piece Unit Price', 'Price per piece unit (decimal)'),
        ('Is Active', 'Yes/No - Whether product is active for sales'),
    ]

    for field, desc in efris_fields:
        reference_sheet.write(row, 0, field, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_products_only.xlsx"'

    return response


def validate_product_row_data(row_data, mapped_columns, row_number):
    """
    Validate a single row of product data
    Returns: (is_valid, errors_list, cleaned_data)
    """
    errors = []
    cleaned = {}

    # Required fields for product import
    required_fields = ['product_name', 'sku', 'selling_price', 'cost_price', 'tax_rate', 'unit_of_measure']

    # Check required fields
    for field in required_fields:
        value = None
        for file_header, standard_name in mapped_columns.items():
            if standard_name == field:
                value = row_data.get(file_header)
                break

        if not value or (isinstance(value, str) and not value.strip()):
            field_display = field.replace('_', ' ').title()
            errors.append(f"{field_display} is required")
        else:
            cleaned[field] = value

    # If required fields missing, return early
    if errors:
        return False, errors, cleaned

    # Validate and clean optional fields
    for file_header, standard_name in mapped_columns.items():
        if standard_name in cleaned:
            continue  # Already processed

        value = row_data.get(file_header)
        if value and isinstance(value, str):
            value = value.strip()

        if not value:
            continue

        # Type-specific validation
        try:
            if standard_name in ['selling_price', 'cost_price', 'discount_percentage', 'excise_duty_rate', 'min_stock_level', 'efris_piece_unit_price']:
                cleaned[standard_name] = Decimal(str(value))
                if cleaned[standard_name] < 0:
                    errors.append(f"{standard_name.replace('_', ' ').title()} cannot be negative")

            elif standard_name == 'tax_rate':
                value = str(value).upper().strip()
                if value not in ['A', 'B', 'C', 'D', 'E']:
                    errors.append(f"Invalid tax rate '{value}'. Must be A, B, C, D, or E")
                else:
                    cleaned[standard_name] = value

            elif standard_name in ['efris_auto_sync', 'efris_has_piece_unit', 'is_active']:
                value_lower = str(value).lower().strip()
                cleaned[standard_name] = value_lower in ['yes', 'true', '1', 'y']

            elif standard_name == 'min_stock_level':
                cleaned[standard_name] = int(float(value))  # Handle both int and decimal strings

            else:
                cleaned[standard_name] = value

        except (ValueError, InvalidOperation) as e:
            errors.append(f"Invalid {standard_name.replace('_', ' ')}: {value}")

    is_valid = len(errors) == 0
    return is_valid, errors, cleaned


def process_product_only_import(cleaned_data, conflict_resolution, user, session, row_number, raw_data):
    """Process product-only import (no stock updates)"""
    from django.db import IntegrityError
    import uuid

    sku = cleaned_data['sku']
    product = None
    created = False
    updated_fields = []

    try:
        # Check if product exists
        product = Product.objects.get(sku=sku)

        if conflict_resolution == 'skip':
            ImportResult.objects.create(
                session=session,
                result_type='skipped',
                row_number=row_number,
                product_name=cleaned_data.get('product_name', ''),
                sku=sku,
                raw_data=raw_data
            )
            return {'status': 'skipped', 'message': 'Product already exists'}

        # Overwrite mode - update product
        if product.name != cleaned_data['product_name']:
            product.name = cleaned_data['product_name']
            updated_fields.append('name')

        if product.selling_price != cleaned_data['selling_price']:
            product.selling_price = cleaned_data['selling_price']
            updated_fields.append('selling_price')

        if product.cost_price != cleaned_data['cost_price']:
            product.cost_price = cleaned_data['cost_price']
            updated_fields.append('cost_price')

        if product.tax_rate != cleaned_data.get('tax_rate', 'A'):
            product.tax_rate = cleaned_data.get('tax_rate', 'A')
            updated_fields.append('tax_rate')

        if product.unit_of_measure != cleaned_data.get('unit_of_measure', '103'):
            product.unit_of_measure = cleaned_data.get('unit_of_measure', '103')
            updated_fields.append('unit_of_measure')

        # Optional fields
        if 'barcode' in cleaned_data:
            new_barcode = cleaned_data['barcode']
            current_barcode = product.barcode

            # Handle empty barcode properly
            if new_barcode == '':
                new_barcode = None

            # Only update if different
            if new_barcode != current_barcode:
                # Check for duplicate barcode if new_barcode is not None
                if new_barcode:
                    existing_barcode = Product.objects.filter(
                        barcode=new_barcode
                    ).exclude(id=product.id).exists()
                    if existing_barcode:
                        # Generate unique barcode instead of raising error
                        new_barcode = f"{new_barcode}_{uuid.uuid4().hex[:8]}"
                        logger.warning(
                            f"Barcode conflict for product {product.name}, using generated barcode: {new_barcode}")

                product.barcode = new_barcode
                updated_fields.append('barcode')

        if 'description' in cleaned_data and product.description != cleaned_data['description']:
            product.description = cleaned_data['description']
            updated_fields.append('description')

        if 'discount_percentage' in cleaned_data and product.discount_percentage != cleaned_data['discount_percentage']:
            product.discount_percentage = cleaned_data['discount_percentage']
            updated_fields.append('discount_percentage')

        if 'excise_duty_rate' in cleaned_data and product.excise_duty_rate != cleaned_data['excise_duty_rate']:
            product.excise_duty_rate = cleaned_data['excise_duty_rate']
            updated_fields.append('excise_duty_rate')

        if 'min_stock_level' in cleaned_data:
            try:
                new_min_stock = int(cleaned_data['min_stock_level'])
                if product.min_stock_level != new_min_stock:
                    product.min_stock_level = new_min_stock
                    updated_fields.append('min_stock_level')
            except (ValueError, TypeError):
                pass  # Keep existing value if invalid

        if 'is_active' in cleaned_data and product.is_active != cleaned_data['is_active']:
            product.is_active = cleaned_data['is_active']
            updated_fields.append('is_active')

        # Handle category
        if 'category' in cleaned_data:
            category, _ = Category.objects.get_or_create(
                name=cleaned_data['category'],
                defaults={'is_active': True}
            )
            if product.category != category:
                product.category = category
                updated_fields.append('category')

        # Handle supplier
        if 'supplier' in cleaned_data:
            supplier, _ = Supplier.objects.get_or_create(
                name=cleaned_data['supplier'],
                defaults={
                    'phone': '0000000000',
                    'is_active': True
                }
            )
            if product.supplier != supplier:
                product.supplier = supplier
                updated_fields.append('supplier')

        # EFRIS fields
        if 'efris_commodity_code' in cleaned_data:
            try:
                category = Category.objects.get(efris_commodity_category_code=cleaned_data['efris_commodity_code'])
                if product.category != category:
                    product.category = category
                    updated_fields.append('efris_category')
            except Category.DoesNotExist:
                pass

        if 'efris_excise_duty_code' in cleaned_data and product.efris_excise_duty_code != cleaned_data[
            'efris_excise_duty_code']:
            product.efris_excise_duty_code = cleaned_data['efris_excise_duty_code']
            updated_fields.append('efris_excise_duty_code')

        if 'efris_auto_sync' in cleaned_data and product.efris_auto_sync_enabled != cleaned_data['efris_auto_sync']:
            product.efris_auto_sync_enabled = cleaned_data['efris_auto_sync']
            updated_fields.append('efris_auto_sync')

        if 'efris_item_code' in cleaned_data and product.efris_item_code != cleaned_data['efris_item_code']:
            product.efris_item_code = cleaned_data['efris_item_code']
            updated_fields.append('efris_item_code')

        if 'efris_has_piece_unit' in cleaned_data and product.efris_has_piece_unit != cleaned_data[
            'efris_has_piece_unit']:
            product.efris_has_piece_unit = cleaned_data['efris_has_piece_unit']
            updated_fields.append('efris_has_piece_unit')

        if 'efris_piece_measure_unit' in cleaned_data and product.efris_piece_measure_unit != cleaned_data[
            'efris_piece_measure_unit']:
            product.efris_piece_measure_unit = cleaned_data['efris_piece_measure_unit']
            updated_fields.append('efris_piece_measure_unit')

        if 'efris_piece_unit_price' in cleaned_data and product.efris_piece_unit_price != cleaned_data[
            'efris_piece_unit_price']:
            product.efris_piece_unit_price = cleaned_data['efris_piece_unit_price']
            updated_fields.append('efris_piece_unit_price')

        if 'efris_goods_code' in cleaned_data and product.efris_goods_code_field != cleaned_data['efris_goods_code']:
            product.efris_goods_code_field = cleaned_data['efris_goods_code']
            updated_fields.append('efris_goods_code_field')

        product.import_session = session

        # Save product
        try:
            product.save()
        except IntegrityError as e:
            if 'barcode' in str(e).lower() and 'unique' in str(e).lower():
                # Handle duplicate barcode by generating a unique one
                if product.barcode:
                    product.barcode = f"{product.barcode}_{uuid.uuid4().hex[:8]}"
                else:
                    product.barcode = f"BC_{sku}_{uuid.uuid4().hex[:6]}"
                product.save()
                updated_fields.append('barcode_auto_fixed')
            else:
                raise

        if updated_fields:
            logger.info(f"✅ Product updated: {product.name} - Fields: {', '.join(updated_fields)}")

    except Product.DoesNotExist:
        # Create new product
        created = True

        # Handle barcode for new product
        barcode = cleaned_data.get('barcode', '')
        if barcode == '':
            barcode = None
        elif barcode:  # Only check duplicates if barcode is not None/empty
            existing_barcode = Product.objects.filter(barcode=barcode).exists()
            if existing_barcode:
                # Generate unique barcode instead of raising error
                barcode = f"{barcode}_{uuid.uuid4().hex[:8]}"
                logger.warning(f"Duplicate barcode for new product {cleaned_data['product_name']}, using: {barcode}")

        try:
            min_stock_level = int(cleaned_data.get('min_stock_level', 5))
        except (ValueError, TypeError):
            min_stock_level = 5

        product = Product(
            sku=sku,
            name=cleaned_data['product_name'],
            selling_price=cleaned_data['selling_price'],
            cost_price=cleaned_data['cost_price'],
            tax_rate=cleaned_data.get('tax_rate', 'A'),
            unit_of_measure=cleaned_data.get('unit_of_measure', '103'),
            discount_percentage=cleaned_data.get('discount_percentage', 0),
            excise_duty_rate=cleaned_data.get('excise_duty_rate', 0),
            min_stock_level=min_stock_level,
            description=cleaned_data.get('description', ''),
            barcode=barcode,  # Can be None
            is_active=cleaned_data.get('is_active', True),
            import_session=session,
            imported_at=timezone.now()
        )

        # Handle category
        if 'category' in cleaned_data:
            category, _ = Category.objects.get_or_create(
                name=cleaned_data['category'],
                defaults={'is_active': True}
            )
            product.category = category
        elif 'efris_commodity_code' in cleaned_data:
            try:
                category = Category.objects.get(efris_commodity_category_code=cleaned_data['efris_commodity_code'])
                product.category = category
            except Category.DoesNotExist:
                pass

        # Handle supplier
        if 'supplier' in cleaned_data:
            supplier, _ = Supplier.objects.get_or_create(
                name=cleaned_data['supplier'],
                defaults={
                    'phone': '0000000000',
                    'is_active': True
                }
            )
            product.supplier = supplier

        # EFRIS fields
        if 'efris_excise_duty_code' in cleaned_data:
            product.efris_excise_duty_code = cleaned_data['efris_excise_duty_code']

        if 'efris_auto_sync' in cleaned_data:
            product.efris_auto_sync_enabled = cleaned_data['efris_auto_sync']

        if 'efris_item_code' in cleaned_data:
            product.efris_item_code = cleaned_data['efris_item_code']

        if 'efris_has_piece_unit' in cleaned_data:
            product.efris_has_piece_unit = cleaned_data['efris_has_piece_unit']

        if 'efris_piece_measure_unit' in cleaned_data:
            product.efris_piece_measure_unit = cleaned_data['efris_piece_measure_unit']

        if 'efris_piece_unit_price' in cleaned_data:
            product.efris_piece_unit_price = cleaned_data['efris_piece_unit_price']

        if 'efris_goods_code' in cleaned_data:
            product.efris_goods_code_field = cleaned_data['efris_goods_code']

        # Save with error handling
        try:
            product.save()
        except IntegrityError as e:
            if 'barcode' in str(e).lower() and 'unique' in str(e).lower():
                # Last resort: generate completely new barcode
                product.barcode = f"IMP_{sku}_{uuid.uuid4().hex[:10]}"
                product.save()
            else:
                raise

        logger.info(f"✅ Product created: {product.name} (SKU: {sku})")

    # Create import result
    result_type = 'created' if created else 'updated'
    try:
        ImportResult.objects.create(
            session=session,
            result_type=result_type,
            row_number=row_number,
            product_name=product.name,
            sku=sku,
            raw_data=raw_data
        )
    except Exception as e:
        logger.error(f"Failed to create ImportResult for row {row_number}: {str(e)}")
        # Don't fail the whole import if result logging fails

    return {
        'status': result_type,
        'product': product,
        'updated_fields': updated_fields if not created else []
    }


@transaction.atomic
def process_product_import_file(file_obj, conflict_resolution, user, column_mapping=None, has_header=True):
    """
    Process product-only import file

    Args:
        file_obj: Uploaded file object
        conflict_resolution: 'overwrite' or 'skip'
        user: User performing the import
        column_mapping: Optional manual column mapping dict
        has_header: Whether file has header row

    Returns:
        dict with import results
    """

    # Create import session
    session = ImportSession.objects.create(
        user=user,
        filename=file_obj.name,
        file_size=file_obj.size,
        import_mode='product_only',
        conflict_resolution=conflict_resolution,
        has_header=has_header,
        status='processing',
        started_at=timezone.now()
    )

    try:
        # Parse file
        headers, data = parse_uploaded_file(file_obj)
        session.total_rows = len(data)
        session.save()

        # Map columns
        if column_mapping:
            mapped_columns = column_mapping
        else:
            mapped_columns = ColumnMapper.map_columns(headers)

        session.column_mapping = mapped_columns
        session.save()

        # Log column mapping
        ImportLog.objects.create(
            session=session,
            level='info',
            message=f'Column mapping completed: {len(mapped_columns)} columns mapped'
        )

        # Process each row
        results = {
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'error_count': 0,
            'errors': [],
            'warnings': []
        }

        for idx, row_data in enumerate(data, start=2):  # Start at 2 for Excel row numbers
            try:
                # Validate row
                is_valid, errors, cleaned_data = validate_product_row_data(row_data, mapped_columns, idx)

                if not is_valid:
                    ImportResult.objects.create(
                        session=session,
                        result_type='error',
                        row_number=idx,
                        error_message='; '.join(errors),
                        raw_data=row_data
                    )
                    results['error_count'] += 1
                    results['errors'].append({
                        'row': idx,
                        'errors': errors
                    })
                    continue

                # Process the row
                result = process_product_only_import(
                    cleaned_data=cleaned_data,
                    conflict_resolution=conflict_resolution,
                    user=user,
                    session=session,
                    row_number=idx,
                    raw_data=row_data
                )

                # Update counts
                if result['status'] == 'created':
                    results['created_count'] += 1
                elif result['status'] == 'updated':
                    results['updated_count'] += 1
                elif result['status'] == 'skipped':
                    results['skipped_count'] += 1

                session.processed_rows += 1

            except ValueError as e:
                # Handle validation errors (like duplicate barcode)
                error_msg = str(e)
                logger.warning(f"Validation error on row {idx}: {error_msg}")

                ImportResult.objects.create(
                    session=session,
                    result_type='error',
                    row_number=idx,
                    error_message=error_msg,
                    raw_data=row_data
                )
                results['error_count'] += 1
                results['errors'].append({
                    'row': idx,
                    'errors': [error_msg]
                })

            except Exception as e:
                logger.error(f"Error processing row {idx}: {str(e)}", exc_info=True)
                results['error_count'] += 1
                results['errors'].append({
                    'row': idx,
                    'errors': [f"Unexpected error: {str(e)}"]
                })

                ImportLog.objects.create(
                    session=session,
                    level='error',
                    message=f'Row {idx}: Unexpected error',
                    row_number=idx,
                    details={'error': str(e)}
                )

            # Update session periodically
            if idx % 10 == 0:
                session.created_count = results['created_count']
                session.updated_count = results['updated_count']
                session.skipped_count = results['skipped_count']
                session.error_count = results['error_count']
                session.save()

        # Final session update
        session.created_count = results['created_count']
        session.updated_count = results['updated_count']
        session.skipped_count = results['skipped_count']
        session.error_count = results['error_count']
        session.status = 'completed'
        session.completed_at = timezone.now()
        session.save()

        return results

    except Exception as e:
        logger.error(f"Product import failed: {str(e)}", exc_info=True)
        session.status = 'failed'
        session.error_message = str(e)
        session.completed_at = timezone.now()
        session.save()

        raise