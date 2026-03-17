from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO, StringIO
import base64
import csv
import hashlib
import io
import json
import logging
import pyotp
import qrcode
import secrets
import string
import zipfile

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import login, logout, update_session_auth_hash, authenticate
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.sites.shortcuts import get_current_site
from django.core.cache import cache
from django.core.exceptions import PermissionDenied, ValidationError
from django.core.files.base import ContentFile
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.db import connection, models, transaction
from django.db.models import Count, Q, Avg, Max
from django.http import HttpResponse, Http404, JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.encoding import force_str
from django.utils.translation import gettext as _
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.views.decorators.http import require_http_methods, require_POST
from django.views.generic import (
    ListView, DetailView, CreateView, UpdateView, DeleteView, FormView
)

from django_otp.plugins.otp_totp.models import TOTPDevice
from django_tenants.utils import schema_context, tenant_context

# For PDF generation
from reportlab.lib import colors
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image

# For Excel export
import openpyxl
from openpyxl.styles import Font, PatternFill

from PIL import Image as PILImage

from company.decorator import check_user_limit
from company.email import send_tenant_email
from company.models import Company, SubscriptionPlan

from .models import CustomUser, UserSignature, Role, RoleHistory, AuditLog, LoginHistory, DataExportLog, APIToken, UserSession
from .forms import (
    CustomUserCreationForm, CustomUserChangeForm, CustomAuthenticationForm, UserRoleAssignForm,
    UserProfileForm, PasswordChangeForm, UserSignatureForm, UserSearchForm,
    UserNotificationForm, UserPreferencesForm, BulkUserActionForm, TwoFactorSetupForm,
    RoleForm, BulkRoleAssignmentForm, BulkUserRoleAssignForm, RoleFilterForm,
    ReviewAuditLogForm,
)
from .utils import (
    require_saas_admin,
    export_audit_logs,
    get_client_ip,
    parse_user_agent,
    get_location_from_ip,
    get_visible_users,
    get_company_user_count,
    can_access_company,
    get_accessible_companies,
    require_company_access,
)
from accounts.middleware import register_session, clear_session_registry
from accounts.sharing_detection import SharingDetectionEngine, DetectionContext

logger = logging.getLogger(__name__)

DASHBOARD_MAPPING = [
    {'permission': 'accounts.view_customuser', 'url_name': 'user_dashboard'},
    {'permission': 'company.view_company', 'url_name': 'dashboard'},
    {'permission': 'stores.view_product', 'url_name': 'stores:dashboard'},
    {'permission': 'inventory.view_product', 'url_name': 'inventory:dashboard'},
    {'permission': 'reports.view_savedreport', 'url_name': 'reports:dashboard'},
]


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


# ============================================================
# RATE LIMITING HELPERS
# All limits use Django's cache backend — no extra packages.
# ============================================================

def _rl_key(scope, identifier):
    """Build a namespaced cache key for rate limiting."""
    h = hashlib.sha256(f"{scope}:{identifier}".encode()).hexdigest()[:16]
    return f"rl:{scope}:{h}"


def _is_rate_limited(scope, identifier, max_attempts, window_seconds):
    """
    Return True when `identifier` has exceeded `max_attempts` within
    the rolling `window_seconds` window for the given `scope`.
    """
    key = _rl_key(scope, identifier)
    attempts = cache.get(key, 0)
    return attempts >= max_attempts


def _increment_rate_limit(scope, identifier, window_seconds):
    """Atomically increment the attempt counter and (re)set the TTL."""
    key = _rl_key(scope, identifier)
    try:
        cache.incr(key)
    except ValueError:
        cache.set(key, 1, timeout=window_seconds)


def _clear_rate_limit(scope, identifier):
    """Clear the rate limit counter after a successful action."""
    cache.delete(_rl_key(scope, identifier))


def _rate_limit_response(request, message, is_ajax=False):
    """Return the appropriate response when rate limit is hit."""
    if is_ajax or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse(
            {'success': False, 'error_type': 'rate_limited', 'error_message': message},
            status=429
        )
    messages.error(request, message)
    return None  # caller should redirect


# ============================================================
# INVITATION TOKEN HELPERS
# Tokens are signed with HMAC so they can't be forged or reused
# without a matching DB record.
# ============================================================

def _generate_invitation_token():
    """Return a URL-safe 48-char token."""
    return secrets.token_urlsafe(36)


def _build_setup_url(request, user, token):
    """Build the full password-setup URL for an invited user."""
    path = reverse('accept_invitation', kwargs={'token': token})
    scheme = 'https' if request.is_secure() else 'http'

    # Build host from tenant schema + BASE_DOMAIN from settings.
    # settings.BASE_DOMAIN is already correct for each environment:
    #   dev:  'localhost:8000'  → host = 'rem.localhost:8000'
    #   prod: 'primebooks.sale' → host = 'rem.primebooks.sale'
    # This avoids relying on tenant.domain_url (no port) or request.get_host()
    # (stripped by middleware) — both of which lose the port on local dev.
    tenant = getattr(connection, 'tenant', None)
    base_domain = getattr(settings, 'BASE_DOMAIN', None)
    if tenant and base_domain:
        host = f"{tenant.schema_name}.{base_domain}"
    else:
        # Last resort fallback
        host = request.get_host()

    return f"{scheme}://{host}{path}"


def get_dashboard_url(user):
    """Return the correct dashboard URL for the user."""
    if getattr(user, 'is_saas_admin', False):
        return reverse('saas_admin_dashboard')

    if getattr(user, 'is_superuser', False) or getattr(user, 'company_admin', False):
        if hasattr(user, 'company') and user.company:
            return reverse('companies:company_detail', kwargs={'company_id': user.company.company_id})

    for item in DASHBOARD_MAPPING:
        if user.has_perm(item['permission']):
            return reverse(item['url_name'])

    return reverse('login')


def custom_login(request):
    """Enhanced login view with proper 2FA enforcement"""
    if request.user.is_authenticated:
        return redirect(get_dashboard_url(request.user))

    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest'

    if request.method == 'POST':
        ip = get_client_ip(request)
        # 10 attempts per IP per 10 minutes; 20 per email per 15 minutes
        email_attempt = request.POST.get('email', '').lower().strip()

        if _is_rate_limited('login_ip', ip, max_attempts=10, window_seconds=600):
            resp = _rate_limit_response(
                request,
                'Too many login attempts from this IP. Please wait 10 minutes.',
                is_ajax=is_ajax
            )
            if resp:
                return resp
            return redirect('login')

        if email_attempt and _is_rate_limited('login_email', email_attempt, max_attempts=20, window_seconds=900):
            resp = _rate_limit_response(
                request,
                'Too many login attempts for this account. Please wait 15 minutes.',
                is_ajax=is_ajax
            )
            if resp:
                return resp
            return redirect('login')

        if is_ajax:
            return _handle_ajax_login(request)
        return _handle_regular_login(request)

    # GET request - show login form
    form = CustomAuthenticationForm(request)

    # Check if we're in 2FA step (user already authenticated credentials)
    show_2fa = False
    pending_user_id = request.session.get('pending_2fa_user_id')
    if pending_user_id:
        try:
            pending_user = CustomUser.objects.get(id=pending_user_id)
            show_2fa = TOTPDevice.objects.filter(user=pending_user, confirmed=True).exists()
        except CustomUser.DoesNotExist:
            request.session.pop('pending_2fa_user_id', None)

    context = {
        'form': form,
        'show_2fa': show_2fa,
    }
    return render(request, 'accounts/login.html', context)


def _handle_ajax_login(request):
    """Handle AJAX login requests with 2FA enforcement"""
    pending_user_id = request.session.get('pending_2fa_user_id')
    code = request.POST.get('code', '').strip()

    if pending_user_id and code:
        return _handle_2fa_verification_ajax(request, pending_user_id, code)

    form = CustomAuthenticationForm(request, data=request.POST)

    if not form.is_valid():
        # Count failed attempts against IP and submitted email
        ip = get_client_ip(request)
        email_attempt = request.POST.get('email', '').lower().strip()
        _increment_rate_limit('login_ip', ip, window_seconds=600)
        if email_attempt:
            _increment_rate_limit('login_email', email_attempt, window_seconds=900)

        logger.warning(f"Form validation failed from {ip}: {form.errors}")
        error_messages = {}
        for field, errors in form.errors.items():
            if field == '__all__':
                error_messages['general'] = list(errors)
            else:
                error_messages[field] = list(errors)

        return JsonResponse({
            'success': False,
            'error_type': 'validation_error',
            'error_message': 'Please correct the errors below',
            'field_errors': error_messages,
        }, status=400)

    user = form.get_user()

    if not user:
        ip = get_client_ip(request)
        _increment_rate_limit('login_ip', ip, window_seconds=600)
        return JsonResponse({
            'success': False,
            'error_type': 'authentication_failed',
            'error_message': 'Invalid credentials'
        }, status=401)

    two_factor_enabled = TOTPDevice.objects.filter(user=user, confirmed=True).exists()

    if two_factor_enabled:
        request.session['pending_2fa_user_id'] = user.id
        request.session['pending_2fa_email'] = user.email
        request.session['pending_2fa_remember'] = form.cleaned_data.get('remember_me', False)
        logger.info(f"2FA required for user: {user.email}")
        return JsonResponse({
            'success': False,
            'two_factor_required': True,
            'error_type': '2fa_required',
            'message': 'Please enter your 6-digit authentication code'
        }, status=200)

    return _complete_login_ajax(request, user, form.cleaned_data.get('remember_me', False))


def _handle_regular_login(request):
    """Handle regular (non-AJAX) form submission with 2FA enforcement"""
    # Check if this is a 2FA verification
    pending_user_id = request.session.get('pending_2fa_user_id')
    code = request.POST.get('code', '').strip()

    if pending_user_id and code:
        # Verify 2FA code
        try:
            user = CustomUser.objects.get(id=pending_user_id, is_active=True)
        except CustomUser.DoesNotExist:
            request.session.pop('pending_2fa_user_id', None)
            messages.error(request, 'Session expired. Please log in again.')
            return redirect('login')

        # Check rate limiting
        if _is_2fa_rate_limited(user):
            messages.error(request, 'Too many failed attempts. Please wait 5 minutes.')
            return redirect('login')

        # Verify code
        if _verify_2fa_code(user, code):
            # Clear session data
            _clear_2fa_rate_limit(user)
            request.session.pop('pending_2fa_user_id', None)
            remember_me = request.session.pop('pending_2fa_remember', False)
            request.session.pop('pending_2fa_email', None)

            # Complete login
            return _complete_login_regular(request, user, remember_me)
        else:
            # Invalid code
            _increment_2fa_attempts(user)
            user.record_login_attempt(success=False, ip_address=get_client_ip(request))
            remaining = _get_remaining_attempts(user)
            messages.error(request, f'Invalid authentication code. {remaining} attempts remaining.')

            form = CustomAuthenticationForm(request)
            return render(request, 'accounts/login.html', {
                'form': form,
                'show_2fa': True,
            })

    # Initial credential submission
    form = CustomAuthenticationForm(request, data=request.POST)

    if form.is_valid():
        user = form.get_user()

        if not user:
            ip = get_client_ip(request)
            _increment_rate_limit('login_ip', ip, window_seconds=600)
            messages.error(request, 'Invalid credentials.')
            return render(request, 'accounts/login.html', {
                'form': form,
                'show_2fa': False,
            })

        two_factor_enabled = TOTPDevice.objects.filter(user=user, confirmed=True).exists()

        if two_factor_enabled:
            request.session['pending_2fa_user_id'] = user.id
            request.session['pending_2fa_email'] = user.email
            request.session['pending_2fa_remember'] = form.cleaned_data.get('remember_me', False)
            messages.info(request, 'Please enter your 6-digit authentication code.')
            return render(request, 'accounts/login.html', {
                'form': form,
                'show_2fa': True,
            })

        return _complete_login_regular(request, user, form.cleaned_data.get('remember_me', False))

    # Form has errors — count against IP and email
    ip = get_client_ip(request)
    email_attempt = request.POST.get('email', '').lower().strip()
    _increment_rate_limit('login_ip', ip, window_seconds=600)
    if email_attempt:
        _increment_rate_limit('login_email', email_attempt, window_seconds=900)

    for field, errors in form.errors.items():
        for error in errors:
            if field == '__all__':
                messages.error(request, error)
            else:
                messages.error(request, f"{field.title()}: {error}")

    return render(request, 'accounts/login.html', {
        'form': form,
        'show_2fa': False,
    })


def _handle_2fa_verification_ajax(request, pending_user_id, code):
    """Handle 2FA code verification for AJAX requests"""
    try:
        user = CustomUser.objects.get(id=pending_user_id, is_active=True)
    except CustomUser.DoesNotExist:
        request.session.pop('pending_2fa_user_id', None)
        return JsonResponse({
            'success': False,
            'error_type': 'invalid_session',
            'error_message': 'Session expired. Please log in again.'
        }, status=401)

    # Check rate limiting
    if _is_2fa_rate_limited(user):
        logger.warning(f"2FA rate limit exceeded for user: {user.email}")
        return JsonResponse({
            'success': False,
            'error_type': 'rate_limited',
            'error_message': 'Too many failed attempts. Please wait 5 minutes.'
        }, status=429)

    # Verify the 2FA code
    if _verify_2fa_code(user, code):
        # Clear 2FA session data
        _clear_2fa_rate_limit(user)
        request.session.pop('pending_2fa_user_id', None)
        remember_me = request.session.pop('pending_2fa_remember', False)
        request.session.pop('pending_2fa_email', None)

        # Complete login
        return _complete_login_ajax(request, user, remember_me)
    else:
        # Invalid 2FA code
        _increment_2fa_attempts(user)
        logger.warning(f"Invalid 2FA code for user: {user.email}")
        user.record_login_attempt(success=False, ip_address=get_client_ip(request))

        return JsonResponse({
            'success': False,
            'error_type': 'invalid_2fa',
            'error_message': 'Invalid authentication code. Please try again.',
            'attempts_remaining': _get_remaining_attempts(user)
        }, status=401)


def _complete_login_ajax(request, user, remember_me):
    """Complete the login process for AJAX requests"""
    backend = getattr(user, 'backend', 'django.contrib.auth.backends.ModelBackend')
    login(request, user, backend=backend)
    # ✅Strict-session: make this the only valid session
    register_session(user, request.session.session_key)

    # ✅ Sharing detection: fingerprint + travel checks run at login time
    _run_sharing_detection(request, user)
    # Clear rate limits on successful login
    ip = get_client_ip(request)
    _clear_rate_limit('login_ip', ip)
    _clear_rate_limit('login_email', user.email.lower())

    if remember_me:
        request.session.set_expiry(60 * 60 * 24 * 30)
    else:
        request.session.set_expiry(0)

    user.record_login_attempt(success=True, ip_address=ip)
    user.last_activity_at = timezone.now()
    user.save(update_fields=['last_activity_at'])

    next_url = request.GET.get('next') or get_dashboard_url(user)
    welcome_message = (
        f'Welcome SaaS Admin: {user.get_short_name()}!'
        if getattr(user, 'is_saas_admin', False)
        else f'Welcome back, {user.get_short_name()}!'
    )
    logger.info(f"Successful login for user: {user.email}")
    return JsonResponse({
        'success': True,
        'two_factor_required': False,
        'message': welcome_message,
        'redirect_url': next_url
    })


def _complete_login_regular(request, user, remember_me):
    """Complete login for regular requests"""
    backend = getattr(user, 'backend', 'django.contrib.auth.backends.ModelBackend')
    login(request, user, backend=backend)
    register_session(user, request.session.session_key)
    _run_sharing_detection(request, user)

    # Clear rate limits on successful login
    ip = get_client_ip(request)
    _clear_rate_limit('login_ip', ip)
    _clear_rate_limit('login_email', user.email.lower())

    if remember_me:
        request.session.set_expiry(60 * 60 * 24 * 30)
    else:
        request.session.set_expiry(0)

    user.record_login_attempt(success=True, ip_address=ip)
    user.last_activity_at = timezone.now()
    user.save(update_fields=['last_activity_at'])

    if getattr(user, 'is_saas_admin', False):
        messages.success(request, f'Welcome SaaS Admin: {user.get_short_name()}!')
    else:
        messages.success(request, f'Welcome back, {user.get_short_name()}!')

    next_url = request.GET.get('next') or get_dashboard_url(user)
    logger.info(f"Successful login for user: {user.email}")
    return redirect(next_url)


# 2FA Helper Functions
def _verify_2fa_code(user, code):
    """
    Verify a 2FA code at login time.
    Accepts either a 6-digit TOTP code or an 8-char backup code.
    Backup codes are single-use: the matching code is removed on success.
    """
    code = code.strip()

    # Try TOTP first (django-otp device, preferred)
    try:
        device = TOTPDevice.objects.get(user=user, confirmed=True)
        if device.verify_token(code):
            logger.info(f"2FA TOTP verification succeeded for {user.email}")
            return True
    except TOTPDevice.DoesNotExist:
        pass  # No django-otp device — fall through to metadata-based TOTP

    # Fallback: metadata-based TOTP secret (set by enable_two_factor)
    secret = (user.metadata or {}).get('totp_secret')
    if secret:
        totp = pyotp.TOTP(secret)
        if totp.verify(code, valid_window=1):
            logger.info(f"2FA metadata TOTP verification succeeded for {user.email}")
            return True

    # Check backup codes (case-insensitive, single-use)
    backup_codes = list(user.backup_codes or [])
    code_upper = code.upper()
    if code_upper in backup_codes:
        backup_codes.remove(code_upper)
        user.backup_codes = backup_codes
        user.save(update_fields=['backup_codes'])
        logger.info(f"2FA backup code used for {user.email}. {len(backup_codes)} codes remaining.")
        return True

    logger.warning(f"2FA verification failed for {user.email}")
    return False


def _is_2fa_rate_limited(user):
    """Check if user is rate limited for 2FA attempts"""
    cache_key = f"2fa_attempts_{user.id}"
    attempts = cache.get(cache_key, 0)
    logger.debug(f"2FA attempts for user {user.id}: {attempts}/5")
    return attempts >= 5


def _increment_2fa_attempts(user):
    """Increment 2FA attempt counter"""
    cache_key = f"2fa_attempts_{user.id}"
    attempts = cache.get(cache_key, 0) + 1
    cache.set(cache_key, attempts, timeout=300)  # 5 minutes
    logger.info(f"2FA attempts for user {user.id}: {attempts}/5")


def _clear_2fa_rate_limit(user):
    """Clear 2FA rate limiting for user"""
    cache_key = f"2fa_attempts_{user.id}"
    cache.delete(cache_key)
    logger.info(f"Cleared 2FA rate limit for user {user.id}")


def _get_remaining_attempts(user):
    """Get remaining 2FA attempts"""
    cache_key = f"2fa_attempts_{user.id}"
    attempts = cache.get(cache_key, 0)
    return max(0, 5 - attempts)

def _run_sharing_detection(request, user):
    """
    Build a DetectionContext from the current request and run all detectors.
    Fingerprint hash comes from a hidden field posted by the login form
    (populated by FingerprintJS or a lightweight server-side fallback).
    """
    try:
        from accounts.utils import get_client_ip, get_location_from_ip

        ip = get_client_ip(request)
        ua = request.META.get('HTTP_USER_AGENT', '')

        # Fingerprint: prefer JS-generated value, fall back to server-side hash
        fp_raw = request.POST.get('fp', '') or request.data.get('fp', '')
        if not fp_raw:
            import hashlib
            accept_lang = request.META.get('HTTP_ACCEPT_LANGUAGE', '')
            fp_raw = hashlib.sha256(f"{ua}:{accept_lang}".encode()).hexdigest()

        lat, lon = None, None
        try:
            loc = get_location_from_ip(ip)
            if loc:
                lat = loc.get('latitude')
                lon = loc.get('longitude')
        except Exception:
            pass

        ctx = DetectionContext(
            user_id=user.pk,
            user_email=user.email,
            ip_address=ip or '0.0.0.0',
            user_agent=ua,
            fingerprint_hash=fp_raw,
            latitude=lat,
            longitude=lon,
            timestamp=timezone.now(),
        )

        SharingDetectionEngine().run(user, ctx, request)

    except Exception as exc:
        import logging
        logging.getLogger(__name__).error(
            f"[SharingDetection] _run_sharing_detection failed: {exc}"
        )

# Logout
def custom_logout(request):
    """Instant logout without intermediate allauth page."""
    if request.user.is_authenticated:
        user = request.user
        clear_session_registry(user.pk)
        user.last_activity_at = timezone.now()
        user.save(update_fields=['last_activity_at'])
        logger.info(f"User {user.email} logged out")
        user_name = user.get_short_name() or user.email.split('@')[0]
        messages.info(request, f"Goodbye, {user_name}! You have been logged out.")
    else:
        logger.info("Anonymous user attempted logout")

    logout(request)
    return redirect('login')

@never_cache
def token_login_complete(request):
    """
    Complete login using token from public router.

    This is the final step of the cross-subdomain bridge:
      public.localhost → login_bridge → mbale.localhost/accounts/login/complete/?token=...

    IMPORTANT: never call login() when the user is already authenticated in
    this tenant — doing so rotates the session key, which immediately triggers
    StrictSingleSessionMiddleware's session_superseded redirect on the very
    next request.
    """
    from public_router.tenant_lookup import verify_login_token

    # Guard: already authenticated → just go to dashboard.
    # Handles browser back/refresh after the token was already consumed.
    if request.user.is_authenticated:
        return redirect(get_dashboard_url(request.user))

    token = request.GET.get('token')

    if not token:
        messages.error(request, 'Invalid login link')
        return redirect('login')

    email, tenant_schema = verify_login_token(token)

    if not email:
        messages.error(request, 'Login link expired or invalid. Please try again.')
        return redirect('login')

    current_schema = request.tenant.schema_name if hasattr(request, 'tenant') else 'public'

    if current_schema != tenant_schema:
        logger.error(f"Schema mismatch: expected {tenant_schema}, got {current_schema}")
        messages.error(request, 'Authentication error. Please try again.')
        from django.conf import settings
        base_url = f"http{'s' if settings.USE_HTTPS else ''}://{settings.BASE_DOMAIN}"
        return redirect(f"{base_url}/accounts/login/")

    try:
        user = CustomUser.objects.get(email=email, is_active=True)
    except CustomUser.DoesNotExist:
        messages.error(request, 'User not found')
        return redirect('login')

    two_factor_enabled = TOTPDevice.objects.filter(user=user, confirmed=True).exists()

    if two_factor_enabled:
        request.session['pending_2fa_user_id'] = user.id
        request.session['pending_2fa_email'] = user.email
        messages.info(request, 'Please complete two-factor authentication.')
        return redirect('login')

    login(request, user, backend='django.contrib.auth.backends.ModelBackend')
    register_session(user, request.session.session_key)
    _run_sharing_detection(request, user)

    remember_me = request.GET.get('remember')
    if remember_me:
        request.session.set_expiry(60 * 60 * 24 * 30)
    else:
        request.session.set_expiry(0)

    user.record_login_attempt(success=True, ip_address=get_client_ip(request))
    user.last_activity_at = timezone.now()
    user.save(update_fields=['last_activity_at'])

    if getattr(user, 'is_saas_admin', False):
        messages.success(request, f'Welcome SaaS Admin: {user.get_short_name()}!')
    else:
        messages.success(request, f'Welcome back, {user.get_short_name()}!')

    next_url = request.GET.get('next') or get_dashboard_url(user)
    logger.info(f"User {user.email} logged in via token to tenant {tenant_schema}")

    return redirect(next_url)

@require_saas_admin
def saas_admin_dashboard(request):
    """Dashboard specifically for SaaS administrators"""
    from company.models import Company, SubscriptionPlan

    # Global statistics across all tenants
    total_companies = Company.objects.count()
    active_companies = Company.objects.filter(status='ACTIVE').count() if hasattr(Company,
                                                                                  'status') else Company.objects.count()
    trial_companies = Company.objects.filter(is_trial=True).count() if hasattr(Company, 'is_trial') else 0

    # User statistics - use visible users to get accurate counts
    total_users = get_visible_users().count()
    active_users = get_visible_users().filter(is_active=True).count()

    # Recent activity
    recent_companies = Company.objects.order_by('-created_at')[:10]
    recent_users = get_visible_users().order_by('-date_joined')[:10]

    # Plan distribution
    plan_stats = []
    if hasattr(Company, 'plan'):
        plan_stats = list(
            Company.objects.values('plan__name', 'plan__display_name')
            .annotate(count=Count('company_id'))
            .order_by('-count')
        )

    # Companies expiring soon
    expiring_soon = []
    if hasattr(Company, 'subscription_ends_at'):
        expiring_soon = Company.objects.filter(
            subscription_ends_at__lte=timezone.now().date() + timedelta(days=30),
            subscription_ends_at__gte=timezone.now().date()
        ).order_by('subscription_ends_at')[:10]

    # SaaS-specific metrics
    context = {
        'total_companies': total_companies,
        'active_companies': active_companies,
        'trial_companies': trial_companies,
        'total_users': total_users,
        'active_users': active_users,
        'recent_companies': recent_companies,
        'recent_users': recent_users,
        'plan_stats': plan_stats,
        'expiring_soon': expiring_soon,
        'accessible_companies': get_accessible_companies(request.user),
        'is_saas_admin': True,
    }

    return render(request, 'accounts/saas_admin_dashboard.html', context)


class RolePermissionMixin(PermissionRequiredMixin):
    """Base mixin for role management permissions"""
    permission_required = 'accounts.add_role'

    def handle_no_permission(self):
        messages.error(
            self.request,
            "You don't have permission to access role management."
        )
        return super().handle_no_permission()


class RoleListView(LoginRequiredMixin, RolePermissionMixin, ListView):
    """
    Advanced role listing with search, filtering, and analytics
    """
    model = Role
    template_name = 'accounts/roles/role_list.html'
    context_object_name = 'roles'
    paginate_by = 20

    def get_queryset(self):
        queryset = Role.objects.select_related('group', 'company', 'created_by') \
            .prefetch_related('group__permissions', 'group__user')

        form = RoleFilterForm(self.request.GET)
        if form.is_valid():
            search = form.cleaned_data.get('search')
            company = form.cleaned_data.get('company')
            is_system_role = form.cleaned_data.get('is_system_role')
            is_active = form.cleaned_data.get('is_active')

            if search:
                queryset = queryset.filter(
                    Q(group__name__icontains=search) |
                    Q(description__icontains=search)
                )

            if company:
                queryset = queryset.filter(Q(company=company) | Q(company__isnull=True))

            if is_system_role is not None:
                queryset = queryset.filter(is_system_role=(is_system_role == 'true'))

            if is_active is not None:
                queryset = queryset.filter(is_active=(is_active == 'true'))

        return queryset.order_by('-priority', 'group__name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add filter form
        context['filter_form'] = RoleFilterForm(self.request.GET)

        # Add analytics data
        total_roles = Role.objects.count()
        system_roles = Role.objects.filter(is_system_role=True).count()
        active_roles = Role.objects.filter(is_active=True).count()

        context.update({
            'total_roles': total_roles,
            'system_roles': system_roles,
            'custom_roles': total_roles - system_roles,
            'active_roles': active_roles,
            'inactive_roles': total_roles - active_roles,
            'can_create_role': self.request.user.has_perm('auth.add_group'),
            'can_manage_system_roles': (
                self.request.user.has_perm('accounts.can_manage_system_roles') or
                self.request.user.is_superuser
            )
        })

        return context

class RoleDetailView(LoginRequiredMixin, RolePermissionMixin, DetailView):
    model = Role
    template_name = 'accounts/roles/role_detail.html'
    context_object_name = 'role'

    def get_queryset(self):
        return Role.objects.select_related(
            'group', 'company', 'created_by'
        ).prefetch_related(
            'group__permissions__content_type',
            'history__user'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        role = self.object

        # Users with this role — select related company instead of non-existent profile
        users_with_role = role.group.user_set.select_related('company').all()

        permission_groups = role.get_permission_groups()
        recent_history = role.history.select_related('user')[:10]

        capacity_info = {
            'current': role.user_count,
            'maximum': role.max_users,
            'percentage': role.capacity_percentage,
            'is_at_capacity': role.is_at_capacity,
            'can_assign': role.can_assign_to_user()[0]
        }

        context.update({
            'users_with_role': users_with_role,
            'permission_groups': permission_groups,
            'recent_history': recent_history,
            'capacity_info': capacity_info,
            'can_edit': self.request.user.has_perm('auth.change_group'),
            'can_delete': (
                self.request.user.has_perm('auth.delete_group') and
                not role.is_system_role
            ),
        })
        return context


class UserRoleAssignView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Bulk assign multiple users to a role"""
    template_name = 'accounts/roles/assign_users.html'
    form_class = BulkUserRoleAssignForm
    permission_required = 'accounts.add_customuser'

    def dispatch(self, request, *args, **kwargs):
        """Override to add better error handling"""
        # Check if user has required permission
        if not request.user.has_perm('accounts.add_customuser'):
            messages.error(request, "You don't have permission to assign user roles.")
            return redirect('role_list')

        # Check if user has a company
        if not request.user.company and not getattr(request.user, 'is_saas_admin', False):
            messages.error(request, "You must be associated with a company to assign roles.")
            return redirect('role_list')

        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['company'] = self.request.user.company
        kwargs['requesting_user'] = self.request.user  # Pass current user for filtering
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_user = self.request.user
        company = current_user.company

        # Get accessible roles for this user
        accessible_roles = Role.objects.accessible_by_user(current_user)

        # Annotate with actual user counts (only active, non-hidden users)
        roles_with_counts = accessible_roles.annotate(
            active_user_count=Count(
                'group__user',
                filter=Q(
                    group__user__is_hidden=False,
                    group__user__is_active=True
                ),
                distinct=True
            )
        ).order_by('-priority', 'group__name')

        # Create role list with computed properties
        roles_list = []
        for role in roles_with_counts:
            # Calculate capacity info
            capacity_percentage = 0
            is_at_capacity = False
            available_slots = None

            if role.max_users:
                capacity_percentage = min(100, (role.active_user_count / role.max_users) * 100)
                is_at_capacity = role.active_user_count >= role.max_users
                available_slots = max(0, role.max_users - role.active_user_count)

            # Add computed attributes to role object
            role.computed_user_count = role.active_user_count
            role.computed_capacity_percentage = round(capacity_percentage, 1)
            role.computed_is_at_capacity = is_at_capacity
            role.computed_available_slots = available_slots

            roles_list.append(role)

        # Get manageable users for current user
        manageable_users = current_user.get_manageable_users().order_by('first_name', 'last_name')

        # Prepare user list with display information
        users_list = []
        for user in manageable_users:
            # Get user's primary role for display
            primary_role = user.primary_role

            # Add display attributes
            user.computed_display_role = primary_role.group.name if primary_role else "No Role"
            user.computed_role_count = user.groups.filter(role__isnull=False).count()
            user.computed_role_priority = user.highest_role_priority

            users_list.append(user)

        # Statistics
        context.update({
            'roles': roles_list,
            'available_users': users_list,
            'total_users': len(users_list),
            'total_roles': len(roles_list),
            'company': company,
            'is_saas_admin': getattr(current_user, 'is_saas_admin', False),
            'can_create_roles': current_user.has_perm('auth.add_group'),
        })

        return context

    def form_valid(self, form):
        users = form.cleaned_data['users']
        role = form.cleaned_data['role']
        current_user = self.request.user

        # Verify user has permission to assign this role
        accessible_roles = Role.objects.accessible_by_user(current_user)
        if role not in accessible_roles:
            messages.error(
                self.request,
                f"You don't have permission to assign the role '{role.group.name}'."
            )
            return self.form_invalid(form)

        # Check role capacity
        if role.max_users:
            current_count = role.group.user_set.filter(
                is_hidden=False,
                is_active=True
            ).count()
            new_count = current_count + len(users)

            if new_count > role.max_users:
                messages.error(
                    self.request,
                    f"Cannot assign {len(users)} users. Role '{role.group.name}' "
                    f"has capacity for only {role.max_users - current_count} more users."
                )
                return self.form_invalid(form)

        # Assign users to role with proper error handling
        success_count = 0
        already_assigned = []
        errors = []

        try:
            with transaction.atomic():
                for user in users:
                    # Verify user is manageable
                    if not current_user.can_manage_user(user):
                        errors.append(f"{user.get_full_name()}: No permission to manage this user")
                        continue

                    # Check if already assigned
                    if user.groups.filter(pk=role.group.pk).exists():
                        already_assigned.append(user.get_full_name())
                        continue

                    # Assign role
                    try:
                        role.group.user_set.add(user)
                        success_count += 1

                        # Log role assignment
                        RoleHistory.objects.create(
                            role=role,
                            action='assigned',
                            user=current_user,
                            affected_user=user,
                            notes=f"Role assigned via bulk assignment"
                        )

                        # Create audit log
                        AuditLog.objects.create(
                            user=current_user,
                            action='user_updated',
                            action_description=f"Assigned role '{role.group.name}' to user {user.email}",
                            content_object=user,
                            resource_name=user.get_full_name() or user.email,
                            ip_address=get_client_ip(self.request),
                            success=True,
                            metadata={
                                'role_id': role.id,
                                'role_name': role.group.name,
                                'assignment_type': 'bulk'
                            }
                        )
                    except Exception as e:
                        errors.append(f"{user.get_full_name()}: {str(e)}")
                        logger.error(f"Error assigning role to user {user.id}: {e}")

        except Exception as e:
            logger.error(f"Error during bulk role assignment: {e}")
            messages.error(self.request, f"An error occurred during assignment: {str(e)}")
            return self.form_invalid(form)

        # Success messages
        if success_count > 0:
            messages.success(
                self.request,
                f"Successfully assigned {success_count} user(s) to role '{role.group.name}'."
            )

        # Warning for already assigned users
        if already_assigned:
            user_names = ', '.join(already_assigned[:3])
            if len(already_assigned) > 3:
                user_names += f" and {len(already_assigned) - 3} more"
            messages.warning(
                self.request,
                f"{len(already_assigned)} user(s) already had this role: {user_names}"
            )

        # Errors
        if errors:
            for error in errors[:5]:
                messages.error(self.request, error)
            if len(errors) > 5:
                messages.warning(
                    self.request,
                    f"...and {len(errors) - 5} more errors. Check logs for details."
                )

        # Redirect to role detail if successful, back to form if all failed
        if success_count > 0:
            return redirect('role_detail', pk=role.pk)
        else:
            return self.form_invalid(form)

    def form_invalid(self, form):
        if form.errors:
            messages.error(
                self.request,
                "Please correct the errors in the form."
            )
        return super().form_invalid(form)

class RoleCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """
    Advanced role creation with enhanced form handling
    """
    model = Role
    form_class = RoleForm
    template_name = 'accounts/roles/role_form.html'
    permission_required = 'accounts.add_role'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'title': 'Create New Role',
            'submit_text': 'Create Role',
            'breadcrumb_title': 'Create Role'
        })
        return context

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Role "{form.cleaned_data["name"]}" has been created successfully.'
        )
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(
            self.request,
            'Please correct the errors below and try again.'
        )
        return super().form_invalid(form)


class RoleUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """
    Advanced role editing with change tracking
    """
    model = Role
    form_class = RoleForm
    template_name = 'accounts/roles/role_form.html'
    permission_required = 'accounts.change_role'

    def dispatch(self, request, *args, **kwargs):
        # Prevent editing system roles unless user has special permission
        role = self.get_object()
        if (role.is_system_role and
                not request.user.has_perm('accounts.can_manage_system_roles') and
                not request.user.is_superuser):
            messages.error(
                request,
                "You don't have permission to edit system roles."
            )
            return HttpResponseRedirect(reverse('role_detail', kwargs={'pk': role.pk}))

        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        role = self.object
        context.update({
            'title': f'Edit Role: {role.group.name}',
            'submit_text': 'Update Role',
            'breadcrumb_title': f'Edit {role.group.name}',
            'role': role
        })
        return context

    def form_valid(self, form):
        messages.success(
            self.request,
            f'Role "{self.object.group.name}" has been updated successfully.'
        )
        return super().form_valid(form)


class RoleDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """
    Safe role deletion with confirmation and validation
    """
    model = Role
    template_name = 'accounts/roles/role_confirm_delete.html'
    success_url = reverse_lazy('role_list')
    permission_required = 'accounts.delete_role'

    def dispatch(self, request, *args, **kwargs):
        role = self.get_object()

        # Prevent deletion of system roles
        if role.is_system_role:
            messages.error(request, "System roles cannot be deleted.")
            return HttpResponseRedirect(reverse('role_detail', kwargs={'pk': role.pk}))

        # Check if role has users assigned
        if role.user_count > 0:
            messages.warning(
                request,
                f"Cannot delete role '{role.group.name}' because {role.user_count} "
                f"user(s) are assigned to it. Please reassign these users first."
            )
            return HttpResponseRedirect(reverse('role_detail', kwargs={'pk': role.pk}))

        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        role = self.get_object()
        role_name = role.group.name

        # Create history record before deletion
        RoleHistory.objects.create(
            role=role,
            action='deleted',
            user=request.user,
            notes=f"Role deleted via web interface"
        )

        # Delete the underlying group (will cascade to role)
        role.group.delete()

        messages.success(
            request,
            f'Role "{role_name}" has been deleted successfully.'
        )

        return HttpResponseRedirect(self.success_url)


class RoleBulkAssignmentView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    permission_required = 'accounts.add_role'
    form_class = BulkRoleAssignmentForm
    template_name = 'accounts/roles/bulk_assignment.html'
    success_url = reverse_lazy('role_list')

    def has_permission(self):
        tenant = getattr(connection, 'tenant', None)
        if not tenant or tenant.schema_name == 'public':
            return False
        with tenant_context(tenant):
            return self.request.user.has_perm(self.permission_required)

    def dispatch(self, request, *args, **kwargs):
        tenant = getattr(connection, 'tenant', None)
        if not tenant:
            raise PermissionDenied("Not in tenant schema")
        with tenant_context(tenant):
            return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        tenant = getattr(connection, 'tenant', None)
        if tenant:
            kwargs['company'] = getattr(tenant, 'company', None)
        return kwargs


class RoleAnalyticsView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """
    Role analytics and reporting with charts and insights
    """
    model = Role
    template_name = 'accounts/roles/role_analytics.html'
    permission_required = 'accounts.change_role'
    context_object_name = 'role'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        role = self.object

        # Date ranges
        thirty_days_ago = timezone.now() - timedelta(days=30)
        seven_days_ago = timezone.now() - timedelta(days=7)

        # Get history data for charts
        history_data = role.history.filter(
            timestamp__gte=thirty_days_ago
        ).values('action', 'timestamp').order_by('timestamp')

        # Count actions by type
        action_counts = {}
        for entry in history_data:
            action = entry['action']
            action_counts[action] = action_counts.get(action, 0) + 1

        # Recent activity (last 7 days)
        recent_history = role.history.filter(
            timestamp__gte=seven_days_ago
        ).select_related('user').order_by('-timestamp')[:10]

        # Permission usage analytics
        permissions_by_app = role.get_permission_groups()

        # Calculate permission distribution
        permission_stats = {
            app: {
                'total': len(perms),
                'view': sum(1 for p in perms if 'view' in p.codename),
                'add': sum(1 for p in perms if 'add' in p.codename),
                'change': sum(1 for p in perms if 'change' in p.codename),
                'delete': sum(1 for p in perms if 'delete' in p.codename),
            }
            for app, perms in permissions_by_app.items()
        }

        # Users with this role grouped by company (if applicable)
        users_by_company = {}
        if role.company:
            users_by_company[role.company.name] = role.user_count
        else:
            # Group users by their companies for system roles
            from django.contrib.auth import get_user_model
            User = get_user_model()

            users_with_role = User.objects.filter(
                groups=role.group,
                is_hidden=False
            ).select_related('company')

            for user in users_with_role:
                company_name = user.company.name if user.company else 'No Company'
                users_by_company[company_name] = users_by_company.get(company_name, 0) + 1

        # User activity analysis
        active_users_count = role.group.user_set.filter(
            is_active=True,
            is_hidden=False
        ).count()

        context.update({
            'history_data': list(history_data),
            'action_counts': action_counts,
            'recent_history': recent_history,
            'permissions_by_app': permissions_by_app,
            'permission_stats': permission_stats,
            'users_by_company': users_by_company,
            'active_users_count': active_users_count,
            'inactive_users_count': role.user_count - active_users_count,
            'capacity_usage': {
                'current': role.user_count,
                'maximum': role.max_users or 'Unlimited',
                'percentage': role.capacity_percentage,
                'available': (role.max_users - role.user_count) if role.max_users else 'Unlimited'
            },
            'timeline_days': 30,
        })

        return context


class RoleHistoryView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Role change history listing with filters
    """
    model = RoleHistory
    template_name = 'accounts/roles/role_history.html'
    context_object_name = 'history_entries'
    paginate_by = 50
    permission_required = 'accounts.add_role'

    def get_queryset(self):
        role_pk = self.kwargs.get('pk')
        queryset = RoleHistory.objects.select_related('role__group', 'user')

        if role_pk:
            # History for specific role
            queryset = queryset.filter(role_id=role_pk)

        # Filter by action if provided
        action = self.request.GET.get('action')
        if action:
            queryset = queryset.filter(action=action)

        # Filter by date range
        days = self.request.GET.get('days', 30)
        if days != 'all':
            try:
                days_ago = timezone.now() - timedelta(days=int(days))
                queryset = queryset.filter(timestamp__gte=days_ago)
            except ValueError:
                pass

        return queryset.order_by('-timestamp')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        role_pk = self.kwargs.get('pk')

        if role_pk:
            role = get_object_or_404(Role, pk=role_pk)
            context.update({
                'role': role,
                'title': f'History for {role.group.name}',
                'breadcrumb_title': f'{role.group.name} History'
            })
        else:
            context.update({
                'title': 'All Role Changes',
                'breadcrumb_title': 'Role History'
            })

        # Get available actions for filter
        context['available_actions'] = RoleHistory.ACTION_CHOICES
        context['current_action'] = self.request.GET.get('action', '')
        context['current_days'] = self.request.GET.get('days', '30')

        return context


class RoleToggleActiveView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """
    Toggle role active status via AJAX
    """
    model = Role
    permission_required = 'accounts.add_role'

    def post(self, request, *args, **kwargs):
        role = self.get_object()

        # Check if user can modify this role
        if role.company and role.company != request.user.company:
            return JsonResponse({
                'success': False,
                'message': 'You do not have permission to modify this role'
            }, status=403)

        # Don't allow deactivating system roles with users
        if role.is_active and role.is_system_role and role.user_count > 0:
            return JsonResponse({
                'success': False,
                'message': 'Cannot deactivate system role with assigned users. Remove users first.'
            }, status=400)

        # Toggle active status
        old_status = role.is_active
        role.is_active = not role.is_active
        role.save()

        # Create history record
        action = 'activated' if role.is_active else 'deactivated'
        RoleHistory.objects.create(
            role=role,
            action=action,
            user=request.user,
            changes={
                'old_status': old_status,
                'new_status': role.is_active
            },
            notes=f"Role {action} via web interface"
        )

        return JsonResponse({
            'success': True,
            'is_active': role.is_active,
            'message': f'Role "{role.group.name}" has been {"activated" if role.is_active else "deactivated"}'
        })


class RolePermissionsAPIView(LoginRequiredMixin,RolePermissionMixin, DetailView):
    """
    API endpoint for role permissions (used by frontend)
    """
    model = Role

    def get(self, request, *args, **kwargs):
        role = self.get_object()

        # Check access
        if role.company and role.company != request.user.company:
            return JsonResponse({
                'error': 'Access denied'
            }, status=403)

        permissions_data = {}
        for app_name, permissions in role.get_permission_groups().items():
            permissions_data[app_name] = [
                {
                    'id': perm.id,
                    'name': perm.name,
                    'codename': perm.codename,
                    'content_type': perm.content_type.model,
                    'action': perm.codename.split('_')[0]  # add, change, delete, view
                }
                for perm in permissions
            ]

        return JsonResponse({
            'role_id': role.id,
            'role_name': role.group.name,
            'description': role.description or '',
            'is_system_role': role.is_system_role,
            'is_active': role.is_active,
            'permissions': permissions_data,
            'total_permissions': role.permission_count,
            'user_count': role.user_count
        })


class RoleAutocompleteView(LoginRequiredMixin,RolePermissionMixin, ListView):
    """
    Autocomplete API for role selection (for Select2, etc.)
    """
    model = Role

    def get_queryset(self):
        query = self.request.GET.get('q', '')
        company = self.request.user.company

        queryset = Role.objects.filter(
            Q(company=company) | Q(is_system_role=True),
            is_active=True
        ).select_related('group')

        if query:
            queryset = queryset.filter(
                Q(group__name__icontains=query) |
                Q(description__icontains=query)
            )

        return queryset.order_by('-priority', 'group__name')[:10]  # Limit results

    def render_to_response(self, context):
        roles_data = [
            {
                'id': role.id,
                'text': role.group.name,
                'description': role.description or '',
                'is_system': role.is_system_role,
                'user_count': role.user_count,
                'color': role.color_code,
                'capacity': {
                    'current': role.user_count,
                    'max': role.max_users,
                    'available': (role.max_users - role.user_count) if role.max_users else None
                }
            }
            for role in context['object_list']
        ]

        return JsonResponse({
            'results': roles_data,
            'pagination': {
                'more': False
            }
        })


class RoleCompareView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """
    Compare permissions between multiple roles
    """
    model = Role
    template_name = 'accounts/roles/role_compare.html'
    permission_required = 'accounts.add_role'

    def get_queryset(self):
        role_ids = self.request.GET.getlist('roles')
        company = self.request.user.company

        return Role.objects.filter(
            Q(id__in=role_ids),
            Q(company=company) | Q(is_system_role=True)
        ).select_related('group').prefetch_related('group__permissions')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        roles = list(context['object_list'])

        if not roles:
            context['error'] = 'Please select at least one role to compare'
            return context

        # Get all permissions from all roles
        all_permissions = set()
        role_permissions = {}

        for role in roles:
            perms = set(role.group.permissions.all())
            role_permissions[role.id] = perms
            all_permissions.update(perms)

        # Group permissions by app and model
        from collections import defaultdict
        permission_matrix = defaultdict(lambda: defaultdict(dict))

        for perm in all_permissions:
            app = perm.content_type.app_label
            model = perm.content_type.model
            action = perm.codename.split('_')[0]

            for role in roles:
                has_perm = perm in role_permissions[role.id]
                if role.id not in permission_matrix[app][model]:
                    permission_matrix[app][model][role.id] = {}
                permission_matrix[app][model][role.id][action] = has_perm

        context['roles'] = roles
        context['permission_matrix'] = dict(permission_matrix)

        return context


@login_required
@permission_required('accounts.view_customuser', raise_exception=True)
def user_dashboard(request):
    """Enhanced user dashboard with SaaS admin support"""
    user = request.user

    # Basic user statistics
    context = {
        'user': user,
        'total_login_count': user.login_count,
        'last_login_ip': user.last_login_ip,
        'account_age': (timezone.now() - user.date_joined).days,
        'is_locked': user.is_locked,
        'two_factor_enabled': user.two_factor_enabled,
        'email_verified': user.email_verified,
        'phone_verified': user.phone_verified,
        'is_saas_admin': getattr(user, 'is_saas_admin', False),
    }

    # Handle company context for different user types
    if getattr(user, 'is_saas_admin', False):
        # SaaS admin can see all companies
        accessible_companies = get_accessible_companies(user)
        context.update({
            'accessible_companies': accessible_companies,
            'can_switch_companies': True,
            'company_count': accessible_companies.count(),
        })
    else:
        # Regular user logic
        owned_company = getattr(user, 'owned_company', None)
        user_company = getattr(user, 'company', None)

        company_memberships = []
        if user_company and user_company.is_active:
            company_memberships = [user_company]

        context.update({
            'owned_company': owned_company,
            'company_memberships': company_memberships,
        })

    # Add admin statistics if user has permissions
    if user.has_perm('accounts.add_customuser') or getattr(user, 'is_saas_admin', False):
        accessible_users = _get_accessible_users(user)
        user_stats = {
            'total': accessible_users.filter(is_hidden=False).count(),
            'active': accessible_users.filter(is_active=True, is_hidden=False).count(),
            'new_today': accessible_users.filter(is_hidden=False,date_joined__date=timezone.now().date()).count(),
            'locked': accessible_users.filter(is_hidden=False,locked_until__gt=timezone.now()).count(),
        }
        context['user_stats'] = user_stats

    return render(request, 'accounts/analytics.html', context)


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def company_user_list(request):
    """List users for the current tenant company."""
    company = getattr(request, 'tenant', None)
    if not company:
        raise PermissionDenied("No active tenant company found.")

    # SaaS admins can access any tenant; others only their own
    if hasattr(request.user, 'company') and request.user.company != company and not getattr(request.user, 'is_saas_admin', False):
        raise PermissionDenied("You don't have access to this company.")

    with schema_context(company.schema_name):
        if getattr(request.user, 'is_saas_admin', False):
            company_users = CustomUser.objects.filter(company=company).order_by('-date_joined')
        else:
            company_users = get_visible_users().filter(company=company).order_by('-date_joined')

        # Search filter
        search_query = request.GET.get('search', '')
        if search_query:
            company_users = company_users.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(email__icontains=search_query) |
                Q(username__icontains=search_query)
            )

        paginator = Paginator(company_users, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        can_add_users = getattr(company, "can_add_employee", lambda: False)()

        context = {
            'company': company,
            'company_users': page_obj,
            'search_query': search_query,
            'can_add_users': can_add_users,
            'is_saas_admin': getattr(request.user, 'is_saas_admin', False),
            'visible_user_count': get_company_user_count(company),
            'user': request.user,
        }

    return render(request, 'accounts/company_user_list.html', context)


@login_required
@permission_required('accounts.add_customuser', raise_exception=True)
def assign_user_to_company(request):
    """Assign existing user to the current tenant company."""
    company = getattr(request, 'tenant', None)
    if not company:
        raise PermissionDenied("No active tenant company found.")

    if hasattr(company, 'can_add_employee') and not company.can_add_employee():
        messages.error(request, 'Company has reached the maximum user limit.')
        return redirect('company_user_list')

    with schema_context(company.schema_name):
        if request.method == 'POST':
            user_email = request.POST.get('user_email')
            is_admin = request.POST.get('is_admin') == 'on'

            try:
                user = CustomUser.objects.get(email=user_email)

                if user.company:
                    if user.company == company:
                        messages.error(request, f'User {user.get_full_name()} is already in this company.')
                    else:
                        messages.error(request, f'User {user.get_full_name()} belongs to another company.')
                else:
                    user.company = company
                    user.company_admin = is_admin
                    user.save()
                    messages.success(request, f'User {user.get_full_name()} assigned successfully.')

            except CustomUser.DoesNotExist:
                messages.error(request, 'User with this email does not exist.')

            return redirect('company_user_list')

        available_users = CustomUser.objects.filter(company__isnull=True, is_active=True)[:100]

    return render(request, 'accounts/assign_user_to_company.html', {
        'company': company,
        'available_users': available_users,
    })


@login_required
@permission_required('accounts.delete_customuser', raise_exception=True)
def remove_user_from_company(request, user_id):
    """Remove user from the current tenant company."""
    company = getattr(request, 'tenant', None)
    if not company:
        raise PermissionDenied("No active tenant company found.")

    with schema_context(company.schema_name):
        user = get_object_or_404(CustomUser, id=user_id)

        if user.company != company:
            messages.error(request, 'User does not belong to this company.')
            return redirect('company_user_list')

        if hasattr(company, 'owner') and company.owner == user:
            messages.error(request, 'Cannot remove company owner.')
            return redirect('company_user_list')

        if request.user == user:
            messages.error(request, 'Cannot remove yourself.')
            return redirect('company_user_list')

        if user.is_saas_admin and not getattr(request.user, 'is_saas_admin', False):
            messages.error(request, 'Cannot remove SaaS administrators.')
            return redirect('company_user_list')

        try:
            with transaction.atomic():
                user_name = user.get_full_name() or user.email
                user.company_admin = False
                user.company = None
                user.save()
                logger.info(f"User {user_name} removed from company {company.name} by {request.user}")
                messages.success(request, f'{user_name} removed successfully.')
        except Exception as e:
            logger.error(f"Error removing user: {e}")
            messages.error(request, f'Error removing user: {str(e)}')

    return redirect('company_user_list')


@login_required
@permission_required('accounts.change_customuser', raise_exception=True)
def toggle_company_admin(request, user_id):
    """Toggle admin status for user in the current tenant company."""
    company = getattr(request, 'tenant', None)
    if not company:
        raise PermissionDenied("No active tenant company found.")

    with schema_context(company.schema_name):
        user = get_object_or_404(CustomUser, id=user_id)

        if user.company != company:
            messages.error(request, 'User does not belong to this company.')
            return redirect('company_user_list')

        if user.is_saas_admin and not getattr(request.user, 'is_saas_admin', False):
            messages.error(request, 'Cannot modify SaaS admins.')
            return redirect('company_user_list')

        if request.user == user:
            messages.error(request, 'Cannot modify your own admin status.')
            return redirect('company_user_list')

        try:
            with transaction.atomic():
                if user.company_admin:
                    admin_count = CustomUser.objects.filter(
                        company=company, company_admin=True, is_active=True
                    ).exclude(id=user.id).count()
                    if admin_count == 0:
                        messages.error(request, 'At least one admin is required.')
                        return redirect('company_user_list')

                    user.company_admin = False
                    action = 'removed'
                else:
                    user.company_admin = True
                    action = 'granted'

                user.save()
                messages.success(request, f'Admin status {action} for {user.get_full_name()}.')

        except Exception as e:
            logger.error(f"Error toggling admin: {e}")
            messages.error(request, f'Error updating admin status: {str(e)}')

    return redirect('company_user_list')


def _get_accessible_users(user):
    """
    Returns a queryset of users accessible by the current user.
    UPDATED for your model structure.
    """
    if getattr(user, 'is_saas_admin', False):
        # SaaS admins can see all users including hidden ones
        return CustomUser.objects.all()

    if getattr(user, 'can_access_all_companies', False):
        # Users with cross-company access
        return CustomUser.objects.filter(is_hidden=False)

    # Company-specific access
    if hasattr(user, 'company'):
        # Get users in same company, excluding hidden users
        base_qs = CustomUser.objects.filter(company=user.company, is_hidden=False)

        # Apply role hierarchy filtering
        user_priority = user.highest_role_priority
        if user_priority > 0:
            # Filter out users with higher role priority
            from django.db.models import Max, Q
            base_qs = base_qs.annotate(
                max_role_priority=Max('groups__role__priority')
            ).filter(
                Q(max_role_priority__lte=user_priority) | Q(max_role_priority__isnull=True)
            )

        return base_qs

    # Default: only see themselves
    return CustomUser.objects.filter(id=user.id)


@login_required
@permission_required('accounts.can_view_reports')
def user_analytics(request):
    """Enhanced user analytics with proper user filtering"""
    # Get date range from request
    days = int(request.GET.get('days', 30))
    end_date = timezone.now()
    start_date = end_date - timedelta(days=days)

    # Get accessible users (properly filtered)
    accessible_users = _get_accessible_users(request.user)

    # User registration analytics
    registration_data = []
    current_date = start_date.date()
    while current_date <= end_date.date():
        count = accessible_users.filter(date_joined__date=current_date).count()
        registration_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'count': count
        })
        current_date += timedelta(days=1)

    role_stats = []
    company = request.user.company if hasattr(request.user, 'company') else None

    if company:
        # Get role statistics for this company
        from django.db.models import Count, Q
        role_stats = list(
            Role.objects.filter(
                Q(company=company) | Q(is_system_role=True)
            ).annotate(
                user_count=Count(
                    'group__user',
                    filter=Q(
                        group__user__in=accessible_users,
                        group__user__is_hidden=False,
                        group__user__is_active=True
                    ),
                    distinct=True
                )
            ).filter(user_count__gt=0).values(
                'group__name',
                'user_count',
                'priority',
                'color_code'
            ).order_by('-priority')
        )

    # Active vs Inactive users
    active_inactive_data = [
        {'status': 'Active', 'count': accessible_users.filter(is_active=True).count()},
        {'status': 'Inactive', 'count': accessible_users.filter(is_active=False).count()},
    ]

    # Company distribution (for high-level admins)
    company_data = []
    is_high_level = (
            (request.user.primary_role and request.user.primary_role.priority >= 90) or
            getattr(request.user, 'is_saas_admin', False)
    )

    if is_high_level:
        from company.models import Company
        for company_obj in Company.objects.all()[:10]:
            user_count = get_company_user_count(company_obj)
            if user_count > 0:
                company_data.append({
                    'name': company_obj.name,
                    'user_count': user_count
                })
        company_data.sort(key=lambda x: x['user_count'], reverse=True)

    context = {
        'registration_data': json.dumps(registration_data),
        'user_type_data': json.dumps(role_stats),  # Keep variable name for template compatibility
        'active_inactive_data': json.dumps(active_inactive_data),
        'company_data': json.dumps(company_data),
        'days': days,
        'total_users': accessible_users.count(),
        'new_users_period': accessible_users.filter(date_joined__gte=start_date).count(),
        'is_super_admin': is_high_level,
        'is_saas_admin': getattr(request.user, 'is_saas_admin', False),
    }

    return render(request, 'accounts/analytics/user_analytics.html', context)


@login_required
@permission_required('accounts.can_view_reports')
def export_analytics_data(request):
    """Export analytics data in various formats"""
    export_format = request.GET.get('format', 'pdf')
    days = int(request.GET.get('days', 30))

    # Get the same data as the analytics view
    end_date = timezone.now()
    start_date = end_date - timedelta(days=days)
    accessible_users = _get_accessible_users(request.user)

    # Prepare data
    registration_data = []
    current_date = start_date.date()
    while current_date <= end_date.date():
        count = accessible_users.filter(date_joined__date=current_date).count()
        registration_data.append({
            'date': current_date.strftime('%Y-%m-%d'),
            'count': count
        })
        current_date += timedelta(days=1)

    company = request.user.company if hasattr(request.user, 'company') else None
    role_stats = []

    if company:
        from django.db.models import Count, Q
        role_stats = list(
            Role.objects.filter(
                Q(company=company) | Q(is_system_role=True)
            ).annotate(
                user_count=Count(
                    'group__user',
                    filter=Q(
                        group__user__in=accessible_users,
                        group__user__is_hidden=False,
                        group__user__is_active=True
                    ),
                    distinct=True
                )
            ).filter(user_count__gt=0).values(
                'group__name',
                'user_count',
                'priority',
                'color_code'
            ).order_by('-priority')
        )

    active_inactive_data = [
        {'status': 'Active', 'count': accessible_users.filter(is_active=True).count()},
        {'status': 'Inactive', 'count': accessible_users.filter(is_active=False).count()},
    ]

    # Generate export based on format
    if export_format == 'pdf':
        return generate_pdf_report(request, {
            'registration_data': registration_data,
            'user_type_data': role_stats,  # Keep variable name for compatibility
            'active_inactive_data': active_inactive_data,
            'days': days,
            'total_users': accessible_users.count(),
            'new_users_period': accessible_users.filter(date_joined__gte=start_date).count(),
            'start_date': start_date,
            'end_date': end_date,
        })
    elif export_format == 'excel':
        return generate_excel_report(request, {
            'registration_data': registration_data,
            'user_type_data': role_stats,
            'active_inactive_data': active_inactive_data,
            'days': days,
            'total_users': accessible_users.count(),
            'new_users_period': accessible_users.filter(date_joined__gte=start_date).count(),
        })
    elif export_format == 'csv':
        return generate_csv_report(registration_data, role_stats, active_inactive_data)

    return JsonResponse({'error': 'Invalid format'}, status=400)


def generate_pdf_report(request, data):
    """Generate comprehensive PDF analytics report"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)

    # Container for the 'Flowable' objects
    elements = []

    # Get styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1,  # Center alignment
        textColor=HexColor('#667eea')
    )

    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=12,
        textColor=HexColor('#4a5568')
    )

    # Title
    title = Paragraph("User Analytics Report", title_style)
    elements.append(title)

    # Report metadata
    report_info = [
        ['Report Generated:', timezone.now().strftime('%B %d, %Y at %I:%M %p')],
        ['Period:', f'Last {data["days"]} days'],
        ['Date Range:', f'{data["start_date"].strftime("%B %d, %Y")} - {data["end_date"].strftime("%B %d, %Y")}'],
        ['Generated by:', request.user.get_full_name() or request.user.username],
    ]

    info_table = Table(report_info, colWidths=[2 * inch, 3 * inch])
    info_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 20))

    # Summary Statistics
    elements.append(Paragraph("Executive Summary", heading_style))

    summary_data = [
        ['Metric', 'Value'],
        ['Total Users', f"{data['total_users']:,}"],
        ['New Users (Period)', f"{data['new_users_period']:,}"],
        ['Active Users', f"{data['active_inactive_data'][0]['count']:,}"],
        ['Inactive Users', f"{data['active_inactive_data'][1]['count']:,}"],
    ]

    summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#f7fafc')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))

    # Registration Trends Chart (simplified table representation)
    elements.append(Paragraph("User Registration Trends", heading_style))

    # Group registration data by week for better readability
    weekly_data = {}
    for item in data['registration_data']:
        date_obj = timezone.datetime.strptime(item['date'], '%Y-%m-%d').date()
        week_start = date_obj - timedelta(days=date_obj.weekday())
        week_key = week_start.strftime('%b %d, %Y')

        if week_key not in weekly_data:
            weekly_data[week_key] = 0
        weekly_data[week_key] += item['count']

    reg_data = [['Week Starting', 'New Registrations']]
    for week, count in weekly_data.items():
        reg_data.append([week, str(count)])

    reg_table = Table(reg_data, colWidths=[3 * inch, 2 * inch])
    reg_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#667eea')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#f7fafc')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(reg_table)
    elements.append(Spacer(1, 20))

    # User Roles Distribution
    elements.append(Paragraph("User Roles Distribution", heading_style))

    user_types_data = [['Role', 'Count', 'Percentage']]
    total_typed_users = sum(item.get('user_count', 0) for item in data['user_type_data'])

    for item in data['user_type_data']:
        role_name = item.get('group__name', 'No Role')
        count = item.get('user_count', 0)
        percentage = f"{(count / total_typed_users * 100):.1f}%" if total_typed_users > 0 else "0%"
        user_types_data.append([role_name, str(count), percentage])

    types_table = Table(user_types_data, colWidths=[2.5 * inch, 1.5 * inch, 1 * inch])
    types_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#f0fff4')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(types_table)
    elements.append(Spacer(1, 20))

    # User Status Distribution
    elements.append(Paragraph("User Status Distribution", heading_style))

    status_data = [['Status', 'Count', 'Percentage']]
    total_status_users = sum(item['count'] for item in data['active_inactive_data'])

    for item in data['active_inactive_data']:
        count = item['count']
        percentage = f"{(count / total_status_users * 100):.1f}%" if total_status_users > 0 else "0%"
        status_data.append([item['status'], str(count), percentage])

    status_table = Table(status_data, colWidths=[2.5 * inch, 1.5 * inch, 1 * inch])
    status_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#f59e0b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), HexColor('#fffbeb')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(status_table)

    # Footer
    elements.append(Spacer(1, 40))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=1
    )
    footer = Paragraph(
        f"Generated on {timezone.now().strftime('%B %d, %Y at %I:%M %p')} | User Analytics Report",
        footer_style
    )
    elements.append(footer)

    # Build PDF
    doc.build(elements)

    # Get the value of the BytesIO buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type='application/pdf')
    response[
        'Content-Disposition'] = f'attachment; filename="user_analytics_{data["days"]}days_{timezone.now().strftime("%Y%m%d")}.pdf"'
    response.write(pdf)

    return response


def generate_excel_report(request, data):
    """Generate Excel analytics report"""
    wb = openpyxl.Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Summary Sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.title = "Executive Summary"

    # Headers and styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="667EEA", end_color="667EEA", fill_type="solid")

    # Title
    summary_ws['A1'] = "User Analytics Report - Executive Summary"
    summary_ws['A1'].font = Font(bold=True, size=16)
    summary_ws.merge_cells('A1:B1')

    # Summary data
    summary_data = [
        ["Metric", "Value"],
        ["Total Users", data['total_users']],
        ["New Users (Period)", data['new_users_period']],
        ["Active Users", data['active_inactive_data'][0]['count']],
        ["Inactive Users", data['active_inactive_data'][1]['count']],
    ]

    for row_idx, row_data in enumerate(summary_data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 3:  # Header row
                cell.font = header_font
                cell.fill = header_fill

    # Registration Data Sheet
    reg_ws = wb.create_sheet("Registration Trends")
    reg_ws['A1'] = "Date"
    reg_ws['B1'] = "New Registrations"
    reg_ws['A1'].font = header_font
    reg_ws['B1'].font = header_font
    reg_ws['A1'].fill = header_fill
    reg_ws['B1'].fill = header_fill

    for idx, item in enumerate(data['registration_data'], start=2):
        reg_ws[f'A{idx}'] = item['date']
        reg_ws[f'B{idx}'] = item['count']

    # User Roles Sheet
    types_ws = wb.create_sheet("User Roles")
    types_ws['A1'] = "Role"
    types_ws['B1'] = "Count"
    types_ws['A1'].font = header_font
    types_ws['B1'].font = header_font
    types_ws['A1'].fill = header_fill
    types_ws['B1'].fill = header_fill

    for idx, item in enumerate(data['user_type_data'], start=2):
        types_ws[f'A{idx}'] = item.get('group__name', 'No Role')
        types_ws[f'B{idx}'] = item.get('user_count', 0)

    # Auto-adjust column widths
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response[
        'Content-Disposition'] = f'attachment; filename="user_analytics_{data["days"]}days_{timezone.now().strftime("%Y%m%d")}.xlsx"'

    return response


def generate_csv_report(registration_data, user_type_data, active_inactive_data):
    """Generate CSV analytics report"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="user_analytics_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)

    # Write different sections
    writer.writerow(['USER ANALYTICS REPORT'])
    writer.writerow(['Generated:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow([])

    # Registration data
    writer.writerow(['REGISTRATION TRENDS'])
    writer.writerow(['Date', 'New Registrations'])
    for item in registration_data:
        writer.writerow([item['date'], item['count']])
    writer.writerow([])

    # User roles
    writer.writerow(['USER ROLES DISTRIBUTION'])
    writer.writerow(['Role', 'Count'])
    for item in user_type_data:
        writer.writerow([item.get('group__name', 'No Role'), item.get('user_count', 0)])
    writer.writerow([])

    # User status
    writer.writerow(['USER STATUS DISTRIBUTION'])
    writer.writerow(['Status', 'Count'])
    for item in active_inactive_data:
        writer.writerow([item['status'], item['count']])

    return response


@login_required
def user_profile(request):
    """Display user profile - READ ONLY"""

    user_company = getattr(request.user, 'company', None)
    company_memberships = [user_company] if user_company and user_company.is_active else []

    profile_completion = calculate_profile_completion(request.user)
    recent_activity = get_recent_user_activity(request.user)

    # Get social accounts

    context = {
        'user': request.user,
        'owned_company': getattr(request.user, 'owned_company', None),
        'company_memberships': company_memberships,
        'profile_completion': profile_completion,
        'recent_activity': recent_activity,
        'has_password': request.user.has_usable_password(),
    }

    return render(request, 'accounts/profile.html', context)


@login_required
def edit_profile(request):
    """Edit user profile - EDIT FORM"""

    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=request.user)

        if form.is_valid():
            user = form.save()

            # Handle avatar upload if present
            if 'avatar' in request.FILES:
                avatar = request.FILES['avatar']
                processed_avatar = process_avatar_image(avatar)
                if processed_avatar:
                    user.avatar.save(
                        f"avatar_{user.id}_{avatar.name}",
                        processed_avatar,
                        save=True
                    )

            messages.success(request, 'Your profile has been updated successfully!')
            return redirect('user_profile')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = UserProfileForm(instance=request.user)

    context = {
        'form': form,
        'user': request.user,
    }

    return render(request, 'accounts/edit_profile.html', context)

def process_avatar_image(image_file):
    """
    Process and optimize uploaded avatar images
    """
    try:
        # Open and process the image
        img = PILImage.open(image_file)

        # Convert RGBA to RGB if necessary
        if img.mode in ("RGBA", "P"):
            img = img.convert("RGB")

        # Resize image to standard avatar size
        max_size = (400, 400)
        img.thumbnail(max_size, PILImage.Resampling.LANCZOS)

        # Create a square image (crop to center if needed)
        width, height = img.size
        if width != height:
            # Crop to square
            size = min(width, height)
            left = (width - size) // 2
            top = (height - size) // 2
            right = left + size
            bottom = top + size
            img = img.crop((left, top, right, bottom))

        # Save processed image to BytesIO
        output = BytesIO()
        img.save(output, format='JPEG', quality=85, optimize=True)
        output.seek(0)

        return ContentFile(output.getvalue())

    except Exception as e:
        logger.error(f"Error processing avatar image: {e}")
        return None


def calculate_profile_completion(user):
    """
    Calculate profile completion percentage
    """
    total_fields = 10
    completed_fields = 0

    # Check required fields
    if user.first_name:
        completed_fields += 1
    if user.last_name:
        completed_fields += 1
    if user.email:
        completed_fields += 1
    if user.phone_number:
        completed_fields += 1
    if user.bio:
        completed_fields += 1
    if user.avatar:
        completed_fields += 1
    if user.timezone:
        completed_fields += 1
    if user.language:
        completed_fields += 1
    if user.email_verified:
        completed_fields += 1
    if user.phone_verified:
        completed_fields += 1

    return round((completed_fields / total_fields) * 100)


def get_recent_user_activity(user, limit=5):
    """
    Get recent user activity (customize based on your activity tracking)
    """
    # This is a placeholder - implement based on your activity tracking system
    activities = []

    # Example activities you might track
    if user.last_login:
        activities.append({
            'type': 'login',
            'description': f'Logged in from {user.last_login_ip or "unknown IP"}',
            'timestamp': user.last_login,
            'icon': 'bi-box-arrow-in-right'
        })

    if user.password_changed_at:
        activities.append({
            'type': 'security',
            'description': 'Password was updated',
            'timestamp': user.password_changed_at,
            'icon': 'bi-key'
        })

    # Add more activity types as needed

    # Sort by timestamp and limit
    activities.sort(key=lambda x: x['timestamp'], reverse=True)
    return activities[:limit]


@login_required
def upload_avatar_ajax(request):
    """
    Handle avatar upload via AJAX for immediate preview
    """
    if request.method == 'POST' and request.FILES.get('avatar'):
        avatar_file = request.FILES['avatar']

        # Validate file
        if avatar_file.size > 5 * 1024 * 1024:  # 5MB limit
            return JsonResponse({
                'success': False,
                'error': 'File size must be less than 5MB'
            }, status=400)

        if not avatar_file.content_type.startswith('image/'):
            return JsonResponse({
                'success': False,
                'error': 'File must be an image'
            }, status=400)

        # Process and save avatar
        processed_avatar = process_avatar_image(avatar_file)
        if processed_avatar:
            # Save to user
            request.user.avatar.save(
                f"avatar_{request.user.id}_{avatar_file.name}",
                processed_avatar,
                save=True
            )

            return JsonResponse({
                'success': True,
                'avatar_url': request.user.avatar.url,
                'message': 'Avatar updated successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'Failed to process image'
            }, status=400)

    return JsonResponse({
        'success': False,
        'error': 'Invalid request'
    }, status=400)



@login_required
def export_profile_data(request):
    """
    Export user profile data as JSON
    """
    user_data = {
        'personal_info': {
            'username': request.user.username,
            'email': request.user.email,
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'middle_name': request.user.middle_name,
            'phone_number': request.user.phone_number,
            'bio': request.user.bio,
            'date_joined': request.user.date_joined.isoformat(),
        },
        'preferences': {
            'timezone': request.user.timezone,
            'language': request.user.language,
        },
        'security': {
            'email_verified': request.user.email_verified,
            'phone_verified': request.user.phone_verified,
            'two_factor_enabled': request.user.two_factor_enabled,
            'login_count': request.user.login_count,
            'last_login': request.user.last_login.isoformat() if request.user.last_login else None,
        },
        'account': {
            'primary_role': request.user.display_role,  # ✅ FIXED
            'all_roles': request.user.role_names,  # ✅ FIXED
            'role_priority': request.user.highest_role_priority,  # ✅ ADDED
            'is_active': request.user.is_active,
            'company': request.user.company.name if hasattr(request.user, 'company') and request.user.company else None,
        }
    }

    response = JsonResponse(user_data, json_dumps_params={'indent': 2})
    response['Content-Disposition'] = f'attachment; filename="profile_data_{request.user.username}.json"'
    return response



@login_required
def delete_avatar(request):
    """
    Delete user avatar via AJAX
    """
    if request.method == 'POST':
        if request.user.avatar:
            request.user.avatar.delete(save=True)
            return JsonResponse({
                'success': True,
                'message': 'Avatar deleted successfully!'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': 'No avatar to delete'
            }, status=400)

    return JsonResponse({
        'success': False,
        'error': 'Invalid request method'
    }, status=405)


@login_required
def user_notification_settings(request):
    """
    Handle user notification preferences
    """
    if request.method == 'POST':
        form = UserNotificationForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Notification preferences updated!'
                })

            messages.success(request, 'Notification preferences updated!')
            return redirect('user_notification_settings')
    else:
        form = UserNotificationForm(instance=request.user)

    context = {
        'form': form,
        'user': request.user,
    }

    return render(request, 'accounts/notification_settings.html', context)


@login_required
def user_preferences(request):
    """
    Handle user preferences and customizations
    """
    if request.method == 'POST':
        form = UserPreferencesForm(request.user, request.POST)
        if form.is_valid():
            form.save()

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Preferences updated!',
                    'preferences': request.user.metadata.get('preferences', {})
                })

            messages.success(request, 'Preferences updated successfully!')
            return redirect('user_preferences')
    else:
        form = UserPreferencesForm(request.user)

    context = {
        'form': form,
        'user': request.user,
    }

    return render(request, 'accounts/preferences.html', context)


@login_required
@require_POST
def send_verification(request):
    """
    Send email or phone verification
    """
    verification_type = request.POST.get('type')  # 'email' or 'phone'

    if verification_type == 'email':
        # Generate verification token
        token = secrets.token_urlsafe(32)

        # Save token to user metadata
        if not request.user.metadata:
            request.user.metadata = {}
        request.user.metadata['email_verification_token'] = token
        request.user.metadata['email_verification_expires'] = (
                timezone.now() + timezone.timedelta(hours=24)
        ).isoformat()
        request.user.save()

        # Send verification email
        verification_link = request.build_absolute_uri(
            f"/accounts/profile/verify-email/?token={token}"
        )

        send_mail(
            'Verify Your Email Address',
            render_to_string('emails/verify_email.txt', {
                'user': request.user,
                'verification_link': verification_link
            }),
            settings.DEFAULT_FROM_EMAIL,
            [request.user.email],
            html_message=render_to_string('emails/verify_email.html', {
                'user': request.user,
                'verification_link': verification_link
            }),
            fail_silently=False,
        )

        return JsonResponse({
            'success': True,
            'message': 'Verification email sent! Check your inbox.'
        })

    elif verification_type == 'phone':
        # Generate SMS verification code
        code = ''.join([str(secrets.randbelow(10)) for _ in range(6)])

        # Save code to user metadata
        if not request.user.metadata:
            request.user.metadata = {}
        request.user.metadata['phone_verification_code'] = code
        request.user.metadata['phone_verification_expires'] = (
                timezone.now() + timezone.timedelta(minutes=10)
        ).isoformat()
        request.user.save()

        # Here you would integrate with SMS service (Twilio, etc.)
        # For now, we'll just return the code (remove in production)
        return JsonResponse({
            'success': True,
            'message': f'Verification code sent to {request.user.phone_number}',
            'debug_code': code if settings.DEBUG else None
        })

    return JsonResponse({
        'success': False,
        'error': 'Invalid verification type'
    }, status=400)


@login_required
def verify_email(request):
    """
    Verify email address with token
    """
    token = request.GET.get('token')
    if not token:
        messages.error(request, 'Invalid verification link.')
        return redirect('user_profile')

    # Check token validity
    user_token = request.user.metadata.get('email_verification_token')
    expires = request.user.metadata.get('email_verification_expires')

    if not user_token or user_token != token:
        messages.error(request, 'Invalid or expired verification link.')
        return redirect('user_profile')

    # Check expiration
    if expires:
        expire_time = datetime.fromisoformat(expires)
        # Make aware if stored as naive ISO string
        if timezone.is_naive(expire_time):
            expire_time = timezone.make_aware(expire_time)
        if timezone.now() > expire_time:
            messages.error(request, 'Verification link has expired.')
            return redirect('user_profile')

    # Verify email
    request.user.email_verified = True
    request.user.metadata.pop('email_verification_token', None)
    request.user.metadata.pop('email_verification_expires', None)
    request.user.save()

    messages.success(request, 'Email address verified successfully!')
    return redirect('user_profile')


@login_required
@require_POST
def verify_phone(request):
    """
    Verify phone number with code
    """
    code = request.POST.get('code')
    if not code:
        return JsonResponse({
            'success': False,
            'error': 'Verification code is required'
        }, status=400)

    # Check code validity
    user_code = request.user.metadata.get('phone_verification_code')
    expires = request.user.metadata.get('phone_verification_expires')

    if not user_code or user_code != code:
        return JsonResponse({
            'success': False,
            'error': 'Invalid verification code'
        }, status=400)

    # Check expiration
    if expires:
        expire_time = datetime.fromisoformat(expires)
        # Make aware if stored as naive ISO string
        if timezone.is_naive(expire_time):
            expire_time = timezone.make_aware(expire_time)
        if timezone.now() > expire_time:
            return JsonResponse({
                'success': False,
                'error': 'Verification code has expired'
            }, status=400)

    # Verify phone
    request.user.phone_verified = True
    request.user.metadata.pop('phone_verification_code', None)
    request.user.metadata.pop('phone_verification_expires', None)
    request.user.save()

    return JsonResponse({
        'success': True,
        'message': 'Phone number verified successfully!'
    })


@login_required
def enable_two_factor(request):
    """
    GET  — generate a new TOTP secret, return QR code and manual key.
    POST — verify the submitted code and activate 2FA.

    The secret lives in the session only until POST verification succeeds,
    at which point it is moved to user.metadata and the session copy removed.
    """
    import base64
    user = request.user

    if user.two_factor_enabled:
        return JsonResponse(
            {'success': False, 'error': 'Two-factor authentication is already enabled'},
            status=400
        )

    if request.method == 'POST':
        totp_code = request.POST.get('totp_code', '').strip()
        secret_key = request.session.get('temp_2fa_secret')

        if not secret_key:
            return JsonResponse(
                {'success': False, 'error': 'Setup session expired. Please refresh and try again.'},
                status=400
            )
        if not totp_code:
            return JsonResponse({'success': False, 'error': 'Please enter the 6-digit code.'}, status=400)

        totp = pyotp.TOTP(secret_key)
        if not totp.verify(totp_code, valid_window=1):
            return JsonResponse(
                {'success': False, 'error': 'Invalid or expired code. Please try again.'},
                status=400
            )

        # Persist the secret and mark 2FA as enabled
        if not user.metadata:
            user.metadata = {}
        user.metadata['totp_secret'] = secret_key
        user.two_factor_enabled = True
        user.backup_codes = [secrets.token_hex(4).upper() for _ in range(10)]
        user.save()

        # Remove the temporary session secret
        request.session.pop('temp_2fa_secret', None)

        logger.info(f"2FA enabled for user: {user.email}")
        return JsonResponse({
            'success': True,
            'message': 'Two-factor authentication enabled successfully!',
            'backup_codes': user.backup_codes,
        })

    # GET — generate secret and QR code
    secret_key = pyotp.random_base32()
    request.session['temp_2fa_secret'] = secret_key

    totp_uri = pyotp.TOTP(secret_key).provisioning_uri(
        name=user.email or user.username,
        issuer_name=getattr(settings, 'TWO_FACTOR_ISSUER', 'PrimeBooks')
    )

    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(totp_uri)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    buffer = BytesIO()
    img.save(buffer, format="PNG")
    qr_code_data = base64.b64encode(buffer.getvalue()).decode()

    return JsonResponse({
        'success': True,
        'qr_code': f"data:image/png;base64,{qr_code_data}",
        'manual_entry_key': secret_key,
        'message': 'Scan the QR code with your authenticator app, then enter the 6-digit code to confirm.',
    })


@login_required
@require_POST
def disable_two_factor(request):
    """
    Disable 2FA after verifying the user's current password.
    Clears the TOTP secret, backup codes, and any confirmed TOTPDevice records.
    """
    user = request.user

    if not user.two_factor_enabled:
        return JsonResponse(
            {'success': False, 'error': 'Two-factor authentication is not currently enabled'},
            status=400
        )

    current_password = request.POST.get('current_password', '')
    if not current_password or not user.check_password(current_password):
        return JsonResponse({'success': False, 'error': 'Incorrect password'}, status=400)

    # Clear all 2FA state
    user.two_factor_enabled = False
    user.backup_codes = []
    if user.metadata:
        user.metadata.pop('totp_secret', None)
    user.save()

    # Also remove any django-otp TOTP devices so login flow stays consistent
    TOTPDevice.objects.filter(user=user).delete()

    logger.info(f"2FA disabled for user: {user.email}")
    return JsonResponse({'success': True, 'message': 'Two-factor authentication has been disabled.'})


@login_required
@require_POST
def generate_backup_codes(request):
    """Regenerate 2FA backup codes. Requires 2FA to be active."""
    user = request.user

    if not user.two_factor_enabled:
        return JsonResponse(
            {'success': False, 'error': 'Two-factor authentication is not enabled'},
            status=400
        )

    user.backup_codes = [secrets.token_hex(4).upper() for _ in range(10)]
    user.save(update_fields=['backup_codes'])

    logger.info(f"Backup codes regenerated for user: {user.email}")
    return JsonResponse({
        'success': True,
        'backup_codes': user.backup_codes,
        'message': 'New backup codes generated. Store them somewhere safe.',
    })


# verify_2fa is no longer used — enable_two_factor handles verification inline.
# verify_two_factor_login is no longer used — _handle_2fa_verification_ajax /
# _handle_regular_login handle 2FA during the login flow.




@login_required
def user_activity_log(request):
    """
    Display user activity log
    """
    # This would typically come from a separate activity tracking system
    # For now, we'll create some sample data
    activities = [
        {
            'type': 'login',
            'description': 'Logged in successfully',
            'ip_address': request.user.last_login_ip,
            'timestamp': request.user.last_login,
            'status': 'success'
        },
        # Add more activities as needed
    ]

    context = {
        'activities': activities,
        'user': request.user,
    }

    return render(request, 'accounts/activity_log.html', context)


# Utility functions

def get_user_active_sessions(user):
    """
    Get active sessions for a user (placeholder implementation)
    """
    # This would typically integrate with Django's session framework
    # or a custom session tracking system
    return []


def get_recent_login_attempts(user, limit=10):
    """
    Get recent login attempts for a user (placeholder implementation)
    """
    # This would come from your login attempt tracking system
    return []

@login_required
def change_password(request):
    """Enhanced password change view"""
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            user.password_changed_at = timezone.now()
            user.save(update_fields=['password_changed_at'])

            update_session_auth_hash(request, user)
            messages.success(request, 'Your password has been changed successfully!')
            return redirect('user_profile')
    else:
        form = PasswordChangeForm(request.user)

    return render(request, 'accounts/change_password.html', {'form': form})


@login_required
def user_signature(request):
    """Enhanced user signature management"""
    signature, created = UserSignature.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        form = UserSignatureForm(request.POST, request.FILES, instance=signature)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your signature has been updated successfully!')
            return redirect('user_signature')
    else:
        form = UserSignatureForm(instance=signature)

    return render(request, 'accounts/signature.html', {'form': form, 'signature': signature})


@login_required
def two_factor_setup(request):
    """Enhanced two-factor authentication setup"""
    if request.method == 'POST':
        form = TwoFactorSetupForm(request.POST)
        if form.is_valid():
            # In a real implementation, you would verify the code here
            request.user.two_factor_enabled = True
            request.user.save(update_fields=['two_factor_enabled'])
            messages.success(request, 'Two-factor authentication has been enabled!')
            return redirect('user_profile')
    else:
        form = TwoFactorSetupForm()

    return render(request, 'accounts/two_factor_setup.html', {'form': form})


@login_required
@permission_required('accounts.view_customuser')
def user_quick_stats(request):
    """Enhanced AJAX endpoint for user statistics"""
    accessible_users = _get_accessible_users(request.user)

    company = request.user.company if hasattr(request.user, 'company') else None
    role_stats = []

    if company:
        from django.db.models import Count, Q
        role_stats = list(
            Role.objects.filter(
                Q(company=company) | Q(is_system_role=True)
            ).annotate(
                user_count=Count(
                    'group__user',
                    filter=Q(
                        group__user__in=accessible_users,
                        group__user__is_hidden=False,
                        group__user__is_active=True
                    ),
                    distinct=True
                )
            ).filter(user_count__gt=0).values(
                'group__name',
                'user_count',
                'priority'
            ).order_by('-priority')
        )

    stats = {
        'total_users': accessible_users.count(),
        'active_users': accessible_users.filter(is_active=True).count(),
        'new_users_today': accessible_users.filter(
            date_joined__date=timezone.now().date()
        ).count(),
        'locked_users': accessible_users.filter(
            locked_until__gt=timezone.now()
        ).count(),
        'user_types': role_stats,  # Keep variable name for API compatibility
        'is_saas_admin': getattr(request.user, 'is_saas_admin', False),
    }

    # Add SaaS admin specific stats
    if getattr(request.user, 'is_saas_admin', False):
        stats['hidden_users'] = CustomUser.objects.filter(is_hidden=True).count()
        stats['saas_admins'] = CustomUser.objects.filter(is_saas_admin=True).count()

    return JsonResponse(stats)


@require_saas_admin
def switch_tenant_view(request):
    """Allow SaaS admin to switch between tenants"""
    tenant_id = request.GET.get('tenant_id')

    if tenant_id:
        try:
            from company.models import Company
            company = get_object_or_404(Company, company_id=tenant_id)

            # Store the target company in session for the middleware to handle
            request.session['saas_admin_target_company'] = company.company_id

            messages.success(request, f'Switching to {company.name}...')
            return JsonResponse({
                'success': True,
                'message': f'Switching to {company.name}',
                'redirect_url': f'/?switch_tenant={tenant_id}'
            })

        except Company.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Company not found'
            }, status=404)

    # Return list of available companies for switching
    accessible_companies = get_accessible_companies(request.user)
    companies_data = [
        {
            'company_id': company.company_id,
            'name': company.name,
            'schema_name': getattr(company, 'schema_name', ''),
            'user_count': get_company_user_count(company),
            'status': getattr(company, 'status', 'ACTIVE')
        }
        for company in accessible_companies
    ]

    return JsonResponse({
        'success': True,
        'companies': companies_data
    })


@require_saas_admin
def saas_admin_user_impersonate(request, user_id):
    """Allow SaaS admin to impersonate a user (for support purposes)"""
    target_user = get_object_or_404(CustomUser, id=user_id)

    # Don't allow impersonating other SaaS admins
    if getattr(target_user, 'is_saas_admin', False):
        messages.error(request, 'Cannot impersonate other SaaS administrators.')
        return redirect('user_detail', pk=user_id)

    # Store original user info in session
    request.session['saas_admin_original_user_id'] = request.user.id
    request.session['saas_admin_impersonating'] = True
    request.session['saas_admin_impersonated_user_id'] = target_user.id

    # Log the impersonation for audit purposes
    logger.info(f"SaaS admin {request.user.email} started impersonating user {target_user.email}")

    messages.success(request, f'Now impersonating user: {target_user.get_full_name() or target_user.email}')
    return redirect('user_dashboard')


@require_saas_admin
def saas_admin_stop_impersonation(request):
    """Stop impersonating a user"""
    if not request.session.get('saas_admin_impersonating'):
        messages.error(request, 'No active impersonation session.')
        return redirect('saas_admin_dashboard')

    impersonated_user_id = request.session.get('saas_admin_impersonated_user_id')

    # Clean up session
    request.session.pop('saas_admin_original_user_id', None)
    request.session.pop('saas_admin_impersonating', None)
    request.session.pop('saas_admin_impersonated_user_id', None)

    # Log the end of impersonation
    logger.info(f"SaaS admin {request.user.email} stopped impersonating user ID {impersonated_user_id}")

    messages.success(request, 'Impersonation session ended.')
    return redirect('saas_admin_dashboard')


@require_saas_admin
def saas_admin_system_settings(request):
    """System-wide settings management for SaaS admins"""
    if request.method == 'POST':
        # Handle system settings updates
        setting_type = request.POST.get('setting_type')

        if setting_type == 'maintenance_mode':
            # Toggle maintenance mode
            maintenance_mode = request.POST.get('maintenance_mode') == 'on'
            # Store in cache or database
            from django.core.cache import cache
            cache.set('system_maintenance_mode', maintenance_mode, timeout=None)

            status = 'enabled' if maintenance_mode else 'disabled'
            messages.success(request, f'Maintenance mode {status}.')

        elif setting_type == 'user_registration':
            # Control user registration
            allow_registration = request.POST.get('allow_registration') == 'on'
            cache.set('allow_user_registration', allow_registration, timeout=None)

            status = 'enabled' if allow_registration else 'disabled'
            messages.success(request, f'User registration {status}.')

        elif setting_type == 'email_settings':
            # Update email settings
            smtp_host = request.POST.get('smtp_host')
            smtp_port = request.POST.get('smtp_port')
            smtp_user = request.POST.get('smtp_user')

            # In a real implementation, you'd update these in settings or database
            messages.success(request, 'Email settings updated.')

        return redirect('saas_admin_system_settings')

    # Get current settings
    from django.core.cache import cache
    context = {
        'maintenance_mode': cache.get('system_maintenance_mode', False),
        'allow_registration': cache.get('allow_user_registration', True),
        'system_stats': {
            'total_companies': Company.objects.count(),
            'total_users': get_visible_users().count(),
            'active_sessions': Session.objects.count(),
        }
    }

    return render(request, 'accounts/saas_admin_system_settings.html', context)



def _user_has_company_access(user, company):
    """Enhanced company access check with SaaS admin support - UPDATED"""
    # SaaS admins have access to all companies
    if getattr(user, 'is_saas_admin', False):
        return True

    # Users with can_access_all_companies flag
    if getattr(user, 'can_access_all_companies', False):
        return True

    # Company owners/admins have access
    if hasattr(user, 'company_admin') and user.company_admin and user.company == company:
        return True

    # Regular users in the company have access
    if hasattr(user, 'company') and user.company == company:
        return True

    return False

def _user_has_management_access(current_user, target_user):
    """Enhanced management access check with SaaS admin support - UPDATED"""
    if getattr(current_user, 'is_saas_admin', False):
        return True

    # Users can't manage themselves
    if current_user.id == target_user.id:
        return False

    # Users can only manage users in the same company
    if current_user.company_id != target_user.company_id:
        return False

    # Check role hierarchy — must be strictly higher priority, not equal
    current_priority = current_user.highest_role_priority
    target_priority = target_user.highest_role_priority

    return current_priority > target_priority

@require_saas_admin
@login_required
def system_admin_dashboard(request):
    """System admin dashboard with global statistics"""
    # Check if user has system admin role (priority >= 90)
    if not ((request.user.primary_role and request.user.primary_role.priority >= 90) or getattr(request.user, 'is_saas_admin', False)):
        raise PermissionDenied

    from company.models import Company

    # Global statistics
    total_companies = Company.objects.count()
    active_companies = Company.objects.filter(is_active=True).count()
    trial_companies = Company.objects.filter(is_trial=True).count() if hasattr(Company, 'is_trial') else 0

    total_users = CustomUser.objects.count()
    active_users = CustomUser.objects.filter(is_active=True).count()

    # Recent activity
    recent_companies = Company.objects.order_by('-created_at')[:10]
    recent_users = CustomUser.objects.order_by('-date_joined')[:10]

    # Plan distribution
    plan_stats = []
    if hasattr(Company, 'plan'):
        plan_stats = list(
            Company.objects.values('plan__name', 'plan__display_name')
            .annotate(count=Count('id'))
            .order_by('-count')
        )

    # Companies expiring soon
    expiring_soon = []
    if hasattr(Company, 'subscription_ends_at'):
        expiring_soon = Company.objects.filter(
            subscription_ends_at__lte=timezone.now().date() + timedelta(days=30),
            subscription_ends_at__gte=timezone.now().date()
        ).order_by('subscription_ends_at')[:10]

    context = {
        'total_companies': total_companies,
        'active_companies': active_companies,
        'trial_companies': trial_companies,
        'expired_companies': 0,  # You'll need to implement this logic
        'total_users': total_users,
        'active_users': active_users,
        'recent_companies': recent_companies,
        'recent_users': recent_users,
        'plan_stats': plan_stats,
        'expiring_soon': expiring_soon,
    }

    return render(request, 'accounts/system_admin_dashboard.html', context)


class UserListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """Enhanced user list view - UPDATED FOR YOUR MODEL STRUCTURE"""
    model = CustomUser
    template_name = 'accounts/user_list.html'
    context_object_name = 'users'
    paginate_by = 25
    permission_required = 'accounts.view_customuser'

    def get_queryset(self):
        user = self.request.user

        # Use the updated accessible users function
        queryset = _get_accessible_users(user)

        # Apply search filters
        search_form = UserSearchForm(self.request.GET)
        if search_form.is_valid():
            search_query = search_form.cleaned_data.get('search_query')
            if search_query:
                queryset = queryset.filter(
                    Q(first_name__icontains=search_query) |
                    Q(last_name__icontains=search_query) |
                    Q(email__icontains=search_query) |
                    Q(username__icontains=search_query)
                )

            # Filter by role
            role_filter = search_form.cleaned_data.get('role')
            if role_filter:
                queryset = queryset.filter(groups__role=role_filter).distinct()

            is_active = search_form.cleaned_data.get('is_active')
            if is_active:
                queryset = queryset.filter(is_active=is_active == 'true')

            email_verified = search_form.cleaned_data.get('email_verified')
            if email_verified:
                queryset = queryset.filter(email_verified=email_verified == 'true')

            date_from = search_form.cleaned_data.get('date_joined_from')
            date_to = search_form.cleaned_data.get('date_joined_to')
            if date_from:
                queryset = queryset.filter(date_joined__date__gte=date_from)
            if date_to:
                queryset = queryset.filter(date_joined__date__lte=date_to)

        return queryset.select_related('company').prefetch_related('groups__role').order_by('-date_joined')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['search_form'] = UserSearchForm(self.request.GET)
        context['bulk_form'] = BulkUserActionForm()

        # User statistics
        queryset = self.get_queryset()
        context['user_stats'] = {
            'total': queryset.count(),
            'active': queryset.filter(is_active=True).count(),
            'inactive': queryset.filter(is_active=False).count(),
            'locked': queryset.filter(locked_until__gt=timezone.now()).count(),
        }

        # Add context for template
        context.update({
            'is_saas_admin': getattr(self.request.user, 'is_saas_admin', False),
            'is_company_admin': getattr(self.request.user, 'company_admin', False),
            'available_roles': Role.objects.accessible_by_user(self.request.user),
        })

        return context

class UserDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Enhanced user detail view - UPDATED FOR YOUR MODEL STRUCTURE"""
    model = CustomUser
    template_name = 'accounts/user_detail.html'
    context_object_name = 'user_profile'
    permission_required = 'accounts.view_customuser'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not self._user_has_access(obj):
            raise Http404("User not found")
        return obj

    def _user_has_access(self, target_user):
        return _user_has_management_access(self.request.user, target_user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_profile = self.object  # Already fetched by DetailView — avoids a second DB query
        current_user = self.request.user

        # Get user's roles
        user_roles = user_profile.all_roles

        can_edit = _user_has_management_access(current_user, user_profile)

        context.update({
            'account_age': (timezone.now() - user_profile.date_joined).days,
            'is_locked': user_profile.is_locked,
            'can_unlock': user_profile.is_locked and current_user.has_perm('accounts.change_customuser'),
            'company': user_profile.company,
            'user_roles': user_roles,
            'can_edit': can_edit,
            'can_manage_users': current_user.has_perm('accounts.add_customuser'),
            'is_saas_admin': getattr(current_user, 'is_saas_admin', False),
            'is_hidden_user': getattr(user_profile, 'is_hidden', False),
            'can_access_all_companies': getattr(user_profile, 'can_access_all_companies', False),
            'available_roles': Role.objects.accessible_by_user(current_user),
        })

        return context

class UserCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Enhanced user creation - UPDATED FOR YOUR MODEL STRUCTURE"""
    model = CustomUser
    form_class = CustomUserCreationForm
    template_name = 'accounts/user_create.html'
    permission_required = 'accounts.add_customuser'
    success_url = reverse_lazy('user_list')

    @method_decorator(check_user_limit)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add available roles to context
        context['available_roles'] = Role.objects.accessible_by_user(self.request.user)
        return context

    def form_valid(self, form):
        current_user = self.request.user
        logger.debug(f"Attempting to create new user by {current_user}")

        try:
            # Set company automatically based on current user's company
            if not form.instance.company:
                form.instance.company = current_user.company

            user = form.save()

            # Handle role assignment if provided
            role_id = self.request.POST.get('role')
            if role_id:
                try:
                    role = Role.objects.get(id=role_id)
                    if role in Role.objects.accessible_by_user(current_user):
                        user.groups.add(role.group)

                        # Log role assignment
                        RoleHistory.objects.create(
                            role=role,
                            action='assigned',
                            user=current_user,
                            affected_user=user,
                            notes=f"Role assigned during user creation"
                        )
                except Role.DoesNotExist:
                    pass

            messages.success(
                self.request,
                f'User {user.get_full_name() or user.email} created successfully!'
            )
            logger.info(f"✅ User {user} created in company {user.company} by {current_user}")

            return super().form_valid(form)

        except ValidationError as e:
            logger.error(f"❌ Validation error during user creation: {str(e)}")
            messages.error(self.request, f"Error creating user: {str(e)}")
            return self.form_invalid(form)
        except Exception as e:
            logger.error(f"❌ Unexpected error during user creation: {str(e)}")
            messages.error(self.request, "An unexpected error occurred while creating the user.")
            return self.form_invalid(form)

class UserUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """User update view with hierarchy checks"""
    model = CustomUser
    form_class = CustomUserChangeForm
    template_name = 'accounts/user_update.html'
    permission_required = 'accounts.change_customuser'

    def get_queryset(self):
        """Only show users this person can manage"""
        return self.request.user.get_manageable_users()

    def get_object(self, queryset=None):
        """Ensure user can manage this specific user"""
        obj = super().get_object(queryset)

        if not self.request.user.can_manage_user(obj):
            raise PermissionDenied("You cannot edit this user.")

        return obj

    def get_form_kwargs(self):
        """Pass request to form"""
        kwargs = super().get_form_kwargs()
        kwargs['request'] = self.request
        return kwargs

    def get_success_url(self):
        return reverse('user_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Capture old roles BEFORE super().form_valid() writes the new state to the DB
        old_roles = set(self.object.all_roles) if 'roles' in form.changed_data else None

        response = super().form_valid(form)

        # Log role changes if roles were modified
        if old_roles is not None:
            new_roles = set(form.cleaned_data.get('roles', []))

            # Log removed roles
            for role in old_roles - new_roles:
                RoleHistory.objects.create(
                    role=role,
                    action='removed',
                    user=self.request.user,
                    affected_user=self.object,
                    notes=f"Role removed during user update"
                )

            # Log added roles
            for role in new_roles - old_roles:
                RoleHistory.objects.create(
                    role=role,
                    action='assigned',
                    user=self.request.user,
                    affected_user=self.object,
                    notes=f"Role assigned during user update"
                )

        messages.success(
            self.request,
            f'User {self.object.get_full_name()} updated successfully!'
        )
        return response

class UserDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = CustomUser
    template_name = 'accounts/user_delete_options.html'
    permission_required = 'accounts.delete_customuser'
    success_url = reverse_lazy('user_list')

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not self._user_has_access(obj):
            raise Http404("User not found")
        return obj

    def _user_has_access(self, target_user):
        current_user = self.request.user
        # Never allow a user to delete themselves via this admin view
        if current_user == target_user:
            return False
        if current_user.primary_role and current_user.primary_role.priority >= 90:
            return True
        return False

    def post(self, request, *args, **kwargs):
        user = self.get_object()
        action = request.POST.get('action')

        if action == 'deactivate':
            user.is_active = False
            user.save()
            messages.success(request, f'User {user.get_full_name()} deactivated successfully!')
        elif action == 'delete':
            user.delete()
            messages.success(request, f'User {user.get_full_name()} deleted successfully!')
        else:
            messages.error(request, 'No action selected.')

        return redirect(self.success_url)


# NOTE: APIToken and UserSession models have been moved to models.py where they belong.
# They are imported at the top of this file via the standard .models import.


@login_required
def deactivate_account(request):
    """
    Deactivate user account (soft delete)
    """
    if request.method == 'POST':
        password = request.POST.get('password')
        reason = request.POST.get('reason', '')
        feedback = request.POST.get('feedback', '')

        # Verify password
        if not request.user.check_password(password):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Incorrect password'
                }, status=400)
            messages.error(request, 'Incorrect password.')
            return render(request, 'accounts/deactivate_account.html')

        # Store deactivation info in metadata
        if not request.user.metadata:
            request.user.metadata = {}

        request.user.metadata['deactivation'] = {
            'date': timezone.now().isoformat(),
            'reason': reason,
            'feedback': feedback,
            'ip_address': get_client_ip(request)
        }

        # Deactivate account
        request.user.is_active = False
        request.user.save()

        # Send confirmation email
        send_mail(
            'Account Deactivated',
            render_to_string('emails/account_deactivated.txt', {
                'user': request.user,
                'reason': reason
            }),
            settings.DEFAULT_FROM_EMAIL,
            [request.user.email],
            html_message=render_to_string('emails/account_deactivated.html', {
                'user': request.user,
                'reason': reason
            }),
            fail_silently=True,
        )

        # Log out user
        logout(request)

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Account deactivated successfully',
                'redirect_url': '/'
            })

        messages.success(request, 'Your account has been deactivated successfully.')
        return redirect('/')

    return render(request, 'accounts/deactivate_account.html')


@login_required
def download_user_data(request):
    """
    Download user data in JSON format
    """
    # Prepare user data
    user_data = {
        'profile': {
            'username': request.user.username,
            'email': request.user.email,
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'middle_name': request.user.middle_name,
            'phone_number': request.user.phone_number,
            'bio': request.user.bio,
            'date_joined': request.user.date_joined.isoformat(),
            'last_login': request.user.last_login.isoformat() if request.user.last_login else None,
            'timezone': request.user.timezone,
            'language': request.user.language,
            'primary_role': request.user.display_role,  # ✅ FIXED
            'all_roles': request.user.role_names,  # ✅ FIXED
            'role_priority': request.user.highest_role_priority,  # ✅ ADDED
        },
        'preferences': request.user.metadata.get('preferences', {}),
        'notifications': request.user.metadata.get('notifications', {}),
        'security': {
            'email_verified': request.user.email_verified,
            'phone_verified': request.user.phone_verified,
            'two_factor_enabled': request.user.two_factor_enabled,
            'login_count': request.user.login_count,
        },
        'activity': {
            'last_activity_at': request.user.last_activity_at.isoformat() if request.user.last_activity_at else None,
            'last_login_ip': request.user.last_login_ip,
        }
    }

    # Create JSON response
    response = JsonResponse(user_data, json_dumps_params={'indent': 2})
    response['Content-Disposition'] = f'attachment; filename="user_data_{request.user.username}_{datetime.now().strftime("%Y%m%d")}.json"'
    return response


@login_required
def privacy_settings(request):
    """
    Handle user privacy settings
    """
    if request.method == 'POST':
        # Get privacy preferences
        privacy_data = {
            'profile_visibility': request.POST.get('profile_visibility', 'private'),
            'show_email': request.POST.get('show_email') == 'on',
            'show_phone': request.POST.get('show_phone') == 'on',
            'show_last_login': request.POST.get('show_last_login') == 'on',
            'allow_search': request.POST.get('allow_search') == 'on',
            'data_processing': request.POST.get('data_processing') == 'on',
            'marketing_consent': request.POST.get('marketing_consent') == 'on',
            'analytics_consent': request.POST.get('analytics_consent') == 'on',
        }

        # Save to user metadata
        if not request.user.metadata:
            request.user.metadata = {}
        request.user.metadata['privacy'] = privacy_data
        request.user.save()

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Privacy settings updated successfully!'
            })

        messages.success(request, 'Privacy settings updated successfully!')
        return redirect('privacy_settings')

    # Get current privacy settings
    privacy_settings = request.user.metadata.get('privacy', {
        'profile_visibility': 'private',
        'show_email': False,
        'show_phone': False,
        'show_last_login': False,
        'allow_search': True,
        'data_processing': True,
        'marketing_consent': False,
        'analytics_consent': True,
    })

    context = {
        'user': request.user,
        'privacy_settings': privacy_settings,
    }

    return render(request, 'accounts/privacy_settings.html', context)


def get_role_statistics_for_company(company, users_queryset=None):
    from django.db.models import Count, Q

    if users_queryset is None:
        users_queryset = CustomUser.objects.filter(
            company=company,
            is_hidden=False,
            is_active=True
        )

    role_stats = Role.objects.filter(
        Q(company=company) | Q(is_system_role=True)
    ).annotate(
        user_count=Count(
            'group__user',
            filter=Q(
                group__user__in=users_queryset,
                group__user__is_hidden=False
            ),
            distinct=True
        )
    ).filter(user_count__gt=0).values(
        'id',
        'group__name',
        'user_count',
        'priority',
        'color_code',
        'is_system_role'
    ).order_by('-priority')

    return list(role_stats)


def get_user_type_display_from_role(user):
    if not user.primary_role:
        return "No Role Assigned"

    # Direct return of role name (cleaner than mapping)
    return user.primary_role.group.name

@login_required
def export_all_data(request):
    """
    Export all user data in a comprehensive ZIP file
    """
    if request.method == 'POST':
        # Create a ZIP file in memory
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            # 1. Profile data as JSON - ✅ FIXED
            profile_data = {
                'username': request.user.username,
                'email': request.user.email,
                'first_name': request.user.first_name,
                'last_name': request.user.last_name,
                'middle_name': request.user.middle_name,
                'phone_number': request.user.phone_number,
                'bio': request.user.bio,
                'date_joined': request.user.date_joined.isoformat(),
                'last_login': request.user.last_login.isoformat() if request.user.last_login else None,
                'timezone': request.user.timezone,
                'language': request.user.language,
                'primary_role': request.user.display_role,  # ✅ FIXED
                'all_roles': request.user.role_names,  # ✅ FIXED
                'role_priority': request.user.highest_role_priority,  # ✅ ADDED
                'metadata': request.user.metadata,
            }
            zip_file.writestr('profile.json', json.dumps(profile_data, indent=2))

            # 2. API tokens as CSV
            api_tokens = APIToken.objects.filter(user=request.user)
            if api_tokens.exists():
                csv_buffer = StringIO()
                csv_writer = csv.writer(csv_buffer)
                csv_writer.writerow(['Name', 'Created', 'Last Used', 'Expires', 'Status'])

                for token in api_tokens:
                    csv_writer.writerow([
                        token.name,
                        token.created_at.isoformat(),
                        token.last_used.isoformat() if token.last_used else 'Never',
                        token.expires_at.isoformat() if token.expires_at else 'Never',
                        'Active' if token.is_active else 'Inactive'
                    ])

                zip_file.writestr('api_tokens.csv', csv_buffer.getvalue())

            # 3. Sessions data
            sessions = UserSession.objects.filter(user=request.user)
            if sessions.exists():
                sessions_data = []
                for session in sessions:
                    sessions_data.append({
                        'ip_address': session.ip_address,
                        'user_agent': session.user_agent,
                        'location': session.location,
                        'created_at': session.created_at.isoformat(),
                        'last_activity': session.last_activity.isoformat(),
                        'is_current': session.is_current,
                    })

                zip_file.writestr('sessions.json', json.dumps(sessions_data, indent=2))

            # 4. Avatar image if exists
            if request.user.avatar:
                try:
                    avatar_content = request.user.avatar.read()
                    avatar_name = f"avatar{request.user.avatar.name[request.user.avatar.name.rfind('.'):]}"
                    zip_file.writestr(avatar_name, avatar_content)
                except Exception:
                    pass  # Skip if avatar file is not accessible

            # 5. Export summary
            summary = {
                'export_date': timezone.now().isoformat(),
                'user_id': request.user.id,
                'username': request.user.username,
                'email': request.user.email,
                'files_included': [
                    'profile.json - Complete profile information',
                    'api_tokens.csv - API tokens history' if api_tokens.exists() else None,
                    'sessions.json - Login sessions data' if sessions.exists() else None,
                    'avatar file - Profile picture' if request.user.avatar else None,
                ],
                'notes': 'This export contains all your personal data stored in our system.'
            }
            # Remove None values
            summary['files_included'] = [f for f in summary['files_included'] if f is not None]

            zip_file.writestr('README.json', json.dumps(summary, indent=2))

        # Prepare response
        zip_buffer.seek(0)
        response = HttpResponse(
            zip_buffer.getvalue(),
            content_type='application/zip'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="user_data_export_{request.user.username}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip"'

        return response

    context = {
        'user': request.user,
        'api_tokens_count': APIToken.objects.filter(user=request.user).count(),
        'sessions_count': UserSession.objects.filter(user=request.user).count(),
    }

    return render(request, 'accounts/export_all_data.html', context)


@login_required
def delete_account_request(request):
    """
    Request account deletion (GDPR compliance)
    """
    if request.method == 'POST':
        password = request.POST.get('password')
        reason = request.POST.get('reason', '')
        feedback = request.POST.get('feedback', '')
        confirm_deletion = request.POST.get('confirm_deletion') == 'on'

        # Verify password
        if not request.user.check_password(password):
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': 'Incorrect password'
                }, status=400)
            messages.error(request, 'Incorrect password.')
            return render(request, 'accounts/delete_account_request.html')

        if not confirm_deletion:
            messages.error(request, 'Please confirm that you want to delete your account.')
            return render(request, 'accounts/delete_account_request.html')

        # Store deletion request in metadata
        if not request.user.metadata:
            request.user.metadata = {}

        deletion_token = secrets.token_urlsafe(32)
        request.user.metadata['deletion_request'] = {
            'date': timezone.now().isoformat(),
            'reason': reason,
            'feedback': feedback,
            'token': deletion_token,
            'ip_address': get_client_ip(request),
            'scheduled_deletion': (timezone.now() + timedelta(days=30)).isoformat(),  # 30-day grace period
        }
        request.user.save()

        # Send confirmation email with cancellation link
        cancellation_link = request.build_absolute_uri(
            f"/accounts/profile/cancel-deletion/?token={deletion_token}"
        )

        send_mail(
            'Account Deletion Request Received',
            render_to_string('emails/deletion_request.txt', {
                'user': request.user,
                'cancellation_link': cancellation_link,
                'deletion_date': (timezone.now() + timedelta(days=30)).strftime('%B %d, %Y')
            }),
            settings.DEFAULT_FROM_EMAIL,
            [request.user.email],
            html_message=render_to_string('emails/deletion_request.html', {
                'user': request.user,
                'cancellation_link': cancellation_link,
                'deletion_date': (timezone.now() + timedelta(days=30)).strftime('%B %d, %Y')
            }),
            fail_silently=False,
        )

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': 'Deletion request submitted. Check your email for confirmation.'
            })

        messages.success(request,
                         'Your account deletion request has been submitted. You will receive an email with further instructions.')
        return redirect('user_profile')

    return render(request, 'accounts/delete_account_request.html')


@login_required
def active_sessions(request):
    """
    Display and manage active user sessions
    """
    # Get all active sessions for the user
    user_sessions = UserSession.objects.filter(
        user=request.user,
        last_activity__gte=timezone.now() - timedelta(days=30)
    ).order_by('-last_activity')

    # Get current session
    current_session_key = request.session.session_key

    # Mark current session
    for session in user_sessions:
        session.is_current_session = session.session_key == current_session_key

    context = {
        'user_sessions': user_sessions,
        'current_session_key': current_session_key,
        'user': request.user,
    }

    return render(request, 'accounts/user_sessions.html', context)


@login_required
def api_tokens(request):
    """
    Manage user API tokens
    """
    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'create':
            name = request.POST.get('name', '').strip()
            permissions = request.POST.getlist('permissions')
            expires_in_days = request.POST.get('expires_in_days')

            if not name:
                messages.error(request, 'Token name is required.')
                return redirect('api_tokens')

            # Generate token
            token = secrets.token_urlsafe(32)

            # Set expiration
            expires_at = None
            if expires_in_days and expires_in_days != 'never':
                expires_at = timezone.now() + timedelta(days=int(expires_in_days))

            # Create token
            api_token = APIToken.objects.create(
                user=request.user,
                name=name,
                token=token,
                permissions=permissions,
                expires_at=expires_at
            )

            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'API token created successfully!',
                    'token': {
                        'id': api_token.id,
                        'name': api_token.name,
                        'token': api_token.token,  # Only show once
                        'created_at': api_token.created_at.isoformat()
                    }
                })

            messages.success(request, f'API token "{name}" created successfully!')
            return redirect('api_tokens')

        elif action == 'delete':
            token_id = request.POST.get('token_id')
            try:
                api_token = APIToken.objects.get(id=token_id, user=request.user)
                token_name = api_token.name
                api_token.delete()

                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'API token "{token_name}" deleted successfully!'
                    })

                messages.success(request, f'API token "{token_name}" deleted successfully!')
            except APIToken.DoesNotExist:
                messages.error(request, 'API token not found.')

            return redirect('api_tokens')

    # Get user's API tokens
    user_tokens = APIToken.objects.filter(user=request.user).order_by('-created_at')

    # Available permissions (customize based on your needs)
    available_permissions = [
        ('read', 'Read Access'),
        ('write', 'Write Access'),
        ('delete', 'Delete Access'),
        ('admin', 'Admin Access'),
    ]

    context = {
        'user_tokens': user_tokens,
        'available_permissions': available_permissions,
        'user': request.user,
    }

    return render(request, 'accounts/api_tokens.html', context)


@login_required
def user_integrations(request):
    """
    Manage user integrations with external services
    """
    # Get user's integrations from metadata
    integrations = request.user.metadata.get('integrations', {})

    available_integrations = {
        'google': {
            'name': 'Google',
            'description': 'Connect with Google services',
            'icon': 'bi-google',
            'status': integrations.get('google', {}).get('status', 'disconnected')
        },
        'microsoft': {
            'name': 'Microsoft',
            'description': 'Connect with Microsoft Office 365',
            'icon': 'bi-microsoft',
            'status': integrations.get('microsoft', {}).get('status', 'disconnected')
        },
        'slack': {
            'name': 'Slack',
            'description': 'Receive notifications in Slack',
            'icon': 'bi-slack',
            'status': integrations.get('slack', {}).get('status', 'disconnected')
        },
        'zapier': {
            'name': 'Zapier',
            'description': 'Automate workflows with Zapier',
            'icon': 'bi-lightning',
            'status': integrations.get('zapier', {}).get('status', 'disconnected')
        }
    }

    if request.method == 'POST':
        action = request.POST.get('action')
        integration = request.POST.get('integration')

        if action == 'connect':
            # In a real implementation, you would redirect to OAuth flow
            # For now, we'll simulate connection
            if not request.user.metadata:
                request.user.metadata = {}
            if 'integrations' not in request.user.metadata:
                request.user.metadata['integrations'] = {}

            request.user.metadata['integrations'][integration] = {
                'status': 'connected',
                'connected_at': timezone.now().isoformat(),
                'access_token': 'simulated_token_' + secrets.token_urlsafe(16)
            }
            request.user.save()

            messages.success(request, f'{available_integrations[integration]["name"]} connected successfully!')

        elif action == 'disconnect':
            if request.user.metadata and 'integrations' in request.user.metadata:
                if integration in request.user.metadata['integrations']:
                    del request.user.metadata['integrations'][integration]
                    request.user.save()

                    messages.success(request,
                                     f'{available_integrations[integration]["name"]} disconnected successfully!')

        return redirect('user_integrations')

    # Update statuses
    for key, integration in available_integrations.items():
        integration['status'] = integrations.get(key, {}).get('status', 'disconnected')
        if integration['status'] == 'connected':
            integration['connected_at'] = integrations.get(key, {}).get('connected_at')

    context = {
        'available_integrations': available_integrations,
        'user': request.user,
    }

    return render(request, 'accounts/user_integrations.html', context)


@login_required
@permission_required('accounts.add_customuser', raise_exception=True)
def bulk_invite_users(request, company_id):
    """Bulk invite users via CSV upload"""
    company = get_object_or_404(Company, company_id=company_id)

    if not _user_has_company_access(request.user, company):
        raise PermissionDenied("You don't have access to this company.")

    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']

        # Validate file type
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a CSV file.')
            return redirect('bulk_invite_users', company_id=company_id)

        try:
            import csv
            from io import TextIOWrapper

            # Process CSV file
            csv_file_wrapper = TextIOWrapper(csv_file.file, encoding='utf-8')
            reader = csv.DictReader(csv_file_wrapper)

            required_columns = ['email']
            optional_columns = ['first_name', 'last_name', 'phone_number', 'role_id']

            # Validate CSV structure
            if not all(col in reader.fieldnames for col in required_columns):
                messages.error(request, f'CSV must contain columns: {", ".join(required_columns)}')
                return redirect('bulk_invite_users', company_id=company_id)

            success_count = 0
            error_count = 0
            errors = []

            for row_num, row in enumerate(reader, start=2):  # start=2 to account for header
                try:
                    email = row['email'].strip()

                    if not email:
                        errors.append(f"Row {row_num}: Email is required")
                        error_count += 1
                        continue

                    # Check if user exists
                    existing_user = CustomUser.objects.filter(email=email).first()
                    if existing_user:
                        if existing_user.company == company:
                            errors.append(f"Row {row_num}: User {email} already exists in company")
                            error_count += 1
                            continue
                        else:
                            existing_user.company = company
                            existing_user.save()
                    else:
                        # Create account with unusable password — user sets it via setup link
                        base_username = email.split('@')[0]
                        username = base_username
                        counter = 1
                        while CustomUser.objects.filter(username=username).exists():
                            username = f"{base_username}{counter}"
                            counter += 1

                        new_user = CustomUser.objects.create_user(
                            email=email,
                            username=username,
                            password=None,
                            company=company,
                            first_name=row.get('first_name', ''),
                            last_name=row.get('last_name', ''),
                            phone_number=row.get('phone_number', ''),
                            is_active=True,
                        )
                        new_user.set_unusable_password()

                        # Generate invitation token (72-hour expiry)
                        token = _generate_invitation_token()
                        expiry = timezone.now() + timedelta(hours=72)
                        if not new_user.metadata:
                            new_user.metadata = {}
                        new_user.metadata['invitation_token'] = token
                        new_user.metadata['invitation_expires'] = expiry.isoformat()
                        new_user.save()

                        setup_url = _build_setup_url(request, new_user, token)
                        first_name = row.get('first_name', '')

                        plain_message = (
                            f"Hello {first_name or 'there'},\n\n"
                            f"You've been invited to join {company.name}.\n\n"
                            f"Click the link below to set your password:\n{setup_url}\n\n"
                            f"This link expires in 72 hours.\n\n"
                            f"Best regards,\nThe {company.name} Team"
                        )

                        html_message = f"""<!DOCTYPE html>
<html><body style="font-family:Arial,sans-serif;background:#f6f9fc;margin:0;padding:0;">
  <div style="max-width:600px;margin:0 auto;background:#fff;padding:40px 30px;">
    <h2 style="color:#667eea;">You've been invited to {company.name}</h2>
    <p>Hello <strong>{first_name or 'there'}</strong>,</p>
    <p>Click below to set your password and activate your account:</p>
    <div style="text-align:center;margin:30px 0;">
      <a href="{setup_url}"
         style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);color:#fff;
                padding:14px 30px;text-decoration:none;border-radius:5px;font-weight:600;">
        Set My Password
      </a>
    </div>
    <p style="color:#718096;font-size:13px;">Link expires in 72 hours.</p>
    <p style="color:#a0aec0;font-size:12px;">
      If you weren't expecting this, you can safely ignore this email.
    </p>
  </div>
</body></html>"""

                        try:
                            send_tenant_email(
                                subject=f"You've been invited to {company.name} — set up your account",
                                message=plain_message,
                                recipient_list=[email],
                                html_message=html_message,
                                tenant=getattr(connection, 'tenant', None),
                            )
                        except Exception as email_err:
                            logger.error(f"Failed to send invitation email to {email}: {email_err}")

                    # Assign role if specified
                    role_id = row.get('role_id')
                    if role_id:
                        try:
                            role = Role.objects.get(id=role_id)
                            if role in Role.objects.accessible_by_user(request.user):
                                target = existing_user if existing_user else new_user
                                target.groups.add(role.group)
                        except (Role.DoesNotExist, ValueError):
                            pass

                    success_count += 1

                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")
                    error_count += 1

            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully processed {success_count} users.')
            if error_count > 0:
                messages.warning(request, f'Failed to process {error_count} users. Check the errors below.')
                for error in errors[:10]:  # Show first 10 errors
                    messages.error(request, error)
                if len(errors) > 10:
                    messages.info(request, f'... and {len(errors) - 10} more errors.')

            return redirect('user_list')

        except Exception as e:
            messages.error(request, f'Error processing CSV file: {str(e)}')
            return redirect('bulk_invite_users', company_id=company_id)

    context = {
        'company': company,
        'available_roles': Role.objects.accessible_by_user(request.user),
    }

    return render(request, 'accounts/bulk_invite_users.html', context)

@login_required
@require_POST
def revoke_session(request, session_id):
    """
    Revoke a specific user session
    """
    try:
        user_session = UserSession.objects.get(id=session_id, user=request.user)

        # Don't allow revoking current session
        if user_session.session_key == request.session.session_key:
            return JsonResponse({
                'success': False,
                'error': 'Cannot revoke current session'
            }, status=400)

        # Delete the Django session
        try:
            Session.objects.get(session_key=user_session.session_key).delete()
        except Session.DoesNotExist:
            pass  # Session already expired or deleted

        # Delete our session record
        session_info = f"{user_session.ip_address} - {user_session.user_agent[:50]}"
        user_session.delete()

        return JsonResponse({
            'success': True,
            'message': f'Session revoked successfully: {session_info}'
        })

    except UserSession.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Session not found'
        }, status=404)


# Utility functions (get_client_ip is imported from .utils above)

@login_required
@permission_required('accounts.change_customuser')
def unlock_user(request, pk):
    """Unlock a locked user account with access control"""
    user = get_object_or_404(CustomUser, pk=pk)

    # Check access
    if not _user_has_management_access(request.user, user):
        raise PermissionDenied

    if user.is_locked:
        user.unlock_account()
        messages.success(request, f'User {user.get_full_name()} has been unlocked successfully!')
    else:
        messages.info(request, f'User {user.get_full_name()} is not locked.')

    return redirect('user_detail', pk=pk)


@login_required
@permission_required('accounts.add_customuser')
@require_http_methods(["POST"])
def bulk_user_actions(request):
    """Enhanced bulk user actions with SaaS admin support"""
    form = BulkUserActionForm(request.POST)

    if form.is_valid():
        action = form.cleaned_data['action']
        user_ids = form.cleaned_data['selected_users']

        # Filter users based on access
        accessible_users = _get_accessible_users(request.user).filter(id__in=user_ids)

        # Don't allow actions on hidden SaaS admin users unless user is SaaS admin
        if not getattr(request.user, 'is_saas_admin', False):
            accessible_users = accessible_users.filter(is_hidden=False)

        count = accessible_users.count()

        if count == 0:
            messages.error(request, 'No accessible users selected.')
            return redirect('user_list')

        if action == 'activate':
            accessible_users.update(is_active=True)
            messages.success(request, f'{count} users activated successfully!')

        elif action == 'deactivate':
            # Don't deactivate the current user or other SaaS admins
            deactivate_qs = accessible_users.exclude(id=request.user.id)
            if not getattr(request.user, 'is_saas_admin', False):
                deactivate_qs = deactivate_qs.exclude(is_saas_admin=True)

            deactivated_count = deactivate_qs.update(is_active=False)
            messages.success(request, f'{deactivated_count} users deactivated successfully!')

        elif action == 'delete':
            # Soft delete - deactivate instead, with same restrictions
            delete_qs = accessible_users.exclude(id=request.user.id)
            if not getattr(request.user, 'is_saas_admin', False):
                delete_qs = delete_qs.exclude(is_saas_admin=True)

            deleted_count = delete_qs.update(is_active=False)
            messages.success(request, f'{deleted_count} users deactivated successfully!')

        elif action == 'export':
            return export_users(request, list(accessible_users.values_list('id', flat=True)))

    return redirect('user_list')


@login_required
@permission_required('accounts.can_export_data')
def export_users(request, user_ids=None):
    """Enhanced export users with role information"""
    if user_ids:
        if isinstance(user_ids, str):
            user_ids_list = [int(uid) for uid in user_ids.split(',')]
        else:
            user_ids_list = user_ids
        users = _get_accessible_users(request.user).filter(id__in=user_ids_list)
    else:
        users = _get_accessible_users(request.user)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="users_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    )

    writer = csv.writer(response)

    # ✅ FIXED: Replace 'User Type' with 'Primary Role' and 'All Roles'
    writer.writerow([
        'ID', 'Email', 'Username', 'First Name', 'Last Name',
        'Primary Role', 'All Roles', 'Role Priority',
        'Is Active', 'Email Verified', 'Phone Number',
        'Date Joined', 'Last Login',
        'Company', 'Is Company Admin', 'Is Hidden', 'Is SaaS Admin'
    ])

    for user in users.select_related('company').prefetch_related('groups__role'):
        company = user.company
        is_admin = getattr(user, 'company_admin', False)

        # Get role information
        primary_role = user.primary_role
        primary_role_name = primary_role.group.name if primary_role else 'No Role'
        all_roles = ', '.join(user.role_names) if user.role_names else 'No Roles'
        role_priority = primary_role.priority if primary_role else 0

        writer.writerow([
            user.id,
            user.email,
            user.username,
            user.first_name,
            user.last_name,
            primary_role_name,
            all_roles,
            role_priority,
            user.is_active,
            getattr(user, 'email_verified', False),
            user.phone_number or '',
            timezone.localtime(user.date_joined).strftime('%Y-%m-%d %H:%M:%S'),
            timezone.localtime(user.last_login).strftime('%Y-%m-%d %H:%M:%S') if user.last_login else '',
            company.name if company else '',
            is_admin,
            getattr(user, 'is_hidden', False),
            getattr(user, 'is_saas_admin', False)
        ])

    return response


@login_required
def check_username_availability(request):
    """AJAX endpoint to check username availability"""
    username = request.GET.get('username', '')
    user_id = request.GET.get('user_id', None)

    if not username:
        return JsonResponse({'available': False, 'message': 'Username is required'})

    query = CustomUser.objects.filter(username=username)
    if user_id:
        query = query.exclude(id=user_id)

    available = not query.exists()
    message = 'Username is available' if available else 'Username is already taken'

    return JsonResponse({'available': available, 'message': message})


@login_required
def check_email_availability(request):
    """AJAX endpoint to check email availability"""
    email = request.GET.get('email', '')
    user_id = request.GET.get('user_id', None)

    if not email:
        return JsonResponse({'available': False, 'message': 'Email is required'})

    query = CustomUser.objects.filter(email=email)
    if user_id:
        query = query.exclude(id=user_id)

    available = not query.exists()
    message = 'Email is available' if available else 'Email is already registered'

    return JsonResponse({'available': available, 'message': message})


# Company switching for multi-tenant users
@require_saas_admin
@login_required
def switch_company(request, company_id):
    """Switch active company context"""
    company = get_object_or_404(Company, id=company_id)

    # Check if user has access to this company
    if not (hasattr(company, 'owner') and company.owner == request.user or
            company in request.user.companies.filter(is_active=True)):
        raise PermissionDenied

    # Store active company in session
    request.session['active_company_id'] = company_id
    messages.success(request, f'Switched to {company.display_name}')

    return redirect('companies:dashboard')


@login_required
@permission_required('accounts.add_customuser', raise_exception=True)
def invite_user(request):
    """
    Invite a new user to the current tenant's company.
    Creates the account in an unusable-password state and emails a
    one-time setup link. The user sets their own password on first visit.
    """
    tenant = getattr(connection, 'tenant', None)
    if not tenant or tenant.schema_name == 'public':
        raise PermissionDenied("You must be in a tenant schema to invite users.")

    # Always resolve company from the logged-in user — this guarantees the
    # correct company is used in the invitation email regardless of DB order.
    # getattr(tenant, 'company') can return None or the wrong company when
    # multiple companies exist and the tenant relationship isn't set.
    company = request.user.company
    if not company:
        # Fallback: try the tenant relationship, but never use .first()
        company = getattr(tenant, 'company', None)
    if not company:
        raise PermissionDenied("No company associated with this tenant.")

    if hasattr(company, 'can_add_employee') and not company.can_add_employee():
        messages.error(request, 'Company has reached the maximum user limit.')
        return redirect('company_user_list')

    with tenant_context(tenant):
        available_roles = Role.objects.accessible_by_user(request.user)

    if request.method == 'POST':
        email = request.POST.get('email', '').strip().lower()
        role_id = request.POST.get('role')
        is_admin = request.POST.get('is_admin') == 'on'
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone_number = request.POST.get('phone_number', '').strip()

        if not email:
            messages.error(request, 'Email is required.')
            return redirect('invite_user')

        with tenant_context(tenant):
            existing_user = CustomUser.objects.filter(email=email).first()

            if existing_user:
                if existing_user.company == company:
                    messages.error(request, 'User already belongs to this company.')
                else:
                    existing_user.company = company
                    existing_user.first_name = first_name or existing_user.first_name
                    existing_user.last_name = last_name or existing_user.last_name
                    existing_user.phone_number = phone_number or existing_user.phone_number
                    existing_user.company_admin = is_admin
                    if role_id:
                        try:
                            role = Role.objects.get(id=role_id)
                            if role in available_roles:
                                existing_user.groups.add(role.group)
                        except Role.DoesNotExist:
                            pass
                    existing_user.save()
                    messages.success(request, f"User {existing_user.email} added to company.")
                return redirect('company_user_list')

            # Create account with unusable password — user sets it via the setup link
            try:
                with transaction.atomic():
                    base_username = email.split('@')[0]
                    username = base_username
                    counter = 1
                    while CustomUser.objects.filter(username=username).exists():
                        username = f"{base_username}{counter}"
                        counter += 1

                    new_user = CustomUser.objects.create_user(
                        email=email,
                        username=username,
                        password=None,          # unusable password until setup link used
                        company=company,
                        first_name=first_name,
                        last_name=last_name,
                        phone_number=phone_number,
                        company_admin=is_admin,
                        is_active=True,
                    )
                    new_user.set_unusable_password()
                    new_user.save(update_fields=['password'])

                    if role_id:
                        try:
                            role = Role.objects.get(id=role_id)
                            if role in available_roles:
                                new_user.groups.add(role.group)
                                RoleHistory.objects.create(
                                    role=role,
                                    action='assigned',
                                    user=request.user,
                                    affected_user=new_user,
                                    notes='Role assigned during invitation',
                                )
                        except Role.DoesNotExist:
                            pass

                    # Generate a one-time setup token (expires in 72 h)
                    token = _generate_invitation_token()
                    expiry = timezone.now() + timedelta(hours=72)
                    if not new_user.metadata:
                        new_user.metadata = {}
                    new_user.metadata['invitation_token'] = token
                    new_user.metadata['invitation_expires'] = expiry.isoformat()
                    new_user.save(update_fields=['metadata'])

                    setup_url = _build_setup_url(request, new_user, token)

                    # Build login URL for the email footer
                    scheme = 'https' if request.is_secure() else 'http'
                    # Reuse _build_setup_url's host logic so port is included on local dev
                    _dummy = _build_setup_url(request, new_user, 'x')
                    _base = _dummy.rsplit('/invite/', 1)[0]
                    login_url = f"{_base}{reverse('login')}"

                    subject = f"You've been invited to {company.name} — set up your account"

                    plain_message = (
                        f"Hello {first_name or 'there'},\n\n"
                        f"You've been invited to join {company.name}.\n\n"
                        f"Click the link below to set your password and activate your account:\n"
                        f"{setup_url}\n\n"
                        f"This link expires in 72 hours.\n\n"
                        f"If you weren't expecting this invitation, you can safely ignore this email.\n\n"
                        f"Best regards,\nThe {company.name} Team"
                    )

                    html_message = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1.0"></head>
<body style="margin:0;padding:0;font-family:'Segoe UI',Arial,sans-serif;background:#f6f9fc;color:#333;">
  <div style="max-width:600px;margin:0 auto;background:#fff;">
    <div style="background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);padding:40px 20px;text-align:center;color:#fff;">
      <h1 style="margin:0;font-size:26px;font-weight:300;">Welcome to {company.name}</h1>
      <p style="margin:10px 0 0;font-size:15px;opacity:.9;">Your account is ready — set your password to get started</p>
    </div>
    <div style="padding:40px 30px;">
      <p>Hello <strong>{first_name or 'there'}</strong>,</p>
      <p>You've been invited to join <strong style="color:#667eea;">{company.name}</strong>.</p>
      <p>Click the button below to set your password and activate your account:</p>
      <div style="text-align:center;margin:35px 0;">
        <a href="{setup_url}"
           style="display:inline-block;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);
                  color:#fff;padding:14px 35px;text-decoration:none;border-radius:5px;
                  font-weight:600;font-size:16px;">
          Set My Password &amp; Activate Account
        </a>
      </div>
      <p style="color:#718096;font-size:13px;text-align:center;">
        Or copy this link into your browser:<br>
        <code style="background:#f7fafc;padding:6px 10px;border-radius:3px;font-size:11px;word-break:break-all;">{setup_url}</code>
      </p>
      <div style="background:#fffbeb;border:1px solid #f6e05e;padding:14px;border-radius:6px;margin:25px 0;">
        <p style="margin:0;font-size:13px;color:#744210;">
          ⏰ This link expires in <strong>72 hours</strong>. After that, ask your administrator to resend the invitation.
        </p>
      </div>
      <p style="color:#718096;font-size:13px;">
        If you weren't expecting this invitation you can safely ignore this email.
      </p>
    </div>
    <div style="background:#f8f9fa;padding:20px 30px;text-align:center;border-top:1px solid #e2e8f0;">
      <p style="margin:0;color:#a0aec0;font-size:12px;">
        Best regards, <strong>The {company.name} Team</strong>
      </p>
    </div>
  </div>
</body>
</html>"""

                    send_tenant_email(
                        subject=subject,
                        message=plain_message,
                        recipient_list=[email],
                        html_message=html_message,
                        tenant=tenant,
                    )

                    log_action = AuditLog.log if hasattr(AuditLog, 'log') else None
                    if log_action:
                        try:
                            log_action(
                                action='user_invited',
                                user=request.user,
                                description=f"Invitation sent to {email}",
                            )
                        except Exception:
                            pass

                    messages.success(request, f"Invitation sent to {email}. They'll receive a setup link valid for 72 hours.")
                    return redirect('company_user_list')

            except Exception as e:
                logger.error(f"Error creating invited user {email}: {e}")
                messages.error(request, f"Error sending invitation: {e}")
                return redirect('invite_user')

    context = {
        'company': company,
        'available_roles': available_roles,
    }
    return render(request, 'accounts/invite_user.html', context)


def accept_invitation(request, token):
    """
    Password-setup view reached via the one-time link emailed to new invitees.
    On GET: show the set-password form.
    On POST: validate, set password, mark account active, log user in.
    """
    # Find the user whose invitation token matches
    try:
        user = CustomUser.objects.get(
            metadata__invitation_token=token,
            is_active=True,
        )
    except CustomUser.DoesNotExist:
        messages.error(request, 'This invitation link is invalid or has already been used.')
        return redirect('login')

    # Check expiry
    expiry_str = (user.metadata or {}).get('invitation_expires')
    if expiry_str:
        expiry = datetime.fromisoformat(expiry_str)
        if timezone.is_naive(expiry):
            expiry = timezone.make_aware(expiry)
        if timezone.now() > expiry:
            messages.error(request, 'This invitation link has expired. Please ask your administrator to resend it.')
            return redirect('login')

    if request.method == 'POST':
        password1 = request.POST.get('password1', '')
        password2 = request.POST.get('password2', '')

        if not password1:
            messages.error(request, 'Please enter a password.')
        elif password1 != password2:
            messages.error(request, 'Passwords do not match.')
        elif len(password1) < 8:
            messages.error(request, 'Password must be at least 8 characters.')
        else:
            user.set_password(password1)
            # Clear invitation token so the link can't be reused
            user.metadata.pop('invitation_token', None)
            user.metadata.pop('invitation_expires', None)
            user.save()

            # Log the user in automatically
            backend = 'django.contrib.auth.backends.ModelBackend'
            login(request, user, backend=backend)
            messages.success(request, f'Welcome to {user.company.name if user.company else "the platform"}! Your account is now active.')
            return redirect(get_dashboard_url(user))

    return render(request, 'accounts/accept_invitation.html', {'invited_user': user})


@login_required
def user_security_settings(request):
    """User security settings page"""
    context = {
        'user': request.user,
        'two_factor_enabled': request.user.two_factor_enabled,
        'backup_codes_count': len(request.user.backup_codes),
        'password_age': (timezone.now() - request.user.password_changed_at).days,
    }

    return render(request, 'accounts/user_security_settings.html', context)


@login_required
@require_saas_admin
def system_companies_list(request):
    """System admin view of all companies"""
    from company.models import Company,SubscriptionPlan

    # Check if user is allowed to access system-level companies
    if not getattr(request.user, "is_saas_admin", False) and not request.user.is_superuser:
        raise PermissionDenied("You are not allowed to view system companies.")

    # Base queryset
    companies = Company.objects.select_related("plan").order_by("-created_at")

    # Apply filters
    status_filter = request.GET.get("status")
    if status_filter and hasattr(Company, "status"):
        companies = companies.filter(status=status_filter)

    plan_filter = request.GET.get("plan")
    if plan_filter and hasattr(Company, "plan"):
        companies = companies.filter(plan__name=plan_filter)

    search_query = request.GET.get("search")
    if search_query:
        filter_q = Q(name__icontains=search_query) | Q(email__icontains=search_query)
        if hasattr(Company, "trading_name"):
            filter_q |= Q(trading_name__icontains=search_query)
        if hasattr(Company, "company_id"):
            filter_q |= Q(company_id__icontains=search_query)
        companies = companies.filter(filter_q)

    # Pagination
    paginator = Paginator(companies, 25)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "companies": page_obj,
        "status_choices": getattr(Company, "STATUS_CHOICES", []),
        "subscription_plans": SubscriptionPlan.objects.filter(is_active=True),
        "current_filters": {
            "status": status_filter,
            "plan": plan_filter,
            "search": search_query,
        },
    }

    return render(request, "accounts/system_companies_list.html", context)



@require_saas_admin
def saas_admin_audit_log(request):
    """Enhanced system audit logs with advanced filtering"""
    logs = AuditLog.objects.select_related('user', 'company', 'store').all()

    # Apply filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    action_type = request.GET.get('action_type')
    user_id = request.GET.get('user')
    company_id = request.GET.get('company')
    severity = request.GET.get('severity')
    success = request.GET.get('success')
    search = request.GET.get('search')

    if date_from:
        logs = logs.filter(timestamp__gte=date_from)
    if date_to:
        logs = logs.filter(timestamp__lte=date_to)
    if action_type:
        logs = logs.filter(action=action_type)
    if user_id:
        logs = logs.filter(user_id=user_id)
    if company_id:
        logs = logs.filter(company_id=company_id)
    if severity:
        logs = logs.filter(severity=severity)
    if success:
        logs = logs.filter(success=success == 'true')
    if search:
        logs = logs.filter(
            Q(action_description__icontains=search) |
            Q(resource_name__icontains=search) |
            Q(user__username__icontains=search) |
            Q(user__email__icontains=search)
        )

    # Export functionality
    if request.GET.get('export') == 'csv':
        csv_data = export_audit_logs(logs, format='csv')
        response = HttpResponse(csv_data, content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="audit_logs_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        return response

    # Statistics
    stats = {
        'total_logs': logs.count(),
        'successful_actions': logs.filter(success=True).count(),
        'failed_actions': logs.filter(success=False).count(),
        'critical_actions': logs.filter(severity='critical').count(),
        'requires_review': logs.filter(requires_review=True, reviewed=False).count(),
    }

    # Pagination
    paginator = Paginator(logs.order_by('-timestamp'), 50)
    page = request.GET.get('page', 1)
    logs_page = paginator.get_page(page)

    # Get filter options
    from company.models import Company
    companies = Company.objects.all()[:100]

    context = {
        'logs': logs_page,
        'stats': stats,
        'action_types': AuditLog.ACTION_TYPES,
        'severity_levels': AuditLog.SEVERITY_LEVELS,
        'users': CustomUser.objects.filter(is_hidden=False)[:100],
        'companies': companies,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'action_type': action_type,
            'user': user_id,
            'company': company_id,
            'severity': severity,
            'success': success,
            'search': search,
        }
    }

    return render(request, 'accounts/saas_admin_audit_log.html', context)


@login_required
def user_activity_log(request):
    """User's personal activity log"""
    # Get user's audit logs
    audit_logs = AuditLog.objects.filter(user=request.user).order_by('-timestamp')

    # Apply date filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    action_type = request.GET.get('action_type')

    if date_from:
        audit_logs = audit_logs.filter(timestamp__gte=date_from)
    if date_to:
        audit_logs = audit_logs.filter(timestamp__lte=date_to)
    if action_type:
        audit_logs = audit_logs.filter(action=action_type)

    # Pagination
    paginator = Paginator(audit_logs, 20)
    page = request.GET.get('page', 1)
    logs_page = paginator.get_page(page)

    # Get login history
    login_history = LoginHistory.objects.filter(user=request.user).order_by('-timestamp')[:10]

    # Calculate statistics
    total_actions = audit_logs.count()
    last_30_days = timezone.now() - timedelta(days=30)
    recent_actions = audit_logs.filter(timestamp__gte=last_30_days).count()

    stats = {
        'total_actions': total_actions,
        'recent_actions': recent_actions,
        'successful_logins': LoginHistory.objects.filter(
            user=request.user,
            status='success'
        ).count(),
        'failed_logins': LoginHistory.objects.filter(
            user=request.user,
            status='failed'
        ).count(),
        'last_login': login_history.first() if login_history else None,
        'most_common_action': audit_logs.values('action').annotate(
            count=Count('action')
        ).order_by('-count').first()
    }

    context = {
        'audit_logs': logs_page,
        'login_history': login_history,
        'stats': stats,
        'action_types': AuditLog.ACTION_TYPES,
        'filters': {
            'date_from': date_from,
            'date_to': date_to,
            'action_type': action_type,
        }
    }

    return render(request, 'accounts/user_activity_log.html', context)


@login_required
def login_history_view(request):
    """Detailed login history with security insights"""
    login_history = LoginHistory.objects.filter(user=request.user).order_by('-timestamp')

    # Apply filters
    status = request.GET.get('status')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    if status:
        login_history = login_history.filter(status=status)
    if date_from:
        login_history = login_history.filter(timestamp__gte=date_from)
    if date_to:
        login_history = login_history.filter(timestamp__lte=date_to)

    # Pagination
    paginator = Paginator(login_history, 25)
    page = request.GET.get('page', 1)
    history_page = paginator.get_page(page)

    # Security insights
    last_30_days = timezone.now() - timedelta(days=30)
    insights = {
        'total_logins': login_history.count(),
        'recent_logins': login_history.filter(timestamp__gte=last_30_days).count(),
        'failed_attempts': login_history.filter(status='failed').count(),
        'unique_locations': login_history.exclude(location='').values('location').distinct().count(),
        'unique_devices': login_history.exclude(device_type='').values('device_type').distinct().count(),
        'recent_failures': login_history.filter(
            status='failed',
            timestamp__gte=timezone.now() - timedelta(days=7)
        ).count()
    }

    # Get unique locations for map visualization
    locations = login_history.exclude(
        latitude__isnull=True
    ).values('location', 'latitude', 'longitude').distinct()[:20]

    context = {
        'login_history': history_page,
        'insights': insights,
        'locations': list(locations),
        'status_choices': LoginHistory.STATUS_CHOICES,
        'filters': {
            'status': status,
            'date_from': date_from,
            'date_to': date_to,
        }
    }

    return render(request, 'accounts/login_history.html', context)


@login_required
def data_export_history(request):
    """View user's data export history"""
    exports = DataExportLog.objects.filter(user=request.user).order_by('-timestamp')

    # Apply filters
    export_type = request.GET.get('export_type')
    resource_type = request.GET.get('resource_type')
    date_from = request.GET.get('date_from')

    if export_type:
        exports = exports.filter(export_type=export_type)
    if resource_type:
        exports = exports.filter(resource_type=resource_type)
    if date_from:
        exports = exports.filter(timestamp__gte=date_from)

    # Pagination
    paginator = Paginator(exports, 20)
    page = request.GET.get('page', 1)
    exports_page = paginator.get_page(page)

    # Statistics
    stats = {
        'total_exports': exports.count(),
        'total_records': exports.aggregate(total=Count('record_count'))['total'] or 0,
        'most_exported': exports.values('resource_type').annotate(
            count=Count('id')
        ).order_by('-count').first(),
    }

    context = {
        'exports': exports_page,
        'stats': stats,
        'export_types': DataExportLog.EXPORT_TYPES,
        'filters': {
            'export_type': export_type,
            'resource_type': resource_type,
            'date_from': date_from,
        }
    }

    return render(request, 'accounts/data_export_history.html', context)


@require_saas_admin
def audit_log_detail(request, log_id):
    """Detailed view of a specific audit log entry"""
    log = get_object_or_404(
        AuditLog.objects.select_related('user', 'company', 'store', 'impersonated_by'),
        id=log_id
    )

    # Parse user agent if available
    user_agent_info = None
    if log.user_agent:
        user_agent_info = parse_user_agent(log.user_agent)

    # Get location info if available
    location_info = None
    if log.ip_address:
        location_info = get_location_from_ip(log.ip_address)

    context = {
        'log': log,
        'user_agent_info': user_agent_info,
        'location_info': location_info,
    }

    return render(request, 'accounts/audit_log_detail.html', context)


@require_saas_admin
def mark_log_reviewed(request, log_id):
    """Mark an audit log as reviewed"""
    if request.method == 'POST':
        log = get_object_or_404(AuditLog, id=log_id)
        log.reviewed = True
        log.reviewed_by = request.user
        log.reviewed_at = timezone.now()
        log.save()

        messages.success(request, _('Audit log marked as reviewed'))
        return redirect('audit_log_detail', log_id=log_id)

    return redirect('saas_admin_audit_log')


@login_required
def security_dashboard(request):
    """Security overview dashboard for user"""
    # Recent login attempts
    recent_logins = LoginHistory.objects.filter(
        user=request.user
    ).order_by('-timestamp')[:10]

    # Failed login attempts in last 24 hours
    last_24h = timezone.now() - timedelta(hours=24)
    recent_failures = LoginHistory.objects.filter(
        user=request.user,
        status='failed',
        timestamp__gte=last_24h
    ).count()

    # Recent critical actions
    last_7_days = timezone.now() - timedelta(days=7)
    critical_actions = AuditLog.objects.filter(
        user=request.user,
        severity='critical',
        timestamp__gte=last_7_days
    ).order_by('-timestamp')[:5]

    # Suspicious activity
    suspicious = AuditLog.objects.filter(
        user=request.user,
        action='suspicious_activity',
        timestamp__gte=last_7_days
    ).count()

    # Active sessions (you may need to implement session tracking)
    active_sessions = LoginHistory.objects.filter(
        user=request.user,
        status='success',
        logout_timestamp__isnull=True
    ).order_by('-timestamp')[:5]

    context = {
        'recent_logins': recent_logins,
        'recent_failures': recent_failures,
        'critical_actions': critical_actions,
        'suspicious_count': suspicious,
        'active_sessions': active_sessions,
    }

    return render(request, 'accounts/security_dashboard.html', context)

@require_saas_admin
def audit_statistics(request):
    """Comprehensive audit statistics page"""
    now = timezone.now()
    period = request.GET.get('period', '7d')

    # Determine date range
    if period == '7d':
        start_date = now - timedelta(days=7)
    elif period == '30d':
        start_date = now - timedelta(days=30)
    elif period == '90d':
        start_date = now - timedelta(days=90)
    elif period == '1y':
        start_date = now - timedelta(days=365)
    else:
        start_date = now - timedelta(days=7)

    # Get logs for current period
    current_logs = AuditLog.objects.filter(timestamp__gte=start_date)

    # Calculate overview metrics
    total_logs = current_logs.count()
    successful_actions = current_logs.filter(success=True).count()
    failed_actions = current_logs.filter(success=False).count()
    success_rate = round((successful_actions / total_logs * 100) if total_logs > 0 else 0, 1)
    failure_rate = round((failed_actions / total_logs * 100) if total_logs > 0 else 0, 1)

    critical_events = current_logs.filter(severity='critical').count()

    active_users = current_logs.values('user').distinct().count()

    avg_response_time = current_logs.filter(
        duration_ms__isnull=False
    ).aggregate(avg=Avg('duration_ms'))['avg']
    avg_response_time = round(avg_response_time) if avg_response_time else 0

    # Calculate growth (compare to previous period)
    previous_start = start_date - (now - start_date)
    previous_logs_count = AuditLog.objects.filter(
        timestamp__gte=previous_start,
        timestamp__lt=start_date
    ).count()

    logs_growth = 0
    if previous_logs_count > 0:
        logs_growth = round(((total_logs - previous_logs_count) / previous_logs_count) * 100, 1)

    # Activity over time (daily breakdown)
    activity_labels = []
    activity_data = []

    days_range = (now - start_date).days
    for i in range(days_range):
        day = start_date + timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)

        count = AuditLog.objects.filter(
            timestamp__gte=day_start,
            timestamp__lt=day_end
        ).count()

        activity_labels.append(day.strftime('%Y-%m-%d'))
        activity_data.append(count)

    # Top actions
    top_actions = current_logs.values('action').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    # Add display names for actions
    action_dict = dict(AuditLog.ACTION_TYPES)
    for action in top_actions:
        action['action_display'] = action_dict.get(action['action'], action['action'])

    # Top users
    top_users = current_logs.filter(
        user__isnull=False
    ).values('user__email').annotate(
        action_count=Count('id')
    ).order_by('-action_count')[:10]

    # Format for template
    top_users_list = [{'email': u['user__email'], 'action_count': u['action_count']} for u in top_users]

    # Action distribution for pie chart
    action_distribution = current_logs.values('action').annotate(
        count=Count('id')
    ).order_by('-count')[:5]

    action_labels = [
        force_str(action_dict.get(a['action'], a['action']))
        for a in action_distribution
    ]
    action_data = [a['count'] for a in action_distribution]

    # User activity for bar chart
    user_activity = current_logs.filter(
        user__isnull=False
    ).values('user__email').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    user_labels = [u['user__email'][:20] for u in user_activity]  # Truncate long emails
    user_data = [u['count'] for u in user_activity]

    # Error rate by action
    actions_with_errors = []
    for action_code, action_name in AuditLog.ACTION_TYPES[:10]:
        total = current_logs.filter(action=action_code).count()
        if total > 0:
            failed = current_logs.filter(action=action_code, success=False).count()
            error_rate = round((failed / total) * 100, 1)
            if error_rate > 0:
                actions_with_errors.append({
                    'name': action_name,
                    'error_rate': error_rate,
                    'total': total,
                    'failed': failed
                })

    actions_with_errors.sort(key=lambda x: x['error_rate'], reverse=True)

    # Activity heatmap (last 7 days, hourly)
    heatmap_data = []
    days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    for i in range(7):
        day = now - timedelta(days=6 - i)
        day_name = days[day.weekday()]
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)

        hours = []
        for hour in range(24):
            hour_start = day_start + timedelta(hours=hour)
            hour_end = hour_start + timedelta(hours=1)

            count = AuditLog.objects.filter(
                timestamp__gte=hour_start,
                timestamp__lt=hour_end
            ).count()

            # Calculate intensity (0-5 scale)
            intensity = min(5, count // 5) if count > 0 else 0

            hours.append({
                'hour': hour,
                'count': count,
                'intensity': intensity
            })

        heatmap_data.append({
            'name': day_name[:3],  # Short name
            'hours': hours
        })

    # Period comparison
    current_period_logs = total_logs
    last_period_logs = previous_logs_count

    current_period_users = active_users
    last_period_users = AuditLog.objects.filter(
        timestamp__gte=previous_start,
        timestamp__lt=start_date
    ).values('user').distinct().count()

    context = {
        # Overview metrics
        'total_logs': total_logs,
        'successful_actions': successful_actions,
        'failed_actions': failed_actions,
        'success_rate': success_rate,
        'failure_rate': failure_rate,
        'critical_events': critical_events,
        'active_users': active_users,
        'avg_response_time': avg_response_time,
        'logs_growth': logs_growth,

        # Charts data
        'activity_labels': json.dumps(activity_labels),
        'activity_data': json.dumps(activity_data),
        'action_labels': json.dumps(action_labels),
        'action_data': json.dumps(action_data),
        'user_labels': json.dumps(user_labels),
        'user_data': json.dumps(user_data),

        # Lists
        'top_actions': top_actions,
        'top_users': top_users_list,
        'actions_with_errors': actions_with_errors,

        # Heatmap
        'heatmap_data': heatmap_data,

        # Period comparison
        'current_period_logs': current_period_logs,
        'last_period_logs': last_period_logs,
        'current_period_users': current_period_users,
        'last_period_users': last_period_users,

        # Settings
        'selected_period': period,
    }

    return render(request, 'accounts/audit_statistics.html', context)

@require_saas_admin
def audit_dashboard(request):
    """Dashboard showing audit statistics and insights"""
    now = timezone.now()
    last_24h = now - timedelta(hours=24)
    last_7d = now - timedelta(days=7)
    last_30d = now - timedelta(days=30)

    # Key metrics
    metrics = {
        'total_logs': AuditLog.objects.count(),
        'logs_24h': AuditLog.objects.filter(timestamp__gte=last_24h).count(),
        'failed_actions_24h': AuditLog.objects.filter(
            timestamp__gte=last_24h,
            success=False
        ).count(),
        'critical_alerts': AuditLog.objects.filter(
            severity='critical',
            timestamp__gte=last_7d
        ).count(),
        'pending_reviews': AuditLog.objects.filter(
            requires_review=True,
            reviewed=False
        ).count(),
    }

    # Activity by hour (last 24 hours)
    activity_by_hour = []
    for i in range(24):
        hour = now - timedelta(hours=i)
        hour_start = hour.replace(minute=0, second=0, microsecond=0)
        hour_end = hour_start + timedelta(hours=1)
        count = AuditLog.objects.filter(
            timestamp__gte=hour_start,
            timestamp__lt=hour_end
        ).count()
        activity_by_hour.append({
            'hour': hour_start.isoformat(),
            'count': count
        })
    activity_by_hour.reverse()

    # Top actions (last 7 days)
    top_actions = AuditLog.objects.filter(
        timestamp__gte=last_7d
    ).values('action').annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    # Top users (last 7 days)
    top_users = AuditLog.objects.filter(
        timestamp__gte=last_7d,
        user__isnull=False
    ).values(
        'user__email', 'user__first_name', 'user__last_name'
    ).annotate(
        count=Count('id')
    ).order_by('-count')[:10]

    # Failed actions (last 24 hours)
    failed_actions = AuditLog.objects.filter(
        success=False,
        timestamp__gte=last_24h
    ).select_related('user', 'company').order_by('-timestamp')[:10]

    # Security events (last 7 days)
    security_events = AuditLog.objects.filter(
        Q(action__in=[
            'login_failed',
            'suspicious_activity',
            'account_locked',
            'password_reset',
            'impersonation_started'
        ]) | Q(severity='critical'),
        timestamp__gte=last_7d
    ).select_related('user', 'company').order_by('-timestamp')[:10]

    context = {
        'metrics': metrics,
        'activity_by_hour': json.dumps(activity_by_hour),
        'top_actions': top_actions,
        'top_users': top_users,
        'failed_actions': failed_actions,
        'security_events': security_events,
        'action_types': AuditLog.ACTION_TYPES,
    }

    return render(request, 'accounts/audit_dashboard.html', context)


@login_required
def security_overview(request):
    """Security overview for user's account"""
    now = timezone.now()
    last_30d = now - timedelta(days=30)
    last_7d = now - timedelta(days=7)

    # Recent security events
    security_events = AuditLog.objects.filter(
        user=request.user,
        action__in=[
            'login_success',
            'login_failed',
            'password_changed',
            'password_reset',
            '2fa_enabled',
            '2fa_disabled',
            'email_verified'
        ],
        timestamp__gte=last_30d
    ).order_by('-timestamp')[:20]

    # Login statistics
    login_stats = {
        'successful': LoginHistory.objects.filter(
            user=request.user,
            status='success',
            timestamp__gte=last_30d
        ).count(),
        'failed': LoginHistory.objects.filter(
            user=request.user,
            status='failed',
            timestamp__gte=last_30d
        ).count(),
        'unique_ips': LoginHistory.objects.filter(
            user=request.user,
            timestamp__gte=last_30d
        ).values('ip_address').distinct().count(),
    }

    # Active sessions (from login history without logout)
    active_sessions = LoginHistory.objects.filter(
        user=request.user,
        status='success',
        logout_timestamp__isnull=True,
        timestamp__gte=now - timedelta(days=7)  # Last 7 days
    ).order_by('-timestamp')[:10]

    # Recent failed login attempts
    failed_attempts = LoginHistory.objects.filter(
        user=request.user,
        status='failed',
        timestamp__gte=last_7d
    ).order_by('-timestamp')[:5]

    # Password last changed
    password_changed = AuditLog.objects.filter(
        user=request.user,
        action='password_changed'
    ).order_by('-timestamp').first()

    # Two-factor authentication status
    tfa_enabled = AuditLog.objects.filter(
        user=request.user,
        action='2fa_enabled'
    ).exists()

    tfa_disabled = AuditLog.objects.filter(
        user=request.user,
        action='2fa_disabled'
    ).order_by('-timestamp').first()

    # Check if 2FA is currently active
    tfa_status = tfa_enabled and (
            not tfa_disabled or
            (password_changed and tfa_disabled.timestamp < password_changed.timestamp)
    )

    # Security score calculation
    security_score = 0
    security_recommendations = []

    # Check password age
    if password_changed:
        days_since_change = (now - password_changed.timestamp).days
        if days_since_change < 90:
            security_score += 20
        else:
            security_recommendations.append(
                _('Your password is over 90 days old. Consider changing it.')
            )
    else:
        security_recommendations.append(
            _('No password change recorded. Consider updating your password.')
        )

    # Check 2FA
    if tfa_status:
        security_score += 30
    else:
        security_recommendations.append(
            _('Enable two-factor authentication for better security.')
        )

    # Check failed login attempts
    if login_stats['failed'] == 0:
        security_score += 25
    elif login_stats['failed'] > 5:
        security_recommendations.append(
            _('Multiple failed login attempts detected. Review your login history.')
        )
    else:
        security_score += 15

    # Check login from multiple IPs
    if login_stats['unique_ips'] <= 3:
        security_score += 25
    elif login_stats['unique_ips'] > 10:
        security_recommendations.append(
            _('Logins from many different locations detected. Ensure all sessions are yours.')
        )
    else:
        security_score += 15

    context = {
        'security_events': security_events,
        'login_stats': login_stats,
        'active_sessions': active_sessions,
        'failed_attempts': failed_attempts,
        'password_changed': password_changed,
        'tfa_status': tfa_status,
        'security_score': min(security_score, 100),
        'security_recommendations': security_recommendations,
    }

    return render(request, 'accounts/security_overview.html', context)


@require_saas_admin
def review_audit_log(request, log_id):
    """Review flagged audit log"""
    audit_log = get_object_or_404(AuditLog, id=log_id)

    # Check if already reviewed
    if audit_log.reviewed:
        messages.info(
            request,
            _('This audit log has already been reviewed by %(user)s on %(date)s') % {
                'user': audit_log.reviewed_by.get_full_name(),
                'date': audit_log.reviewed_at.strftime('%Y-%m-%d %H:%M:%S')
            }
        )

    if request.method == 'POST':
        form = ReviewAuditLogForm(request.POST)
        if form.is_valid():
            # Mark as reviewed
            audit_log.reviewed = True
            audit_log.reviewed_by = request.user
            audit_log.reviewed_at = timezone.now()

            # Store review notes in metadata
            if not audit_log.metadata:
                audit_log.metadata = {}

            audit_log.metadata['review_notes'] = form.cleaned_data['notes']
            audit_log.metadata['review_action'] = form.cleaned_data.get('action', '')
            audit_log.metadata['reviewed_by_email'] = request.user.email

            audit_log.save()

            # Log the review action
            AuditLog.objects.create(
                user=request.user,
                action='other',
                action_description=f'Reviewed audit log #{audit_log.id}',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                request_path=request.path,
                request_method=request.method,
                metadata={
                    'reviewed_log_id': audit_log.id,
                    'review_action': form.cleaned_data.get('action', '')
                }
            )

            messages.success(request, _('Audit log reviewed successfully'))

            # Redirect based on action
            if form.cleaned_data.get('action') == 'investigate':
                messages.warning(
                    request,
                    _('This entry has been marked for investigation')
                )

            return redirect('saas_admin_audit_log')
    else:
        # Pre-fill form if there are existing review notes
        initial_data = {}
        if audit_log.metadata and 'review_notes' in audit_log.metadata:
            initial_data['notes'] = audit_log.metadata['review_notes']

        form = ReviewAuditLogForm(initial=initial_data)

    # Get related logs (same user, similar time)
    related_logs = AuditLog.objects.filter(
        user=audit_log.user,
        timestamp__gte=audit_log.timestamp - timedelta(minutes=5),
        timestamp__lte=audit_log.timestamp + timedelta(minutes=5)
    ).exclude(id=audit_log.id).order_by('-timestamp')[:5]

    # Get user's recent activity
    user_recent_activity = None
    if audit_log.user:
        user_recent_activity = AuditLog.objects.filter(
            user=audit_log.user,
            timestamp__gte=audit_log.timestamp - timedelta(hours=1),
            timestamp__lte=audit_log.timestamp
        ).order_by('-timestamp')[:10]

    # Parse user agent
    user_agent_info = None
    if audit_log.user_agent:
        user_agent_info = parse_user_agent(audit_log.user_agent)

    # Get location info
    location_info = None
    if audit_log.ip_address:
        location_info = get_location_from_ip(audit_log.ip_address)

    context = {
        'audit_log': audit_log,
        'form': form,
        'related_logs': related_logs,
        'user_recent_activity': user_recent_activity,
        'user_agent_info': user_agent_info,
        'location_info': location_info,
    }

    return render(request, 'accounts/review_audit_log.html', context)


@require_saas_admin
def bulk_review_audit_logs(request):
    """Bulk review multiple audit logs"""
    if request.method == 'POST':
        log_ids = request.POST.getlist('log_ids')
        review_notes = request.POST.get('review_notes', '')
        review_action = request.POST.get('review_action', '')

        if not log_ids:
            messages.error(request, _('No logs selected for review'))
            return redirect('saas_admin_audit_log')

        # Update all selected logs
        updated_count = 0
        for log_id in log_ids:
            try:
                audit_log = AuditLog.objects.get(id=log_id)
                audit_log.reviewed = True
                audit_log.reviewed_by = request.user
                audit_log.reviewed_at = timezone.now()

                if not audit_log.metadata:
                    audit_log.metadata = {}

                audit_log.metadata['review_notes'] = review_notes
                audit_log.metadata['review_action'] = review_action
                audit_log.metadata['bulk_review'] = True
                audit_log.save()

                updated_count += 1
            except AuditLog.DoesNotExist:
                continue

        messages.success(
            request,
            _('Successfully reviewed %(count)d audit logs') % {'count': updated_count}
        )

        return redirect('saas_admin_audit_log')

    return redirect('saas_admin_audit_log')


@login_required
def revoke_session(request, session_id):
    """Revoke/logout a specific session"""
    try:
        login_history = get_object_or_404(
            LoginHistory,
            id=session_id,
            user=request.user
        )

        # Mark session as logged out
        if not login_history.logout_timestamp:
            login_history.logout_timestamp = timezone.now()
            login_history.save()

            # Log the action
            AuditLog.objects.create(
                user=request.user,
                action='logout',
                action_description=f'Session revoked manually',
                ip_address=get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', ''),
                metadata={
                    'revoked_session_id': session_id,
                    'original_ip': login_history.ip_address
                }
            )

            messages.success(request, _('Session revoked successfully'))
        else:
            messages.info(request, _('This session was already logged out'))

    except Exception as e:
        messages.error(request, _('Failed to revoke session: %(error)s') % {'error': str(e)})

    return redirect('security_overview')


@require_saas_admin
def export_audit_dashboard_data(request):
    """Export dashboard data as JSON for external analysis"""
    now = timezone.now()
    last_30d = now - timedelta(days=30)

    # Compile dashboard data
    data = {
        'generated_at': now.isoformat(),
        'period': '30_days',
        'metrics': {
            'total_logs': AuditLog.objects.count(),
            'logs_30d': AuditLog.objects.filter(timestamp__gte=last_30d).count(),
            'unique_users': AuditLog.objects.values('user').distinct().count(),
            'failed_actions': AuditLog.objects.filter(success=False).count(),
            'critical_events': AuditLog.objects.filter(severity='critical').count(),
        },
        'top_actions': list(
            AuditLog.objects.filter(timestamp__gte=last_30d)
            .values('action')
            .annotate(count=Count('id'))
            .order_by('-count')[:20]
        ),
        'daily_activity': [],
    }

    # Daily activity for last 30 days
    for i in range(30):
        day = now - timedelta(days=i)
        day_start = day.replace(hour=0, minute=0, second=0, microsecond=0)
        day_end = day_start + timedelta(days=1)

        data['daily_activity'].append({
            'date': day_start.date().isoformat(),
            'total': AuditLog.objects.filter(
                timestamp__gte=day_start,
                timestamp__lt=day_end
            ).count(),
            'failed': AuditLog.objects.filter(
                timestamp__gte=day_start,
                timestamp__lt=day_end,
                success=False
            ).count(),
        })

    # Log the export
    AuditLog.objects.create(
        user=request.user,
        action='report_exported',
        action_description='Exported audit dashboard data',
        ip_address=get_client_ip(request),
        metadata={'export_format': 'json'}
    )

    response = JsonResponse(data)
    response['Content-Disposition'] = f'attachment; filename="audit_dashboard_{now.strftime("%Y%m%d_%H%M%S")}.json"'
    return response