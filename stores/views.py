from decimal import Decimal
from django.http import request
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib import messages
from django.views.generic import ListView, View, TemplateView
from django.core.exceptions import PermissionDenied
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy, reverse
from django.db.models import Sum, F, Avg, Count, Q
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required, permission_required
from django.http import HttpResponse, JsonResponse, request
from django.utils import timezone
from django.utils.translation import gettext as _
from datetime import datetime, timedelta
import json
import csv
from company.mixins import CompanyFieldLockMixin
from django.db.models import (
    Case, When, Value, F, Q, Sum, ExpressionWrapper,
    FloatField, CharField, DecimalField
)
from io import BytesIO
import logging

# Optional dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from django.http import Http404
from accounts.models import AuditLog
from inventory.models import Product, Stock, Category, StockMovement
from customers.models import Customer
from sales.models import Sale, SaleItem
from core.mixins import CompanyRestrictedFormMixin
from .models import Store, StoreOperatingHours, StoreDevice, DeviceOperatorLog, UserDeviceSession, SecurityAlert, \
    DeviceFingerprint, StoreAccess
from .forms import (
    StoreForm, StoreOperatingHoursForm, StoreDeviceForm,
    StoreFilterForm, BulkStoreActionForm, StoreStaffAssignmentForm, EnhancedStoreReportForm
)
from django.core.exceptions import ValidationError, PermissionDenied
from company.decorator import check_branch_limit
from django.utils.decorators import method_decorator
from django.contrib.auth import get_user_model
from django.core.serializers.json import DjangoJSONEncoder
from .utils import (
    get_user_active_sessions,
    force_terminate_user_sessions,
    generate_session_report,
    generate_security_report,
    get_user_accessible_stores,
    get_visible_users_for_store,
    filter_stores_by_permissions,
    validate_store_access,
    filter_session_queryset,
    filter_security_alerts,
    get_store_performance_metrics
)

User = get_user_model()
logger = logging.getLogger(__name__)


@login_required
@permission_required('stores.view_devicefingerprint', raise_exception=True)
def user_sessions_view(request):
    """
    View to display user's active sessions
    """
    active_sessions = get_user_active_sessions(request.user)

    # Get all sessions history
    all_sessions = UserDeviceSession.objects.filter(
        user=request.user
    ).order_by('-created_at')[:20]

    # Get security alerts
    recent_alerts = SecurityAlert.objects.filter(
        user=request.user,
        created_at__gte=timezone.now() - timedelta(days=30)
    ).order_by('-created_at')[:10]

    # Get known devices
    known_devices = DeviceFingerprint.objects.filter(
        user=request.user,
        is_active=True
    ).order_by('-last_seen_at')

    context = {
        'active_sessions': active_sessions,
        'all_sessions': all_sessions,
        'recent_alerts': recent_alerts,
        'known_devices': known_devices,
        'active_count': active_sessions.count(),
    }

    return render(request, 'accounts/user_sessions.html', context)


@login_required
@require_http_methods(["POST"])
def terminate_session_view(request, session_id):
    """
    Terminate a specific session
    """
    session = get_object_or_404(
        UserDeviceSession,
        id=session_id,
        user=request.user,
        is_active=True
    )

    from .utils import terminate_device_session
    terminate_device_session(session, reason='LOGGED_OUT', request=request)

    messages.success(request, 'Session terminated successfully.')
    return redirect('stores:user_sessions')


@login_required
@require_http_methods(["POST"])
def terminate_all_sessions_view(request):
    """
    Terminate all user sessions except current one
    """
    current_session = request.session.get('device_session_id')

    count = force_terminate_user_sessions(
        request.user,
        except_session_id=current_session,
        terminated_by=request.user
    )

    messages.success(request, f'{count} session(s) terminated successfully.')
    return redirect('stores:user_sessions')


@login_required
@require_http_methods(["POST"])
def trust_device_view(request, pk):  # Changed from fingerprint_id to pk
    """
    Mark a device as trusted
    """
    device = get_object_or_404(
        DeviceFingerprint,
        id=pk,  # Changed from fingerprint_id to pk
        user=request.user
    )

    device.is_trusted = True
    device.trust_score = 100
    device.save(update_fields=['is_trusted', 'trust_score'])

    messages.success(request, f'Device "{device.device_name}" marked as trusted.')
    return redirect('stores:device_fingerprints')  # Redirect to device fingerprints page


@login_required
@require_http_methods(["POST"])
def remove_device_view(request, pk):  # Changed from fingerprint_id to pk
    """
    Remove/deactivate a device
    """
    device = get_object_or_404(
        DeviceFingerprint,
        id=pk,  # Changed from fingerprint_id to pk
        user=request.user
    )

    device.is_active = False
    device.save(update_fields=['is_active'])

    # Also terminate any active sessions from this device
    UserDeviceSession.objects.filter(
        user=request.user,
        device_fingerprint=device.fingerprint_hash,
        is_active=True
    ).update(is_active=False, status='FORCE_CLOSED')

    messages.success(request, f'Device "{device.device_name}" removed.')
    return redirect('stores:device_fingerprints')  # Redirect to device fingerprints page


# API Endpoints for AJAX requests

@login_required
def api_active_sessions(request):
    """
    API endpoint to get active sessions
    """
    sessions = get_user_active_sessions(request.user)

    data = [{
        'id': session.id,
        'device_name': f"{session.browser_name} on {session.os_name}",
        'ip_address': session.ip_address,
        'location': session.location_string,
        'created_at': session.created_at.isoformat(),
        'last_activity': session.last_activity_at.isoformat(),
        'is_current': session.id == request.session.get('device_session_id'),
        'is_suspicious': session.is_suspicious,
        'store': session.store.name if session.store else None,
        'store_device': session.store_device.name if session.store_device else None,
        'status': session.status,
        'expires_at': session.expires_at.isoformat(),
    } for session in sessions]

    return JsonResponse({
        'sessions': data,
        'count': len(data),
        'max_allowed': 3,
    })


@login_required
def api_security_alerts(request):
    """
    API endpoint to get security alerts
    """
    alerts = SecurityAlert.objects.filter(
        user=request.user,
        status='OPEN'
    ).order_by('-created_at')[:10]

    data = [{
        'id': alert.id,
        'type': alert.get_alert_type_display(),
        'severity': alert.severity,
        'title': alert.title,
        'description': alert.description,
        'created_at': alert.created_at.isoformat(),
    } for alert in alerts]

    return JsonResponse({
        'alerts': data,
        'count': len(data),
    })


@login_required
@require_http_methods(["POST"])
def api_extend_session(request, session_id):
    """
    API endpoint to extend a session
    """
    session = get_object_or_404(
        UserDeviceSession,
        id=session_id,
        user=request.user,
        is_active=True
    )

    hours = int(request.POST.get('hours', 24))
    session.extend_session(hours=hours)

    return JsonResponse({
        'success': True,
        'new_expiry': session.expires_at.isoformat(),
    })


def device_sessions_dashboard(request):
    """Dashboard for viewing all device sessions across stores"""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    # Filter options
    store_id = request.GET.get('store')
    status = request.GET.get('status', 'ACTIVE')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    # Filter sessions
    sessions = filter_session_queryset(request.user)

    # Apply additional filters
    if store_id:
        try:
            # Validate store access
            store = get_object_or_404(Store, id=store_id)
            if store not in stores:
                messages.error(request, 'Access denied to selected store')
                return redirect('stores:device_sessions_dashboard')
            sessions = sessions.filter(store=store)
        except Store.DoesNotExist:
            messages.error(request, 'Invalid store selection')
            return redirect('stores:device_sessions_dashboard')

    if status and status != 'ALL':
        if status == 'ACTIVE':
            sessions = sessions.filter(is_active=True, expires_at__gt=timezone.now())
        else:
            sessions = sessions.filter(status=status)

    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            sessions = sessions.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass

    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            sessions = sessions.filter(created_at__lt=date_to_obj)
        except ValueError:
            pass

    # Order and paginate
    sessions = sessions.order_by('-created_at')

    from django.core.paginator import Paginator
    paginator = Paginator(sessions, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    stats = {
        'total_sessions': sessions.count(),
        'active_sessions': sessions.filter(is_active=True, expires_at__gt=timezone.now()).count(),
        'suspicious_sessions': sessions.filter(is_suspicious=True).count(),
        'new_device_logins': sessions.filter(
            is_new_device=True,
            created_at__gte=timezone.now() - timedelta(days=7)
        ).count(),
    }

    context = {
        'sessions': page_obj,
        'stores': stores,
        'stats': stats,
        'selected_store': store_id,
        'selected_status': status,
        'date_from': date_from,
        'date_to': date_to,
    }

    return render(request, 'stores/device_sessions_dashboard.html', context)


@login_required
@permission_required('stores.view_storedevice')
def security_alerts_view(request):
    """View for managing security alerts"""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    # Filter options
    store_id = request.GET.get('store')
    severity = request.GET.get('severity')
    status = request.GET.get('status', 'OPEN')
    alert_type = request.GET.get('alert_type')

    # Filter alerts
    alerts = filter_security_alerts(request.user)

    # Apply additional filters
    if store_id:
        try:
            store = get_object_or_404(Store, id=store_id)
            if store not in stores:
                messages.error(request, 'Access denied to selected store')
                return redirect('stores:security_alerts')
            alerts = alerts.filter(store=store)
        except Store.DoesNotExist:
            messages.error(request, 'Invalid store selection')
            return redirect('stores:security_alerts')

    if severity:
        alerts = alerts.filter(severity=severity)

    if status and status != 'ALL':
        alerts = alerts.filter(status=status)

    if alert_type:
        alerts = alerts.filter(alert_type=alert_type)

    # Order by severity and date
    alerts = alerts.order_by('-created_at')

    # Paginate
    from django.core.paginator import Paginator
    paginator = Paginator(alerts, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    stats = {
        'total_alerts': alerts.count(),
        'open_alerts': alerts.filter(status='OPEN').count(),
        'high_severity': alerts.filter(severity='HIGH', status='OPEN').count(),
        'critical_severity': alerts.filter(severity='CRITICAL', status='OPEN').count(),
    }

    context = {
        'alerts': page_obj,
        'stores': stores,
        'stats': stats,
        'selected_store': store_id,
        'selected_severity': severity,
        'selected_status': status,
        'selected_type': alert_type,
        'severity_choices': SecurityAlert.SEVERITY_LEVELS,
        'status_choices': SecurityAlert.STATUS_CHOICES,
        'alert_type_choices': SecurityAlert.ALERT_TYPES,
    }

    return render(request, 'stores/security_alerts.html', context)


@login_required
@permission_required('stores.change_storedevice')
@require_http_methods(["POST"])
def resolve_security_alert(request, alert_id):
    """Resolve a security alert"""
    alert = get_object_or_404(SecurityAlert, id=alert_id)

    # Validate store access
    stores = get_user_accessible_stores(request.user)
    if alert.store not in stores:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    action = request.POST.get('action', 'resolve')
    notes = request.POST.get('notes', '')

    if action == 'resolve':
        alert.resolve(resolved_by=request.user, notes=notes)
        messages.success(request, 'Alert resolved successfully.')
    elif action == 'false_positive':
        alert.mark_false_positive(resolved_by=request.user, notes=notes)
        messages.success(request, 'Alert marked as false positive.')
    elif action == 'investigating':
        alert.status = 'INVESTIGATING'
        alert.save(update_fields=['status'])
        messages.success(request, 'Alert marked as investigating.')

    return redirect('stores:security_alerts')


@login_required
@permission_required('stores.view_storedevice')
def device_fingerprints_view(request):
    """
    View for managing device fingerprints
    """
    # Get accessible stores first
    stores = get_user_accessible_stores(request.user)

    # Get filter parameters
    user_id = request.GET.get('user')
    trusted = request.GET.get('trusted')
    active = request.GET.get('active', 'true')

    # Base queryset - limit to users in accessible stores
    fingerprints = DeviceFingerprint.objects.filter(
        user__stores__in=stores
    ).select_related('user').distinct()

    # Apply filters
    if user_id:
        fingerprints = fingerprints.filter(user_id=user_id)

    if trusted and trusted != 'all':
        fingerprints = fingerprints.filter(is_trusted=(trusted == 'true'))

    if active and active != 'all':
        fingerprints = fingerprints.filter(is_active=(active == 'true'))

    # Order by most recent
    fingerprints = fingerprints.order_by('-last_seen_at')

    # Paginate
    from django.core.paginator import Paginator
    paginator = Paginator(fingerprints, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Statistics
    stats = {
        'total_devices': fingerprints.count(),
        'trusted_devices': fingerprints.filter(is_trusted=True).count(),
        'active_devices': fingerprints.filter(is_active=True).count(),
        'recent_logins': fingerprints.filter(last_seen_at__gte=timezone.now() - timedelta(days=7)).count(),
    }

    context = {
        'fingerprints': page_obj,
        'stats': stats,
        'selected_user': user_id,
        'selected_trusted': trusted,
        'selected_active': active,
    }

    return render(request, 'stores/device_fingerprints.html', context)


@login_required
@require_http_methods(["GET"])
def device_session_report(request):
    """Generate device session report"""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    # Date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    export_format = request.GET.get('format', 'json')

    # Generate report
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        date_from = (timezone.now() - timedelta(days=30)).date()

    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        date_to = timezone.now().date()

    store = None
    if store_id:
        store = get_object_or_404(Store, id=store_id)
        if store not in stores:
            return JsonResponse({'error': 'Permission denied'}, status=403)

    # Get report data
    report = generate_session_report(
        store=store,
        date_from=date_from,
        date_to=date_to
    )

    # Handle CSV export
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="session_report_{date_from}_{date_to}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Device Session Report'])
        writer.writerow([f'Period: {date_from} to {date_to}'])
        writer.writerow([f'Store: {store.name if store else "All Stores"}'])
        writer.writerow([])

        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Sessions', report.get('total_sessions', 0)])
        writer.writerow(['Active Sessions', report.get('active_sessions', 0)])
        writer.writerow(['Suspicious Sessions', report.get('suspicious_sessions', 0)])
        writer.writerow(['New Device Sessions', report.get('new_device_sessions', 0)])
        writer.writerow(['Unique Users', report.get('unique_users', 0)])
        writer.writerow(['Unique Devices', report.get('unique_devices', 0)])

        writer.writerow([])
        writer.writerow(['Sessions by Browser'])
        writer.writerow(['Browser', 'Count'])
        for item in report.get('sessions_by_browser', []):
            writer.writerow([item['browser_name'], item['count']])

        writer.writerow([])
        writer.writerow(['Sessions by OS'])
        writer.writerow(['OS', 'Count'])
        for item in report.get('sessions_by_os', []):
            writer.writerow([item['os_name'], item['count']])

        return response

    return JsonResponse(report)


@login_required
@require_http_methods(["GET"])
def security_report(request):
    """Generate security report"""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    # Date range
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    store_id = request.GET.get('store')
    severity = request.GET.get('severity')
    export_format = request.GET.get('format', 'json')

    # Parse dates
    if date_from:
        date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
    else:
        date_from = (timezone.now() - timedelta(days=30)).date()

    if date_to:
        date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
    else:
        date_to = timezone.now().date()

    store = None
    if store_id:
        store = get_object_or_404(Store, id=store_id)
        if store not in stores:
            return JsonResponse({'error': 'Permission denied'}, status=403)

    # Generate report data
    report = generate_security_report(
        store=store,
        severity=severity,
        date_from=date_from,
        date_to=date_to
    )

    # Handle CSV export
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="security_report_{date_from}_{date_to}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Security Alerts Report'])
        writer.writerow([f'Period: {date_from} to {date_to}'])
        writer.writerow([f'Store: {store.name if store else "All Stores"}'])
        writer.writerow([])

        writer.writerow(['Metric', 'Value'])
        writer.writerow(['Total Alerts', report.get('total_alerts', 0)])
        writer.writerow(['Open Alerts', report.get('open_alerts', 0)])
        writer.writerow(['Resolved Alerts', report.get('resolved_alerts', 0)])
        writer.writerow(['High Severity Open', report.get('high_severity_open', 0)])
        writer.writerow(['Critical Severity Open', report.get('critical_severity_open', 0)])

        writer.writerow([])
        writer.writerow(['Alerts by Severity'])
        writer.writerow(['Severity', 'Count'])
        for item in report.get('by_severity', []):
            writer.writerow([item['severity'], item['count']])

        writer.writerow([])
        writer.writerow(['Alerts by Type'])
        writer.writerow(['Type', 'Count'])
        for item in report.get('by_type', []):
            writer.writerow([item['alert_type'], item['count']])

        return response

    return JsonResponse(report)


# --- Store Selection and POS Views ---

@login_required
def pos_interface(request):
    """POS Terminal Interface - Store-specific sales interface."""
    store_id = request.GET.get('store')
    if not store_id:
        messages.error(request, 'Please select a store to access POS terminal')
        return redirect('stores:select_store')

    store = get_object_or_404(Store, id=store_id)

    # Validate store access
    try:
        validate_store_access(request.user, store, action='view', raise_exception=True)
    except PermissionDenied:
        messages.error(request, 'You do not have access to this store')
        return redirect('stores:select_store')

    context = {
        'store': store,
        'recent_customers': Customer.objects.filter(sales__store=store).distinct()[:10],
        'categories': Category.objects.filter(products__store_inventories__store=store).distinct(),
        'payment_methods': Sale.PAYMENT_METHODS,
        'tax_rates': [('A', '18%'), ('B', '12%'), ('C', '0%')],
    }
    return render(request, 'sales/pos_terminal.html', context)


class SelectStoreView(LoginRequiredMixin, ListView):
    """
    View for users to select which store they want to work with
    """
    template_name = 'stores/select_store.html'
    context_object_name = 'stores'

    def get_queryset(self):
        return get_user_accessible_stores(self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['current_store_id'] = self.request.session.get('current_store_id')
        return context


class SwitchStoreView(LoginRequiredMixin, View):
    """
    Switch to a different store
    """

    def post(self, request, store_id):
        try:
            store = Store.objects.get(id=store_id, is_active=True)

            # Validate store access
            try:
                validate_store_access(request.user, store, action='view', raise_exception=True)
            except PermissionDenied:
                messages.error(request, "You do not have access to that store.")
                return redirect('stores:select_store')

            request.session['current_store_id'] = store.id
            messages.success(request, f"Switched to {store.name}")

            # Log the switch
            from accounts.models import AuditLog
            AuditLog.log(
                action='other',
                user=request.user,
                description=f"Switched to store: {store.name}",
                store=store
            )

            return redirect(request.GET.get('next', 'user_dashboard'))

        except Store.DoesNotExist:
            messages.error(request, "Store not found.")
            return redirect('stores:select_store')


class NoStoreAccessView(LoginRequiredMixin, TemplateView):
    """
    View shown when user has no store access
    """
    template_name = 'stores/no_store_access.html'

    def dispatch(self, request, *args, **kwargs):
        # Allow access to this view even without store access
        # We'll handle the check manually
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Add company info if available
        if hasattr(self.request, 'company'):
            context['company'] = self.request.company

        # Add user info
        context['user'] = self.request.user

        # Check if user might have access now
        from stores.models import Store, StoreAccess

        user = self.request.user
        has_access = False

        if hasattr(user, 'stores') and user.stores.filter(is_active=True).exists():
            has_access = True
        elif hasattr(user, 'managed_stores') and user.managed_stores.filter(is_active=True).exists():
            has_access = True
        elif StoreAccess.objects.filter(user=user, is_active=True, store__is_active=True).exists():
            has_access = True
        elif hasattr(self.request, 'company'):
            if Store.objects.filter(
                    company=self.request.company,
                    is_active=True,
                    accessible_by_all=True
            ).exists():
                has_access = True

        context['has_access'] = has_access

        return context

class CheckStoreAccessView(LoginRequiredMixin, View):
    """
    API endpoint to check if user has gained store access
    """

    def get(self, request):
        from django.contrib.auth import get_user_model
        User = get_user_model()

        user = request.user
        has_access = False
        accessible_stores = []

        try:
            # Check through multiple access methods
            if hasattr(user, 'stores'):
                # Direct store assignment
                accessible_stores = user.stores.filter(is_active=True)

            if not accessible_stores and hasattr(user, 'managed_stores'):
                # Store manager assignment
                accessible_stores = user.managed_stores.filter(is_active=True)

            if not accessible_stores and hasattr(request, 'company'):
                # Company-wide access
                accessible_stores = Store.objects.filter(
                    company=request.company,
                    is_active=True,
                    accessible_by_all=True
                )

            # Check store access permissions
            if not accessible_stores:
                accessible_stores = Store.objects.filter(
                    access_permissions__user=user,
                    access_permissions__is_active=True,
                    is_active=True
                ).distinct()

            has_access = accessible_stores.exists()

            return JsonResponse({
                'has_access': has_access,
                'store_count': accessible_stores.count(),
                'message': _('Access granted') if has_access else _('No store access')
            })

        except Exception as e:
            return JsonResponse({
                'has_access': False,
                'error': str(e),
                'message': _('Error checking access')
            }, status=500)


@login_required
@require_http_methods(["GET"])
def pos_product_search(request):
    """API endpoint for POS product search with inventory data."""
    query = request.GET.get('q', '').strip()
    store_id = request.GET.get('store_id')
    category_id = request.GET.get('category_id')

    if not query and not category_id:
        return JsonResponse({'products': []})

    try:
        store = Store.objects.get(id=store_id)
        # Validate store access
        try:
            validate_store_access(request.user, store, action='view', raise_exception=True)
        except PermissionDenied:
            return JsonResponse({'error': 'Access denied to store'}, status=403)
    except Store.DoesNotExist:
        return JsonResponse({'error': 'Invalid store'}, status=400)

    products_query = Product.objects.filter(
        store_inventories__store=store, is_active=True
    ).select_related().prefetch_related('store_inventories')

    if query:
        products_query = products_query.filter(
            Q(name__icontains=query) | Q(sku__icontains=query) | Q(barcode__icontains=query)
        )
    if category_id:
        products_query = products_query.filter(category_id=category_id)

    products = [{
        'id': product.id,
        'name': product.name,
        'sku': product.sku,
        'barcode': product.barcode,
        'price': float(product.selling_price),
        'cost_price': float(product.cost_price),
        'tax_rate': product.tax_rate,
        'unit_of_measure': product.unit_of_measure or 'pcs',
        'category': product.category.name if product.category else None,
        'image_url': product.image.url if product.image else None,
        'stock': {
            'available': product.store_inventories.get(store=store).quantity if product.store_inventories.filter(
                store=store).exists() else 0,
            'unit': product.unit_of_measure or 'pcs',
            'in_stock': product.store_inventories.get(store=store).quantity > 0 if product.store_inventories.filter(
                store=store).exists() else False
        }
    } for product in products_query[:20]]

    return JsonResponse({'products': products})


@login_required
@require_http_methods(["GET"])
def pos_customer_search(request):
    """API endpoint for POS customer search."""
    query = request.GET.get('q', '').strip()
    if len(query) < 2:
        return JsonResponse({'customers': []})

    customers = Customer.objects.filter(
        Q(name__icontains=query) | Q(phone__icontains=query) | Q(email__icontains=query),
        is_active=True
    )[:15]

    return JsonResponse({
        'customers': [{
            'id': customer.id,
            'name': customer.name,
            'phone': customer.phone,
            'email': customer.email,
            'address': customer.address,
            'total_purchases': customer.total_purchases or 0,
            'last_purchase': customer.last_purchase_date.isoformat() if customer.last_purchase_date else None
        } for customer in customers]
    })


@login_required
@require_http_methods(["POST"])
def pos_create_sale(request):
    """Create sale from POS terminal."""
    try:
        data = json.loads(request.body)
        store = get_object_or_404(Store, id=data.get('store_id'))

        # Validate store access
        try:
            validate_store_access(request.user, store, action='view', raise_exception=True)
        except PermissionDenied:
            return JsonResponse({'success': False, 'error': 'Access denied to store'}, status=403)

        if not data.get('items') or len(data['items']) == 0:
            return JsonResponse({'success': False, 'error': 'No items in cart'}, status=400)
        if not data.get('payment_method'):
            return JsonResponse({'success': False, 'error': 'Payment method is required'}, status=400)

        sale = Sale.objects.create(
            store=store,
            user=request.user,
            customer_id=data.get('customer_id'),
            payment_method=data['payment_method'],
            transaction_type=data.get('transaction_type', 'SALE'),
            document_type=data.get('document_type', 'RECEIPT'),
            currency=data.get('currency', 'UGX'),
            discount_amount=float(data.get('discount_amount', 0)),
            notes=data.get('notes', ''),
            payment_amount=float(data.get('payment_amount', 0)),
            payment_reference=data.get('payment_reference', '')
        )

        total_amount = 0
        total_tax = 0
        for item_data in data['items']:
            product = get_object_or_404(Product, id=item_data['product_id'])
            quantity = float(item_data['quantity'])
            unit_price = float(item_data['unit_price'])
            discount = float(item_data.get('discount', 0))
            tax_rate = item_data.get('tax_rate', 'A')
            tax_percentage = 18 if tax_rate == 'A' else (12 if tax_rate == 'B' else 0)

            line_total = quantity * unit_price
            tax_amount = line_total * (tax_percentage / 100)
            item_total = line_total + tax_amount - discount

            SaleItem.objects.create(
                sale=sale,
                product=product,
                quantity=quantity,
                unit_price=unit_price,
                tax_rate=tax_rate,
                tax_amount=tax_amount,
                discount=discount,
                total_amount=item_total
            )

            total_amount += item_total
            total_tax += tax_amount

            stock, created = Stock.objects.get_or_create(
                store=store,
                product=product,
                defaults={'quantity': 0, 'low_stock_threshold': 5, 'reorder_quantity': 10}
            )
            stock.quantity -= quantity
            stock.save()

        sale.subtotal = total_amount - total_tax
        sale.tax_amount = total_tax
        sale.total_amount = total_amount - float(data.get('discount_amount', 0))
        sale.save()

        change_amount = float(data.get('payment_amount', 0)) - sale.total_amount

        return JsonResponse({
            'success': True,
            'sale_id': sale.id,
            'invoice_number': sale.invoice_number,
            'total_amount': sale.total_amount,
            'change_amount': max(change_amount, 0),
            'print_url': f'/sales/{sale.id}/print-receipt/'
        })
    except Exception as e:
        logger.error(f"Error creating sale: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def pos_quick_customer(request):
    """Create a quick customer from POS."""
    try:
        data = json.loads(request.body)
        customer = Customer.objects.create(
            name=data['name'],
            phone=data.get('phone', ''),
            email=data.get('email', ''),
            address=data.get('address', ''),
            created_by=request.user
        )
        return JsonResponse({
            'success': True,
            'customer': {
                'id': customer.id,
                'name': customer.name,
                'phone': customer.phone,
                'email': customer.email,
                'address': customer.address
            }
        })
    except Exception as e:
        logger.error(f"Error creating quick customer: {str(e)}", exc_info=True)
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


# ==========================================
# UPDATED CLASS-BASED VIEWS
# ==========================================

class StoreListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List stores with filtering and bulk actions."""
    model = Store
    template_name = 'stores/store_list.html'
    context_object_name = 'stores'
    paginate_by = 20
    permission_required = 'stores.view_store'

    def get_queryset(self):
        # Get accessible stores
        queryset = get_user_accessible_stores(self.request.user).select_related(
            'company'
        ).prefetch_related('staff', 'store_managers', 'devices', 'inventory_items')

        # Apply form filters
        form = StoreFilterForm(self.request.GET)
        if form.is_valid():
            search = form.cleaned_data.get('search')
            region = form.cleaned_data.get('region')
            status = form.cleaned_data.get('status')
            efris_status = form.cleaned_data.get('efris_status')
            store_type = form.cleaned_data.get('store_type')

            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(code__icontains=search) |
                    Q(physical_address__icontains=search) |
                    Q(region__icontains=search) |
                    Q(manager_name__icontains=search)
                )

            if region:
                queryset = queryset.filter(region__icontains=region)

            if status == 'active':
                queryset = queryset.filter(is_active=True)
            elif status == 'inactive':
                queryset = queryset.filter(is_active=False)

            if efris_status == 'enabled':
                queryset = queryset.filter(efris_enabled=True)
            elif efris_status == 'disabled':
                queryset = queryset.filter(efris_enabled=False)

            if store_type:
                queryset = queryset.filter(store_type=store_type)

        return queryset.order_by('-is_main_branch', 'name')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = StoreFilterForm(self.request.GET)
        context['total_stores'] = self.get_queryset().count()

        # Add statistics
        queryset = self.get_queryset()
        context['stats'] = {
            'active_stores': queryset.filter(is_active=True).count(),
            'efris_enabled': queryset.filter(efris_enabled=True).count(),
            'main_stores': queryset.filter(is_main_branch=True).count(),
        }
        return context


class StoreDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Detailed view of a store with related information."""
    model = Store
    template_name = 'stores/store_detail.html'
    context_object_name = 'store'
    permission_required = 'stores.view_store'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)

        # Validate access
        try:
            validate_store_access(self.request.user, obj, action='view', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            raise

        return obj

    def get_queryset(self):
        # Get accessible stores
        return get_user_accessible_stores(self.request.user).select_related(
            'company'
        ).prefetch_related('devices', 'inventory_items__product', 'staff', 'store_managers')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        store = self.object

        # Filter staff to show only visible users
        visible_staff = get_visible_users_for_store(store, self.request.user)

        # Get store performance metrics
        try:
            performance_metrics = get_store_performance_metrics(store, days=30)
            context['performance_metrics'] = performance_metrics
        except Exception as e:
            logger.error(f"Error getting performance metrics: {e}")
            context['performance_metrics'] = {}

        context.update({
            'operating_hours': [
                {'day': day.capitalize(), **details}
                for day, details in sorted(store.operating_hours.items())
            ] if isinstance(store.operating_hours, dict) else [],
            'devices': store.devices.filter(is_active=True).order_by('-registered_at'),
            'inventory': store.inventory_items.select_related('product').all()[:20],
            'low_stock_items': store.inventory_items.filter(
                quantity__lte=F('low_stock_threshold')
            ).count(),
            'visible_staff': visible_staff,
            'store_managers': store.store_managers.filter(is_active=True, is_hidden=False),
            'staff_form': StoreStaffAssignmentForm(store_instance=store, user=self.request.user),
            'recent_logs': DeviceOperatorLog.objects.filter(
                device__store=store,
                user__is_hidden=False
            ).select_related('user', 'device').order_by('-timestamp')[:10],
            'store_open_now': store.is_open_now(),
            # Check if user is a manager for this store
            'is_store_manager': store.store_managers.filter(id=self.request.user.id).exists(),
            # EFRIS configuration
            'efris_config': store.effective_efris_config,
            'can_fiscalize': store.can_fiscalize,
        })
        return context


class StoreCreateView(CompanyFieldLockMixin, LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new store"""

    model = Store
    form_class = StoreForm
    template_name = 'stores/store_form.html'
    permission_required = 'stores.add_store'
    success_url = reverse_lazy('stores:store_list')

    @method_decorator(check_branch_limit)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['tenant'] = getattr(self.request, 'tenant', None)
        return kwargs

    def form_valid(self, form):
        # Set company if user has one
        if hasattr(self.request.user, 'company') and self.request.user.company:
            form.instance.company = self.request.user.company

        try:
            response = super().form_valid(form)
            messages.success(self.request, f'Store "{form.instance.name}" created successfully!')
            return response
        except ValidationError as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)
        except Exception as e:
            messages.error(self.request, f'Error creating store: {str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        """Add branch limit and EFRIS status context to template"""
        context = super().get_context_data(**kwargs)

        # Check EFRIS status - check company first, then tenant
        efris_is_enabled = False

        # Check company level EFRIS
        if hasattr(self.request.user, 'company') and self.request.user.company:
            company = self.request.user.company
            efris_is_enabled = getattr(company, 'efris_enabled', False)

            current_stores = Store.objects.filter(company=company).count()
            context.update({
                'current_stores': current_stores,
                'branch_limit': getattr(company.plan, 'branch_limit', 0) if hasattr(company, 'plan') else 0,
                'can_create_more': current_stores < getattr(company.plan, 'branch_limit', 0) if hasattr(company,
                                                                                                        'plan') else True
            })

        # If not enabled at company level, check tenant level
        if not efris_is_enabled and hasattr(self.request, 'tenant'):
            efris_is_enabled = getattr(self.request.tenant, 'efris_enabled', False)

        context['efris_is_enabled'] = efris_is_enabled

        # Pass current company for reference
        if hasattr(self.request.user, 'company') and self.request.user.company:
            context['company'] = self.request.user.company

        return context


class StoreUpdateView(CompanyFieldLockMixin, LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update an existing store"""

    model = Store
    form_class = StoreForm
    template_name = 'stores/store_form.html'
    permission_required = 'stores.change_store'

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        kwargs['tenant'] = getattr(self.request, 'tenant', None)
        return kwargs

    def get_success_url(self):
        return reverse('stores:store_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        # Check if user has permission to update this store
        try:
            validate_store_access(self.request.user, self.object, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, f'Store "{form.instance.name}" updated successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)

    def get_context_data(self, **kwargs):
        """Add EFRIS status context to template"""
        context = super().get_context_data(**kwargs)

        # Check EFRIS status - check company first, then tenant
        efris_is_enabled = False

        # Check company level EFRIS
        if self.object and self.object.company:
            efris_is_enabled = getattr(self.object.company, 'efris_enabled', False)

        # If not enabled at company level, check tenant level
        if not efris_is_enabled and hasattr(self.request, 'tenant'):
            efris_is_enabled = getattr(self.request.tenant, 'efris_enabled', False)

        context['efris_is_enabled'] = efris_is_enabled

        # Pass the current store instance for reference
        context['object'] = self.object

        return context

class StoreDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a store."""
    model = Store
    template_name = 'stores/store_delete.html'
    permission_required = 'stores.delete_store'
    success_url = reverse_lazy('stores:store_list')

    def delete(self, request, *args, **kwargs):
        store = self.get_object()

        # Check if user has permission to delete this store
        try:
            validate_store_access(self.request.user, store, action='delete', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return redirect('stores:store_detail', pk=store.pk)

        store_name = store.name
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'Store "{store_name}" deleted successfully!')
        return response


@login_required
@permission_required('stores.view_store')
def store_dashboard(request):
    """Store management dashboard with analytics."""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)
    active_stores = stores.filter(is_active=True)

    # Get current store from session
    current_store_id = request.session.get('current_store_id')
    current_store = None
    if current_store_id:
        try:
            current_store = Store.objects.get(id=current_store_id, is_active=True)
            # Validate access to current store
            try:
                validate_store_access(request.user, current_store, action='view', raise_exception=True)
            except PermissionDenied:
                current_store = None
        except Store.DoesNotExist:
            current_store = None

    context = {
        'stats': {
            'total_stores': stores.count(),
            'active_stores': active_stores.count(),
            'inactive_stores': stores.filter(is_active=False).count(),
            'efris_enabled': stores.filter(efris_enabled=True).count(),
            'total_devices': StoreDevice.objects.filter(store__in=stores, is_active=True).count(),
            'active_devices': StoreDevice.objects.filter(
                store__in=stores,
                is_active=True
            ).count(),
        },
        'recent_stores': stores.order_by('-created_at')[:5],
        'low_stock_count': Stock.objects.filter(
            store__in=active_stores,
            quantity__lte=F('low_stock_threshold')
        ).count(),
        'stores_by_region': list(
            stores.values('region').annotate(count=Count('id')).order_by('-count')[:10]
        ),
        'recent_activity': DeviceOperatorLog.objects.filter(
            device__store__in=stores,
            user__is_hidden=False
        ).select_related('user', 'device__store').order_by('-timestamp')[:10],
        # Current store context
        'current_store': current_store,
        'can_switch_stores': stores.count() > 1,
        'accessible_stores': stores,
    }
    return render(request, 'stores/dashboard.html', context)


@login_required
@permission_required('stores.change_store')
def bulk_store_actions(request):
    """Handle bulk actions on stores."""
    if request.method == 'POST':
        form = BulkStoreActionForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            store_select = form.cleaned_data['selected_stores']

            # Get selected stores with access validation
            accessible_stores = get_user_accessible_stores(request.user)

            if store_select == 'all':
                stores = accessible_stores
            else:
                try:
                    store_ids = [int(id.strip()) for id in store_select.split(',')]
                    stores = accessible_stores.filter(id__in=store_ids)
                except (ValueError, AttributeError):
                    messages.error(request, 'Invalid store selection.')
                    return redirect('stores:store_list')

            count = stores.count()

            actions = {
                'activate': lambda: stores.update(is_active=True),
                'deactivate': lambda: stores.update(is_active=False),
                'enable_efris': lambda: stores.update(efris_enabled=True),
                'disable_efris': lambda: stores.update(efris_enabled=False),
                'delete': lambda: stores.delete() if request.user.has_perm('stores.delete_store') else None
            }

            if action in actions:
                if action == 'delete' and not request.user.has_perm('stores.delete_store'):
                    messages.error(request, 'Permission denied for delete action.')
                else:
                    actions[action]()
                    messages.success(request, f'{count} stores {action}d successfully!')
            else:
                messages.error(request, 'Invalid action.')
        else:
            messages.error(request, 'Invalid form data.')
    return redirect('stores:store_list')


@login_required
@permission_required('stores.change_store')
def manage_store_staff(request, pk):
    """Manage staff assignments for a store."""
    store = get_object_or_404(Store, pk=pk)

    # Validate store access
    try:
        validate_store_access(request.user, store, action='change', raise_exception=True)
    except PermissionDenied as e:
        messages.error(request, str(e))
        return redirect('stores:store_detail', pk=pk)

    # Check if user is a store manager or has higher permissions
    is_store_manager = store.store_managers.filter(id=request.user.id).exists()
    if not (request.user.is_company_owner or request.user.company_admin or is_store_manager):
        messages.error(request, 'You do not have permission to manage staff for this store')
        return redirect('stores:store_detail', pk=pk)

    # Get visible staff
    visible_current_staff = get_visible_users_for_store(store, request.user)

    form = StoreStaffAssignmentForm(
        store_instance=store,
        user=request.user,
        data=request.POST if request.method == 'POST' else None
    )

    if request.method == 'POST' and form.is_valid():
        if form.cleaned_data.get('add_staff'):
            added_count = 0
            for staff_member in form.cleaned_data['add_staff']:
                # Check if user can manage this staff member
                if request.user.can_manage_user(staff_member):
                    store.staff.add(staff_member)

                    # Create StoreAccess record
                    StoreAccess.objects.get_or_create(
                        user=staff_member,
                        store=store,
                        defaults={
                            'access_level': 'staff',
                            'can_view_sales': True,
                            'can_create_sales': True,
                            'can_view_inventory': True,
                            'granted_by': request.user,
                        }
                    )
                    added_count += 1
                else:
                    messages.warning(
                        request,
                        f'Insufficient privileges to add {staff_member.get_full_name()}'
                    )
            if added_count > 0:
                messages.success(request, f'{added_count} staff member(s) added successfully!')

        if form.cleaned_data.get('remove_staff'):
            removed_count = 0
            for staff_member in form.cleaned_data['remove_staff']:
                if request.user.can_manage_user(staff_member):
                    store.staff.remove(staff_member)

                    # Revoke StoreAccess
                    access = StoreAccess.objects.filter(
                        user=staff_member,
                        store=store
                    ).first()
                    if access:
                        access.revoke(revoked_by=request.user)
                    removed_count += 1
                else:
                    messages.warning(
                        request,
                        f'Insufficient privileges to remove {staff_member.get_full_name()}'
                    )
            if removed_count > 0:
                messages.success(request, f'{removed_count} staff member(s) removed successfully!')

        return redirect('stores:store_detail', pk=pk)

    context = {
        'store': store,
        'form': form,
        'current_staff': visible_current_staff,
        'store_managers': store.store_managers.filter(is_active=True),
        'is_store_manager': is_store_manager,
    }
    return render(request, 'stores/manage_staff.html', context)


# Add these views to your stores/views.py file

class ManageStoreAccessView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """View for managing store access permissions"""
    model = StoreAccess
    template_name = 'stores/manage_access.html'
    context_object_name = 'access_permissions'
    permission_required = 'stores.view_storeaccess'
    paginate_by = 20

    def get_queryset(self):
        store = get_object_or_404(Store, pk=self.kwargs['store_id'])

        # Validate store access
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied:
            return StoreAccess.objects.none()

        # Get store managers to check if user is a manager
        is_store_manager = store.store_managers.filter(id=self.request.user.id).exists()

        # If user is not company admin or store manager, show only their own access
        if not (self.request.user.is_company_owner or
                self.request.user.company_admin or
                is_store_manager):
            return store.access_permissions.filter(user=self.request.user, is_active=True)

        # Show all access permissions for the store
        return store.access_permissions.filter(is_active=True).select_related(
            'user', 'granted_by'
        )

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        store = get_object_or_404(Store, pk=self.kwargs['store_id'])

        context.update({
            'store': store,
            'store_id': store.id,
            'can_manage_access': self._can_manage_access(store),
            'available_users': self._get_available_users(store),
            'access_level_choices': StoreAccess.ACCESS_LEVELS,
        })
        return context

    def _can_manage_access(self, store):
        """Check if user can manage access for this store"""
        if self.request.user.is_saas_admin:
            return True
        if not store.company or store.company != self.request.user.company:
            return False

        is_store_manager = store.store_managers.filter(id=self.request.user.id).exists()
        return (self.request.user.is_company_owner or
                self.request.user.company_admin or
                is_store_manager)

    def _get_available_users(self, store):
        """Get users who can be granted access to this store"""
        # Get all active users in the company
        company_users = User.objects.filter(
            company=store.company,
            is_active=True,
            is_hidden=False
        ).exclude(
            id__in=store.access_permissions.filter(is_active=True).values('user')
        )

        # Filter based on user's permissions
        if not self.request.user.is_saas_admin:
            # Regular users can only grant access to users they can manage
            from accounts.utils import can_manage_user
            company_users = [
                user for user in company_users
                if can_manage_user(self.request.user, user)
            ]

        return company_users


class CreateStoreAccessView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for creating new store access"""
    permission_required = 'stores.add_storeaccess'

    def post(self, request, store_id):
        store = get_object_or_404(Store, pk=store_id)

        # Validate store access
        try:
            validate_store_access(request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(request, str(e))
            return redirect('stores:manage_store_access', store_id=store_id)

        # Check if user can manage access
        if not self._can_manage_access(store, request.user):
            messages.error(request, 'You do not have permission to manage access for this store')
            return redirect('stores:store_detail', pk=store_id)

        user_id = request.POST.get('user_id')
        access_level = request.POST.get('access_level', 'staff')

        try:
            user = User.objects.get(pk=user_id, company=store.company, is_active=True)

            # Check if user can manage the target user
            if not request.user.can_manage_user(user):
                messages.error(request, f'You cannot grant access to {user.get_full_name()}')
                return redirect('stores:manage_store_access', store_id=store_id)

            # Check if access already exists
            existing_access = StoreAccess.objects.filter(
                user=user,
                store=store,
                is_active=True
            ).first()

            if existing_access:
                messages.warning(request, f'{user.get_full_name()} already has access to this store')
                return redirect('stores:manage_store_access', store_id=store_id)

            # Create new access
            store_access = StoreAccess.objects.create(
                user=user,
                store=store,
                access_level=access_level,
                granted_by=request.user,
                # Set default permissions based on access level
                can_view_sales=True,
                can_create_sales=access_level in ['staff', 'manager', 'admin'],
                can_view_inventory=True,
                can_manage_inventory=access_level in ['manager', 'admin'],
                can_view_reports=access_level in ['manager', 'admin'],
                can_fiscalize=access_level in ['manager', 'admin'],
                can_manage_staff=access_level == 'admin',
            )

            # Also add user to store staff if not already there
            if not store.staff.filter(id=user.id).exists():
                store.staff.add(user)

            # Log the action using your AuditLog model
            from accounts.models import AuditLog
            AuditLog.log(
                action='store_access_granted',
                user=request.user,
                description=f"Granted {access_level} access to {store.name} for {user.get_full_name()}",
                store=store,
                metadata={
                    'granted_to_user_id': user.id,
                    'access_level': access_level,
                }
            )

            messages.success(
                request,
                f'Access granted to {user.get_full_name()} ({store_access.get_access_level_display()})'
            )

        except User.DoesNotExist:
            messages.error(request, 'User not found')
        except Exception as e:
            logger.error(f"Error creating store access: {str(e)}", exc_info=True)
            messages.error(request, f'Error granting access: {str(e)}')

        return redirect('stores:manage_store_access', store_id=store_id)

    def _can_manage_access(self, store, user):
        """Check if user can manage access for this store"""
        if user.is_saas_admin:
            return True
        if not store.company or store.company != user.company:
            return False

        is_store_manager = store.store_managers.filter(id=user.id).exists()
        return (user.company_admin or
                is_store_manager)  # Changed from company_administrator to company_admin

class EditStoreAccessView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """View for editing store access permissions"""
    model = StoreAccess
    template_name = 'stores/edit_access.html'
    context_object_name = 'store_access'
    permission_required = 'stores.change_storeaccess'
    fields = [
        'access_level',
        'can_view_sales',
        'can_create_sales',
        'can_view_inventory',
        'can_manage_inventory',
        'can_view_reports',
        'can_fiscalize',
        'can_manage_staff',
    ]

    def get_object(self, queryset=None):
        store_id = self.kwargs['store_id']
        user_id = self.kwargs['user_id']

        store_access = get_object_or_404(
            StoreAccess,
            store_id=store_id,
            user_id=user_id,
            is_active=True
        )

        # Validate store access
        store = store_access.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied:
            raise Http404("Access not found")

        # Check if user can edit this access
        if not self._can_edit_access(store_access):
            raise PermissionDenied("You do not have permission to edit this access")

        return store_access

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        store_access = self.object
        context.update({
            'store': store_access.store,
            'user': store_access.user,
            'access_level_choices': StoreAccess.ACCESS_LEVELS,
        })
        return context

    def form_valid(self, form):
        store_access = form.instance

        # Log the changes
        changes = []
        for field in form.changed_data:
            old_value = getattr(store_access, field)
            new_value = form.cleaned_data[field]
            changes.append(f"{field}: {old_value} → {new_value}")

        response = super().form_valid(form)

        # Log the action
        from accounts.models import AuditLog
        AuditLog.log(
            action='store_access_updated',
            user=self.request.user,
            description=f"Updated access permissions for {store_access.user.get_full_name()} in {store_access.store.name}",
            store=store_access.store,
            metadata={
                'updated_user_id': store_access.user.id,
                'changes': changes,
                'new_access_level': store_access.access_level,
            }
        )

        messages.success(
            self.request,
            f'Access permissions updated for {store_access.user.get_full_name()}'
        )

        return response

    def get_success_url(self):
        return reverse('stores:manage_store_access', kwargs={'store_id': self.object.store.id})

    def _can_edit_access(self, store_access):
        """Check if user can edit this access record"""
        user = self.request.user
        store = store_access.store

        if user.is_saas_admin:
            return True

        # User must be in the same company
        if not store.company or store.company != user.company:
            return False

        # Check if user is a store manager
        is_store_manager = store.store_managers.filter(id=user.id).exists()

        # Company admins and store managers can edit access
        if user.is_company_owner or user.company_admin or is_store_manager:
            # But they cannot edit their own access if they're not company owner
            if store_access.user == user and not user.is_company_owner:
                return False
            return True

        return False


class RevokeStoreAccessView(LoginRequiredMixin, PermissionRequiredMixin, View):
    """View for revoking store access"""
    permission_required = 'stores.change_storeaccess'

    def post(self, request, store_id, user_id):
        store = get_object_or_404(Store, pk=store_id)

        # Validate store access
        try:
            validate_store_access(request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(request, str(e))
            return redirect('stores:store_detail', pk=store_id)

        try:
            store_access = StoreAccess.objects.get(
                store_id=store_id,
                user_id=user_id,
                is_active=True
            )

            # Check if user can revoke this access
            if not self._can_revoke_access(store_access, request.user):
                messages.error(request, 'You do not have permission to revoke this access')
                return redirect('stores:manage_store_access', store_id=store_id)

            # Prevent self-revocation (unless company owner)
            if store_access.user == request.user and not request.user.is_company_owner:
                messages.error(request, 'You cannot revoke your own access')
                return redirect('stores:manage_store_access', store_id=store_id)

            user_name = store_access.user.get_full_name()

            # Revoke the access
            store_access.revoke(revoked_by=request.user)

            # Also remove user from store staff if they're not a store manager
            if not store.store_managers.filter(id=user_id).exists():
                store.staff.remove(store_access.user)

            messages.success(request, f'Access revoked for {user_name}')

        except StoreAccess.DoesNotExist:
            messages.error(request, 'Access record not found')
        except Exception as e:
            logger.error(f"Error revoking store access: {str(e)}", exc_info=True)
            messages.error(request, f'Error revoking access: {str(e)}')

        return redirect('stores:manage_store_access', store_id=store_id)

    def _can_revoke_access(self, store_access, user):
        """Check if user can revoke this access"""
        store = store_access.store

        if user.is_saas_admin:
            return True

        # User must be in the same company
        if not store.company or store.company != user.company:
            return False

        # Check if user is a store manager
        is_store_manager = store.store_managers.filter(id=user.id).exists()

        # Company admins and store managers can revoke access
        if user.is_company_owner or user.company_admin or is_store_manager:
            # But they cannot revoke their own access if they're not company owner
            if store_access.user == user and not user.is_company_owner:
                return False
            return True

        return False


# --- Operating Hours Views ---

class StoreOperatingHoursCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create store operating hours."""
    model = StoreOperatingHours
    form_class = StoreOperatingHoursForm
    template_name = 'stores/operating_hours_form.html'
    permission_required = 'stores.add_storeoperatinghours'

    def get_success_url(self):
        return reverse('stores:store_detail', kwargs={'pk': self.object.store.pk})

    def form_valid(self, form):
        # Validate store access
        store = form.instance.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, 'Operating hours added successfully!')
        return super().form_valid(form)


class StoreOperatingHoursUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update store operating hours."""
    model = StoreOperatingHours
    form_class = StoreOperatingHoursForm
    template_name = 'stores/operating_hours_form.html'
    permission_required = 'stores.change_storeoperatinghours'

    def get_success_url(self):
        return reverse('stores:store_detail', kwargs={'pk': self.object.store.pk})

    def form_valid(self, form):
        # Validate store access
        store = form.instance.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, 'Operating hours updated successfully!')
        return super().form_valid(form)


# --- Device Views ---

class StoreDeviceListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List store devices."""
    model = StoreDevice
    template_name = 'stores/device_list.html'
    context_object_name = 'devices'
    paginate_by = 25
    permission_required = 'stores.view_storedevice'

    def get_queryset(self):
        # Get accessible stores
        stores = get_user_accessible_stores(self.request.user)
        return StoreDevice.objects.filter(
            store__in=stores
        ).select_related('store__company').order_by('-registered_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        devices = self.get_queryset()
        context['stats'] = {
            'total_devices': devices.count(),
            'active_devices': devices.filter(is_active=True).count(),
            'pos_devices': devices.filter(device_type='POS').count(),
            'needs_maintenance': devices.filter(last_maintenance__lt=timezone.now() - timedelta(days=90)).count(),
        }
        return context


class StoreDeviceCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new store device."""
    model = StoreDevice
    form_class = StoreDeviceForm
    template_name = 'stores/device_form.html'
    permission_required = 'stores.add_storedevice'
    success_url = reverse_lazy('stores:device_list')

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Validate store access
        store = form.instance.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, f'Device "{form.instance.name}" created successfully!')
        return super().form_valid(form)


class StoreDeviceUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update a store device."""
    model = StoreDevice
    form_class = StoreDeviceForm
    template_name = 'stores/device_form.html'
    permission_required = 'stores.change_storedevice'

    def get_success_url(self):
        return reverse('stores:device_detail', kwargs={'pk': self.object.pk})

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        # Validate store access
        store = form.instance.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, f'Device "{form.instance.name}" updated successfully!')
        return super().form_valid(form)


class StoreDeviceDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Detailed view of a store device."""
    model = StoreDevice
    template_name = 'stores/device_detail.html'
    context_object_name = 'device'
    permission_required = 'stores.view_storedevice'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)

        # Validate store access
        try:
            validate_store_access(self.request.user, obj.store, action='view', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            raise

        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recent_logs'] = DeviceOperatorLog.objects.filter(
            device=self.object
        ).select_related('user').order_by('-timestamp')[:20]
        return context


@login_required
@permission_required('stores.change_storedevice')
def device_maintenance_update(request, device_id):
    """Update device maintenance date."""
    device = get_object_or_404(StoreDevice, pk=device_id)

    # Validate store access
    try:
        validate_store_access(request.user, device.store, action='change', raise_exception=True)
    except PermissionDenied:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    if request.method == 'POST':
        device.last_maintenance = timezone.now()
        device.save()
        DeviceOperatorLog.objects.create(
            user=request.user,
            action='MAINTENANCE',
            device=device,
            store=device.store,
            details={'updated_by': request.user.username}
        )
        messages.success(request, f'Maintenance date updated for {device.name}')
        return JsonResponse({'status': 'success'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=400)


# --- Inventory Views ---

class StoreInventoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List store inventory items - optimized for template compatibility."""
    model = Stock
    template_name = 'stores/inventory_list.html'
    context_object_name = 'inventory_items'
    paginate_by = 30
    permission_required = 'inventory.view_stock'

    def get_queryset(self):
        # Get accessible stores
        stores = get_user_accessible_stores(self.request.user)

        queryset = Stock.objects.filter(
            store__in=stores
        ).select_related(
            'store',
            'product'
        ).order_by('store__name', 'product__name')

        # Apply filters
        if store_id := self.request.GET.get('store'):
            queryset = queryset.filter(store_id=store_id)
        if search := self.request.GET.get('search'):
            queryset = queryset.filter(
                Q(product__name__icontains=search) |
                Q(product__sku__icontains=search) |
                Q(store__name__icontains=search)
            )
        if self.request.GET.get('low_stock') == 'true':
            queryset = queryset.filter(quantity__lte=F('low_stock_threshold'))

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        queryset = self.get_queryset()

        # Get low stock items for the summary section
        low_stock_queryset = queryset.filter(quantity__lte=F('low_stock_threshold'))

        context.update({
            'stores': get_user_accessible_stores(self.request.user).filter(is_active=True).order_by('name'),
            'selected_store': self.request.GET.get('store'),
            'search_query': self.request.GET.get('search', ''),
            'low_stock_filter': self.request.GET.get('low_stock'),
            'low_stock_items': low_stock_queryset,
            'total_items': queryset.count(),
            'low_stock_count': low_stock_queryset.count(),
            'total_quantity': queryset.aggregate(total=Sum('quantity'))['total'] or 0,
            'total_value': queryset.aggregate(total=Sum(F('quantity') * F('product__cost_price')))['total'] or 0,
        })
        return context


class StoreInventoryDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Detailed view of a store inventory item."""
    model = Stock
    template_name = 'stores/inventory_detail.html'
    context_object_name = 'inventory_item'
    permission_required = 'inventory.view_stock'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)

        # Validate store access
        try:
            validate_store_access(self.request.user, obj.store, action='view', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            raise

        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recent_movements'] = StockMovement.objects.filter(
            product=self.object.product, store=self.object.store
        ).select_related('created_by').order_by('-created_at')[:10]
        return context


class StoreInventoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, CreateView):
    """Create a new store inventory item."""
    model = Stock
    template_name = 'stores/inventory_form.html'
    success_url = reverse_lazy('stores:inventory_list')
    permission_required = 'inventory.add_stock'
    fields = ['product', 'store', 'quantity', 'low_stock_threshold', 'reorder_quantity']

    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Filter stores to only accessible ones
        form.fields['store'].queryset = get_user_accessible_stores(self.request.user)
        return form

    def form_valid(self, form):
        # Validate store access
        store = form.instance.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, 'Store inventory item created successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


class StoreInventoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, UpdateView):
    """Update a store inventory item."""
    model = Stock
    template_name = 'stores/inventory_form.html'
    success_url = reverse_lazy('stores:inventory_list')
    permission_required = 'inventory.change_stock'
    fields = ['quantity', 'low_stock_threshold', 'reorder_quantity']

    def form_valid(self, form):
        # Validate store access
        store = self.object.store
        try:
            validate_store_access(self.request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(self.request, str(e))
            return self.form_invalid(form)

        messages.success(self.request, 'Store inventory item updated successfully!')
        return super().form_valid(form)

    def form_invalid(self, form):
        messages.error(self.request, 'Please correct the errors below.')
        return super().form_invalid(form)


class StoreInventoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    """Delete a store inventory item."""
    model = Stock
    template_name = 'stores/inventory_confirm_delete.html'
    success_url = reverse_lazy('stores:inventory_list')
    permission_required = 'inventory.delete_stock'

    def delete(self, request, *args, **kwargs):
        # Validate store access
        store = self.get_object().store
        try:
            validate_store_access(request.user, store, action='change', raise_exception=True)
        except PermissionDenied as e:
            messages.error(request, str(e))
            return redirect('stores:inventory_detail', pk=self.get_object().pk)

        messages.success(request, 'Store inventory item deleted successfully!')
        return super().delete(request, *args, **kwargs)


# --- API Views ---

@login_required
@permission_required('inventory.view_stock')
@require_http_methods(["GET"])
def inventory_search_api(request):
    """AJAX endpoint for inventory search."""
    query = request.GET.get('q', '')
    store_id = request.GET.get('store_id')

    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    inventory_items = Stock.objects.filter(
        product__name__icontains=query,
        store__in=stores
    ).select_related('product', 'store')

    if store_id:
        # Validate store access
        try:
            store = get_object_or_404(Store, id=store_id)
            if store not in stores:
                return JsonResponse({'error': 'Access denied'}, status=403)
            inventory_items = inventory_items.filter(store_id=store_id)
        except Store.DoesNotExist:
            return JsonResponse({'error': 'Store not found'}, status=404)

    return JsonResponse({
        'results': [{
            'id': item.id,
            'product_name': item.product.name,
            'store_name': item.store.name,
            'quantity': str(item.quantity),
            'is_low_stock': item.quantity <= item.low_stock_threshold,
            'url': reverse('stores:inventory_detail', kwargs={'pk': item.pk})
        } for item in inventory_items[:10]]
    })


@login_required
@permission_required('inventory.change_stock')
@require_http_methods(["POST"])
def quick_quantity_update(request, pk):
    """AJAX endpoint for quick quantity updates."""
    inventory_item = get_object_or_404(Stock, pk=pk)

    # Validate store access
    try:
        validate_store_access(request.user, inventory_item.store, action='change', raise_exception=True)
    except PermissionDenied:
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        new_quantity = float(request.POST.get('quantity', 0))
        if new_quantity < 0:
            return JsonResponse({'error': 'Quantity cannot be negative'}, status=400)
        inventory_item.quantity = new_quantity
        inventory_item.save()
        return JsonResponse({
            'success': True,
            'new_quantity': str(inventory_item.quantity),
            'is_low_stock': inventory_item.quantity <= inventory_item.low_stock_threshold,
            'message': 'Quantity updated successfully'
        })
    except (ValueError, TypeError):
        return JsonResponse({'error': 'Invalid quantity value'}, status=400)
    except Exception as e:
        logger.error(f"Error updating quantity: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=500)


@login_required
@permission_required('inventory.view_stock')
def low_stock_alert_api(request):
    """API endpoint for low stock alerts."""
    store_id = request.GET.get('store_id')

    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    queryset = Stock.objects.filter(
        quantity__lte=F('low_stock_threshold'),
        store__in=stores
    ).select_related('product', 'store')

    if store_id:
        # Validate store access
        try:
            store = get_object_or_404(Store, id=store_id)
            if store not in stores:
                return JsonResponse({'error': 'Access denied'}, status=403)
            queryset = queryset.filter(store_id=store_id)
        except Store.DoesNotExist:
            return JsonResponse({'error': 'Store not found'}, status=404)

    return JsonResponse({
        'alerts': [{
            'id': item.id,
            'product_name': item.product.name,
            'store_name': item.store.name,
            'current_quantity': str(item.quantity),
            'threshold': str(item.low_stock_threshold),
            'reorder_quantity': str(item.reorder_quantity),
            'shortage': str(max(0, item.low_stock_threshold - item.quantity)),
        } for item in queryset[:20]],
        'count': len(queryset[:20]),
        'total_low_stock': queryset.count()
    })


@login_required
@permission_required('inventory.view_stock')
def low_stock_alert(request):
    """View for low stock alerts across all stores."""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    low_stock_items = Stock.objects.filter(
        quantity__lte=F('low_stock_threshold'),
        store__in=stores
    ).select_related('store', 'product', 'product__category').order_by('quantity')

    # Convert to the format expected by your template
    items_with_computations = []
    for stock_item in low_stock_items:
        # Calculate computed fields
        total_cost = stock_item.quantity * stock_item.product.cost_price
        reorder_gap = stock_item.low_stock_threshold - stock_item.quantity
        stock_percentage = (
                stock_item.quantity / stock_item.low_stock_threshold * 100
        ) if stock_item.low_stock_threshold > 0 else 0
        recommended_order_qty = max(0, (stock_item.low_stock_threshold * Decimal('1.5')) - stock_item.quantity)

        items_with_computations.append({
            'stock': stock_item,
            'total_cost': total_cost,
            'reorder_gap': reorder_gap,
            'half_reorder': stock_item.low_stock_threshold / 2,
            'stock_percentage': min(100, max(0, round(stock_percentage, 1))),
            'recommended_order_qty': recommended_order_qty.quantize(Decimal('0.01'))
        })

    # Group by store for display
    stores_with_alerts = {}
    for item in items_with_computations:
        store_name = item['stock'].store.name
        stores_with_alerts.setdefault(store_name, {'store': item['stock'].store, 'items': []})
        stores_with_alerts[store_name]['items'].append(item)

    context = {
        'stores_with_alerts': stores_with_alerts,
        'total_low_stock_items': len(items_with_computations),
        'low_stock_items': items_with_computations,
    }
    return render(request, 'stores/low_stock_alert.html', context)


# --- Analytics and Reporting Views ---

@login_required
@permission_required('stores.view_store')
def store_analytics(request):
    """Advanced analytics view for stores."""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user)

    analytics_data = {
        'store_performance': [{
            'name': store.name,
            'inventory_value': float(
                store.inventory_items.aggregate(total=Sum(F('quantity') * F('product__cost_price')))['total'] or 0
            ),
            'device_count': store.devices.filter(is_active=True).count(),
            'staff_count': store.staff.filter(is_hidden=False).count(),
            'manager_count': store.store_managers.count(),
            'low_stock_items': store.inventory_items.filter(quantity__lte=F('low_stock_threshold')).count(),
            'efris_enabled': store.efris_enabled,
            'is_main_branch': store.is_main_branch,
        } for store in stores],
        'inventory_summary': Stock.objects.filter(store__in=stores).aggregate(
            total_items=Sum('quantity'),
            low_stock_count=Count('id', filter=Q(quantity__lte=F('low_stock_threshold')))
        ),
        'device_status': StoreDevice.objects.filter(store__in=stores).aggregate(
            total=Count('id'),
            active=Count('id', filter=Q(is_active=True)),
            pos_devices=Count('id', filter=Q(device_type='POS')),
            invoice_printers=Count('id', filter=Q(device_type='INVOICE'))
        ),
        'regional_distribution': list(
            stores.values('region').annotate(store_count=Count('id')).order_by('-store_count'))
    }

    context = {
        'analytics_data': analytics_data,
        'stores': stores,
        'total_stores': stores.count(),
    }
    return render(request, 'stores/analytics.html', context)


# --- Map and API Views ---

@login_required
def store_map_view(request):
    """Interactive map view of all store locations with enhanced features."""

    accessible_stores = get_user_accessible_stores(request.user)

    stores_with_coordinates = accessible_stores.filter(
        latitude__isnull=False,
        longitude__isnull=False
    ).select_related('company').annotate(
        inventory_count=Count('inventory_items'),
        low_stock_count=Count('inventory_items', filter=Q(
            inventory_items__quantity__lte=F('inventory_items__low_stock_threshold')
        )),
        total_inventory_value=Sum(
            F('inventory_items__quantity') * F('inventory_items__product__cost_price')
        ),
        device_count=Count('devices', filter=Q(devices__is_active=True)),
        staff_count=Count('staff', filter=Q(
            staff__is_active=True,
            staff__is_hidden=False
        ))
    ).values(
        'id', 'name', 'code', 'physical_address', 'latitude', 'longitude',
        'phone', 'email', 'region', 'store_type', 'is_main_branch',
        'manager_name', 'manager_phone', 'efris_enabled', 'efris_device_number',
        'inventory_count', 'low_stock_count', 'total_inventory_value',
        'device_count', 'staff_count'
    )

    # Get stores without coordinates for the unmapped list
    stores_without_coordinates = accessible_stores.filter(
        Q(latitude__isnull=True) | Q(longitude__isnull=True)
    ).select_related('company').values(
        'id', 'name', 'code', 'region', 'physical_address', 'phone',
        'store_type', 'is_main_branch', 'manager_name'
    )

    # Get region statistics
    region_stats = accessible_stores.filter(
        region__isnull=False
    ).exclude(region='').values('region').annotate(
        store_count=Count('id'),
        mapped_count=Count('id', filter=Q(
            latitude__isnull=False,
            longitude__isnull=False
        )),
        main_branch_count=Count('id', filter=Q(is_main_branch=True))
    ).order_by('-store_count')

    # Get store type statistics
    store_type_stats = accessible_stores.values('store_type').annotate(
        count=Count('id'),
        mapped_count=Count('id', filter=Q(
            latitude__isnull=False,
            longitude__isnull=False
        ))
    ).order_by('-count')

    # Calculate map center (average of all coordinates)
    if stores_with_coordinates:
        avg_lat = stores_with_coordinates.aggregate(Avg('latitude'))['latitude__avg']
        avg_lng = stores_with_coordinates.aggregate(Avg('longitude'))['longitude__avg']
        map_center = [float(avg_lat) if avg_lat else 0.3476, float(avg_lng) if avg_lng else 32.5825]
    else:
        map_center = [0.3476, 32.5825]  # Default to Kampala, Uganda

    # Count EFRIS enabled stores
    efris_enabled_count = accessible_stores.filter(
        efris_enabled=True,
        is_registered_with_efris=True
    ).count()

    context = {
        'stores_data': json.dumps(list(stores_with_coordinates), cls=DjangoJSONEncoder),
        'unmapped_stores': list(stores_without_coordinates),
        'total_stores': accessible_stores.count(),
        'mapped_stores': len(stores_with_coordinates),
        'unmapped_count': len(stores_without_coordinates),
        'regions_count': region_stats.count(),
        'region_stats': list(region_stats),
        'store_type_stats': list(store_type_stats),
        'map_center': map_center,
        'efris_enabled_count': efris_enabled_count,
        'main_branch_count': accessible_stores.filter(is_main_branch=True).count(),
    }
    return render(request, 'stores/store_map.html', context)


@login_required
def nearest_stores_api(request):
    try:
        # Get required parameters
        lat = float(request.GET.get('lat'))
        lon = float(request.GET.get('lon'))

        # Get optional parameters
        limit = int(request.GET.get('limit', 5))
        max_distance = request.GET.get('max_distance')
        store_type = request.GET.get('store_type')
        efris_only = request.GET.get('efris_only', '').lower() == 'true'

        if max_distance:
            max_distance = float(max_distance)

        # Validate parameters
        if not (-90 <= lat <= 90) or not (-180 <= lon <= 180):
            return JsonResponse({
                'status': 'error',
                'message': 'Invalid latitude or longitude'
            }, status=400)

        if limit < 1 or limit > 100:
            return JsonResponse({
                'status': 'error',
                'message': 'Limit must be between 1 and 100'
            }, status=400)

        # Get accessible stores
        accessible_stores = get_user_accessible_stores(request.user)

        # Build base queryset
        base_queryset = accessible_stores.filter(
            latitude__isnull=False,
            longitude__isnull=False
        ).select_related('company').annotate(
            inventory_count=Count('inventory_items'),
            low_stock_count=Count('inventory_items', filter=Q(
                inventory_items__quantity__lte=F('inventory_items__low_stock_threshold')
            )),
            device_count=Count('devices', filter=Q(devices__is_active=True))
        )

        # Apply store type filter
        if store_type:
            base_queryset = base_queryset.filter(store_type=store_type)

        # Apply EFRIS filter
        if efris_only:
            base_queryset = base_queryset.filter(
                efris_enabled=True,
                is_registered_with_efris=True
            )

        # Calculate distances for all stores
        stores_with_distance = []
        for store in base_queryset:
            distance = store.distance_to(lat, lon)
            if distance is not None:
                # Apply distance filter if specified
                if max_distance is None or distance <= max_distance:
                    stores_with_distance.append({
                        'id': store.id,
                        'name': store.name,
                        'code': store.code,
                        'store_type': store.get_store_type_display(),
                        'store_type_code': store.store_type,
                        'is_main_branch': store.is_main_branch,
                        'address': store.physical_address,
                        'region': store.region,
                        'phone': store.phone,
                        'email': store.email,
                        'latitude': float(store.latitude),
                        'longitude': float(store.longitude),
                        'distance_km': distance,
                        'distance_miles': round(distance * 0.621371, 2),
                        'manager_name': store.manager_name,
                        'manager_phone': store.manager_phone,
                        'efris_enabled': store.efris_enabled,
                        'efris_status': store.efris_status,
                        'is_open_now': store.is_open_now(),
                        'inventory': {
                            'total_products': store.inventory_count,
                            'low_stock_count': store.low_stock_count,
                        },
                        'devices': {
                            'total': store.device_count
                        },
                        'map_url': store.get_map_url(),
                        'directions_url': store.get_directions_url(lat, lon),
                    })

        # Sort by distance
        stores_with_distance.sort(key=lambda x: x['distance_km'])

        # Limit results
        nearest_stores = stores_with_distance[:limit]

        # Calculate some statistics
        if stores_with_distance:
            avg_distance = sum(s['distance_km'] for s in stores_with_distance) / len(stores_with_distance)
            closest_distance = stores_with_distance[0]['distance_km'] if stores_with_distance else None
        else:
            avg_distance = None
            closest_distance = None

        return JsonResponse({
            'status': 'success',
            'search_location': {
                'latitude': lat,
                'longitude': lon
            },
            'filters': {
                'max_distance_km': max_distance,
                'store_type': store_type,
                'efris_only': efris_only
            },
            'results': {
                'stores': nearest_stores,
                'total_found': len(stores_with_distance),
                'returned': len(nearest_stores),
                'average_distance_km': round(avg_distance, 2) if avg_distance else None,
                'closest_distance_km': closest_distance
            }
        })

    except (ValueError, TypeError) as e:
        return JsonResponse({
            'status': 'error',
            'message': f'Invalid parameter: {str(e)}'
        }, status=400)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': 'An error occurred while processing your request'
        }, status=500)


@login_required
@csrf_exempt
def store_api_data(request):
    """Enhanced API endpoint for store data (for AJAX requests)."""
    # Get accessible stores
    accessible_stores = get_user_accessible_stores(request.user)

    action = request.GET.get('action', '')

    if action == 'store_locations':
        stores = accessible_stores.filter(
            latitude__isnull=False,
            longitude__isnull=False
        ).select_related('company').annotate(
            inventory_count=Count('inventory_items'),
            low_stock_count=Count('inventory_items', filter=Q(
                inventory_items__quantity__lte=F('inventory_items__low_stock_threshold')
            )),
            device_count=Count('devices', filter=Q(devices__is_active=True))
        ).values(
            'id', 'name', 'code', 'latitude', 'longitude',
            'physical_address', 'region', 'phone', 'store_type',
            'is_main_branch', 'efris_enabled', 'manager_name',
            'inventory_count', 'low_stock_count', 'device_count'
        )
        return JsonResponse({'stores': list(stores)})

    elif action == 'store_inventory_summary':
        store_id = request.GET.get('store_id')
        try:
            store = accessible_stores.get(pk=store_id)
        except Store.DoesNotExist:
            return JsonResponse({'error': 'Store not found or access denied'}, status=403)

        inventory = store.inventory_items.select_related('product')

        total_value = inventory.aggregate(
            total=Sum(F('quantity') * F('product__cost_price'))
        )['total'] or 0

        low_stock_items = inventory.filter(
            quantity__lte=F('low_stock_threshold')
        ).values('product__name', 'quantity', 'low_stock_threshold')[:5]

        # Get device information
        devices = store.devices.filter(is_active=True).values(
            'name', 'device_type', 'device_number', 'is_efris_linked'
        )

        return JsonResponse({
            'store_name': store.name,
            'store_code': store.code,
            'store_type': store.get_store_type_display(),
            'is_main_branch': store.is_main_branch,
            'manager_name': store.manager_name,
            'manager_phone': store.manager_phone,
            'efris_enabled': store.efris_enabled,
            'efris_status': store.efris_status,
            'total_products': inventory.count(),
            'low_stock_count': len(low_stock_items),
            'low_stock_items': list(low_stock_items),
            'total_value': float(total_value),
            'devices': list(devices),
            'last_updated': inventory.order_by('-last_updated').first().last_updated.isoformat()
            if inventory.exists() else None
        })

    elif action == 'search_stores':
        query = request.GET.get('q', '')
        stores = accessible_stores.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(region__icontains=query) |
            Q(physical_address__icontains=query) |
            Q(manager_name__icontains=query)
        ).select_related('company').values(
            'id', 'name', 'code', 'region', 'is_active',
            'latitude', 'longitude', 'store_type', 'is_main_branch'
        )[:15]
        return JsonResponse({'stores': list(stores)})

    elif action == 'region_stores':
        region = request.GET.get('region', '')
        stores = accessible_stores.filter(
            region=region,
            latitude__isnull=False,
            longitude__isnull=False
        ).values(
            'id', 'name', 'code', 'latitude', 'longitude',
            'physical_address', 'store_type', 'is_main_branch'
        )
        return JsonResponse({'stores': list(stores)})

    elif action == 'store_type_filter':
        store_type = request.GET.get('store_type', '')
        stores = accessible_stores.filter(
            store_type=store_type,
            latitude__isnull=False,
            longitude__isnull=False
        ).values(
            'id', 'name', 'code', 'latitude', 'longitude',
            'physical_address', 'region', 'is_main_branch'
        )
        return JsonResponse({'stores': list(stores)})

    return JsonResponse({'error': 'Invalid action'}, status=400)


@login_required
def store_details_api(request, store_id):
    """Get detailed information about a specific store for the map."""
    # Get accessible stores
    accessible_stores = get_user_accessible_stores(request.user)

    try:
        store = accessible_stores.select_related('company').annotate(
            inventory_count=Count('inventory_items'),
            low_stock_count=Count('inventory_items', filter=Q(
                inventory_items__quantity__lte=F('inventory_items__low_stock_threshold')
            )),
            out_of_stock_count=Count('inventory_items', filter=Q(
                inventory_items__quantity=0
            )),
            total_inventory_value=Sum(
                F('inventory_items__quantity') * F('inventory_items__product__cost_price')
            ),
            device_count=Count('devices', filter=Q(devices__is_active=True)),
            efris_device_count=Count('devices', filter=Q(
                devices__is_active=True,
                devices__is_efris_linked=True
            ))
        ).get(pk=store_id)

        # Get recent sales (if user has permission)
        recent_sales = []
        if request.user.has_perm('sales.view_sale'):
            recent_sales = store.sales.order_by('-date')[:5].values(
                'id', 'invoice_number', 'date', 'total_amount', 'payment_status'
            )

        # Get top products
        top_products = store.inventory_items.select_related('product').order_by(
            '-quantity'
        )[:10].values(
            'product__name', 'quantity', 'low_stock_threshold'
        )

        # Get devices
        devices = store.devices.filter(is_active=True).values(
            'id', 'name', 'device_type', 'device_number',
            'serial_number', 'is_efris_linked'
        )

        # Operating hours
        operating_hours = None
        if store.operating_hours:
            operating_hours = store.operating_hours

        data = {
            'id': store.id,
            'name': store.name,
            'code': store.code,
            'store_type': store.get_store_type_display(),
            'is_main_branch': store.is_main_branch,
            'region': store.region,
            'physical_address': store.physical_address,
            'phone': store.phone,
            'secondary_phone': store.secondary_phone,
            'email': store.email,
            'latitude': float(store.latitude) if store.latitude else None,
            'longitude': float(store.longitude) if store.longitude else None,
            'is_active': store.is_active,
            'manager_name': store.manager_name,
            'manager_phone': store.manager_phone,
            'operating_hours': operating_hours,
            'is_open_now': store.is_open_now(),
            'efris_info': {
                'enabled': store.efris_enabled,
                'device_number': store.efris_device_number,
                'registered': store.is_registered_with_efris,
                'status': store.efris_status,
                'can_fiscalize': store.can_fiscalize,
                'last_sync': store.efris_last_sync.isoformat() if store.efris_last_sync else None
            },
            'inventory_stats': {
                'total_products': store.inventory_count,
                'low_stock': store.low_stock_count,
                'out_of_stock': store.out_of_stock_count,
                'total_value': float(store.total_inventory_value or 0)
            },
            'device_stats': {
                'total_devices': store.device_count,
                'efris_devices': store.efris_device_count,
                'devices': list(devices)
            },
            'recent_sales': list(recent_sales),
            'top_products': list(top_products)
        }

        return JsonResponse(data)
    except Store.DoesNotExist:
        return JsonResponse({'error': 'Store not found or access denied'}, status=404)


# --- Device Log Views ---

class DeviceOperatorLogListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    """List device operator logs."""
    model = DeviceOperatorLog
    template_name = 'stores/device_logs.html'
    context_object_name = 'logs'
    paginate_by = 50
    permission_required = 'stores.view_deviceoperatorlog'

    def get_queryset(self):
        # Get accessible stores
        accessible_stores = get_user_accessible_stores(self.request.user)

        queryset = DeviceOperatorLog.objects.filter(
            device__store__in=accessible_stores
        ).select_related('user', 'device__store').order_by('-timestamp')

        if store_id := self.request.GET.get('store'):
            queryset = queryset.filter(device__store_id=store_id)
        if date_from := self.request.GET.get('date_from'):
            try:
                queryset = queryset.filter(timestamp__gte=datetime.strptime(date_from, '%Y-%m-%d'))
            except ValueError:
                pass
        if date_to := self.request.GET.get('date_to'):
            try:
                queryset = queryset.filter(timestamp__lte=datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1))
            except ValueError:
                pass
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'stores': get_user_accessible_stores(self.request.user).filter(is_active=True).order_by('name'),
            'selected_store': self.request.GET.get('store'),
            'date_from': self.request.GET.get('date_from'),
            'date_to': self.request.GET.get('date_to')
        })
        return context


# --- Utility Views ---

@login_required
@permission_required('stores.view_store')
def export_stores_data(request):
    """Export all store data in CSV format."""
    # Get accessible stores
    stores = get_user_accessible_stores(request.user).select_related('company').prefetch_related('staff', 'devices',
                                                                                                 'inventory_items')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="stores_export_{datetime.now().strftime("%Y%m%d")}.csv"'
    writer = csv.writer(response)
    writer.writerow([
        'ID', 'Name', 'Code', 'Company', 'Address', 'Region', 'Phone',
        'Email', 'EFRIS Enabled', 'Status', 'Staff Count', 'Manager Count', 'Device Count',
        'Inventory Items', 'Created At', 'Updated At'
    ])

    for store in stores:
        writer.writerow([
            store.id,
            store.name,
            store.code,
            store.company.name if store.company else '',
            store.physical_address,
            store.region or '',
            store.phone or '',
            store.email or '',
            'Yes' if store.efris_enabled else 'No',
            'Active' if store.is_active else 'Inactive',
            store.staff.filter(is_hidden=False).count(),
            store.store_managers.count(),
            store.devices.count(),
            store.inventory_items.count(),
            store.created_at.strftime('%Y-%m-%d %H:%M') if store.created_at else '',
            store.updated_at.strftime('%Y-%m-%d %H:%M') if store.updated_at else ''
        ])
    return response


# --- Helper Functions for Reporting (kept for compatibility) ---

def _get_selected_stores(store_select_value, user):
    """Get stores based on user selection and permissions."""
    accessible_stores = get_user_accessible_stores(user)

    if not store_select_value or store_select_value == 'all':
        return accessible_stores

    try:
        store_ids = [int(s.strip()) for s in store_select_value.split(',') if s.strip().isdigit()]
        return accessible_stores.filter(id__in=store_ids) if store_ids else accessible_stores.none()
    except Exception as e:
        logger.error(f"Invalid store selection string '{store_select_value}': {e}")
        return accessible_stores.none()


def _validate_report_request(report_data, user):
    """Validate report request data."""
    required_fields = ['report_type', 'store_select', 'start_date', 'end_date', 'export_format']
    for field in required_fields:
        if not report_data.get(field):
            return {'valid': False, 'error': f'{field.replace("_", " ").title()} is required'}

    try:
        # ✅ FIXED: The dates are already date objects from form cleaning
        # No need to parse them with strptime
        start_date = report_data['start_date']
        end_date = report_data['end_date']

        if start_date > end_date:
            return {'valid': False, 'error': 'Start date must be before end date'}
        if (end_date - start_date).days > 365:
            return {'valid': False, 'error': 'Date range cannot exceed 1 year'}
    except (ValueError, TypeError) as e:
        # Handle cases where dates might not be valid date objects
        logger.error(f"Date validation error: {str(e)}")
        return {'valid': False, 'error': 'Invalid date'}

    # Validate store access
    accessible_stores = get_user_accessible_stores(user)
    if report_data['store_select'] != 'all':
        try:
            store_ids = [int(s.strip()) for s in report_data['store_select'].split(',') if s.strip().isdigit()]
            if not accessible_stores.filter(id__in=store_ids).exists():
                return {'valid': False, 'error': 'Access denied to selected store(s)'}
        except (ValueError, TypeError):
            return {'valid': False, 'error': 'Invalid store selection'}

    return {'valid': True}

def _get_report_statistics(user):
    """Get statistics for the report dashboard."""
    accessible_stores = get_user_accessible_stores(user)

    return {
        'total_stores': accessible_stores.count(),
        'active_stores': accessible_stores.filter(is_active=True).count(),
        'total_devices': StoreDevice.objects.filter(
            store__in=accessible_stores,
            is_active=True
        ).count(),
        'low_stock_items': Stock.objects.filter(
            store__in=accessible_stores,
            quantity__lte=F('low_stock_threshold')
        ).count(),
    }

@login_required
@permission_required('stores.view_store')
def generate_store_report(request):
    """Enhanced store report generation with multiple formats and types."""
    if request.method == 'POST':
        form = EnhancedStoreReportForm(request.POST, user=request.user)
        if form.is_valid():
            report_data = form.cleaned_data
            stores = _get_selected_stores(report_data['store_select'], request.user)

            if not stores.exists():
                messages.error(request, 'No accessible stores found')
                return redirect('stores:generate_report')

            validation = _validate_report_request(report_data, request.user)
            if not validation['valid']:
                messages.error(request, validation['error'])
                return redirect('stores:generate_report')

            try:
                if report_data['export_format'] == 'csv':
                    return _generate_csv_report(stores, report_data)
                elif report_data['export_format'] == 'excel':
                    if not EXCEL_AVAILABLE:
                        messages.error(request, 'Excel export not available. Please install openpyxl.')
                        return redirect('stores:generate_report')
                    return _generate_excel_report(stores, report_data)
                elif report_data['export_format'] == 'pdf':
                    if not PDF_AVAILABLE:
                        messages.error(request, 'PDF export not available. Please install reportlab.')
                        return redirect('stores:generate_report')
                    return _generate_pdf_report(stores, report_data)
                messages.error(request, 'Invalid export format')
            except Exception as e:
                logger.error(f"Error generating report: {str(e)}", exc_info=True)
                messages.error(request, f'Report generation failed: {str(e)}')
        else:
            messages.error(request, 'Invalid form data.')
        return redirect('stores:generate_report')

    context = {
        'form': EnhancedStoreReportForm(user=request.user),
        # ✅ FIXED: Use permission-based filtering
        'available_stores': filter_stores_by_permissions(request.user, action='view'),
        'report_stats': _get_report_statistics(request.user),  # Pass user
        'excel_available': EXCEL_AVAILABLE,
        'pdf_available': PDF_AVAILABLE,
    }
    return render(request, 'stores/generate_report.html', context)


@login_required
@permission_required('stores.view_store')
def export_report_direct(request, report_type):
    """Direct export endpoint for quick reports."""
    # ✅ FIXED: Use permission-based filtering
    stores = filter_stores_by_permissions(request.user, action='view')

    if not stores.exists():
        messages.error(request, "No accessible stores found.")
        return redirect('stores:dashboard')

    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    report_data = {
        'report_type': report_type,
        'stores': stores,
        'start_date': start_date,
        'end_date': end_date,
        'include_charts': True,
        'include_summary': True,
        'include_raw_data': False,
        'include_images': False,
        'detailed_breakdown': False,
        'compare_periods': False,
    }
    return _generate_csv_report(stores, report_data)

def _generate_csv_report(stores, report_data):
    """Generate CSV report."""
    response = HttpResponse(content_type='text/csv')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"store_report_{report_data['report_type']}_{timestamp}.csv"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)

    report_functions = {
        'store_summary': _write_store_summary_csv,
        'inventory': _write_inventory_csv,
        'operating_hours': _write_operating_hours_csv,
        'device_status': _write_device_status_csv,
        'staff_assignment': _write_staff_assignment_csv,
        'comprehensive': _write_comprehensive_csv
    }

    if report_data['report_type'] in report_functions:
        report_functions[report_data['report_type']](writer, stores, report_data)
    else:
        writer.writerow(['Error: Invalid report type'])
    return response


def _write_store_summary_csv(writer, stores, report_data):
    """Write store summary data to CSV."""
    writer.writerow(['=== STORE SUMMARY REPORT ==='])
    writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([f"Period: {report_data['start_date']} to {report_data['end_date']}"])
    writer.writerow([])
    headers = ['Store Name', 'Code', 'Company', 'Region', 'Status', 'Phone', 'Email', 'Address', 'EFRIS Enabled',
               'Staff Count', 'Device Count', 'Inventory Value', 'Created Date']
    writer.writerow(headers)

    for store in stores:
        inventory_value = Stock.objects.filter(store=store).aggregate(
            total=Sum(F('quantity') * F('product__cost_price'))
        )['total'] or 0
        writer.writerow([
            store.name,
            store.code or '',
            store.company.name if store.company else '',
            store.region or '',
            'Active' if store.is_active else 'Inactive',
            store.phone or '',
            store.email or '',
            store.physical_address or '',
            'Yes' if store.efris_enabled else 'No',
            store.staff.filter(is_hidden=False).count(),
            store.devices.filter(is_active=True).count(),
            f"{inventory_value:,.2f}",
            store.created_at.strftime('%Y-%m-%d') if store.created_at else ''
        ])


def _write_inventory_csv(writer, stores, report_data):
    """Write inventory report data to CSV."""
    writer.writerow(['=== INVENTORY REPORT ==='])
    writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([f"Period: {report_data['start_date']} to {report_data['end_date']}"])
    writer.writerow([])
    headers = ['Store Name', 'Product Name', 'SKU', 'Category', 'Current Stock', 'Low Stock Threshold',
               'Reorder Quantity', 'Unit Price', 'Total Value', 'Last Updated', 'Status']
    writer.writerow(headers)

    for store in stores:
        for item in Stock.objects.filter(store=store).select_related('product', 'product__category'):
            status = 'Out of Stock' if item.quantity == 0 else (
                'Low Stock' if item.quantity <= item.low_stock_threshold else 'In Stock')
            total_value = item.quantity * item.product.cost_price
            writer.writerow([
                store.name,
                item.product.name,
                item.product.sku or '',
                item.product.category.name if item.product.category else '',
                item.quantity,
                item.low_stock_threshold,
                item.reorder_quantity,
                f"{item.product.selling_price:,.2f}",
                f"{total_value:,.2f}",
                item.last_updated.strftime('%Y-%m-%d %H:%M') if item.last_updated else '',
                status
            ])


def _write_operating_hours_csv(writer, stores, report_data):
    """Write operating hours data to CSV."""
    writer.writerow(['=== OPERATING HOURS REPORT ==='])
    writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])
    headers = ['Store Name', 'Day of Week', 'Opening Time', 'Closing Time', 'Is Open', 'Break Start', 'Break End',
               'Special Notes']
    writer.writerow(headers)
    days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']

    for store in stores:
        if isinstance(store.operating_hours, dict):
            for day in days:
                day_data = store.operating_hours.get(day, {})
                writer.writerow([
                    store.name,
                    day.capitalize(),
                    day_data.get('open', 'N/A'),
                    day_data.get('close', 'N/A'),
                    'Yes' if day_data.get('is_open', False) else 'No',
                    day_data.get('break_start', 'N/A'),
                    day_data.get('break_end', 'N/A'),
                    day_data.get('notes', '')
                ])
        else:
            for day in days:
                writer.writerow(
                    [store.name, day.capitalize(), 'Not configured', 'Not configured', 'Unknown', '', '', ''])


def _write_device_status_csv(writer, stores, report_data):
    """Write device status data to CSV."""
    writer.writerow(['=== DEVICE STATUS REPORT ==='])
    writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])
    headers = ['Store Name', 'Device Name', 'Device Type', 'Serial Number', 'Status', 'Registered Date',
               'Last Maintenance', 'EFRIS Enabled', 'Device ID', 'Notes']
    writer.writerow(headers)

    for store in stores:
        devices = store.devices.all()
        if not devices.exists():
            writer.writerow([store.name, 'No devices', '-', '-', '-', '-', '-', '-', '-', '-'])
        for device in devices:
            writer.writerow([
                store.name,
                device.name,
                device.get_device_type_display() if hasattr(device, 'get_device_type_display') else device.device_type,
                device.serial_number or '',
                'Active' if device.is_active else 'Inactive',
                device.registered_at.strftime('%Y-%m-%d') if device.registered_at else '',
                device.last_maintenance.strftime('%Y-%m-%d') if device.last_maintenance else 'Never',
                'Yes' if getattr(device, 'efris_enabled', False) else 'No',
                device.device_id or '',
                device.notes or ''
            ])


def _write_staff_assignment_csv(writer, stores, report_data):
    """Write staff assignment data to CSV."""
    writer.writerow(['=== STAFF ASSIGNMENT REPORT ==='])
    writer.writerow([f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    writer.writerow([])
    headers = ['Store Name', 'Staff Name', 'Email', 'User Type', 'Phone', 'Is Active', 'Date Joined', 'Last Login']
    writer.writerow(headers)

    for store in stores:
        staff = store.staff.filter(is_hidden=False)
        if not staff.exists():
            writer.writerow([store.name, 'No staff assigned', '-', '-', '-', '-', '-', '-'])
        for member in staff:
            writer.writerow([
                store.name,
                member.get_full_name() or member.username,
                member.email,
                member.display_role if hasattr(member, 'get_user_type_display') else 'Staff',
                getattr(member, 'phone_number', ''),
                'Active' if member.is_active else 'Inactive',
                member.date_joined.strftime('%Y-%m-%d') if member.date_joined else '',
                member.last_login.strftime('%Y-%m-%d %H:%M') if member.last_login else 'Never'
            ])


def _write_comprehensive_csv(writer, stores, report_data):
    """Write comprehensive report combining all data types."""
    for func in [_write_store_summary_csv, _write_inventory_csv, _write_operating_hours_csv,
                 _write_device_status_csv, _write_staff_assignment_csv]:
        func(writer, stores, report_data)
        writer.writerow([])


def _generate_excel_report(stores, report_data):
    """Generate Excel report with multiple worksheets."""
    if not EXCEL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if report_data['report_type'] == 'comprehensive':
        for sheet_type, populate_func in [
            ('Summary', _populate_excel_summary_sheet),
            ('Inventory', _populate_excel_inventory_sheet),
            ('Devices', _populate_excel_devices_sheet),
            ('Staff', _populate_excel_staff_sheet)
        ]:
            ws = wb.create_sheet(sheet_type)
            populate_func(ws, stores, report_data)
    else:
        ws = wb.create_sheet(report_data['report_type'].replace('_', ' ').title())
        {
            'store_summary': _populate_excel_summary_sheet,
            'inventory': _populate_excel_inventory_sheet,
            'device_status': _populate_excel_devices_sheet,
            'staff_assignment': _populate_excel_staff_sheet
        }.get(report_data['report_type'], lambda x, y, z: None)(ws, stores, report_data)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response[
        'Content-Disposition'] = f'attachment; filename="store_report_{report_data["report_type"]}_{timestamp}.xlsx"'
    return response


def _populate_excel_summary_sheet(ws, stores, report_data):
    """Populate Excel summary sheet."""
    ws['A1'] = 'Store Summary Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A3'] = f"Period: {report_data['start_date']} to {report_data['end_date']}"
    headers = ['Store Name', 'Code', 'Status', 'Region', 'Phone', 'Staff', 'Devices', 'Inventory Value']
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    for store in stores:
        inventory_value = Stock.objects.filter(store=store).aggregate(
            total=Sum(F('quantity') * F('product__cost_price'))
        )['total'] or 0
        ws.append([
            store.name,
            store.code or '',
            'Active' if store.is_active else 'Inactive',
            store.region or '',
            store.phone or '',
            store.staff.filter(is_hidden=False).count(),
            store.devices.filter(is_active=True).count(),
            inventory_value
        ])

    for column in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)


def _populate_excel_inventory_sheet(ws, stores, report_data):
    """Populate Excel inventory sheet."""
    ws['A1'] = 'Inventory Report'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    headers = ['Store', 'Product', 'SKU', 'Category', 'Quantity', 'Threshold', 'Status', 'Value']
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")

    for store in stores:
        for item in Stock.objects.filter(store=store).select_related('product', 'product__category'):
            status = 'Out of Stock' if item.quantity == 0 else (
                'Low Stock' if item.quantity <= item.low_stock_threshold else 'In Stock')
            ws.append([
                store.name,
                item.product.name,
                item.product.sku or '',
                item.product.category.name if item.product.category else '',
                item.quantity,
                item.low_stock_threshold,
                status,
                item.quantity * item.product.cost_price
            ])


def _populate_excel_devices_sheet(ws, stores, report_data):
    """Populate Excel devices sheet."""
    ws['A1'] = 'Device Status Report'
    ws['A1'].font = Font(bold=True, size=16)
    headers = ['Store', 'Device Name', 'Type', 'Serial Number', 'Status', 'Last Maintenance']
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    for store in stores:
        for device in store.devices.all():
            ws.append([
                store.name,
                device.name,
                device.device_type,
                device.serial_number or '',
                'Active' if device.is_active else 'Inactive',
                device.last_maintenance.strftime('%Y-%m-%d') if device.last_maintenance else 'Never'
            ])


def _populate_excel_staff_sheet(ws, stores, report_data):
    """Populate Excel staff sheet."""
    ws['A1'] = 'Staff Assignment Report'
    ws['A1'].font = Font(bold=True, size=16)
    headers = ['Store', 'Name', 'Email', 'User Type', 'Status']
    ws.append([])
    ws.append(headers)

    header_row = ws.max_row
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")

    for store in stores:
        for member in store.staff.filter(is_hidden=False):
            ws.append([
                store.name,
                member.get_full_name() or member.username,
                member.email,
                member.display_role if hasattr(member, 'get_user_type_display') else 'Staff',
                'Active' if member.is_active else 'Inactive'
            ])


def _generate_pdf_report(stores, report_data):
    """Generate PDF report with professional formatting."""
    if not PDF_AVAILABLE:
        raise ImportError("reportlab is required for PDF export")

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    story = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=30, alignment=1)

    story.append(Paragraph(f"Store Report - {report_data['report_type'].replace('_', ' ').title()}", title_style))
    story.append(Spacer(1, 20))
    story.append(Table([
        ['Generated on:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Report type:', report_data['report_type'].replace('_', ' ').title()],
        ['Date range:', f"{report_data['start_date']} to {report_data['end_date']}"],
        ['Total stores:', str(len(stores))]
    ], style=TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0)
    ])))
    story.append(Spacer(1, 20))

    if report_data['report_type'] == 'store_summary':
        _add_pdf_store_summary(story, stores, styles)
    elif report_data['report_type'] == 'inventory':
        _add_pdf_inventory_report(story, stores, styles)
    elif report_data['report_type'] == 'comprehensive':
        _add_pdf_comprehensive_report(story, stores, styles)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response[
        'Content-Disposition'] = f'attachment; filename="store_report_{report_data["report_type"]}_{timestamp}.pdf"'
    return response


def _add_pdf_store_summary(story, stores, styles):
    """Add store summary table to PDF."""
    story.append(Paragraph("Store Summary", styles['Heading2']))
    story.append(Spacer(1, 12))
    data = [['Store Name', 'Status', 'Region', 'Staff', 'Devices']]
    for store in stores:
        data.append([
            Paragraph(store.name[:30], styles['Normal']),
            'Active' if store.is_active else 'Inactive',
            store.region or 'N/A',
            str(store.staff.filter(is_hidden=False).count()),
            str(store.devices.filter(is_active=True).count())
        ])

    table = Table(data, colWidths=[200, 80, 100, 60, 60])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    story.append(table)


def _add_pdf_inventory_report(story, stores, styles):
    """Add inventory report to PDF."""
    story.append(Paragraph("Inventory Status Report", styles['Heading2']))
    story.append(Spacer(1, 12))
    story.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
    story.append(Spacer(1, 20))

    for store in stores:
        story.append(Paragraph(f"Store: {store.name} ({store.code or 'N/A'})", styles['Heading3']))
        story.append(Spacer(1, 10))
        inventory = Stock.objects.filter(store=store).select_related('product', 'product__category').order_by(
            'product__name')[:50]

        if inventory.exists():
            data = [['Product', 'SKU', 'Category', 'Qty', 'Threshold', 'Status', 'Unit Price', 'Value']]
            total_value = 0
            low_stock_count = 0
            out_of_stock_count = 0

            for item in inventory:
                status = 'OUT' if item.quantity == 0 else ('LOW' if item.quantity <= item.low_stock_threshold else 'OK')
                if item.quantity == 0:
                    out_of_stock_count += 1
                elif item.quantity <= item.low_stock_threshold:
                    low_stock_count += 1
                item_value = item.quantity * item.product.cost_price
                total_value += item_value
                data.append([
                    Paragraph(item.product.name[:35] + ('...' if len(item.product.name) > 35 else ''),
                              styles['Normal']),
                    item.product.sku[:15] if item.product.sku else 'N/A',
                    item.product.category.name[:12] if item.product.category else 'N/A',
                    str(int(item.quantity)),
                    str(int(item.low_stock_threshold)),
                    status,
                    f"{item.product.cost_price:,.0f}",
                    f"{item_value:,.0f}"
                ])

            table = Table(data, colWidths=[140, 60, 60, 35, 45, 35, 55, 70])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#70AD47')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('ALIGN', (1, 1), (1, -1), 'LEFT'),
                ('ALIGN', (2, 1), (2, -1), 'LEFT'),
                ('ALIGN', (3, 1), (-1, -1), 'CENTER'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('LEFTPADDING', (0, 0), (-1, -1), 4),
                ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                ('TOPPADDING', (0, 1), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 4)
            ]))
            story.append(table)
            story.append(Spacer(1, 12))

            summary_data = [
                ['Total Items:', str(inventory.count())],
                ['Low Stock Items:', str(low_stock_count)],
                ['Out of Stock Items:', str(out_of_stock_count)],
                ['Displayed Items:', str(len(inventory))],
                ['Total Inventory Value:', f"UGX {total_value:,.2f}"]
            ]
            summary_table = Table(summary_data, colWidths=[150, 200])
            summary_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),
                ('TEXTCOLOR', (1, 1), (1, 1), colors.orange if low_stock_count > 0 else colors.black),
                ('TEXTCOLOR', (1, 2), (1, 2), colors.red if out_of_stock_count > 0 else colors.black),
                ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ('RIGHTPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4)
            ]))
            story.append(summary_table)

            if inventory.count() > 50:
                story.append(Spacer(1, 8))
                story.append(Paragraph(
                    f"Note: Showing first 50 items out of {inventory.count()} total items.",
                    ParagraphStyle('Note', parent=styles['Normal'], fontSize=8, textColor=colors.grey, leftIndent=20)
                ))
        else:
            story.append(Paragraph("No inventory items found for this store.", styles['Normal']))
        story.append(Spacer(1, 25))


def _add_pdf_comprehensive_report(story, stores, styles):
    """Add comprehensive report sections to PDF."""
    _add_pdf_store_summary(story, stores, styles)
    story.append(Spacer(1, 30))
    _add_pdf_inventory_report(story, stores, styles)
    story.append(Spacer(1, 30))
    story.append(Paragraph("Summary Statistics", styles['Heading2']))
    story.append(Spacer(1, 12))

    total_staff = sum(store.staff.filter(is_hidden=False).count() for store in stores)
    total_devices = sum(store.devices.filter(is_active=True).count() for store in stores)
    total_inventory_value = sum(
        Stock.objects.filter(store=store).aggregate(total=Sum(F('quantity') * F('product__cost_price')))['total'] or 0
        for store in stores
    )

    summary_data = [
        ['Total Stores:', str(len(stores))],
        ['Total Staff:', str(total_staff)],
        ['Total Devices:', str(total_devices)],
        ['Total Inventory Value:', f"UGX {total_inventory_value:,.2f}"]
    ]
    summary_table = Table(summary_data, colWidths=[200, 300])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 10),
        ('RIGHTPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8)
    ]))
    story.append(summary_table)

