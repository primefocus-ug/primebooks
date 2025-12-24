import csv
import logging
from io import BytesIO, StringIO
from decimal import Decimal, InvalidOperation
from datetime import datetime

import xlsxwriter
import openpyxl
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db import transaction
from django.core.exceptions import ValidationError
from difflib import get_close_matches

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

from .models import Customer, CustomerGroup
from stores.models import Store

logger = logging.getLogger(__name__)


# ============================================================================
# SAMPLE FILE GENERATION VIEWS
# ============================================================================

@login_required
def download_sample_customers_csv(request):
    """Generate CSV sample file for customers - UPDATED VERSION"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="sample_customers.csv"'

    writer = csv.writer(response)

    # Headers with all fields
    headers = [
        'Name*', 'Customer Type*', 'Phone*', 'Email', 'TIN', 'NIN', 'BRN',
        'Physical Address', 'Postal Address', 'District', 'Country',
        'Is VAT Registered', 'Credit Limit', 'Store Name*',
        'Passport Number', 'Driving License', 'Voter ID', 'Alien ID'
    ]
    writer.writerow(headers)

    # Sample data rows with new requirements
    sample_data = [
        # Individual - only name and phone required
        [
            'John Doe', 'INDIVIDUAL', '+256700123456', 'john.doe@email.com', '',
            'CM1234567890ABC', '', '123 Main Street, Kampala', 'P.O. Box 1234',
            'Kampala', 'Uganda', 'No', '0', 'Main Store',
            '', '', '', ''
        ],
        # Business - requires TIN
        [
            'ABC Technologies Ltd', 'BUSINESS', '+256700987654', 'info@abctech.com',
            '1000123456', '', 'BN2023/12345', 'Plot 45, Industrial Area',
            'P.O. Box 5678', 'Kampala', 'Uganda', 'Yes', '5000000', 'Main Store',
            '', '', '', ''
        ],
        # Government - requires TIN
        [
            'Ministry of Health', 'GOVERNMENT', '+256700111222', 'moh@gov.ug',
            '9000111222', '', 'GOVT-MOH-2023', 'Plot 6, Lourdel Road',
            'P.O. Box 7272', 'Kampala', 'Uganda', 'No', '0', 'Main Store',
            '', '', '', ''
        ],
        # NGO - requires TIN (not BRN)
        [
            'Save The Children Uganda', 'NGO', '+256700333444', 'info@stc.org',
            'NGO-2020/789', '', '', 'Plot 15, Kololo', 'P.O. Box 4444',
            'Kampala', 'Uganda', 'No', '1000000', 'Main Store',
            '', '', '', ''
        ],
        # Individual tourist - only name and phone required
        [
            'Jane Smith (Tourist)', 'INDIVIDUAL', '+447700123456', 'jane@email.com',
            '', '', '', 'Hotel Africana, Kampala', '',
            'Kampala', 'United Kingdom', 'No', '0', 'Downtown Branch',
            'AB1234567', '', '', ''
        ],
        # Business without BRN - allowed, only TIN required
        [
            'Small Business Ltd', 'BUSINESS', '+256700555666', 'info@smallbusiness.com',
            '1000999888', '', '', 'Plot 22, Nakawa', 'P.O. Box 9999',
            'Kampala', 'Uganda', 'No', '100000', 'Downtown Branch',
            '', '', '', ''
        ],
    ]

    for row in sample_data:
        writer.writerow(row)

    return response

@login_required
def download_sample_customers_excel(request):
    """Generate Excel sample file with formatting and instructions"""
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)

    # Create worksheets
    data_sheet = workbook.add_worksheet('Customer Data')
    instructions_sheet = workbook.add_worksheet('Instructions')
    reference_sheet = workbook.add_worksheet('Reference Data')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4F46E5',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    required_format = workbook.add_format({
        'bold': True,
        'bg_color': '#DC2626',
        'font_color': 'white',
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'text_wrap': True
    })

    sample_format = workbook.add_format({
        'border': 1,
        'align': 'left',
        'valign': 'vcenter'
    })

    instruction_header = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'font_color': '#1F2937'
    })

    instruction_text = workbook.add_format({
        'text_wrap': True,
        'valign': 'top'
    })

    # ========================================================================
    # DATA SHEET
    # ========================================================================

    headers = [
        ('Name*', True),
        ('Customer Type*', True),
        ('Phone*', True),
        ('Email', False),
        ('TIN', False),
        ('NIN', False),
        ('BRN', False),
        ('Physical Address', False),
        ('Postal Address', False),
        ('District', False),
        ('Country', False),
        ('Is VAT Registered', False),
        ('Credit Limit', False),
        ('Store Name*', True),
        ('Passport Number', False),
        ('Driving License', False),
        ('Voter ID', False),
        ('Alien ID', False),
        ('EFRIS Customer Type', False),
        ('Auto Sync EFRIS', False),
    ]

    # Write headers
    for col, (header, is_required) in enumerate(headers):
        if is_required:
            data_sheet.write(0, col, header, required_format)
        else:
            data_sheet.write(0, col, header, header_format)

    # Sample data
    sample_data = [
        [
            'John Doe', 'INDIVIDUAL', '+256700123456', 'john.doe@email.com', '',
            'CM1234567890ABC', '', '123 Main Street, Kampala', 'P.O. Box 1234',
            'Kampala', 'Uganda', 'No', 0, 'Main Store',
            '', '', '', '', '1', 'Yes'
        ],
        [
            'ABC Technologies Ltd', 'BUSINESS', '+256700987654', 'info@abctech.com',
            '1000123456', '', 'BN2023/12345', 'Plot 45, Industrial Area',
            'P.O. Box 5678', 'Kampala', 'Uganda', 'Yes', 5000000, 'Main Store',
            '', '', '', '', '2', 'Yes'
        ],
        [
            'Ministry of Health', 'GOVERNMENT', '+256700111222', 'moh@gov.ug',
            '9000111222', '', 'GOVT-MOH-2023', 'Plot 6, Lourdel Road',
            'P.O. Box 7272', 'Kampala', 'Uganda', 'No', 0, 'Main Store',
            '', '', '', '', '3', 'Yes'
        ],
        [
            'Save The Children Uganda', 'NGO', '+256700333444', 'info@stc.org',
            '', '', 'NGO-2020/789', 'Plot 15, Kololo', 'P.O. Box 4444',
            'Kampala', 'Uganda', 'No', 1000000, 'Main Store',
            '', '', '', '', '4', 'Yes'
        ],
        [
            'Jane Smith (Tourist)', 'INDIVIDUAL', '+447700123456', 'jane@email.com',
            '', '', '', 'Hotel Africana, Kampala', '',
            'Kampala', 'United Kingdom', 'No', 0, 'Downtown Branch',
            'AB1234567', '', '', '', '1', 'No'
        ],
    ]

    for row_idx, row_data in enumerate(sample_data, start=1):
        for col_idx, value in enumerate(row_data):
            data_sheet.write(row_idx, col_idx, value, sample_format)

    # Set column widths
    column_widths = [20, 15, 15, 25, 15, 20, 15, 30, 20, 15, 15, 15, 15, 20, 18, 18, 15, 15, 18, 15]
    for col, width in enumerate(column_widths):
        data_sheet.set_column(col, col, width)

    # Freeze first row
    data_sheet.freeze_panes(1, 0)

    # ========================================================================
    # INSTRUCTIONS SHEET
    # ========================================================================

    instructions_sheet.set_column(0, 0, 50)
    instructions_sheet.set_column(1, 1, 60)

    row = 0
    instructions_sheet.write(row, 0, 'CUSTOMER IMPORT INSTRUCTIONS', instruction_header)
    row += 2

    instructions = [
        ('Required Fields:', 'Fields marked with * (red header) are mandatory:'),
        ('', '- Name: Full customer name'),
        ('', '- Customer Type: INDIVIDUAL, BUSINESS, GOVERNMENT, or NGO'),
        ('', '- Phone: Valid phone number (e.g., +256700123456)'),
        ('', '- Store Name: Must match an existing store in your system'),
        ('', ''),
        ('Customer Type Requirements:', ''),
        ('', '- INDIVIDUAL: Only name and phone required'),
        ('', '- BUSINESS: Name, phone, and TIN required'),
        ('', '- GOVERNMENT: Name, phone, and TIN required'),
        ('', '- NGO: Name, phone, and TIN required'),
        ('', ''),
        ('TIN Requirements:', 'TIN is required for BUSINESS, GOVERNMENT, and NGO customers'),
        ('', 'BRN is optional for all customer types'),
        ('', ''),
        ('Identification Numbers:', 'All identification fields are optional:'),
        ('', '- TIN: Tax Identification Number (required for Business/Government/NGO)'),
        ('', '- NIN: National ID Number (optional for individuals)'),
        ('', '- BRN: Business Registration Number (optional)'),
        ('', '- Passport Number: For foreign nationals (optional)'),
        ('', '- Driving License, Voter ID, Alien ID: Alternative IDs (optional)'),
        ('', ''),
        ('VAT Registration:', ''),
        ('', 'Use "Yes" or "No" for VAT registered status'),
        ('', ''),
        ('Credit Limit:', 'Enter numeric value (default: 0)'),
        ('', 'Maximum credit amount allowed for this customer'),
        ('', ''),
        ('Import Behavior:', ''),
        ('', '- Existing customers (matched by phone) can be updated'),
        ('', '- New customers will be created'),
        ('', '- Invalid rows will be skipped with error messages'),
        ('', ''),
        ('Validation Rules:', ''),
        ('', '- All customers: Name and phone required'),
        ('', '- Business/Government/NGO: TIN required'),
        ('', '- All other fields are optional'),
        ('', '- Store names must match exactly (case-insensitive)'),
        ('', ''),
        ('Tips:', ''),
        ('', '- Keep phone numbers unique for each customer'),
        ('', '- Use consistent formatting'),
        ('', '- Test with a few rows first'),
        ('', '- Review validation results before final import'),
    ]

    for instruction in instructions:
        instructions_sheet.write(row, 0, instruction[0],
                                 instruction_header if instruction[0] and instruction[0].endswith(
                                     ':') else instruction_text)
        instructions_sheet.write(row, 1, instruction[1], instruction_text)
        row += 1

    # ========================================================================
    # REFERENCE DATA SHEET
    # ========================================================================

    reference_sheet.set_column(0, 0, 20)
    reference_sheet.set_column(1, 1, 50)

    row = 0
    reference_sheet.write(row, 0, 'CUSTOMER TYPES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Code', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    customer_types = [
        ('INDIVIDUAL', 'Individual customers - requires only name and phone'),
        ('BUSINESS', 'Registered businesses - requires name, phone, and TIN'),
        ('GOVERNMENT', 'Government agencies - requires name, phone, and TIN'),
        ('NGO', 'Non-profit organizations - requires name, phone, and TIN'),
    ]

    for code, desc in customer_types:
        reference_sheet.write(row, 0, code, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    # EFRIS Customer Types
    row += 2
    reference_sheet.write(row, 0, 'EFRIS TYPES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Code', header_format)
    reference_sheet.write(row, 1, 'Description', header_format)
    row += 1

    efris_types = [
        ('1', 'Individual'),
        ('2', 'Business'),
        ('3', 'Government'),
        ('4', 'NGO'),
    ]

    for code, desc in efris_types:
        reference_sheet.write(row, 0, code, sample_format)
        reference_sheet.write(row, 1, desc, sample_format)
        row += 1

    # Yes/No Values
    row += 2
    reference_sheet.write(row, 0, 'YES/NO VALUES', instruction_header)
    row += 1
    reference_sheet.write(row, 0, 'Accepted Values', header_format)
    reference_sheet.write(row, 1, 'Meaning', header_format)
    row += 1

    yes_no = [
        ('Yes, yes, Y, y, 1, True, true', 'Treated as YES/TRUE'),
        ('No, no, N, n, 0, False, false', 'Treated as NO/FALSE'),
    ]

    for value, meaning in yes_no:
        reference_sheet.write(row, 0, value, sample_format)
        reference_sheet.write(row, 1, meaning, sample_format)
        row += 1

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="sample_customers.xlsx"'

    return response


# ============================================================================
# EXPORT VIEWS
# ============================================================================

@login_required
def export_customers_csv(request):
    """Export customers to CSV"""
    # Get filter parameters
    customer_type = request.GET.get('customer_type')
    efris_status = request.GET.get('efris_status')
    is_active = request.GET.get('is_active')
    store_id = request.GET.get('store')

    # Build queryset
    queryset = Customer.objects.select_related('store').all()

    if customer_type:
        queryset = queryset.filter(customer_type=customer_type)
    if efris_status:
        queryset = queryset.filter(efris_status=efris_status)
    if is_active:
        queryset = queryset.filter(is_active=is_active == '1')
    if store_id:
        queryset = queryset.filter(store_id=store_id)

    response = HttpResponse(content_type='text/csv')
    filename = f'customers_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)

    # Headers
    writer.writerow([
        'Customer ID', 'Name', 'Customer Type', 'Phone', 'Email',
        'TIN', 'NIN', 'BRN', 'Physical Address', 'Postal Address',
        'District', 'Country', 'VAT Registered', 'Credit Limit',
        'Store', 'Active', 'EFRIS Status', 'EFRIS Customer ID',
        'EFRIS Registered At', 'Created At'
    ])

    # Data
    for customer in queryset:
        writer.writerow([
            customer.customer_id,
            customer.name,
            customer.get_customer_type_display(),
            customer.phone,
            customer.email or '',
            customer.tin or '',
            customer.nin or '',
            customer.brn or '',
            customer.physical_address or '',
            customer.postal_address or '',
            customer.district or '',
            customer.country,
            'Yes' if customer.is_vat_registered else 'No',
            float(customer.credit_limit),
            customer.store.name,
            'Yes' if customer.is_active else 'No',
            customer.get_efris_status_display(),
            customer.efris_customer_id or '',
            customer.efris_registered_at.strftime('%Y-%m-%d %H:%M') if customer.efris_registered_at else '',
            customer.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    return response


@login_required
def export_customers_excel(request):
    """Export customers to Excel with formatting"""
    # Get filter parameters
    customer_type = request.GET.get('customer_type')
    efris_status = request.GET.get('efris_status')
    is_active = request.GET.get('is_active')
    store_id = request.GET.get('store')

    # Build queryset
    queryset = Customer.objects.select_related('store').all()

    if customer_type:
        queryset = queryset.filter(customer_type=customer_type)
    if efris_status:
        queryset = queryset.filter(efris_status=efris_status)
    if is_active:
        queryset = queryset.filter(is_active=is_active == '1')
    if store_id:
        queryset = queryset.filter(store_id=store_id)

    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)
    worksheet = workbook.add_worksheet('Customers')

    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'bg_color': '#4F46E5',
        'font_color': 'white',
        'border': 1,
        'align': 'center'
    })

    cell_format = workbook.add_format({
        'border': 1,
        'align': 'left'
    })

    number_format = workbook.add_format({
        'border': 1,
        'align': 'right',
        'num_format': '#,##0.00'
    })

    # Headers
    headers = [
        'Customer ID', 'Name', 'Type', 'Phone', 'Email', 'TIN', 'NIN', 'BRN',
        'Physical Address', 'District', 'VAT Reg', 'Credit Limit', 'Store',
        'Active', 'EFRIS Status', 'EFRIS ID', 'Created At'
    ]

    for col, header in enumerate(headers):
        worksheet.write(0, col, header, header_format)

    # Data
    row = 1
    for customer in queryset:
        data = [
            customer.customer_id,
            customer.name,
            customer.get_customer_type_display(),
            customer.phone,
            customer.email or '',
            customer.tin or '',
            customer.nin or '',
            customer.brn or '',
            customer.physical_address or '',
            customer.district or '',
            'Yes' if customer.is_vat_registered else 'No',
            float(customer.credit_limit),
            customer.store.name,
            'Yes' if customer.is_active else 'No',
            customer.get_efris_status_display(),
            customer.efris_customer_id or '',
            customer.created_at.strftime('%Y-%m-%d %H:%M'),
        ]

        for col, value in enumerate(data):
            if col == 11:  # Credit Limit
                worksheet.write(row, col, value, number_format)
            else:
                worksheet.write(row, col, value, cell_format)

        row += 1

    # Auto-adjust column widths
    for col, header in enumerate(headers):
        worksheet.set_column(col, col, len(header) + 5)

    workbook.close()
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'customers_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


@login_required
def export_customers_pdf(request):
    """Export customers to PDF"""
    # Get filter parameters
    customer_type = request.GET.get('customer_type')
    efris_status = request.GET.get('efris_status')
    is_active = request.GET.get('is_active')
    store_id = request.GET.get('store')

    # Build queryset
    queryset = Customer.objects.select_related('store').all()

    if customer_type:
        queryset = queryset.filter(customer_type=customer_type)
    if efris_status:
        queryset = queryset.filter(efris_status=efris_status)
    if is_active:
        queryset = queryset.filter(is_active=is_active == '1')
    if store_id:
        queryset = queryset.filter(store_id=store_id)

    response = HttpResponse(content_type='application/pdf')
    filename = f'customers_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1
    )

    story = []
    story.append(Paragraph("Customer List", title_style))
    story.append(Paragraph(
        f"Generated on: {timezone.now().strftime('%B %d, %Y at %I:%M %p')}",
        styles['Normal']
    ))
    story.append(Spacer(1, 20))

    # Create table data
    data = [['Name', 'Type', 'Phone', 'Store', 'VAT', 'EFRIS Status', 'Active']]

    for customer in queryset[:100]:  # Limit to 100 for PDF
        data.append([
            customer.name[:25] + ('...' if len(customer.name) > 25 else ''),
            customer.get_customer_type_display()[:10],
            customer.phone,
            customer.store.name[:15] + ('...' if len(customer.store.name) > 15 else ''),
            'Yes' if customer.is_vat_registered else 'No',
            customer.get_efris_status_display()[:12],
            'Yes' if customer.is_active else 'No',
        ])

    # Create table
    table = Table(data, colWidths=[2 * inch, 1 * inch, 1.3 * inch, 1.5 * inch, 0.6 * inch, 1.2 * inch, 0.7 * inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    story.append(table)

    # Add summary
    if queryset.count() > 100:
        story.append(Spacer(1, 20))
        story.append(Paragraph(
            f"Note: Showing first 100 of {queryset.count()} total customers",
            styles['Italic']
        ))

    doc.build(story)
    return response


# ============================================================================
# IMPORT PROCESSING LOGIC
# ============================================================================

class CustomerColumnMapper:
    """Smart column mapper for customer imports"""

    COLUMN_MAPPINGS = {
        'name': ['name', 'customer name', 'full name', 'customer', 'client name'],
        'customer_type': ['customer type', 'type', 'customer_type', 'client type', 'category'],
        'phone': ['phone', 'phone number', 'mobile', 'telephone', 'contact', 'phone_number'],
        'email': ['email', 'email address', 'e-mail', 'mail'],
        'tin': ['tin', 'tax id', 'tax identification', 'tax number', 'tax_id'],
        'nin': ['nin', 'national id', 'national_id', 'id number', 'id_number'],
        'brn': ['brn', 'business registration', 'registration number', 'reg number', 'business_registration_number'],
        'physical_address': ['physical address', 'address', 'location', 'street address', 'physical_address'],
        'postal_address': ['postal address', 'postal', 'po box', 'p.o. box', 'postal_address'],
        'district': ['district', 'region', 'area'],
        'country': ['country', 'nation'],
        'is_vat_registered': ['vat registered', 'vat', 'is vat registered', 'vat_registered', 'is_vat_registered'],
        'credit_limit': ['credit limit', 'credit', 'limit', 'credit_limit'],
        'store_name': ['store', 'store name', 'branch', 'location', 'shop', 'store_name'],
        'passport_number': ['passport', 'passport number', 'passport_number', 'passport no'],
        'driving_license': ['driving license', 'license', 'driving_license', 'dl number'],
        'voter_id': ['voter id', 'voter_id', 'voters card', 'voter card'],
        'alien_id': ['alien id', 'alien_id', 'alien number'],
        'efris_customer_type': ['efris type', 'efris customer type', 'efris_customer_type'],
        'auto_sync_efris': ['auto sync', 'sync efris', 'auto_sync_efris', 'efris sync'],
    }

    @classmethod
    def map_columns(cls, file_headers):
        """Map file headers to standardized column names"""
        mapped = {}
        file_headers_lower = [h.lower().strip() for h in file_headers]

        for standard_name, variations in cls.COLUMN_MAPPINGS.items():
            for variation in variations:
                if variation in file_headers_lower:
                    idx = file_headers_lower.index(variation)
                    mapped[file_headers[idx]] = standard_name
                    break

            if standard_name not in mapped.values():
                for header in file_headers:
                    matches = get_close_matches(header.lower(), variations, n=1, cutoff=0.8)
                    if matches:
                        mapped[header] = standard_name
                        break

        return mapped

    @classmethod
    def get_mapping_suggestions(cls, file_headers):
        """Get suggestions for unmapped columns"""
        mapped = cls.map_columns(file_headers)
        suggestions = {}

        for header in file_headers:
            if header not in mapped:
                all_variations = []
                for variations in cls.COLUMN_MAPPINGS.values():
                    all_variations.extend(variations)

                matches = get_close_matches(header.lower(), all_variations, n=3, cutoff=0.6)
                if matches:
                    suggestions[header] = matches

        return suggestions


def parse_uploaded_file(file_obj):
    """Parse CSV or Excel file"""
    file_extension = file_obj.name.split('.')[-1].lower()

    if file_extension == 'csv':
        return parse_csv_file(file_obj)
    elif file_extension in ['xlsx', 'xls']:
        return parse_excel_file(file_obj)
    else:
        raise ValueError(f"Unsupported file type: {file_extension}")


def parse_csv_file(file_obj):
    """Parse CSV file"""
    file_obj.seek(0)
    content = file_obj.read().decode('utf-8-sig')
    reader = csv.DictReader(StringIO(content))

    headers = reader.fieldnames
    data = list(reader)

    return headers, data


def parse_excel_file(file_obj):
    """Parse Excel file"""
    file_obj.seek(0)
    workbook = openpyxl.load_workbook(file_obj, read_only=True)
    sheet = workbook.active

    headers = [cell.value for cell in sheet[1]]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

    workbook.close()
    return headers, data


def validate_customer_row(row_data, mapped_columns, row_number):
    """Validate a single row of customer data"""
    errors = []
    cleaned = {}

    # Required fields for all customers
    required_fields = ['name', 'customer_type', 'phone', 'store_name']

    for field in required_fields:
        value = None
        for file_header, standard_name in mapped_columns.items():
            if standard_name == field:
                value = row_data.get(file_header)
                break

        if value is None or (isinstance(value, str) and not value.strip()):
            field_display = field.replace('_', ' ').title()
            errors.append(f"{field_display} is required")
        else:
            cleaned[field] = value

    if errors:
        return False, errors, cleaned

    # Validate and clean optional fields
    for file_header, standard_name in mapped_columns.items():
        if standard_name in cleaned:
            continue

        value = row_data.get(file_header)

        # Skip None values
        if value is None:
            continue

        try:
            # Convert all values to string for consistency
            if isinstance(value, (int, float, Decimal)):
                value = str(value)

            if standard_name == 'customer_type':
                value = value.upper().strip()
                if value not in ['INDIVIDUAL', 'BUSINESS', 'GOVERNMENT', 'NGO']:
                    errors.append(f"Invalid customer type '{value}'. Must be INDIVIDUAL, BUSINESS, GOVERNMENT, or NGO")
                else:
                    cleaned[standard_name] = value

            elif standard_name == 'phone':
                # Basic phone validation
                phone_str = str(value).strip().replace(' ', '').replace('-', '')
                if not phone_str.startswith('+'):
                    if phone_str.startswith('0'):
                        phone_str = '+256' + phone_str[1:]
                    elif phone_str.startswith('256'):
                        phone_str = '+' + phone_str
                    else:
                        phone_str = '+256' + phone_str
                cleaned[standard_name] = phone_str

            elif standard_name == 'email':
                # Basic email validation
                email_str = str(value).strip()
                if email_str and '@' in email_str:
                    cleaned[standard_name] = email_str.lower()
                elif email_str:
                    errors.append(f"Invalid email format: {value}")
                else:
                    cleaned[standard_name] = ''

            elif standard_name in ['is_vat_registered', 'auto_sync_efris']:
                value_str = str(value).lower().strip()
                cleaned[standard_name] = value_str in ['yes', 'true', '1', 'y', 'on']

            elif standard_name == 'credit_limit':
                try:
                    cleaned[standard_name] = Decimal(str(value))
                    if cleaned[standard_name] < 0:
                        errors.append("Credit limit cannot be negative")
                except (ValueError, InvalidOperation):
                    errors.append(f"Invalid credit limit: {value}")

            elif standard_name == 'efris_customer_type':
                value_str = str(value).strip()
                if value_str and value_str not in ['1', '2', '3', '4']:
                    errors.append(f"Invalid EFRIS customer type '{value}'. Must be 1, 2, 3, or 4")
                else:
                    cleaned[standard_name] = value_str

            elif standard_name in ['tin', 'nin', 'brn', 'passport_number',
                                   'driving_license', 'voter_id', 'alien_id']:
                # Convert identification numbers to string and clean
                value_str = str(value).strip() if value else ''
                cleaned[standard_name] = value_str.upper() if value_str else ''

            elif standard_name in ['physical_address', 'postal_address', 'district', 'country', 'store_name']:
                # Convert to string and strip
                cleaned[standard_name] = str(value).strip() if value else ''

            else:
                # For any other fields, just store as string
                cleaned[standard_name] = str(value).strip() if value else ''

        except Exception as e:
            errors.append(f"Error processing {standard_name}: {str(e)}")

    # Business logic validation according to new requirements
    customer_type = cleaned.get('customer_type')
    tin = cleaned.get('tin', '')

    # Business/Government/NGO require TIN
    if customer_type in ['BUSINESS', 'GOVERNMENT', 'NGO']:
        if not tin or not tin.strip():
            errors.append(f"{customer_type} customers must have TIN")

    # Individual customers - no TIN requirement
    # All other fields optional

    is_valid = len(errors) == 0
    return is_valid, errors, cleaned


@login_required
@permission_required('customers.add_customer', raise_exception=True)
def customer_import(request):
    """Main customer import view"""

    if request.method == 'GET':
        # Clear any previous import errors from session
        if 'import_errors' in request.session:
            del request.session['import_errors']

        context = {
            'stores': Store.objects.filter(is_active=True).order_by('name'),
        }
        return render(request, 'customers/customer_import.html', context)

    elif request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('import_file')
            conflict_resolution = request.POST.get('conflict_resolution', 'overwrite')
            auto_sync_efris = request.POST.get('auto_sync_efris') == 'on'

            logger.info(f"Import started - File: {uploaded_file.name if uploaded_file else 'None'}")
            logger.info(f"Conflict resolution: {conflict_resolution}, Auto sync EFRIS: {auto_sync_efris}")

            if not uploaded_file:
                messages.error(request, 'No file uploaded. Please select a file to import.')
                return redirect('customers:customer_import')

            # Validate file type
            file_extension = uploaded_file.name.split('.')[-1].lower()
            if file_extension not in ['csv', 'xlsx', 'xls']:
                messages.error(request,
                               'Invalid file type. Only CSV and Excel files (.csv, .xlsx, .xls) are supported.')
                return redirect('customers:customer_import')

            # Validate file size (10MB max)
            if uploaded_file.size > 10 * 1024 * 1024:
                messages.error(request, 'File size exceeds 10MB limit.')
                return redirect('customers:customer_import')

            # Process import
            results = process_customer_import(
                file_obj=uploaded_file,
                conflict_resolution=conflict_resolution,
                auto_sync_efris=auto_sync_efris,
                user=request.user
            )

            logger.info(f"Import completed - Results: {results}")

            # Show success messages
            if results['created_count'] > 0:
                messages.success(request, f'Successfully created {results["created_count"]} customers.')
            if results['updated_count'] > 0:
                messages.success(request, f'Successfully updated {results["updated_count"]} customers.')
            if results['efris_synced'] > 0:
                messages.success(request, f'{results["efris_synced"]} customers synced to EFRIS.')
            if results['skipped_count'] > 0:
                messages.info(request, f'{results["skipped_count"]} customers were skipped (already exist).')
            if results['error_count'] > 0:
                messages.warning(request, f'{results["error_count"]} rows had errors and were not imported.')
                # Store errors in session for display
                request.session['import_errors'] = results['errors'][:50]

            # Show warnings
            if results.get('warnings'):
                for warning in results['warnings'][:10]:  # Limit to 10 warnings
                    messages.warning(request, f"Row {warning['row']}: {warning['warning']}")

            return redirect('customers:customer_list')  # Redirect to customer list instead

        except Exception as e:
            error_message = f'Import failed: {str(e)}'
            logger.error(error_message, exc_info=True)
            messages.error(request, error_message)
            return redirect('customers:customer_import')


@transaction.atomic
def process_customer_import(file_obj, conflict_resolution, auto_sync_efris, user):
    """Process customer import file - FIXED VERSION"""
    try:
        # Parse file
        headers, data = parse_uploaded_file(file_obj)

        # Map columns
        mapped_columns = CustomerColumnMapper.map_columns(headers)

        if not mapped_columns:
            raise ValueError(
                "Could not map any columns from the uploaded file. Please ensure the file uses the correct template format.")

        results = {
            'total_rows': len(data),
            'created_count': 0,
            'updated_count': 0,
            'skipped_count': 0,
            'error_count': 0,
            'efris_synced': 0,
            'errors': [],
            'warnings': []
        }

        # Initialize EFRIS service if needed
        service = None
        if auto_sync_efris:
            try:
                from efris.services import EFRISCustomerService
                service = EFRISCustomerService()
                logger.info("Initialized EFRISCustomerService")
            except Exception as e:
                logger.error(f"Failed to initialize EFRISCustomerService: {str(e)}", exc_info=True)
                results['warnings'].append({
                    'row': 0,
                    'warning': f"EFRIS service initialization failed: {str(e)}"
                })
                auto_sync_efris = False

        for idx, row_data in enumerate(data, start=2):
            try:
                logger.debug(f"Processing row {idx}: {row_data}")

                # Skip completely empty rows
                if not any(str(v).strip() for v in row_data.values() if v is not None):
                    logger.debug(f"Row {idx}: Skipping empty row")
                    continue

                # Validate row
                is_valid, errors, cleaned_data = validate_customer_row(
                    row_data,
                    mapped_columns,
                    idx
                )

                if not is_valid:
                    results['error_count'] += 1
                    results['errors'].append({
                        'row': idx,
                        'errors': errors
                    })
                    logger.warning(f"Row {idx} invalid: {errors}")
                    continue

                logger.debug(f"Row {idx} cleaned data: {cleaned_data}")

                # Check if store exists
                store_name = cleaned_data['store_name']
                try:
                    store = Store.objects.get(name__iexact=store_name)
                    logger.debug(f"Row {idx}: Found store '{store_name}'")
                except Store.DoesNotExist:
                    results['error_count'] += 1
                    results['errors'].append({
                        'row': idx,
                        'errors': [f"Store '{store_name}' not found"]
                    })
                    logger.warning(f"Row {idx}: Store '{store_name}' not found")
                    continue

                # Check if customer exists (by phone)
                phone = cleaned_data['phone']
                existing_customer = Customer.objects.filter(phone=phone).first()

                if existing_customer:
                    if conflict_resolution == 'skip':
                        results['skipped_count'] += 1
                        logger.info(f"Row {idx}: Skipped customer with phone {phone}")
                        continue

                    # Update existing customer
                    try:
                        existing_customer.name = cleaned_data['name']
                        existing_customer.customer_type = cleaned_data['customer_type']
                        existing_customer.store = store

                        # Update optional fields
                        for field in ['email', 'tin', 'nin', 'brn', 'physical_address',
                                      'postal_address', 'district', 'country', 'is_vat_registered',
                                      'credit_limit', 'passport_number', 'driving_license',
                                      'voter_id', 'alien_id', 'efris_customer_type']:
                            if field in cleaned_data:
                                setattr(existing_customer, field, cleaned_data[field])

                        existing_customer.save()
                        results['updated_count'] += 1
                        logger.info(f"Row {idx}: Updated customer with phone {phone}")

                        # EFRIS sync for updated customer
                        if auto_sync_efris and service and existing_customer.is_efris_registered:
                            try:
                                result = service.update_customer(existing_customer)
                                if result.get('success'):
                                    results['efris_synced'] += 1
                                    logger.info(f"Row {idx}: Successfully synced customer {phone} to EFRIS")
                                else:
                                    logger.warning(
                                        f"Row {idx}: EFRIS update failed: {result.get('error', 'Unknown error')}")
                                    results['warnings'].append({
                                        'row': idx,
                                        'warning': f"EFRIS update failed: {result.get('error', 'Unknown error')}"
                                    })
                            except Exception as e:
                                logger.error(f"Row {idx}: EFRIS update error: {str(e)}", exc_info=True)
                                results['warnings'].append({
                                    'row': idx,
                                    'warning': f"Failed to sync customer to EFRIS: {str(e)}"
                                })

                    except Exception as e:
                        logger.error(f"Row {idx}: Error updating customer: {str(e)}", exc_info=True)
                        results['error_count'] += 1
                        results['errors'].append({
                            'row': idx,
                            'errors': [f"Error updating customer: {str(e)}"]
                        })
                        continue

                else:
                    # Create new customer
                    try:
                        # In the process_customer_import function, when creating customer_data:
                        customer_data = {
                            'name': cleaned_data['name'],
                            'customer_type': cleaned_data['customer_type'],
                            'phone': cleaned_data['phone'],
                            'store': store,
                            'email': cleaned_data.get('email', ''),
                            # Convert TIN to string if it exists
                            'tin': str(cleaned_data.get('tin', '')) if cleaned_data.get('tin') else '',
                            'nin': cleaned_data.get('nin', ''),
                            'brn': cleaned_data.get('brn', ''),
                            'physical_address': cleaned_data.get('physical_address', ''),
                            'postal_address': cleaned_data.get('postal_address', ''),
                            'district': cleaned_data.get('district', ''),
                            'country': cleaned_data.get('country', 'Uganda'),
                            'is_vat_registered': cleaned_data.get('is_vat_registered', False),
                            'credit_limit': cleaned_data.get('credit_limit', 0),
                            'passport_number': cleaned_data.get('passport_number', ''),
                            'driving_license': cleaned_data.get('driving_license', ''),
                            'voter_id': cleaned_data.get('voter_id', ''),
                            'alien_id': cleaned_data.get('alien_id', ''),
                            'efris_customer_type': cleaned_data.get('efris_customer_type', ''),
                            'created_by': user,
                        }

                        customer = Customer.objects.create(**customer_data)
                        results['created_count'] += 1
                        logger.info(f"Row {idx}: Created customer with phone {phone}")

                        # EFRIS sync for new customer
                        if auto_sync_efris and service and customer.can_sync_to_efris:
                            try:
                                result = service.register_customer(customer)
                                if result.get('success'):
                                    results['efris_synced'] += 1
                                    logger.info(f"Row {idx}: Successfully synced customer {phone} to EFRIS")
                                else:
                                    logger.warning(
                                        f"Row {idx}: EFRIS registration failed: {result.get('error', 'Unknown error')}")
                                    results['warnings'].append({
                                        'row': idx,
                                        'warning': f"EFRIS registration failed: {result.get('error', 'Unknown error')}"
                                    })
                            except Exception as e:
                                logger.error(f"Row {idx}: EFRIS registration error: {str(e)}", exc_info=True)
                                results['warnings'].append({
                                    'row': idx,
                                    'warning': f"Failed to sync customer to EFRIS: {str(e)}"
                                })

                    except Exception as e:
                        logger.error(f"Row {idx}: Error creating customer: {str(e)}", exc_info=True)
                        results['error_count'] += 1
                        results['errors'].append({
                            'row': idx,
                            'errors': [f"Error creating customer: {str(e)}"]
                        })
                        continue

            except Exception as e:
                logger.error(f"Row {idx}: Unexpected error: {str(e)}", exc_info=True)
                results['error_count'] += 1
                results['errors'].append({
                    'row': idx,
                    'errors': [f"Unexpected error: {str(e)}"]
                })

        logger.info(f"Import completed: {results}")
        return results

    except Exception as e:
        logger.error(f"Customer import failed for file {file_obj.name}: {str(e)}", exc_info=True)
        raise

@login_required
def preview_customer_import(request):
    """Preview customer import data"""

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        uploaded_file = request.FILES.get('preview_file')
        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        # Parse file
        headers, data = parse_uploaded_file(uploaded_file)

        # Map columns
        mapped_columns = CustomerColumnMapper.map_columns(headers)
        suggestions = CustomerColumnMapper.get_mapping_suggestions(headers)

        # Get preview data (first 10 rows)
        preview_data = data[:10]

        # Validate preview rows
        preview_results = []
        for idx, row in enumerate(preview_data, start=2):
            is_valid, errors, cleaned = validate_customer_row(
                row,
                mapped_columns,
                idx
            )
            preview_results.append({
                'row_number': idx,
                'data': row,
                'is_valid': is_valid,
                'errors': errors,
                'cleaned': {k: str(v) for k, v in cleaned.items()}
            })

        return JsonResponse({
            'success': True,
            'headers': headers,
            'mapped_columns': mapped_columns,
            'suggestions': suggestions,
            'total_rows': len(data),
            'preview_data': preview_results,
        })

    except Exception as e:
        logger.error(f"Preview error: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=400)


@login_required
def validate_customer_import(request):
    """Validate customer import data without saving"""

    if request.method != 'POST':
        return JsonResponse({'error': 'Method not allowed'}, status=405)

    try:
        uploaded_file = request.FILES.get('validate_file')

        if not uploaded_file:
            return JsonResponse({'error': 'No file uploaded'}, status=400)

        # Parse file
        headers, data = parse_uploaded_file(uploaded_file)

        # Map columns
        mapped_columns = CustomerColumnMapper.map_columns(headers)

        # Validate all rows
        validation_results = {
            'total_rows': len(data),
            'valid_rows': 0,
            'invalid_rows': 0,
            'errors': [],
            'warnings': []
        }

        for idx, row in enumerate(data, start=2):
            is_valid, errors, cleaned = validate_customer_row(
                row,
                mapped_columns,
                idx
            )

            if is_valid:
                validation_results['valid_rows'] += 1

                # Check for warnings
                store_name = cleaned.get('store_name')
                if store_name and not Store.objects.filter(name__iexact=store_name).exists():
                    validation_results['warnings'].append({
                        'row': idx,
                        'warning': f"Store '{store_name}' not found"
                    })

                # Check duplicate phone
                phone = cleaned.get('phone')
                if phone and Customer.objects.filter(phone=phone).exists():
                    validation_results['warnings'].append({
                        'row': idx,
                        'warning': f"Customer with phone '{phone}' already exists"
                    })
            else:
                validation_results['invalid_rows'] += 1
                validation_results['errors'].append({
                    'row': idx,
                    'errors': errors,
                    'data': row
                })

        return JsonResponse({
            'success': True,
            'validation': validation_results,
            'mapped_columns': mapped_columns
        })

    except Exception as e:
        logger.error(f"Validation error: {str(e)}", exc_info=True)
        return JsonResponse({'error': str(e)}, status=400)