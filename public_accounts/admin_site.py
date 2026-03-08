from django.urls import path, reverse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.views import View
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.utils import timezone
from collections import OrderedDict
from django.core.paginator import Paginator
from django.db.models import Q
import csv
import io
from datetime import datetime


class PublicAdminSite:
    """
    Custom admin site for public schema
    Similar to Django's admin but completely separate
    """

    def __init__(self, name='public_admin'):
        self.name = name
        self._registry = OrderedDict()  # app_label -> ModelAdmin instances
        self.app_configs = {}

    def register(self, model_class, admin_class=None, app_label=None):
        """
        Register a model with the public admin site
        """
        if admin_class is None:
            admin_class = PublicModelAdmin

        if app_label is None:
            app_label = model_class._meta.app_label

        if app_label not in self._registry:
            self._registry[app_label] = OrderedDict()

        model_admin = admin_class(model_class, self)
        self._registry[app_label][model_class] = model_admin

        # Store app config
        if app_label not in self.app_configs:
            self.app_configs[app_label] = {
                'name': app_label.replace('_', ' ').title(),
                'models': []
            }

        # Store model_name as a string instead of accessing __name__ in template
        model_name = model_class.__name__.lower()

        self.app_configs[app_label]['models'].append({
            'model': model_class,
            'admin': model_admin,
            'name': model_class._meta.verbose_name_plural,
            'model_name': model_name,  # Add this - store the name directly
            'url_name': f'{app_label}_{model_name}'
        })

    def get_urls(self):
        """Generate URL patterns for the admin site"""
        from django.urls import path
        from django.db import models as django_models
        from . import api_views

        urlpatterns = [
            path('', self.index_view, name='public_admin_index'),
            path('login/', self.login_view, name='public_admin_login'),
            path('logout/', self.logout_view, name='public_admin_logout'),
            path('profile/', self.profile_view, name='public_admin_profile'),
            path('profile/edit/', self.profile_edit_view, name='public_admin_profile_edit'),
            path('change-password/', self.change_password_view, name='public_admin_change_password'),
            path('password-reset/', self.password_reset_request_view, name='public_admin_password_reset'),
            path('password-reset/<str:token>/', self.password_reset_confirm_view,
                 name='public_admin_password_reset_confirm'),

            # API Endpoints
            path('api/users/search/', api_views.user_search_api, name='public_admin_api_user_search'),
            path('api/users/<int:user_id>/unlock/', api_views.unlock_user_api, name='public_admin_api_user_unlock'),
            path('api/users/<int:user_id>/verify-email/', api_views.verify_email_api,
                 name='public_admin_api_verify_email'),
            path('api/activities/stats/', api_views.activity_stats_api, name='public_admin_api_activity_stats'),
        ]

        # Add URLs for each registered model
        for app_label, models_dict in self._registry.items():
            for model_class, model_admin in models_dict.items():
                model_name = model_class.__name__.lower()

                # Check if model uses UUID, string PK, or integer PK
                pk_field = model_class._meta.pk
                if pk_field.get_internal_type() == 'UUIDField':
                    pk_pattern = '<uuid:pk>'
                elif isinstance(pk_field, django_models.CharField):
                    pk_pattern = '<str:pk>'
                else:
                    pk_pattern = '<int:pk>'

                # Get custom URLs from model admin if it has them
                custom_urls = []
                if hasattr(model_admin, 'get_urls'):
                    try:
                        custom_urls = model_admin.get_urls()
                    except Exception as e:
                        import logging
                        logger = logging.getLogger(__name__)
                        logger.error(f"Error getting custom URLs for {app_label}.{model_name}: {e}")

                # Add custom URLs first, then standard URLs
                urlpatterns += custom_urls  # Add custom URLs first

                urlpatterns += [
                    path(f'{app_label}/{model_name}/',
                         model_admin.changelist_view,
                         name=f'public_admin_{app_label}_{model_name}_list'),
                    path(f'{app_label}/{model_name}/add/',
                         model_admin.add_view,
                         name=f'public_admin_{app_label}_{model_name}_add'),
                    path(f'{app_label}/{model_name}/{pk_pattern}/',
                         model_admin.change_view,
                         name=f'public_admin_{app_label}_{model_name}_change'),
                    path(f'{app_label}/{model_name}/{pk_pattern}/delete/',
                         model_admin.delete_view,
                         name=f'public_admin_{app_label}_{model_name}_delete'),
                    path(f'{app_label}/{model_name}/{pk_pattern}/password/',
                         model_admin.password_reset_view,
                         name=f'public_admin_{app_label}_{model_name}_password'),
                    path(f'{app_label}/{model_name}/{pk_pattern}/history/',
                         model_admin.history_view,
                         name=f'public_admin_{app_label}_{model_name}_history'),
                    # Export URLs
                    path(f'{app_label}/{model_name}/export/csv/',
                         model_admin.export_csv_view,
                         name=f'public_admin_{app_label}_{model_name}_export_csv'),
                    path(f'{app_label}/{model_name}/export/excel/',
                         model_admin.export_excel_view,
                         name=f'public_admin_{app_label}_{model_name}_export_excel'),
                    path(f'{app_label}/{model_name}/export/pdf/',
                         model_admin.export_pdf_view,
                         name=f'public_admin_{app_label}_{model_name}_export_pdf'),
                    # Bulk actions
                    path(f'{app_label}/{model_name}/bulk-action/',
                         model_admin.bulk_action_view,
                         name=f'public_admin_{app_label}_{model_name}_bulk_action'),
                ]

        return urlpatterns

    @property
    def urls(self):
        return self.get_urls(), 'public_admin', self.name

    def has_permission(self, request):
        """Check if user has permission to access admin"""
        return request.user.is_authenticated and request.user.is_staff

    def index_view(self, request):
        """Main dashboard view"""
        if not request.user.is_authenticated:
            return redirect('public_admin:public_admin_login')

        if not self.has_permission(request):
            return HttpResponseForbidden("You don't have permission to access this area.")

        # Get statistics
        from .models import PublicUser, PublicUserActivity

        stats = {
            'total_users': PublicUser.objects.count(),
            'active_users': PublicUser.objects.filter(is_active=True).count(),
            'admin_users': PublicUser.objects.filter(is_admin=True).count(),
            'recent_activities': PublicUserActivity.objects.select_related('user')[:10],
        }

        context = {
            'title': 'PrimeBook Public Admin Dashboard',
            'user': request.user,
            'app_configs': self.app_configs,
            'stats': stats,
        }

        return render(request, 'public_admin/index.html', context)

    def admin_view(self, view):
        """
        Decorator to create an admin view with permission checking.
        Similar to Django's admin.site.admin_view()
        """

        def inner(request, *args, **kwargs):
            if not self.has_permission(request):
                return HttpResponseForbidden("You don't have permission to access this area.")
            return view(request, *args, **kwargs)

        return inner

    def login_view(self, request):
        """Login view"""
        if request.user.is_authenticated:
            return redirect('public_admin:public_admin_index')

        if request.method == 'POST':
            from django.contrib.auth import authenticate, login
            from .models import PublicUserActivity
            from django.db import connection

            identifier = request.POST.get('identifier')
            password = request.POST.get('password')

            # CRITICAL: Ensure we're in public schema
            if hasattr(connection, 'schema_name'):
                if connection.schema_name != 'public':
                    messages.error(request, 'This login page is only for public admin users.')
                    return redirect('public_admin:public_admin_login')

            # Authenticate with explicit identifier parameter
            user = authenticate(
                request,
                identifier=identifier,
                password=password
            )

            if user is not None:
                # Double check it's a PublicUser
                from .models import PublicUser
                if not isinstance(user, PublicUser):
                    messages.error(request, 'Invalid login credentials.')
                    return redirect('public_admin:public_admin_login')

                # Login with explicit backend
                login(request, user, backend='public_accounts.backends.PublicIdentifierBackend')

                # Log activity
                PublicUserActivity.objects.create(
                    user=user,
                    action='LOGIN',
                    description=f'User logged in',
                    ip_address=self.get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                )

                messages.success(request, f'Welcome back, {user.get_full_name()}!')

                # Check if password change is required
                if user.force_password_change:
                    messages.warning(request, 'Please change your password.')
                    return redirect('public_admin:public_admin_change_password')

                # Redirect to next or index
                next_url = request.GET.get('next') or request.POST.get('next')
                if next_url:
                    return redirect(next_url)

                return redirect('public_admin:public_admin_index')
            else:
                messages.error(request, 'Invalid identifier or password.')

        return render(request, 'public_admin/login.html', {'title': 'Login'})

    def logout_view(self, request):
        """Logout view"""
        if request.user.is_authenticated:
            from .models import PublicUserActivity

            # Log activity
            PublicUserActivity.objects.create(
                user=request.user,
                action='LOGOUT',
                description=f'User logged out',
                ip_address=self.get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
            )

        logout(request)
        messages.success(request, 'You have been logged out successfully.')
        return redirect('public_admin:public_admin_login')

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def profile_view(self, request):
        """User profile view"""
        context = {
            'title': 'Your Profile',
            'user': request.user,
        }
        return render(request, 'public_admin/profile.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def profile_edit_view(self, request):
        """Edit user profile"""
        from .forms import ProfileUpdateForm

        if request.method == 'POST':
            form = ProfileUpdateForm(request.POST, request.FILES, instance=request.user)
            if form.is_valid():
                form.save()
                messages.success(request, 'Your profile has been updated successfully.')
                return redirect('public_admin:public_admin_profile')
        else:
            form = ProfileUpdateForm(instance=request.user)

        context = {
            'title': 'Edit Profile',
            'form': form,
        }
        return render(request, 'public_admin/profile_edit.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def change_password_view(self, request):
        """Change password view"""
        from .forms import PasswordChangeForm
        from .models import PublicUserActivity

        if request.method == 'POST':
            form = PasswordChangeForm(request.user, request.POST)
            if form.is_valid():
                user = form.save()

                # Log activity
                PublicUserActivity.objects.create(
                    user=user,
                    action='PASSWORD_CHANGE',
                    description='User changed password',
                    ip_address=self.get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                )

                # Update session to prevent logout
                from django.contrib.auth import update_session_auth_hash
                update_session_auth_hash(request, user)

                messages.success(request, 'Your password has been changed successfully.')
                return redirect('public_admin:public_admin_profile')
        else:
            form = PasswordChangeForm(request.user)

        context = {
            'title': 'Change Password',
            'form': form,
        }
        return render(request, 'public_admin/change_password.html', context)

    def password_reset_request_view(self, request):
        """Password reset request view"""
        from .forms import PasswordResetRequestForm
        from .models import PublicUser, PasswordResetToken

        if request.method == 'POST':
            form = PasswordResetRequestForm(request.POST)
            if form.is_valid():
                identifier = form.cleaned_data['identifier']
                email = form.cleaned_data['email']

                try:
                    user = PublicUser.objects.get(identifier=identifier, email=email, is_active=True)

                    # Create reset token
                    token = PasswordResetToken.generate_token()
                    reset_token = PasswordResetToken.objects.create(
                        user=user,
                        token=token,
                        expires_at=timezone.now() + timezone.timedelta(hours=24),
                        ip_address=self.get_client_ip(request)
                    )

                    # Send reset email
                    user.send_password_reset_email(token)

                    messages.success(request, 'Password reset instructions have been sent to your email.')
                    return redirect('public_admin:public_admin_login')

                except PublicUser.DoesNotExist:
                    messages.error(request, 'Invalid identifier or email address.')
        else:
            form = PasswordResetRequestForm()

        context = {
            'title': 'Reset Password',
            'form': form,
        }
        return render(request, 'public_admin/password_reset_request.html', context)

    def password_reset_confirm_view(self, request, token):
        """Password reset confirmation view"""
        from .forms import PasswordResetConfirmForm
        from .models import PasswordResetToken, PublicUserActivity

        try:
            reset_token = PasswordResetToken.objects.get(token=token)

            if not reset_token.is_valid():
                messages.error(request, 'This password reset link has expired or has already been used.')
                return redirect('public_admin:public_admin_password_reset')

            if request.method == 'POST':
                form = PasswordResetConfirmForm(request.POST)
                if form.is_valid():
                    # Set new password
                    user = reset_token.user
                    user.set_password(form.cleaned_data['new_password1'])
                    user.force_password_change = False
                    user.password_changed_at = timezone.now()
                    user.save()

                    # Mark token as used
                    reset_token.mark_as_used(self.get_client_ip(request))

                    # Log activity
                    PublicUserActivity.objects.create(
                        user=user,
                        action='PASSWORD_RESET',
                        description='Password reset completed',
                        ip_address=self.get_client_ip(request),
                        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                    )

                    messages.success(request, 'Your password has been reset successfully. You can now login.')
                    return redirect('public_admin:public_admin_login')
            else:
                form = PasswordResetConfirmForm()

            context = {
                'title': 'Set New Password',
                'form': form,
                'token': token,
            }
            return render(request, 'public_admin/password_reset_confirm.html', context)

        except PasswordResetToken.DoesNotExist:
            messages.error(request, 'Invalid password reset link.')
            return redirect('public_admin:public_admin_password_reset')

    @staticmethod
    def get_client_ip(request):
        """Get client IP address from request"""
        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')
        return ip


class PublicModelAdmin:
    """
    Base class for model admins in public schema
    Similar to Django's ModelAdmin
    """

    list_display = ['__str__']
    list_filter = []
    search_fields = []
    ordering = []
    list_per_page = 25
    form_class = None
    readonly_fields = []
    fieldsets = None
    exclude = []
    actions = []  # List of custom actions

    # Permissions
    has_add_permission_flag = True
    has_change_permission_flag = True
    has_delete_permission_flag = True
    has_view_permission_flag = True
    has_export_permission_flag = True

    def __init__(self, model, admin_site):
        self.model = model
        self.admin_site = admin_site

    def get_url_params(self):
        """
        Get consistent URL parameters for this model
        Returns app_label and model_name that match URL patterns
        """
        # Get correct app_label from registration
        correct_app_label = self.model._meta.app_label
        for app_label, models_dict in self.admin_site._registry.items():
            if self.model in models_dict:
                correct_app_label = app_label
                break

        # Always use lowercase for model name to match URL patterns
        model_name_lower = self.model.__name__.lower()

        return correct_app_label, model_name_lower

    def has_add_permission(self, request):
        """Check if user can add objects"""
        return self.has_add_permission_flag and request.user.is_staff

    def has_change_permission(self, request, obj=None):
        """Check if user can change objects"""
        return self.has_change_permission_flag and request.user.is_staff

    def has_delete_permission(self, request, obj=None):
        """Check if user can delete objects"""
        return self.has_delete_permission_flag and request.user.is_staff

    def has_view_permission(self, request, obj=None):
        """Check if user can view objects"""
        return self.has_view_permission_flag and request.user.is_staff

    def has_export_permission(self, request):
        """Check if user can export data"""
        return self.has_export_permission_flag and request.user.is_staff

    def get_form_class(self):
        """Get form class for add/change views"""
        if self.form_class:
            return self.form_class

        # Auto-generate ModelForm
        from django import forms

        class AutoModelForm(forms.ModelForm):
            class Meta:
                model = self.model
                exclude = self.exclude or []

        return AutoModelForm

    def get_queryset(self, request):
        """Get queryset for list view"""
        return self.model.objects.all()

    def get_list_display(self, request):
        """Get fields to display in list view"""
        return self.list_display

    def get_fieldsets(self, request, obj=None):
        """Get fieldsets for form"""
        if self.fieldsets:
            return self.fieldsets

        # Auto-generate simple fieldset
        fields = [f.name for f in self.model._meta.fields
                  if f.name not in self.exclude and not f.name == 'id']

        return [(None, {'fields': fields})]

    def get_actions(self, request):
        """Get available actions for bulk operations"""
        actions = [
            {
                'name': 'delete_selected',
                'description': 'Delete selected items'
            }
        ]

        # Add custom actions
        for action_name in self.actions:
            action_method = getattr(self, action_name, None)
            if action_method:
                actions.append({
                    'name': action_name,
                    'description': getattr(action_method, 'short_description', action_name.replace('_', ' ').title())
                })

        return actions

    def get_export_fields(self):
        """Get fields to export"""
        if self.list_display and self.list_display != ['__str__']:
            return [field for field in self.list_display if field != '__str__']

        # Default to all fields except id
        return [f.name for f in self.model._meta.fields if f.name != 'id']

    def get_field_value(self, obj, field_name):
        """Get field value for export"""
        try:
            if field_name == '__str__':
                return str(obj)

            value = getattr(obj, field_name)

            # Handle callable methods
            if callable(value):
                value = value()

            # Handle foreign keys
            if hasattr(value, 'pk'):
                return str(value)

            # Handle dates
            if isinstance(value, datetime):
                return value.strftime('%Y-%m-%d %H:%M:%S')

            return str(value) if value is not None else ''
        except:
            return ''

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def changelist_view(self, request):
        """List view"""
        if not self.has_view_permission(request):
            return HttpResponseForbidden("You don't have permission to view this.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        queryset = self.get_queryset(request)

        # Apply ordering
        if self.ordering:
            queryset = queryset.order_by(*self.ordering)

        # Apply search
        search_query = request.GET.get('q')
        if search_query and self.search_fields:
            q_objects = Q()
            for field in self.search_fields:
                q_objects |= Q(**{f'{field}__icontains': search_query})
            queryset = queryset.filter(q_objects)

        # Apply filters
        for filter_field in self.list_filter:
            filter_value = request.GET.get(filter_field)
            if filter_value:
                queryset = queryset.filter(**{filter_field: filter_value})

        # Pagination
        paginator = Paginator(queryset, self.list_per_page)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'title': self.model._meta.verbose_name_plural,
            'model': self.model,
            'model_name': model_name_lower,
            'model_name_lower': model_name_lower,
            'app_label': app_label,
            'verbose_name': str(self.model._meta.verbose_name),
            'verbose_name_plural': str(self.model._meta.verbose_name_plural),

            'objects': page_obj,
            'search_query': search_query,
            'list_display': self.get_list_display(request),
            'has_add_permission': self.has_add_permission(request),
            'has_change_permission': self.has_change_permission(request),
            'has_delete_permission': self.has_delete_permission(request),
            'has_export_permission': self.has_export_permission(request),
            'actions': self.get_actions(request),
        }

        return render(request, 'public_admin/changelist.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def add_view(self, request):
        if not self.has_add_permission(request):
            return HttpResponseForbidden("You don't have permission to add this.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        FormClass = self.get_form_class()

        if request.method == 'POST':
            form = FormClass(request.POST, request.FILES)
            if form.is_valid():
                obj = form.save()

                # Log activity
                from .models import PublicUserActivity
                PublicUserActivity.objects.create(
                    user=request.user,
                    action='CREATE',
                    app_name=app_label,
                    model_name=self.model.__name__,
                    object_id=str(obj.pk),
                    description=f'Created {self.model._meta.verbose_name}: {obj}',
                    ip_address=PublicAdminSite.get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                )

                messages.success(request, f'{self.model._meta.verbose_name} added successfully.')
                return redirect(f'public_admin:public_admin_{app_label}_{model_name_lower}_list')
        else:
            form = FormClass()

        context = {
            'title': f'Add {self.model._meta.verbose_name}',
            'model': self.model,
            'model_name': model_name_lower,
            'model_name_lower': model_name_lower,
            'verbose_name': str(self.model._meta.verbose_name),
            'verbose_name_plural': str(self.model._meta.verbose_name_plural),
            'app_label': app_label,
            'form': form,
            'fieldsets': self.get_fieldsets(request),
        }

        return render(request, 'public_admin/change_form.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def change_view(self, request, pk):
        obj = get_object_or_404(self.model, pk=pk)

        if not self.has_change_permission(request, obj):
            return HttpResponseForbidden("You don't have permission to change this.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        FormClass = self.get_form_class()

        if request.method == 'POST':
            form = FormClass(request.POST, request.FILES, instance=obj)
            if form.is_valid():
                # Capture old status BEFORE saving to detect PENDING/FAILED → PROCESSING
                old_status = getattr(obj, 'status', None)

                obj = form.save()

                # ── Tenant provisioning hook ──────────────────────────────────
                # The generic change_view just calls form.save() — it has no business logic.
                # When a TenantSignupRequest is set to PROCESSING and the tenant doesn't
                # exist yet, we must fire the Celery task here. This is exactly what was
                # missing: changing status to PROCESSING in the admin did nothing because
                # no code was watching for that transition and dispatching the task.
                from public_router.models import TenantSignupRequest as _TSR
                if isinstance(obj, _TSR):
                    new_status = obj.status
                    should_provision = (
                        new_status == 'PROCESSING'
                        and not obj.tenant_created
                        and old_status in ('PENDING', 'FAILED', 'PROCESSING')
                    )
                    if should_provision:
                        import logging as _logging
                        import secrets, string
                        _logger = _logging.getLogger(__name__)
                        try:
                            from public_router.tasks import create_tenant_async
                            from public_router.models import TenantApprovalWorkflow
                            from django.core.cache import cache

                            # Retrieve stored password or generate a fresh one
                            password = None
                            try:
                                wf = getattr(obj, 'approval_workflow', None)
                                if wf and wf.generated_password:
                                    password = wf.generated_password
                            except Exception:
                                pass

                            if not password:
                                alphabet = string.ascii_letters + string.digits + "!@#$%^&*"
                                password = ''.join(secrets.choice(alphabet) for _ in range(12))
                                try:
                                    wf, _ = TenantApprovalWorkflow.objects.get_or_create(
                                        signup_request=obj
                                    )
                                    wf.generated_password = password
                                    wf.save(update_fields=['generated_password', 'updated_at'])
                                except Exception as wf_err:
                                    _logger.error(
                                        f"Could not persist password for {obj.request_id}: {wf_err}"
                                    )

                            task = create_tenant_async.apply_async(
                                args=[str(obj.request_id), password],
                                countdown=2,
                            )
                            cache.set(
                                f'signup_task_{obj.request_id}',
                                {
                                    'task_id': task.id,
                                    'triggered_by': str(request.user),
                                    'triggered_via': 'admin_change_view',
                                },
                                timeout=3600
                            )
                            messages.info(
                                request,
                                f'Tenant creation queued for "{obj.company_name}". '
                                f'Refresh in 30–60 seconds to see the result.'
                            )
                        except Exception as task_err:
                            _logger.error(
                                f"Failed to queue tenant task for {obj.request_id}: {task_err}"
                            )
                            messages.error(
                                request,
                                f'Saved, but tenant creation task failed to queue: {task_err}. '
                                f'Use the "Retry processing" bulk action to try again.'
                            )
                # ── end tenant provisioning hook ──────────────────────────────

                # Log activity
                from .models import PublicUserActivity
                PublicUserActivity.objects.create(
                    user=request.user,
                    action='UPDATE',
                    app_name=app_label,
                    model_name=self.model.__name__,
                    object_id=str(obj.pk),
                    description=f'Updated {self.model._meta.verbose_name}: {obj}',
                    ip_address=PublicAdminSite.get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                )

                messages.success(request, f'{self.model._meta.verbose_name} updated successfully.')
                return redirect(f'public_admin:public_admin_{app_label}_{model_name_lower}_list')
        else:
            form = FormClass(instance=obj)

        context = {
            'title': f'Change {self.model._meta.verbose_name}',
            'model': self.model,
            'object': obj,
            'model_name': model_name_lower,
            'model_name_lower': model_name_lower,
            'verbose_name': str(self.model._meta.verbose_name),
            'verbose_name_plural': str(self.model._meta.verbose_name_plural),
            'app_label': app_label,
            'form': form,
            'fieldsets': self.get_fieldsets(request, obj),
        }

        return render(request, 'public_admin/change_form.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def delete_view(self, request, pk):
        obj = get_object_or_404(self.model, pk=pk)

        if not self.has_delete_permission(request, obj):
            return HttpResponseForbidden("You don't have permission to delete this.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        if request.method == 'POST':
            obj_str = str(obj)
            obj.delete()

            from .models import PublicUserActivity
            PublicUserActivity.objects.create(
                user=request.user,
                action='DELETE',
                app_name=app_label,
                model_name=self.model.__name__,
                object_id=str(pk),
                description=f'Deleted {self.model._meta.verbose_name}: {obj_str}',
                ip_address=PublicAdminSite.get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
            )

            messages.success(request, f'{self.model._meta.verbose_name} deleted successfully.')
            return redirect(f'public_admin:public_admin_{app_label}_{model_name_lower}_list')

        context = {
            'title': f'Delete {self.model._meta.verbose_name}',
            'object': obj,
            'model': self.model,
            'model_name': model_name_lower,
            'model_name_lower': model_name_lower,
            'verbose_name': str(self.model._meta.verbose_name),
            'verbose_name_plural': str(self.model._meta.verbose_name_plural),
            'app_label': app_label,
        }

        return render(request, 'public_admin/delete_confirmation.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def history_view(self, request, pk):
        """View object history"""
        obj = get_object_or_404(self.model, pk=pk)

        if not self.has_view_permission(request, obj):
            return HttpResponseForbidden("You don't have permission to view this.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        from .models import PublicUserActivity

        # Get all activities for this object
        activities = PublicUserActivity.objects.filter(
            app_name=app_label,
            model_name=self.model.__name__,
            object_id=str(pk)
        ).select_related('user').order_by('-timestamp')

        # Pagination
        paginator = Paginator(activities, 25)
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)

        context = {
            'title': f'History - {obj}',
            'object': obj,
            'model': self.model,
            'model_name': model_name_lower,
            'verbose_name': str(self.model._meta.verbose_name),
            'verbose_name_plural': str(self.model._meta.verbose_name_plural),
            'app_label': app_label,
            'activities': page_obj,
        }

        return render(request, 'public_admin/history.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def password_reset_view(self, request, pk):
        """Password reset view (only for PublicUser model)"""
        from .models import PublicUser
        from .forms import AdminPasswordResetForm

        if self.model != PublicUser:
            return HttpResponseForbidden("This action is only available for users.")

        obj = get_object_or_404(self.model, pk=pk)

        if not request.user.is_admin and request.user != obj:
            return HttpResponseForbidden("You don't have permission to reset this password.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        if request.method == 'POST':
            form = AdminPasswordResetForm(obj, request.POST)
            if form.is_valid():
                user, new_password = form.save()

                # Log activity
                from .models import PublicUserActivity
                PublicUserActivity.objects.create(
                    user=request.user,
                    action='PASSWORD_RESET',
                    app_name=app_label,
                    model_name=self.model.__name__,
                    object_id=str(obj.pk),
                    description=f'Reset password for user: {obj}',
                    ip_address=PublicAdminSite.get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                )

                messages.success(request, f'Password reset successfully. New password: {new_password}')
                return redirect(f'public_admin:public_admin_{app_label}_{model_name_lower}_change', pk=pk)
        else:
            form = AdminPasswordResetForm(obj)

        context = {
            'title': f'Reset Password - {obj}',
            'object': obj,
            'form': form,
            'model': self.model,
            'model_name': model_name_lower,
        }

        return render(request, 'public_admin/password_reset_admin.html', context)

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def export_csv_view(self, request):
        """Export to CSV"""
        if not self.has_export_permission(request):
            return HttpResponseForbidden("You don't have permission to export data.")

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        queryset = self.get_queryset(request)

        # Apply search if provided
        search_query = request.GET.get('q')
        if search_query and self.search_fields:
            q_objects = Q()
            for field in self.search_fields:
                q_objects |= Q(**{f'{field}__icontains': search_query})
            queryset = queryset.filter(q_objects)

        # Create the HttpResponse object with CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{self.model.__name__}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        writer = csv.writer(response)

        # Get export fields
        export_fields = self.get_export_fields()

        # Write header
        writer.writerow([field.replace('_', ' ').title() for field in export_fields])

        # Write data
        for obj in queryset:
            row = [self.get_field_value(obj, field) for field in export_fields]
            writer.writerow(row)

        # Log activity
        from .models import PublicUserActivity
        PublicUserActivity.objects.create(
            user=request.user,
            action='EXPORT',
            app_name=app_label,
            model_name=self.model.__name__,
            description=f'Exported {queryset.count()} {self.model._meta.verbose_name_plural} to CSV',
            ip_address=PublicAdminSite.get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
        )

        return response

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def export_excel_view(self, request):
        """Export to Excel"""
        if not self.has_export_permission(request):
            return HttpResponseForbidden("You don't have permission to export data.")

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            messages.error(request, 'openpyxl is not installed. Install it with: pip install openpyxl')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        queryset = self.get_queryset(request)

        # Apply search if provided
        search_query = request.GET.get('q')
        if search_query and self.search_fields:
            q_objects = Q()
            for field in self.search_fields:
                q_objects |= Q(**{f'{field}__icontains': search_query})
            queryset = queryset.filter(q_objects)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self.model._meta.verbose_name_plural[:31]  # Excel sheet name limit

        # Get export fields
        export_fields = self.get_export_fields()

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        # Write header
        for col_num, field_name in enumerate(export_fields, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = field_name.replace('_', ' ').title()
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Write data
        for row_num, obj in enumerate(queryset, 2):
            for col_num, field_name in enumerate(export_fields, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = self.get_field_value(obj, field_name)
                cell.alignment = Alignment(vertical='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.model.__name__}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        wb.save(response)

        # Log activity
        from .models import PublicUserActivity
        PublicUserActivity.objects.create(
            user=request.user,
            action='EXPORT',
            app_name=app_label,
            model_name=self.model.__name__,
            description=f'Exported {queryset.count()} {self.model._meta.verbose_name_plural} to Excel',
            ip_address=PublicAdminSite.get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
        )

        return response

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def export_pdf_view(self, request):
        """Export to PDF"""
        if not self.has_export_permission(request):
            return HttpResponseForbidden("You don't have permission to export data.")

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
        except ImportError:
            messages.error(request, 'reportlab is not installed. Install it with: pip install reportlab')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        queryset = self.get_queryset(request)

        # Apply search if provided
        search_query = request.GET.get('q')
        if search_query and self.search_fields:
            q_objects = Q()
            for field in self.search_fields:
                q_objects |= Q(**{f'{field}__icontains': search_query})
            queryset = queryset.filter(q_objects)

        # Create the HttpResponse object with PDF header
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.model.__name__}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

        # Create the PDF object
        doc = SimpleDocTemplate(response, pagesize=A4)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )

        # Add title
        title = Paragraph(f"{self.model._meta.verbose_name_plural} Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Add metadata
        meta_style = styles['Normal']
        meta_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        meta_text += f"Total Records: {queryset.count()}<br/>"
        meta_text += f"Generated by: {request.user.get_full_name()}"
        meta = Paragraph(meta_text, meta_style)
        elements.append(meta)
        elements.append(Spacer(1, 0.3 * inch))

        # Get export fields
        export_fields = self.get_export_fields()

        # Prepare table data
        table_data = []

        # Header row
        headers = [field.replace('_', ' ').title() for field in export_fields]
        table_data.append(headers)

        # Data rows (limit to prevent huge PDFs)
        max_rows = 100
        for obj in queryset[:max_rows]:
            row = []
            for field_name in export_fields:
                value = self.get_field_value(obj, field_name)
                # Truncate long values
                if len(str(value)) > 50:
                    value = str(value)[:47] + "..."
                row.append(str(value))
            table_data.append(row)

        # Create table
        table = Table(table_data)

        # Style the table
        table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data styling
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),

            # Grid
            ('GRID', (0, 0), (-1, -1), 1, colors.black),

            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)

        # Add note if records were truncated
        if queryset.count() > max_rows:
            elements.append(Spacer(1, 0.2 * inch))
            note = Paragraph(
                f"<i>Note: Showing first {max_rows} of {queryset.count()} records. "
                f"Use Excel export for complete data.</i>",
                styles['Italic']
            )
            elements.append(note)

        # Build PDF
        doc.build(elements)

        # Log activity
        from .models import PublicUserActivity
        PublicUserActivity.objects.create(
            user=request.user,
            action='EXPORT',
            app_name=app_label,
            model_name=self.model.__name__,
            description=f'Exported {queryset.count()} {self.model._meta.verbose_name_plural} to PDF',
            ip_address=PublicAdminSite.get_client_ip(request),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
        )

        return response

    @method_decorator(login_required(login_url='public_admin:public_admin_login'))
    def bulk_action_view(self, request):
        """Handle bulk actions"""
        if request.method != 'POST':
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # GET URL PARAMS FIRST
        app_label, model_name_lower = self.get_url_params()

        action = request.POST.get('action')
        selected_ids = request.POST.getlist('selected_items')

        if not action:
            messages.error(request, 'Please select an action.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        if not selected_ids:
            messages.error(request, 'Please select at least one item.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        queryset = self.model.objects.filter(pk__in=selected_ids)

        # Handle delete action
        if action == 'delete_selected':
            if not self.has_delete_permission(request):
                return HttpResponseForbidden("You don't have permission to delete.")

            count = queryset.count()
            queryset.delete()

            # Log activity
            from .models import PublicUserActivity
            PublicUserActivity.objects.create(
                user=request.user,
                action='DELETE',
                app_name=app_label,
                model_name=self.model.__name__,
                description=f'Bulk deleted {count} {self.model._meta.verbose_name_plural}',
                ip_address=PublicAdminSite.get_client_ip(request),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
            )

            messages.success(request, f'Successfully deleted {count} {self.model._meta.verbose_name_plural}.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Handle custom actions
        if hasattr(self, action):
            action_method = getattr(self, action)
            try:
                result = action_method(request, queryset)

                # Log activity
                from .models import PublicUserActivity
                PublicUserActivity.objects.create(
                    user=request.user,
                    action='UPDATE',
                    app_name=app_label,
                    model_name=self.model.__name__,
                    description=f'Bulk action "{action}" on {queryset.count()} items',
                    ip_address=PublicAdminSite.get_client_ip(request),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255]
                )

                if result:
                    messages.success(request, result)
                else:
                    messages.success(request, f'Action "{action}" completed successfully.')
            except Exception as e:
                messages.error(request, f'Error executing action: {str(e)}')
        else:
            messages.error(request, f'Unknown action: {action}')

        return redirect(request.META.get('HTTP_REFERER', '/'))


# Create global admin site instance
public_admin = PublicAdminSite()