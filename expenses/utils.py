"""
utils.py — Utility functions for the expenses app.

Changes from original:
  • All aggregations use amount_base for currency-correct totals.
  • get_expense_summary also returns by_currency breakdown.
  • generate_chart_data returns an empty list (not NameError) for unknown group_by.
  • validate_expense_approval — new helper used by quick_approve_api.
  • export_to_pdf / export_to_excel updated with vendor, currency, status columns.
  • matplotlib import is deferred inside functions to avoid import-time side-effects.
"""

import io
import logging
from datetime import timedelta
from decimal import Decimal

from django.db.models import Avg, Count, Sum
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek
from django.utils import timezone

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Date range helpers
# ---------------------------------------------------------------------------

def get_date_range(period):
    """Return (start_date, end_date) for a named period, or (None, None)."""
    today = timezone.now().date()

    mapping = {
        'today':     (today, today),
        'week':      (today - timedelta(days=today.weekday()), today),
        'fortnight': (today - timedelta(days=14), today),
        'month':     (today.replace(day=1), today),
        '6months':   (today - timedelta(days=182), today),
        'year':      (today.replace(month=1, day=1), today),
    }

    if period == 'quarter':
        qm = ((today.month - 1) // 3) * 3 + 1
        return today.replace(month=qm, day=1), today

    return mapping.get(period, (None, None))


# ---------------------------------------------------------------------------
# Summary statistics
# ---------------------------------------------------------------------------

def get_expense_summary(expenses):
    """
    Return a dict with:
      total, count, average  — in base currency
      by_tag                 — {tag_name: Decimal}
      by_currency            — {currency_code: Decimal}  (original amounts)
    """
    total = expenses.aggregate(total=Sum('amount_base'))['total'] or Decimal('0.00')
    count = expenses.count()
    avg = expenses.aggregate(avg=Avg('amount_base'))['avg'] or Decimal('0.00')

    by_tag: dict[str, Decimal] = {}
    by_currency: dict[str, Decimal] = {}

    for expense in expenses.prefetch_related('tags'):
        for tag in expense.tags.all():
            by_tag[tag.name] = by_tag.get(tag.name, Decimal('0')) + expense.amount_base

        key = expense.currency
        by_currency[key] = by_currency.get(key, Decimal('0')) + expense.amount

    by_tag = dict(sorted(by_tag.items(), key=lambda x: x[1], reverse=True))

    return {
        'total': total,
        'count': count,
        'average': avg,
        'by_tag': by_tag,
        'by_currency': by_currency,
    }


# ---------------------------------------------------------------------------
# Chart data
# ---------------------------------------------------------------------------

def generate_chart_data(expenses, group_by='date'):
    """
    Return a list of {'period': date, 'total': Decimal} dicts.
    Unknown group_by returns [].
    """
    trunc_map = {
        'date':  TruncDate,
        'week':  TruncWeek,
        'month': TruncMonth,
    }
    trunc_fn = trunc_map.get(group_by)
    if trunc_fn is None:
        logger.warning("generate_chart_data: unknown group_by=%r — returning []", group_by)
        return []

    data = (
        expenses
        .annotate(period=trunc_fn('date'))
        .values('period')
        .annotate(total=Sum('amount_base'))
        .order_by('period')
    )
    return list(data)


# ---------------------------------------------------------------------------
# Approval validation helper
# ---------------------------------------------------------------------------

def validate_expense_approval(expense, approver) -> dict:
    """
    Run pre-approval checks and return {'errors': [...], 'warnings': [...]}.

    Used by quick_approve_api and the approval dashboard before calling
    ExpenseApproval.record().
    """
    errors = []
    warnings = []

    if not approver.has_perm('expenses.approve_expense'):
        errors.append('You do not have permission to approve expenses.')

    if expense.user_id == approver.pk:
        errors.append('You cannot approve your own expense.')

    if expense.status not in ('submitted', 'under_review'):
        errors.append(
            f'Expense is in "{expense.get_status_display()}" state and cannot be approved.'
        )

    if not expense.amount or expense.amount <= 0:
        warnings.append('Expense amount appears to be zero or negative.')

    return {'errors': errors, 'warnings': warnings}


# ---------------------------------------------------------------------------
# Chart image helpers (matplotlib — import deferred to avoid side-effects)
# ---------------------------------------------------------------------------

def create_pie_chart_image(tag_summary: dict):
    """Return a BytesIO PNG of a pie chart, or None if tag_summary is empty."""
    if not tag_summary:
        return None
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(8, 6))
        labels = list(tag_summary.keys())[:10]
        sizes = [float(tag_summary[label]) for label in labels]
        ax.pie(sizes, labels=labels, autopct='%1.1f%%', startangle=90)
        ax.axis('equal')

        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        plt.close()
        return buf
    except ImportError:
        logger.warning("create_pie_chart_image: matplotlib not installed")
        return None


def create_bar_chart_image(chart_data: list):
    """Return a BytesIO PNG of a bar chart, or None if chart_data is empty."""
    if not chart_data:
        return None
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt

        fig, ax = plt.subplots(figsize=(10, 6))
        dates = [item['period'].strftime('%Y-%m-%d') for item in chart_data]
        amounts = [float(item['total']) for item in chart_data]
        ax.bar(dates, amounts, color='#3498db')
        ax.set_xlabel('Date')
        ax.set_ylabel('Amount (base currency)')
        ax.set_title('Expenses Over Time')
        plt.xticks(rotation=45, ha='right')
        plt.tight_layout()

        buf = io.BytesIO()
        plt.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        plt.close()
        return buf
    except ImportError:
        logger.warning("create_bar_chart_image: matplotlib not installed")
        return None


# ---------------------------------------------------------------------------
# PDF export (ReportLab)
# ---------------------------------------------------------------------------

def export_to_pdf(expenses, summary, filters=None):
    """
    Build a PDF expense report using ReportLab.
    Returns a BytesIO buffer ready to stream.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
        )
    except ImportError:
        raise ImportError(
            "reportlab is required for PDF export. Install it with: pip install reportlab"
        )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=24,
    )
    elements.append(Paragraph('Expense Report', title_style))
    elements.append(Spacer(1, 0.15 * inch))

    # Summary table
    summary_data = [
        ['Total (base currency):', f"{summary['total']:,.2f}"],
        ['Number of Expenses:',    str(summary['count'])],
        ['Average (base):',        f"{summary['average']:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[3 * inch, 2 * inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME',   (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0, 0), (-1, -1), 11),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('GRID',       (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.25 * inch))

    # Pie chart
    if summary.get('by_tag'):
        elements.append(Paragraph('Expenses by Tag', styles['Heading2']))
        pie_image = create_pie_chart_image(summary['by_tag'])
        if pie_image:
            elements.append(Image(pie_image, width=5 * inch, height=3.75 * inch))
            elements.append(Spacer(1, 0.25 * inch))

    # Bar chart
    chart_data = generate_chart_data(expenses, group_by='date')
    if chart_data:
        elements.append(Paragraph('Expenses Over Time', styles['Heading2']))
        bar_image = create_bar_chart_image(chart_data)
        if bar_image:
            elements.append(Image(bar_image, width=6 * inch, height=3.6 * inch))
            elements.append(Spacer(1, 0.25 * inch))

    # Detailed table — now includes currency, vendor, status
    elements.append(Paragraph('Detailed Expenses', styles['Heading2']))
    elements.append(Spacer(1, 0.1 * inch))

    expense_data = [['Date', 'Description', 'Vendor', 'Currency', 'Amount', 'Status']]
    for expense in expenses.prefetch_related('tags')[:100]:
        expense_data.append([
            expense.date.strftime('%Y-%m-%d'),
            expense.description[:35],
            (expense.vendor or '')[:25],
            expense.currency,
            f"{expense.amount:,.2f}",
            expense.get_status_display(),
        ])

    col_widths = [0.9 * inch, 2.2 * inch, 1.6 * inch, 0.7 * inch, 0.9 * inch, 0.9 * inch]
    expense_table = Table(expense_data, colWidths=col_widths)
    expense_table.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1,  0), colors.HexColor('#3498db')),
        ('TEXTCOLOR',     (0, 0), (-1,  0), colors.whitesmoke),
        ('FONTNAME',      (0, 0), (-1,  0), 'Helvetica-Bold'),
        ('FONTSIZE',      (0, 0), (-1,  0), 9),
        ('BACKGROUND',    (0, 1), (-1, -1), colors.beige),
        ('FONTSIZE',      (0, 1), (-1, -1), 8),
        ('GRID',          (0, 0), (-1, -1), 0.4, colors.grey),
        ('ALIGN',         (4, 0), (4, -1), 'RIGHT'),
    ]))
    elements.append(expense_table)

    doc.build(elements)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------------------------
# Excel export (openpyxl)
# ---------------------------------------------------------------------------

def export_to_excel(expenses, summary, filters=None):
    """
    Build an Excel workbook with Summary, Detailed, and Trend Analysis sheets.
    Returns a BytesIO buffer.
    """
    try:
        import openpyxl
        from openpyxl.chart import BarChart, PieChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. Install it with: pip install openpyxl"
        )

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------
    # Sheet 1: Summary
    # ------------------------------------------------------------------
    ws = wb.active
    ws.title = 'Summary'

    ws['A1'] = 'Expense Report Summary'
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:C1')

    rows = [
        ('Total (base currency):', float(summary['total'])),
        ('Number of Expenses:',    summary['count']),
        ('Average (base):',        float(summary['average'])),
    ]
    for i, (label, value) in enumerate(rows, start=3):
        ws[f'A{i}'] = label
        ws[f'A{i}'].font = Font(bold=True)
        ws[f'B{i}'] = value
        if isinstance(value, float):
            ws[f'B{i}'].number_format = '#,##0.00'

    # By-tag breakdown
    ws['A7'] = 'Expenses by Tag'
    ws['A7'].font = Font(size=13, bold=True)
    ws['A8'] = 'Tag'
    ws['B8'] = 'Amount (base)'
    for cell in ws['8:8']:
        cell.font = Font(bold=True)

    tag_row = 9
    for tag, amount in summary.get('by_tag', {}).items():
        ws[f'A{tag_row}'] = tag
        ws[f'B{tag_row}'] = float(amount)
        ws[f'B{tag_row}'].number_format = '#,##0.00'
        tag_row += 1

    if summary.get('by_tag') and tag_row > 9:
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=9, max_row=tag_row - 1)
        data = Reference(ws, min_col=2, min_row=8, max_row=tag_row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = 'Expenses by Tag'
        ws.add_chart(pie, 'D3')

    # ------------------------------------------------------------------
    # Sheet 2: Detailed expenses — with vendor, currency, status columns
    # ------------------------------------------------------------------
    ws_detail = wb.create_sheet('Detailed Expenses')
    headers = ['Date', 'Description', 'Vendor', 'Currency', 'Amount', 'Amount (Base)', 'Status', 'Tags', 'Notes']
    ws_detail.append(headers)

    header_fill = PatternFill(start_color='3498db', end_color='3498db', fill_type='solid')
    for cell in ws_detail[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for expense in expenses.prefetch_related('tags'):
        tags = ', '.join(expense.tags.names())
        ws_detail.append([
            expense.date,
            expense.description,
            expense.vendor or '',
            expense.currency,
            float(expense.amount),
            float(expense.amount_base),
            expense.get_status_display(),
            tags,
            expense.notes,
        ])

    for row_idx in range(2, ws_detail.max_row + 1):
        ws_detail[f'E{row_idx}'].number_format = '#,##0.00'
        ws_detail[f'F{row_idx}'].number_format = '#,##0.00'

    col_widths = {'A': 12, 'B': 32, 'C': 22, 'D': 10, 'E': 14, 'F': 14, 'G': 16, 'H': 22, 'I': 30}
    for col, width in col_widths.items():
        ws_detail.column_dimensions[col].width = width

    # ------------------------------------------------------------------
    # Sheet 3: Trend analysis
    # ------------------------------------------------------------------
    chart_data = generate_chart_data(expenses, group_by='month')
    if chart_data:
        ws_chart = wb.create_sheet('Trend Analysis')
        ws_chart['A1'] = 'Period'
        ws_chart['B1'] = 'Total (base)'
        ws_chart['A1'].font = Font(bold=True)
        ws_chart['B1'].font = Font(bold=True)

        for idx, item in enumerate(chart_data, start=2):
            ws_chart[f'A{idx}'] = item['period'].strftime('%Y-%m')
            ws_chart[f'B{idx}'] = float(item['total'])
            ws_chart[f'B{idx}'].number_format = '#,##0.00'

        bar = BarChart()
        bar.title = 'Monthly Expense Trend'
        bar.x_axis.title = 'Period'
        bar.y_axis.title = 'Amount (base currency)'
        data_ref = Reference(ws_chart, min_col=2, min_row=1, max_row=len(chart_data) + 1)
        cats_ref = Reference(ws_chart, min_col=1, min_row=2, max_row=len(chart_data) + 1)
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        ws_chart.add_chart(bar, 'D2')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf